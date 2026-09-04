"""Extract partner identity evidence from the 2026 manual reconciliation files.

The extractor is read-only with respect to the manual workbooks and Snowflake.
It scans reconciliation workbooks for data sheets containing an account/partner
name plus an SF ID or CMS ID, then emits row-level evidence and consensus files.
"""

from __future__ import annotations

import argparse
import re
from collections.abc import Iterable
from pathlib import Path

import openpyxl
import pandas as pd


DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026"
)
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parents[1] / "output" / "partner_map_deep_audit_20260903"

ACTIVE_VENDOR_FOLDERS = {
    "Acronis": "Acronis",
    "Auvik CMS": "Auvik",
    "Auvik CW": "Auvik",
    "Bitdefender": "Bitdefender",
    "ESET": "ESET",
    "Exium": "Exium",
    "KeepIT": "KeepIT",
    "Proofpoint": "Proofpoint",
    "SentinelOne": "SentinelOne",
    "Webroot CMS": "Webroot",
    "Webroot CW": "Webroot",
}

MONTH_FOLDER_RE = re.compile(r"^(?P<month>\d{2})_[A-Z]{3}_(?P<year>\d{4})$", re.IGNORECASE)
ACT_ID_RE = re.compile(r"^ACT-\d+$", re.IGNORECASE)
MAX_DATA_ROWS = 100_000
MAX_CONSECUTIVE_BLANK_ROWS = 250

PARTNER_HEADERS = {
    "account name",
    "account",
    "partner name",
    "tenant name",
    "vendor partner name",
}
SF_HEADERS = {
    "sf id",
    "sfid",
    "salesforce id",
    "salesforce account id",
    "cws account unique identifier",
}
CMS_HEADERS = {"cms id", "cmsid", "partner id"}
ZUORA_HEADERS = {"zuora name", "zuora account name"}
PARENT_HEADERS = {"parent co", "parent company", "parent"}
STATUS_HEADERS = {"status", "account status"}


def normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower())).strip()


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    if not text or text.lower() in {"none", "nan", "n/a", "na", "-", "0"}:
        return None
    return text


def clean_sf_id(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.upper()
    return text if ACT_ID_RE.fullmatch(text) else None


def clean_cms_id(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    return text


def billing_month(path: Path) -> str | None:
    for part in path.parts:
        match = MONTH_FOLDER_RE.match(part)
        if match:
            return f"{match.group('year')}-{match.group('month')}-01"
    return None


def header_index(headers: list[str], candidates: set[str]) -> int | None:
    for index, header in enumerate(headers):
        if header in candidates:
            return index
    return None


def detect_header(rows: Iterable[tuple[object, ...]]) -> tuple[int, dict[str, int | None]] | None:
    for row_number, row in enumerate(rows, start=1):
        if row_number > 15:
            return None
        headers = [normalize_header(value) for value in row]
        columns = {
            "partner": header_index(headers, PARTNER_HEADERS),
            "sf_id": header_index(headers, SF_HEADERS),
            "cms_id": header_index(headers, CMS_HEADERS),
            "zuora_name": header_index(headers, ZUORA_HEADERS),
            "parent_company": header_index(headers, PARENT_HEADERS),
            "status": header_index(headers, STATUS_HEADERS),
        }
        if columns["partner"] is not None and (
            columns["sf_id"] is not None or columns["cms_id"] is not None
        ):
            return row_number, columns
    return None


def cell(row: tuple[object, ...], index: int | None) -> object:
    return row[index] if index is not None and index < len(row) else None


def candidate_workbooks(root: Path) -> Iterable[tuple[str, Path]]:
    for folder_name, vendor in ACTIVE_VENDOR_FOLDERS.items():
        vendor_root = root / folder_name
        if not vendor_root.exists():
            continue
        for path in sorted(vendor_root.rglob("*.xlsx")):
            if path.name.startswith("~$") or "invoice" in {part.lower() for part in path.parts}:
                continue
            filename = path.name.lower()
            if not any(token in filename for token in ("recon", "reconciliation")):
                continue
            yield vendor, path
        for path in sorted(vendor_root.rglob("*.xlsm")):
            if path.name.startswith("~$") or "invoice" in {part.lower() for part in path.parts}:
                continue
            filename = path.name.lower()
            if not any(token in filename for token in ("recon", "reconciliation")):
                continue
            yield vendor, path


def extract_workbook(vendor: str, path: Path) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    evidence: list[dict[str, object]] = []
    scans: list[dict[str, object]] = []
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return evidence, [{"vendor": vendor, "source_file": str(path), "status": f"open_error:{type(exc).__name__}"}]

    try:
        for sheet in workbook.worksheets:
            detected = detect_header(sheet.iter_rows(values_only=True))
            if detected is None:
                continue
            header_row, columns = detected
            extracted = 0
            consecutive_blank_rows = 0
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1,
            ):
                partner_name = clean_text(cell(row, columns["partner"]))
                sf_id = clean_sf_id(cell(row, columns["sf_id"]))
                cms_id = clean_cms_id(cell(row, columns["cms_id"]))
                if not partner_name or (not sf_id and not cms_id):
                    consecutive_blank_rows += 1
                    if consecutive_blank_rows >= MAX_CONSECUTIVE_BLANK_ROWS:
                        break
                    if row_number - header_row >= MAX_DATA_ROWS:
                        break
                    continue
                consecutive_blank_rows = 0
                evidence.append(
                    {
                        "vendor": vendor,
                        "billing_month": billing_month(path),
                        "partner_name": partner_name,
                        "partner_name_normalized": normalize_header(partner_name),
                        "sf_id": sf_id,
                        "cms_id": cms_id,
                        "zuora_name": clean_text(cell(row, columns["zuora_name"])),
                        "parent_company": clean_text(cell(row, columns["parent_company"])),
                        "status": clean_text(cell(row, columns["status"])),
                        "source_file": str(path),
                        "source_sheet": sheet.title,
                        "source_row": row_number,
                    }
                )
                extracted += 1
            scans.append(
                {
                    "vendor": vendor,
                    "source_file": str(path),
                    "source_sheet": sheet.title,
                    "header_row": header_row,
                    "evidence_rows": extracted,
                    "status": "ok",
                }
            )
    finally:
        workbook.close()
    return evidence, scans


def build_consensus(evidence: pd.DataFrame) -> pd.DataFrame:
    if evidence.empty:
        return pd.DataFrame()
    grouped = (
        evidence.groupby(
            [
                "vendor",
                "partner_name_normalized",
                "sf_id",
                "cms_id",
                "zuora_name",
                "parent_company",
            ],
            dropna=False,
        )
        .agg(
            evidence_rows=("partner_name", "size"),
            first_month=("billing_month", "min"),
            last_month=("billing_month", "max"),
            source_file_count=("source_file", "nunique"),
            sample_partner_name=("partner_name", "first"),
        )
        .reset_index()
    )
    grouped["consensus_rank"] = grouped.groupby(["vendor", "partner_name_normalized"])[
        "evidence_rows"
    ].rank(method="first", ascending=False)
    return grouped.sort_values(
        ["vendor", "partner_name_normalized", "consensus_rank"],
        kind="stable",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    all_evidence: list[dict[str, object]] = []
    all_scans: list[dict[str, object]] = []
    for index, (vendor, path) in enumerate(candidate_workbooks(args.source_root), start=1):
        print(f"[{index:>3}] {vendor}: {path.name}", flush=True)
        evidence, scans = extract_workbook(vendor, path)
        all_evidence.extend(evidence)
        all_scans.extend(scans)
        print(f"      extracted {len(evidence):,} rows ({len(all_evidence):,} total)", flush=True)

    evidence_df = pd.DataFrame(all_evidence)
    scans_df = pd.DataFrame(all_scans)
    consensus_df = build_consensus(evidence_df)
    if not consensus_df.empty:
        conflict_keys = (
            consensus_df.groupby(["vendor", "partner_name_normalized"], dropna=False)
            .agg(sf_ids=("sf_id", "nunique"), cms_ids=("cms_id", "nunique"))
            .reset_index()
        )
        conflict_keys = conflict_keys[(conflict_keys["sf_ids"] > 1) | (conflict_keys["cms_ids"] > 1)]
        conflicts_df = consensus_df.merge(
            conflict_keys[["vendor", "partner_name_normalized"]],
            on=["vendor", "partner_name_normalized"],
            how="inner",
        )
    else:
        conflicts_df = pd.DataFrame()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    evidence_df.to_csv(args.output_dir / "manual_mapping_evidence.csv", index=False)
    consensus_df.to_csv(args.output_dir / "manual_mapping_consensus.csv", index=False)
    conflicts_df.to_csv(args.output_dir / "manual_mapping_conflicts.csv", index=False)
    scans_df.to_csv(args.output_dir / "manual_mapping_scan_manifest.csv", index=False)

    print(f"Evidence rows: {len(evidence_df):,}")
    print(f"Consensus rows: {len(consensus_df):,}")
    print(f"Conflict rows: {len(conflicts_df):,}")
    print(f"Output: {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())