"""Build the normalized KeepIT usage table from invoice-aligned workbooks."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import re
import shutil
import sys
import tempfile
from collections.abc import Iterable
from pathlib import Path

import openpyxl
import pandas as pd
import pdfplumber


KEEPIT_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = KEEPIT_ROOT.parent
WORKSPACE_ROOT = next((p for p in (PROJECT_ROOT, *PROJECT_ROOT.parents) if (p / "TEMPLATES").exists()), PROJECT_ROOT)
OUTPUT_DIR = KEEPIT_ROOT / "outputs"

DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026\KeepIT"
)
DEFAULT_MANUAL_ROOT = DEFAULT_SOURCE_ROOT

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
FQN = f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}"
TARGET_VENDOR = "KeepIT"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)
DATACENTER_TABS = ("dk-co", "au-sy", "us-dc", "ca-tr", "uk-ld")
BLANK_ROW_CUTOFF = 50
SUMMARY_MAX_ROW = 600

TARGET_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
)

DESCRIPTION_SKU_MAP = {
    "Microsoft 365 Mailbox OneDrive Total consumption": "KI-M365-FUL",
    "Microsoft 365 Exchange consumption": "KI-M365-FUL",
    "Entra ID consumption": "KI-AZUR-CSP",
    "Google Workspace consumption": "KI-GOOG-FUL",
    "GW Gmail consumption": "KI-GOOG-FUL",
    "Salesforce consumption": "KI-SFDC-FUL",
    "Dynamics 365 consumption": "KI-D365-FUL",
    "Dynamics 365 Light consumption": "KI-D365-LGT",
}


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.replace("$", "").replace(" ", "").strip()
        if text == "":
            return None
        # European format: 1.234,56 -> 1234.56
        if re.match(r"^\d{1,3}(\.\d{3})*(,\d+)$", text):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
        value = text
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed)


def to_keepit_invoice_number(value: object) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".")
    parsed = pd.to_numeric(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed)


def read_file_bytes(path: Path, *, allow_copy_fallback: bool = True) -> bytes:
    try:
        return path.read_bytes()
    except PermissionError:
        if not allow_copy_fallback:
            raise
        with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
            tmp_path = Path(handle.name)
        try:
            shutil.copy2(path, tmp_path)
            return tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)


def normalize_header(value: object) -> str:
    return str(value if value is not None else "").strip().lower()


def billing_month_from_folder(month_folder: Path) -> dt.date:
    match = MONTH_FOLDER_RE.match(month_folder.name)
    if not match:
        raise ValueError(f"Invalid KeepIT month folder name: {month_folder}")
    return dt.date(int(match.group("yyyy")), int(match.group("mm")), 1)


def discover_month_folders(source_root: Path) -> dict[str, Path]:
    folders: dict[str, Path] = {}
    for child in source_root.iterdir():
        if child.is_dir() and (match := MONTH_FOLDER_RE.match(child.name)):
            folders[f"{match.group('yyyy')}-{match.group('mm')}"] = child
    return dict(sorted(folders.items()))


def strip_description_window(description: object) -> str | None:
    text = clean_text(description)
    if not text:
        return None
    return re.sub(r"\s*\(.*$", "", text).strip()


def derive_vendor_sku(description: object) -> str | None:
    base = strip_description_window(description)
    if not base:
        return None
    return DESCRIPTION_SKU_MAP.get(base)


def is_main_summary_file(path: Path) -> bool:
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and "summary" in name
        and ("connectwise" in name or "main" in name)
        and "promo" not in name
        and "takeout" not in name
        and "zuora" not in name
        and "recon" not in name
    )


def is_promo_summary_file(path: Path) -> bool:
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and ("promo" in name or name.startswith("takeout invoice "))
        and "summary" in name
        and "post year" not in name
        and "post 3rd" not in name
        and "recon" not in name
        and "r-saas" not in name
    )


def is_takeout_summary_file(path: Path) -> bool:
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and "takeout" in name
        and ("post year" in name or "post 3rd" in name)
    )


def is_takeout_invoice_file(path: Path) -> bool:
    return path.suffix.lower() == ".pdf" and not path.name.startswith("~$") and "invoice" in path.name.lower()


def pdf_has_takeout_lines(path: Path) -> bool:
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages[:1])
    except Exception:
        return False
    return " - Takeout " in text and "SKU Description" in text


def locate_usage_files(month_folder: Path) -> list[tuple[str, Path]]:
    files = [path for path in month_folder.iterdir() if path.is_file()]
    main = sorted(path for path in files if is_main_summary_file(path))
    takeout_pdf = sorted(path for path in files if is_takeout_invoice_file(path) and pdf_has_takeout_lines(path))
    promo = [] if takeout_pdf else sorted(path for path in files if is_promo_summary_file(path))
    takeout_xlsx = sorted(path for path in files if is_takeout_summary_file(path))
    takeout = takeout_pdf or takeout_xlsx
    if len(takeout) > 1:
        takeout = sorted(takeout, key=lambda p: ("post" not in p.name.lower(), p.name.lower()))[:1]
    return [("MAIN", path) for path in main] + [("PROMO", path) for path in promo] + [("TAKEOUT", path) for path in takeout]


def make_usage_row(
    *,
    billing_month: dt.date,
    partner: object,
    sku: object,
    quantity: object,
    unit_price: object,
    amount: object,
    modifier: object = None,
    currency: object = "USD",
) -> dict[str, object] | None:
    parsed_amount = to_number(amount)
    parsed_quantity = to_number(quantity)
    sku_text = clean_text(sku)
    partner_text = clean_text(partner)
    if parsed_amount is None or parsed_quantity is None or not sku_text:
        return None
    return {
        "BILLING_MONTH": billing_month,
        "VENDOR": "KeepIT",
        "VENDOR_PARTNER_NAME": partner_text,
        "VENDOR_PRODUCT_SKU": sku_text,
        "MODIFIER": clean_text(modifier),
        "QUANTITY": parsed_quantity,
        "UNIT_PRICE": to_number(unit_price),
        "AMOUNT": parsed_amount,
        "CURRENCY": clean_text(currency) or "USD",
    }


def load_workbook(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(read_file_bytes(path)), read_only=True, data_only=True)


def parse_main(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    workbook = load_workbook(path)
    rows: list[dict[str, object]] = []
    audit: list[dict[str, object]] = []
    try:
        sheet_lookup = {sheet.lower(): sheet for sheet in workbook.sheetnames}
        for tab in DATACENTER_TABS:
            sheet_name = sheet_lookup.get(tab)
            if not sheet_name:
                audit.append({"source_family": "MAIN", "source_file": path.name, "source_sheet": tab, "rows": 0, "reason": "missing_sheet"})
                continue
            ws = workbook[sheet_name]
            headers = [normalize_header(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
            header_index = {header: idx for idx, header in enumerate(headers) if header}
            required = ["companyname2", "fullname2", "description1", "units1", "unit-price1", "price1"]
            missing = [col for col in required if col not in header_index]
            if missing:
                audit.append({"source_family": "MAIN", "source_file": path.name, "source_sheet": sheet_name, "rows": 0, "reason": f"missing_columns:{';'.join(missing)}"})
                continue
            start = len(rows)
            blank_rows = 0
            for values in ws.iter_rows(min_row=2, values_only=True):
                if not any(values):
                    blank_rows += 1
                    if blank_rows >= BLANK_ROW_CUTOFF:
                        break
                    continue
                blank_rows = 0
                description = values[header_index["description1"]]
                sku = derive_vendor_sku(description)
                if not sku:
                    if any(values[idx] for idx in [header_index["description1"], header_index["units1"], header_index["price1"]]):
                        raise RuntimeError(f"Unmapped KeepIT Main description in {path.name}/{sheet_name}: {description!r}")
                    continue
                partner = values[header_index["fullname2"]] or values[header_index["companyname2"]]
                row = make_usage_row(
                    billing_month=billing_month,
                    partner=partner,
                    sku=sku,
                    modifier=None,
                    quantity=values[header_index["units1"]],
                    unit_price=values[header_index["unit-price1"]],
                    amount=values[header_index["price1"]],
                )
                if row:
                    rows.append(row)
            audit.append({"source_family": "MAIN", "source_file": path.name, "source_sheet": sheet_name, "rows": len(rows) - start, "reason": "accepted"})
    finally:
        workbook.close()
    return rows, audit


def summary_sheet(workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in ("Summary", "Summay"):
        if name in workbook.sheetnames:
            return workbook[name]
    raise RuntimeError(f"Workbook has no Summary/Summay sheet: {workbook.sheetnames}")


def parse_promo(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    workbook = load_workbook(path)
    rows: list[dict[str, object]] = []
    try:
        ws = summary_sheet(workbook)
        headers = [clean_text(ws.cell(4, col).value) for col in range(1, min(ws.max_column, 60) + 1)]

        def header_col(pattern: str) -> int | None:
            needle = pattern.lower()
            for idx, header in enumerate(headers):
                if header and needle in header.lower():
                    return idx
            return None

        partner_idx = header_col("takeout partners due")
        min_seats_idx = header_col("min comm")
        overage_seats_idx = next((idx for idx, h in enumerate(headers) if h and h.lower() == "overage"), None)
        min_amount_idx = header_col("amt - min comm")
        overage_amount_idx = header_col("amt - overage")
        retention_amount_idx = header_col("extra charge for unlimited retention")
        retention_seats_idx = header_col("number of seats with unlimited retention")
        required = {
            "partner": partner_idx,
            "min_seats": min_seats_idx,
            "overage_seats": overage_seats_idx,
            "min_amount": min_amount_idx,
            "overage_amount": overage_amount_idx,
            "retention_amount": retention_amount_idx,
            "retention_seats": retention_seats_idx,
        }
        missing = [name for name, idx in required.items() if idx is None]
        if missing:
            raise RuntimeError(f"Promo Summary schema mismatch in {path.name}: missing {missing}")

        blank_rows = 0
        max_col = max(idx for idx in required.values() if idx is not None) + 1
        for values in ws.iter_rows(min_row=5, max_row=min(ws.max_row, SUMMARY_MAX_ROW), min_col=1, max_col=max_col, values_only=True):
            partner = clean_text(values[partner_idx])
            if not partner:
                if not any(values):
                    blank_rows += 1
                    if blank_rows >= BLANK_ROW_CUTOFF:
                        break
                else:
                    blank_rows = 0
                continue
            blank_rows = 0
            if partner.lower().startswith(("total", "subtotal")):
                continue
            components = (
                ("KI-M365-FUL", None, values[min_seats_idx], 0.75, values[min_amount_idx]),
                ("KI-M365-FUL", None, values[overage_seats_idx], 0.75, values[overage_amount_idx]),
                ("KI-M365-FUL", None, values[retention_seats_idx], 0.25, values[retention_amount_idx]),
            )
            for sku, modifier, quantity, unit_price, amount in components:
                row = make_usage_row(
                    billing_month=billing_month,
                    partner=partner,
                    sku=sku,
                    modifier=modifier,
                    quantity=quantity,
                    unit_price=unit_price,
                    amount=amount,
                )
                if row and abs(float(row["AMOUNT"])) > 0:
                    rows.append(row)
    finally:
        workbook.close()
    return rows, [{"source_family": "PROMO", "source_file": path.name, "source_sheet": "Summary", "rows": len(rows), "reason": "accepted"}]


TAKEOUT_LINE_RE = re.compile(
    r"^(?P<sku>KI-[A-Z0-9-]+)\s+For\s+.+?\s+-\s+Takeout\s+"
    r"(?P<quantity>[0-9.]+)\s+"
    r"(?P<unit_price>[0-9]+,[0-9]+)\s+"
    r"(?P<net_amount>[0-9.]+,[0-9]{2})\s+"
    r"(?P<total_amount>[0-9.]+,[0-9]{2})$"
)


def parse_takeout_pdf(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rows: list[dict[str, object]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                match = TAKEOUT_LINE_RE.match(line.strip())
                if not match:
                    continue
                row = make_usage_row(
                    billing_month=billing_month,
                    partner="ConnectWise (Continuum) - Consolidated",
                    sku=match.group("sku"),
                    modifier="TAKEOUT",
                    quantity=to_keepit_invoice_number(match.group("quantity")),
                    unit_price=to_keepit_invoice_number(match.group("unit_price")),
                    amount=to_keepit_invoice_number(match.group("total_amount")),
                )
                if row:
                    rows.append(row)
    if not rows:
        raise RuntimeError(f"No takeout invoice lines parsed from {path}")
    return rows, [{"source_family": "TAKEOUT", "source_file": path.name, "source_sheet": "PDF", "rows": len(rows), "reason": "accepted"}]


def parse_takeout_workbook(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    workbook = load_workbook(path)
    rows: list[dict[str, object]] = []
    try:
        ws = summary_sheet(workbook)
        blank_rows = 0
        for values in ws.iter_rows(min_row=5, max_row=min(ws.max_row, SUMMARY_MAX_ROW), min_col=9, max_col=15, values_only=True):
            partner = clean_text(values[0])
            quantity = to_number(values[5])
            amount = to_number(values[6])
            if not partner and quantity is None and amount is None:
                blank_rows += 1
                if blank_rows >= BLANK_ROW_CUTOFF:
                    break
                continue
            blank_rows = 0
            if not partner or quantity is None or amount is None:
                continue
            if partner.lower().startswith(("total", "subtotal")):
                continue
            row = make_usage_row(
                billing_month=billing_month,
                partner=partner,
                sku="KEEPIT_POST_PROMO_USAGE",
                modifier="TAKEOUT",
                quantity=quantity,
                unit_price=None,
                amount=amount,
            )
            if row and abs(float(row["AMOUNT"])) > 0:
                rows.append(row)
    finally:
        workbook.close()
    return rows, [{"source_family": "TAKEOUT", "source_file": path.name, "source_sheet": "Summary", "rows": len(rows), "reason": "accepted"}]


def parse_takeout(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    if path.suffix.lower() == ".pdf":
        return parse_takeout_pdf(path, billing_month)
    return parse_takeout_workbook(path, billing_month)


def load_month(month_folder: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    billing_month = billing_month_from_folder(month_folder)
    records: list[dict[str, object]] = []
    audits: list[dict[str, object]] = []
    usage_files = locate_usage_files(month_folder)
    if not usage_files:
        raise RuntimeError(f"No KeepIT usage workbook found in {month_folder}")
    for source_family, path in usage_files:
        if source_family == "MAIN":
            parsed, audit = parse_main(path, billing_month)
        elif source_family == "PROMO":
            parsed, audit = parse_promo(path, billing_month)
        elif source_family == "TAKEOUT":
            parsed, audit = parse_takeout(path, billing_month)
        else:
            continue
        records.extend(parsed)
        audits.extend({"billing_month": billing_month.isoformat(), **row} for row in audit)
        frame = pd.DataFrame(parsed, columns=TARGET_COLUMNS)
        print(
            f"[{billing_month}] {source_family} {path.name}: rows={len(frame):,}, "
            f"quantity={frame['QUANTITY'].sum(skipna=True) if not frame.empty else 0:,.2f}, "
            f"amount={frame['AMOUNT'].sum(skipna=True) if not frame.empty else 0:,.2f}"
        )
    df = pd.DataFrame(records, columns=TARGET_COLUMNS)
    if df.empty:
        return df, pd.DataFrame(audits)
    grouped = (
        df.groupby(["BILLING_MONTH", "VENDOR", "VENDOR_PARTNER_NAME", "VENDOR_PRODUCT_SKU", "MODIFIER", "UNIT_PRICE", "CURRENCY"], dropna=False)
        .agg(QUANTITY=("QUANTITY", "sum"), AMOUNT=("AMOUNT", "sum"))
        .reset_index()
    )
    return grouped[list(TARGET_COLUMNS)], pd.DataFrame(audits)


def load_all(source_root: Path, months: Iterable[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    month_folders = discover_month_folders(source_root)
    if months:
        requested = {month.removesuffix("-01") for month in months}
        month_folders = {month: path for month, path in month_folders.items() if month in requested}
    if not month_folders:
        raise RuntimeError(f"No month folders selected under {source_root}")
    frames: list[pd.DataFrame] = []
    audits: list[pd.DataFrame] = []
    for _, folder in month_folders.items():
        frame, audit = load_month(folder)
        frames.append(frame)
        audits.append(audit)
    return pd.concat(frames, ignore_index=True), pd.concat(audits, ignore_index=True)


def snowflake_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {FQN} (
    BILLING_MONTH DATE,
    VENDOR VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU VARCHAR,
    MODIFIER VARCHAR,
    QUANTITY NUMBER(38, 6),
    UNIT_PRICE NUMBER(18, 6),
    AMOUNT NUMBER(38, 6),
    CURRENCY VARCHAR
);
"""


def load_snowflake(df: pd.DataFrame, *, reset: bool = False) -> None:
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from snowflake.connector.pandas_tools import write_pandas
    from TEMPLATES.Python.connection import get_snowflake_connection

    load_df = df.loc[:, list(TARGET_COLUMNS)].copy()
    load_df["BILLING_MONTH"] = pd.to_datetime(load_df["BILLING_MONTH"]).dt.date
    conn = get_snowflake_connection(role="DEVELOPER", warehouse="REPORTING_WH", database=TARGET_DATABASE, schema=TARGET_SCHEMA)
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        if reset:
            print(f"RESET: dropping and recreating {FQN}.")
            cur.execute(
                f"DELETE FROM {FQN} WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )
        cur.execute(snowflake_ddl())
        incoming_months = sorted(load_df["BILLING_MONTH"].astype(str).unique().tolist())
        if not reset:
            month_list = ", ".join(f"'{month}'::DATE" for month in incoming_months)
            cur.execute(
                f"DELETE FROM {FQN} "
                f"WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s) "
                f"AND BILLING_MONTH IN ({month_list})",
                (TARGET_VENDOR,),
            )
        load_df = _fill_missing_prices(load_df, TARGET_VENDOR, conn=conn)
        success, chunks, rows, output = write_pandas(
            conn,
            load_df,
            TARGET_TABLE,
            database=TARGET_DATABASE,
            schema=TARGET_SCHEMA,
            quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"write_pandas failed: {output}")
        conn.commit()
        print(f"Loaded {rows:,} rows for {incoming_months} into {FQN} in {chunks} chunk(s).")
    finally:
        conn.close()


def write_ingest_audits(df: pd.DataFrame, scan_df: pd.DataFrame, label: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    scan_path = OUTPUT_DIR / f"keepit_source_scan_{label}.csv"
    scan_df.to_csv(scan_path, index=False, quoting=csv.QUOTE_MINIMAL)
    monthly = (
        df.groupby(["BILLING_MONTH", "VENDOR_PRODUCT_SKU", "MODIFIER", "UNIT_PRICE"], dropna=False)
        .agg(row_count=("VENDOR", "size"), quantity=("QUANTITY", "sum"), amount=("AMOUNT", "sum"))
        .reset_index()
    )
    monthly_path = OUTPUT_DIR / f"keepit_usage_ingest_audit_{label}.csv"
    monthly.to_csv(monthly_path, index=False, quoting=csv.QUOTE_MINIMAL)
    print(f"Wrote source scan: {scan_path}")
    print(f"Wrote usage audit: {monthly_path}")


def validate_usage(df: pd.DataFrame) -> None:
    if df["VENDOR_PRODUCT_SKU"].isna().any():
        raise RuntimeError("KEEPIT_USAGE has null VENDOR_PRODUCT_SKU rows.")
    if df["AMOUNT"].isna().any() or df["QUANTITY"].isna().any():
        raise RuntimeError("KEEPIT_USAGE has null quantity or amount rows.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest invoice-aligned KeepIT usage into Snowflake.")
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--manual-root", type=Path, default=DEFAULT_MANUAL_ROOT)
    parser.add_argument("--month", action="append", help="Month to load as YYYY-MM or YYYY-MM-01. Repeatable.")
    parser.add_argument("--all-months", action="store_true", help="Load every month folder under the source root.")
    parser.add_argument("--reset", action="store_true", help="Drop and recreate KEEPIT_USAGE before loading.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and audit locally without loading Snowflake.")
    parser.add_argument("--validate-manual", action="store_true", help="Retained for CLI compatibility; no extra action.")
    parser.add_argument("--label", default=dt.datetime.now().strftime("%Y%m%d_%H%M%S"))
    return parser.parse_args()



# ---------------------------------------------------------------------------
# Dynamic invoice rate fill (universal safety net)
# ---------------------------------------------------------------------------
def _fill_missing_prices(df, vendor_name, conn=None):
    from invoice_rate_backfill import fill_missing_prices_dynamic

    return fill_missing_prices_dynamic(df=df, vendor_name=vendor_name, conn=conn)

def main() -> None:
    args = parse_args()
    if not args.all_months and not args.month:
        args.all_months = True
    months = None if args.all_months else args.month
    df, scan_df = load_all(args.source_root, months)
    validate_usage(df)
    write_ingest_audits(df, scan_df, args.label)
    if args.dry_run:
        print(f"DRY RUN: parsed {len(df):,} grouped usage rows. Snowflake load skipped.")
        return
    load_snowflake(df, reset=args.reset)


if __name__ == "__main__":
    main()

