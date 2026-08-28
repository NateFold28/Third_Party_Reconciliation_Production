"""Ingest Auvik Connectwise Usage Report workbooks into AUVIK_USAGE.

One ingestion architecture for both Auvik CW and Auvik CMS:

  * Auvik delivers one workbook per month whose name varies across months
    (e.g. "Connectwise Usage Report July.xlsx", "ConnectWise Usage Report -
    June Invoices.xlsx"). The identical workbook is placed in both the
    Auvik CW and Auvik CMS subfolders, differing only in trivial metadata.

  * ONLY the Auvik CW copy is read. Ingesting both copies would double-count
    the entire month. This is the single most important rule in this module.

  * The CW/CMS entity split is taken from the Partner Name column:
      'ConnectWise Inc'  â†’ STREAM = 'CW'
      'ConnectWise, LLC' â†’ STREAM = 'CMS'
    Every row belongs to exactly one entity.

  * BILLING_MONTH is derived from Start Date in the data (the true usage
    period start). The folder label and the filename are both unreliable â€”
    they name different months across 2026 and must not be trusted.

  * Product Tenant ID is read as a text string before any numeric coercion
    to preserve 19-digit precision. Excel stores it as a float and loses
    the last ~4 digits if coerced (e.g. 1139453724789191421 â†’ 1.13945e+18).

  * Load-time assertions (any failure raises RuntimeError):
      1. Within a single file, Start Date / End Date hold exactly one
         distinct pair (single usage period per workbook).
      2. Invoice Date = first day of the month after End Date.
      3. Subtotal + Tax = Total Charge Amount on every data row.
      4. Entity totals from the ingested rows match the Summary tab to
         the cent (hard gate).

  * Content-hash deduplication: if the same file is placed in both folders
    and accidentally scanned, the second copy is silently skipped.

Source layout:
    <SOURCE_ROOT>/Auvik CW/MM_MON_YYYY/<usage workbook>.xlsx
    Workbook: Summary sheet (skipped) + one INV-* sheet per invoice.

Target tables:
    AUVIK_USAGE            â€“ one row per invoice line
    AUVIK_USAGE_FILE_AUDIT â€“ one row per source file with all control totals
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import io
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Paths and Snowflake targets
# ---------------------------------------------------------------------------

AUVIK_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = AUVIK_ROOT.parent
WORKSPACE_ROOT = next((p for p in (PROJECT_ROOT, *PROJECT_ROOT.parents) if (p / "TEMPLATES").exists()), PROJECT_ROOT)

DEFAULT_SOURCE_ROOT_CW = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Auvik CW"
)

OUTPUT_DIR = AUVIK_ROOT / "outputs"

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
AUDIT_TABLE = "AUVIK_USAGE_FILE_AUDIT"
TARGET_VENDOR = "Auvik"

# Folder naming: MM_MON_YYYY
MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)

# ---------------------------------------------------------------------------
# File-name noise exclusions.
# Any .xlsx whose lowercased name contains one of these tokens is NOT a
# Connectwise Usage Report workbook and must be skipped.
# ---------------------------------------------------------------------------
_EXCLUDE_NAME_TOKENS: tuple[str, ...] = (
    "overage",
    "zuora",
    "zoura",        # common Auvik filename typo for 'zuora'
    "template",
    "reconcil",     # "reconciliation", "recon"
    "creditnote",
    "credit note",
)

# ---------------------------------------------------------------------------
# Source column spec (18 columns, 0-indexed).
# Column 0 is an unnamed row counter and is skipped; columns 1â€“17 are the
# named data columns confirmed across all eight June invoice tabs.
# ---------------------------------------------------------------------------
# Canonical header names as they appear in the workbook (cols 1-17).
SOURCE_COLUMNS: tuple[str, ...] = (
    "Invoice Date",
    "Invoice Name",
    "Partner Name",
    "Shipping_Account",
    "Domain Prefix",
    "Product",
    "UnitPrice",
    "Charge Type",
    "Primary Tenant ID",
    "Product Tenant ID",   # MUST remain TEXT â€“ 19-digit float precision risk
    "Start Date",
    "End Date",
    "Quantity",
    "Overage Quantity",
    "Subtotal",
    "Tax",
    "Total Charge Amount",
)

# ---------------------------------------------------------------------------
# Target table columns (source + derived).
# ---------------------------------------------------------------------------
USAGE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
)

# ---------------------------------------------------------------------------
# Audit table columns.
# ---------------------------------------------------------------------------
AUDIT_COLUMNS: tuple[str, ...] = (
    "SOURCE_FOLDER",
    "SOURCE_FILE",
    "SOURCE_CONTENT_HASH",
    "LOAD_STATUS",
    "BILLING_MONTH",
    "INVOICE_MONTH",
    "PERIOD_START",
    "PERIOD_END",
    "TOTAL_DATA_ROWS",
    "CW_ROWS",
    "CMS_ROWS",
    "CHARGEABLE_ROWS",
    "ZERO_AMOUNT_ROWS",
    "TOTAL_ROWS_EXCLUDED",      # summary/total rows dropped
    "INGESTED_CW_TOTAL",        # sum(Total Charge Amount) for CW rows
    "INGESTED_CMS_TOTAL",       # sum(Total Charge Amount) for CMS rows
    "SUMMARY_CW_TOTAL",         # entity total from the Summary tab
    "SUMMARY_CMS_TOTAL",
    "CW_TOTAL_DELTA",
    "CMS_TOTAL_DELTA",
    "PERIOD_UNIQUE_CHECK",      # PASS / WARN / FAIL
    "INVOICE_DATE_CHECK",       # PASS / WARN
    "SUBTOTAL_TAX_CHECK",       # PASS / WARN
    "ENTITY_TOTAL_CHECK",       # PASS / FAIL
    "ERROR_MESSAGE",
    "INGESTED_AT",
)

# ---------------------------------------------------------------------------
# Partner Name â†’ STREAM mapping
# ---------------------------------------------------------------------------
STREAM_MAP: dict[str, str] = {
    "connectwise inc": "CW",
    "connectwise, llc": "CMS",
}


# ---------------------------------------------------------------------------
# Product family classification (spec Â§"The derived columns")
# ---------------------------------------------------------------------------
def classify_product_family(product: str | None) -> str:
    """Return 'Performance', 'ASM', or 'Billable' for a raw product string.

    Rules (case-insensitive, evaluated in priority order):
      1. Performance  â€” contains 'performance'
      2. ASM          â€” contains 'asm' or 'saas management'
      3. Billable     â€” everything else (Essentials, Billable Devices, etc.)

    An unrecognised product that doesn't match any of the three rules
    defaults to Billable and a WARNING is emitted so a genuinely new product
    line is never silently absorbed.
    """
    if not product:
        return "Billable"
    lower = product.lower()
    if "performance" in lower:
        return "Performance"
    if "asm" in lower or "saas management" in lower:
        return "ASM"
    return "Billable"


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SourceFile:
    path: Path
    folder_month: str | None  # YYYY-MM from folder name, may be None


def _read_file_bytes(path: Path) -> bytes:
    """Read bytes, falling back to a PowerShell Copy-Item for OneDrive locks."""
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir) / path.name
            escaped_src = str(path).replace("'", "''")
            escaped_dest = str(tmp).replace("'", "''")
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    f"Copy-Item -LiteralPath '{escaped_src}' -Destination '{escaped_dest}' -Force",
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            return tmp.read_bytes()


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).replace("\xa0", " ").strip()
    return text if text else None


def _to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("$", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_date(value: object) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    ts = pd.to_datetime(value, errors="coerce")
    return None if pd.isna(ts) else ts.date()


def _first_day(d: dt.date | None) -> dt.date | None:
    return dt.date(d.year, d.month, 1) if d else None


def _vendor_product_sku(product: object) -> str | None:
    text = _clean_text(product)
    if text is None:
        return None
    return re.sub(r"^\s*Overage\s*-\s*", "", text, flags=re.IGNORECASE).strip() or None


def _billable_quantity(
    charge_type: object,
    quantity: object,
    overage_quantity: object,
    total_charge_amount: object,
) -> float | None:
    # Auvik usage rows can include negative overage quantities that net to a
    # zero-dollar line; these should contribute 0 billable units downstream.
    total = _to_number(total_charge_amount)
    if total is not None and abs(total) < 1e-9:
        return 0.0

    charge = (_clean_text(charge_type) or "").lower()
    if charge == "usage":
        return _to_number(overage_quantity)
    return _to_number(quantity)


def _month_from_folder(path: Path) -> str | None:
    match = MONTH_FOLDER_RE.match(path.parent.name)
    if not match:
        return None
    return f"{match.group('yyyy')}-{match.group('mm')}"


def _product_tenant_id_text(raw: object) -> str | None:
    """Read Product Tenant ID as a lossless text string.

    Excel coerces the 19-digit integer to a float, destroying the last ~4
    digits. openpyxl exposes it as a Python float (e.g. 1.95544e+17).
    Converting to int before str() restores the full integer form.
    """
    if raw is None:
        return None
    if isinstance(raw, float):
        # Convert via int to avoid scientific notation and trailing .0
        return str(int(raw))
    if isinstance(raw, int):
        return str(raw)
    text = str(raw).strip()
    # If it looks like scientific notation, normalise it
    try:
        return str(int(float(text)))
    except (ValueError, OverflowError):
        return text or None


def _content_hash_rows(all_rows: list[tuple[object, ...]]) -> str:
    """SHA-256 of tab-delimited text representation of every data row."""
    parts: list[str] = []
    for row in all_rows:
        parts.append("\t".join(_clean_text(v) or "" for v in row))
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def _is_usage_report(path: Path) -> bool:
    """Return True if this .xlsx is likely a Connectwise Usage Report workbook.

    Strategy: include any .xlsx that is not a known noise file.
    Noise exclusions: overage template, zuora export, recon workbook, credit
    note, invoice PDF (those are .pdf anyway but guard defensively).
    """
    name_lower = path.name.lower()
    if path.name.startswith("~$"):
        return False
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    for token in _EXCLUDE_NAME_TOKENS:
        if token in name_lower:
            return False
    # Must contain at least one of these keywords to qualify
    # (guards against random reference docs that might be added later)
    if not any(kw in name_lower for kw in ("usage", "report", "connectwise")):
        return False
    return True


def discover_source_files(source_root_cw: Path, month_filter: str | None = None) -> list[SourceFile]:
    """Return all usage-report workbooks under the CW folder.

    Only the CW folder is scanned. CMS is skipped entirely â€” the entity
    split is done by Partner Name in the data, and both folders carry
    the identical file.
    """
    files: list[SourceFile] = []
    if not source_root_cw.exists():
        print(f"WARNING: CW source root does not exist: {source_root_cw}", flush=True)
        return files
    for month_dir in sorted(source_root_cw.iterdir()):
        if not month_dir.is_dir():
            continue
        m = MONTH_FOLDER_RE.match(month_dir.name)
        if not m:
            continue
        folder_month = f"{m.group('yyyy')}-{m.group('mm')}"
        if month_filter and folder_month != month_filter:
            continue
        for path in sorted(month_dir.glob("*.xlsx")):
            if _is_usage_report(path):
                files.append(SourceFile(path=path, folder_month=folder_month))
    return files


# ---------------------------------------------------------------------------
# Summary-tab parser (control totals)
# ---------------------------------------------------------------------------

def _parse_summary_totals(wb: openpyxl.Workbook) -> dict[str, float]:
    """Read entity totals from the Summary sheet.

    The June 2026 Summary sheet layout (confirmed):
        Col 0: None
        Col 1: entity name ('ConnectWise Inc' / 'ConnectWise, LLC' / 'Total')
        Col 2: total amount

    Entity name may be in col 0 or col 1 depending on how Auvik formats the
    sheet — this function checks both positions for robustness.
    Returns dict keyed by 'ConnectWise Inc' and 'ConnectWise, LLC'.
    """
    totals: dict[str, float] = {}
    if "Summary" not in wb.sheetnames:
        return totals
    ws = wb["Summary"]
    _entity_keys = {"connectwise inc", "connectwise, llc"}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Find entity name: check cols 0 and 1 (layout has varied)
        entity: str | None = None
        for col_idx in (0, 1):
            if col_idx < len(row):
                candidate = _clean_text(row[col_idx])
                if candidate and candidate.lower() in _entity_keys:
                    entity = candidate
                    break
        if entity is None:
            continue
        # Last non-None numeric cell in the row is the total amount
        amount = None
        for cell in reversed(row):
            v = _to_number(cell)
            if v is not None:
                amount = v
                break
        if amount is not None:
            totals[entity] = amount
    return totals


# ---------------------------------------------------------------------------
# Workbook parser
# ---------------------------------------------------------------------------

def _validate_header(header_raw: tuple[object, ...]) -> bool:
    """Return True if the header row matches the expected 18-column layout.

    Col 0 may be blank or a number. Cols 1-17 must match SOURCE_COLUMNS.
    """
    if len(header_raw) < len(SOURCE_COLUMNS) + 1:
        return False
    actual = tuple(_clean_text(header_raw[i + 1]) for i in range(len(SOURCE_COLUMNS)))
    return actual == SOURCE_COLUMNS


def parse_usage_workbook(
    source: SourceFile,
    *,
    ingested_at: dt.datetime,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse a single Connectwise Usage Report workbook.

    Returns (usage_df, audit_df).  Raises RuntimeError on hard-gate failures
    (entity total mismatch).  Soft failures (period uniqueness, invoice date
    offset, subtotal arithmetic) are recorded in the audit row and printed as
    WARNINGs but do not abort the load.
    """
    file_bytes = _read_file_bytes(source.path)
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    data_records: list[dict[str, object]] = []
    total_rows_excluded = 0
    unknown_streams: list[str] = []
    all_content_rows: list[tuple[object, ...]] = []

    try:
        summary_totals = _parse_summary_totals(wb)
        invoice_sheets = [s for s in wb.sheetnames if s.lower() != "summary"]

        for sheet_name in invoice_sheets:
            ws = wb[sheet_name]
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue

            header_raw = sheet_rows[0]
            if not _validate_header(header_raw):
                print(
                    f"  WARNING: unexpected header in sheet '{sheet_name}' "
                    f"of {source.path.name} â€” skipping sheet.",
                    flush=True,
                )
                continue

            for sheet_row_num, raw_row in enumerate(sheet_rows[1:], start=2):
                if not raw_row or all(c is None for c in raw_row):
                    continue

                # Column 0 is the row counter (may be None on total row)
                # Columns 1-17 are the source data columns
                row_data = raw_row[1 : len(SOURCE_COLUMNS) + 1]
                row_dict = dict(zip(SOURCE_COLUMNS, row_data))

                # Total rows: Partner Name is null
                partner_name_raw = row_dict.get("Partner Name")
                partner_name = _clean_text(partner_name_raw)
                if partner_name is None:
                    total_rows_excluded += 1
                    continue

                # Track content rows for hash (skip total rows)
                all_content_rows.append(raw_row)

                data_records.append(
                    {
                        **row_dict,
                        "_sheet_name": sheet_name,
                        "_source_row_number": sheet_row_num,
                    }
                )
    finally:
        wb.close()

    content_hash = _content_hash_rows(all_content_rows) if all_content_rows else file_hash

    if not data_records:
        # No data rows found at all â€” audit and return empty
        audit_df = pd.DataFrame(
            [
                {
                    "SOURCE_FOLDER": source.path.parent.name,
                    "SOURCE_FILE": source.path.name,
                    "SOURCE_CONTENT_HASH": content_hash,
                    "LOAD_STATUS": "LOADED_EMPTY",
                    "BILLING_MONTH": None,
                    "INVOICE_MONTH": None,
                    "PERIOD_START": None,
                    "PERIOD_END": None,
                    "TOTAL_DATA_ROWS": 0,
                    "CW_ROWS": 0,
                    "CMS_ROWS": 0,
                    "CHARGEABLE_ROWS": 0,
                    "ZERO_AMOUNT_ROWS": 0,
                    "TOTAL_ROWS_EXCLUDED": total_rows_excluded,
                    "INGESTED_CW_TOTAL": None,
                    "INGESTED_CMS_TOTAL": None,
                    "SUMMARY_CW_TOTAL": summary_totals.get("ConnectWise Inc"),
                    "SUMMARY_CMS_TOTAL": summary_totals.get("ConnectWise, LLC"),
                    "CW_TOTAL_DELTA": None,
                    "CMS_TOTAL_DELTA": None,
                    "PERIOD_UNIQUE_CHECK": "N/A",
                    "INVOICE_DATE_CHECK": "N/A",
                    "SUBTOTAL_TAX_CHECK": "N/A",
                    "ENTITY_TOTAL_CHECK": "N/A",
                    "ERROR_MESSAGE": "No data rows found",
                    "INGESTED_AT": ingested_at,
                }
            ],
            columns=AUDIT_COLUMNS,
        )
        return pd.DataFrame(columns=USAGE_COLUMNS), audit_df

    raw_df = pd.DataFrame(data_records)

    # ------------------------------------------------------------------
    # Numeric coercion.  Product Tenant ID stays as TEXT (lossless).
    # ------------------------------------------------------------------
    for col in ("UnitPrice", "Quantity", "Overage Quantity", "Subtotal", "Tax", "Total Charge Amount"):
        raw_df[col] = raw_df[col].map(_to_number)

    raw_df["_invoice_date"] = raw_df["Invoice Date"].map(_to_date)
    raw_df["_start_date"] = raw_df["Start Date"].map(_to_date)
    raw_df["_end_date"] = raw_df["End Date"].map(_to_date)
    raw_df["_product_tenant_id_text"] = raw_df["Product Tenant ID"].map(_product_tenant_id_text)

    # BILLING_MONTH from Start Date (the authoritative usage period start)
    raw_df["_billing_month"] = raw_df["_start_date"].map(_first_day)
    raw_df["_invoice_month"] = raw_df["_invoice_date"].map(_first_day)

    # STREAM from Partner Name
    def _resolve_stream(partner: object) -> str:
        p = _clean_text(partner)
        if p:
            key = p.lower()
            if key in STREAM_MAP:
                return STREAM_MAP[key]
            unknown_streams.append(p)
        return "UNKNOWN"

    raw_df["_stream"] = raw_df["Partner Name"].map(_resolve_stream)

    if unknown_streams:
        unique_unknown = sorted(set(unknown_streams))
        print(
            f"  WARNING: {source.path.name} â€” unrecognised Partner Name values "
            f"(mapped to STREAM='UNKNOWN'): {unique_unknown}",
            flush=True,
        )

    # PRODUCT_FAMILY
    raw_df["_product_family"] = raw_df["Product"].map(_clean_text).map(classify_product_family)

    # Flag new product family variants that classify to Billable but look
    # unfamiliar (contain neither 'essentials', 'billable', 'license',
    # 'continuum' nor 'plan') so the recon team can spot genuine new SKUs.
    _known_billable_tokens = {"essentials", "billable", "license", "continuum", "plan", "addon", "add-on"}
    for prod in raw_df.loc[raw_df["_product_family"] == "Billable", "Product"].map(_clean_text).dropna().unique():
        if not any(t in prod.lower() for t in _known_billable_tokens):
            print(
                f"  INFO: {source.path.name} â€” product defaulted to Billable "
                f"(consider adding mapping rule): '{prod}'",
                flush=True,
            )

    # CHARGEABLE_FLAG
    raw_df["_chargeable"] = raw_df["Total Charge Amount"].fillna(0.0).ne(0.0)

    # ------------------------------------------------------------------
    # Load-time assertions
    # ------------------------------------------------------------------
    errors: list[str] = []
    warnings: list[str] = []

    # 1. Period uniqueness: one distinct (Start Date, End Date) pair per file
    period_pairs = raw_df[["_start_date", "_end_date"]].dropna().drop_duplicates()
    if len(period_pairs) == 0:
        period_check = "FAIL"
        errors.append("No valid Start Date / End Date pairs found.")
    elif len(period_pairs) == 1:
        period_check = "PASS"
    else:
        period_check = "WARN"
        warnings.append(
            f"Multiple period pairs found ({len(period_pairs)}): "
            + "; ".join(
                f"{row[0]}-{row[1]}" for row in period_pairs.itertuples(index=False, name=None)
            )
        )

    # 2. Invoice Date = first day of month after End Date
    invoice_date_check = "PASS"
    period_end_dates = raw_df["_end_date"].dropna().unique()
    invoice_dates = raw_df["_invoice_date"].dropna().unique()
    if len(period_end_dates) == 1 and len(invoice_dates) == 1:
        end_date = period_end_dates[0]
        inv_date = invoice_dates[0]
        # Expected invoice date: first day of month after end_date's month
        if end_date.month == 12:
            expected_inv_month = dt.date(end_date.year + 1, 1, 1)
        else:
            expected_inv_month = dt.date(end_date.year, end_date.month + 1, 1)
        if inv_date != expected_inv_month:
            invoice_date_check = "WARN"
            warnings.append(
                f"Invoice Date {inv_date} is not the first day of the month "
                f"after End Date {end_date} (expected {expected_inv_month})."
            )
    else:
        invoice_date_check = "WARN"
        warnings.append("Could not validate Invoice Date offset: multiple end/invoice dates.")

    # 3. Subtotal + Tax == Total Charge Amount (tolerance $0.01)
    subtotal_ok = (
        (raw_df["Subtotal"].fillna(0.0) + raw_df["Tax"].fillna(0.0) - raw_df["Total Charge Amount"].fillna(0.0))
        .abs()
        .le(0.011)
        .all()
    )
    subtotal_check = "PASS" if subtotal_ok else "WARN"
    if not subtotal_ok:
        n_bad = int(
            (
                (raw_df["Subtotal"].fillna(0.0) + raw_df["Tax"].fillna(0.0) - raw_df["Total Charge Amount"].fillna(0.0))
                .abs()
                .gt(0.011)
            ).sum()
        )
        warnings.append(f"Subtotal + Tax â‰  Total Charge Amount on {n_bad} row(s) (tolerance $0.01).")

    # 4. Entity totals vs Summary tab (hard gate â€” $0.01 tolerance)
    cw_ingested = float(
        raw_df.loc[raw_df["_stream"] == "CW", "Total Charge Amount"].fillna(0.0).sum()
    )
    cms_ingested = float(
        raw_df.loc[raw_df["_stream"] == "CMS", "Total Charge Amount"].fillna(0.0).sum()
    )
    summary_cw = summary_totals.get("ConnectWise Inc")
    summary_cms = summary_totals.get("ConnectWise, LLC")

    cw_delta = abs(cw_ingested - (summary_cw or 0.0))
    cms_delta = abs(cms_ingested - (summary_cms or 0.0))
    entity_check = "PASS"

    if summary_totals:
        if summary_cw is not None and cw_delta > 0.011:
            entity_check = "FAIL"
            errors.append(
                f"CW entity total mismatch: ingested ${cw_ingested:,.2f}, "
                f"Summary ${summary_cw:,.2f}, delta ${cw_delta:,.2f}."
            )
        if summary_cms is not None and cms_delta > 0.011:
            entity_check = "FAIL"
            errors.append(
                f"CMS entity total mismatch: ingested ${cms_ingested:,.2f}, "
                f"Summary ${summary_cms:,.2f}, delta ${cms_delta:,.2f}."
            )
    else:
        entity_check = "WARN"
        warnings.append("Summary tab not found or totals not parsed â€” entity total check skipped.")

    # ------------------------------------------------------------------
    # Build usage DataFrame
    # ------------------------------------------------------------------
    billing_month = (
        raw_df["_billing_month"].dropna().iloc[0] if not raw_df["_billing_month"].dropna().empty else None
    )
    invoice_month = (
        raw_df["_invoice_month"].dropna().iloc[0] if not raw_df["_invoice_month"].dropna().empty else None
    )

    usage_df = pd.DataFrame(
        {
            "BILLING_MONTH": raw_df["_billing_month"],
            "VENDOR": "Auvik",
            "VENDOR_PARTNER_NAME": raw_df["Shipping_Account"].map(_clean_text),
            "VENDOR_PRODUCT_SKU": raw_df["Product"].map(_vendor_product_sku),
            "MODIFIER": raw_df["_stream"],
            "QUANTITY": [
                _billable_quantity(charge_type, quantity, overage_quantity, total_charge_amount)
                for charge_type, quantity, overage_quantity, total_charge_amount in zip(
                    raw_df["Charge Type"],
                    raw_df["Quantity"],
                    raw_df["Overage Quantity"],
                    raw_df["Total Charge Amount"],
                    strict=True,
                )
            ],
            "UNIT_PRICE": raw_df["UnitPrice"],
            "AMOUNT": raw_df["Total Charge Amount"],
            "CURRENCY": "USD",
        },
        columns=USAGE_COLUMNS,
    )

    # ------------------------------------------------------------------
    # Build audit row
    # ------------------------------------------------------------------
    period_start = raw_df["_start_date"].dropna().min() if not raw_df["_start_date"].dropna().empty else None
    period_end = raw_df["_end_date"].dropna().max() if not raw_df["_end_date"].dropna().empty else None

    load_status = "LOADED"
    if errors:
        load_status = "LOADED_WITH_ERRORS"

    error_msg: str | None = None
    if errors or warnings:
        parts = [f"ERROR: {e}" for e in errors] + [f"WARN: {w}" for w in warnings]
        error_msg = " | ".join(parts)

    audit_df = pd.DataFrame(
        [
            {
                "SOURCE_FOLDER": source.path.parent.name,
                "SOURCE_FILE": source.path.name,
                "SOURCE_CONTENT_HASH": content_hash,
                "LOAD_STATUS": load_status,
                "BILLING_MONTH": billing_month,
                "INVOICE_MONTH": invoice_month,
                "PERIOD_START": period_start,
                "PERIOD_END": period_end,
                "TOTAL_DATA_ROWS": len(usage_df),
                "CW_ROWS": int((usage_df["MODIFIER"] == "CW").sum()),
                "CMS_ROWS": int((usage_df["MODIFIER"] == "CMS").sum()),
                "CHARGEABLE_ROWS": int(usage_df["AMOUNT"].fillna(0).ne(0).sum()),
                "ZERO_AMOUNT_ROWS": int(usage_df["AMOUNT"].fillna(0).eq(0).sum()),
                "TOTAL_ROWS_EXCLUDED": total_rows_excluded,
                "INGESTED_CW_TOTAL": round(cw_ingested, 2),
                "INGESTED_CMS_TOTAL": round(cms_ingested, 2),
                "SUMMARY_CW_TOTAL": summary_cw,
                "SUMMARY_CMS_TOTAL": summary_cms,
                "CW_TOTAL_DELTA": round(cw_delta, 4) if summary_cw is not None else None,
                "CMS_TOTAL_DELTA": round(cms_delta, 4) if summary_cms is not None else None,
                "PERIOD_UNIQUE_CHECK": period_check,
                "INVOICE_DATE_CHECK": invoice_date_check,
                "SUBTOTAL_TAX_CHECK": subtotal_check,
                "ENTITY_TOTAL_CHECK": entity_check,
                "ERROR_MESSAGE": error_msg,
                "INGESTED_AT": ingested_at,
            }
        ],
        columns=AUDIT_COLUMNS,
    )

    if entity_check == "FAIL":
        raise RuntimeError(
            f"Entity total hard-gate FAILED for {source.path.name}: {error_msg}"
        )

    if warnings:
        for w in warnings:
            print(f"  WARNING [{source.path.name}]: {w}", flush=True)

    return usage_df, audit_df


# ---------------------------------------------------------------------------
# Build function â€” scan all files and combine
# ---------------------------------------------------------------------------

def build_usage(
    source_root_cw: Path,
    month_filter: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Scan the CW folder, parse every qualifying workbook, and return
    (usage_df, audit_df).  Content-hash deduplication prevents double-loading
    if identical files are discovered more than once.
    """
    ingested_at = dt.datetime.now(dt.UTC).replace(tzinfo=None)
    files = discover_source_files(source_root_cw, month_filter=month_filter)
    if not files:
        raise FileNotFoundError(
            f"No Connectwise Usage Report workbooks found in {source_root_cw}"
            + (f" for month {month_filter}" if month_filter else "")
        )

    usage_frames: list[pd.DataFrame] = []
    audit_frames: list[pd.DataFrame] = []
    seen_hashes: set[str] = set()

    for source in files:
        print(
            f"Parsing {source.path.parent.name}/{source.path.name} ...",
            flush=True,
        )
        usage_df, audit_df = parse_usage_workbook(source, ingested_at=ingested_at)
        content_hash = audit_df["SOURCE_CONTENT_HASH"].iloc[0]

        if content_hash in seen_hashes:
            audit_df["LOAD_STATUS"] = "SKIPPED_DUPLICATE_CONTENT"
            usage_df = pd.DataFrame(columns=USAGE_COLUMNS)
            print(
                f"  SKIPPED (duplicate content): {source.path.name}",
                flush=True,
            )
        else:
            seen_hashes.add(content_hash)

        audit_frames.append(audit_df)
        usage_frames.append(usage_df)

        if not usage_df.empty:
            status = audit_df["LOAD_STATUS"].iloc[0]
            n_rows = len(usage_df)
            cw_total = float(
                usage_df.loc[usage_df["MODIFIER"] == "CW", "AMOUNT"].fillna(0).sum()
            )
            cms_total = float(
                usage_df.loc[usage_df["MODIFIER"] == "CMS", "AMOUNT"].fillna(0).sum()
            )
            print(
                f"  {status}: rows={n_rows:,} "
                f"CW=${cw_total:,.2f} CMS=${cms_total:,.2f}",
                flush=True,
            )

    non_empty = [f for f in usage_frames if not f.empty]
    usage_all = (
        pd.concat(non_empty, ignore_index=True)
        if non_empty
        else pd.DataFrame(columns=USAGE_COLUMNS)
    )
    if not usage_all.empty:
        usage_all = (
            usage_all.groupby(
                [
                    "BILLING_MONTH",
                    "VENDOR",
                    "VENDOR_PARTNER_NAME",
                    "VENDOR_PRODUCT_SKU",
                    "MODIFIER",
                    "UNIT_PRICE",
                    "CURRENCY",
                ],
                dropna=False,
                as_index=False,
            )
            .agg(QUANTITY=("QUANTITY", "sum"), AMOUNT=("AMOUNT", "sum"))
        )
        usage_all = usage_all[
            [
                "BILLING_MONTH",
                "VENDOR",
                "VENDOR_PARTNER_NAME",
                "VENDOR_PRODUCT_SKU",
                "MODIFIER",
                "QUANTITY",
                "UNIT_PRICE",
                "AMOUNT",
                "CURRENCY",
            ]
        ]
    audit_all = (
        pd.concat(audit_frames, ignore_index=True)
        if audit_frames
        else pd.DataFrame(columns=AUDIT_COLUMNS)
    )
    return usage_all, audit_all


def filter_usage_months(
    usage_df: pd.DataFrame,
    audit_df: pd.DataFrame,
    months: list[str] | None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not months:
        return usage_df, audit_df
    wanted = {dt.date.fromisoformat(f"{m}-01") for m in months}
    fu = usage_df[usage_df["BILLING_MONTH"].isin(wanted)].copy()
    fa = audit_df[audit_df["BILLING_MONTH"].isin(wanted)].copy()
    if fu.empty:
        raise RuntimeError(
            f"No usage rows found for requested month(s): {', '.join(months)}"
        )
    return fu, fa


# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

def usage_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} (
    BILLING_MONTH       DATE,
    VENDOR              VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU  VARCHAR,
    MODIFIER            VARCHAR,
    QUANTITY            NUMBER(18,4),
    UNIT_PRICE          NUMBER(18,6),
    AMOUNT              NUMBER(18,4),
    CURRENCY            VARCHAR
);
"""


def audit_ddl() -> str:
    return f"""
CREATE OR REPLACE TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} (
    SOURCE_FOLDER           VARCHAR,
    SOURCE_FILE             VARCHAR,
    SOURCE_CONTENT_HASH     VARCHAR,
    LOAD_STATUS             VARCHAR,
    BILLING_MONTH           DATE,
    INVOICE_MONTH           DATE,
    PERIOD_START            DATE,
    PERIOD_END              DATE,
    TOTAL_DATA_ROWS         NUMBER,
    CW_ROWS                 NUMBER,
    CMS_ROWS                NUMBER,
    CHARGEABLE_ROWS         NUMBER,
    ZERO_AMOUNT_ROWS        NUMBER,
    TOTAL_ROWS_EXCLUDED     NUMBER,
    INGESTED_CW_TOTAL       NUMBER(18,4),
    INGESTED_CMS_TOTAL      NUMBER(18,4),
    SUMMARY_CW_TOTAL        NUMBER(18,4),
    SUMMARY_CMS_TOTAL       NUMBER(18,4),
    CW_TOTAL_DELTA          NUMBER(18,4),
    CMS_TOTAL_DELTA         NUMBER(18,4),
    PERIOD_UNIQUE_CHECK     VARCHAR,
    INVOICE_DATE_CHECK      VARCHAR,
    SUBTOTAL_TAX_CHECK      VARCHAR,
    ENTITY_TOTAL_CHECK      VARCHAR,
    ERROR_MESSAGE           VARCHAR,
    INGESTED_AT             TIMESTAMP_NTZ
);
"""


# ---------------------------------------------------------------------------
# Snowflake load
# ---------------------------------------------------------------------------

def _snowflake_connection():
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from TEMPLATES.Python.connection import get_snowflake_connection

    return get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )


def load_snowflake(
    usage_df: pd.DataFrame,
    audit_df: pd.DataFrame,
    *,
    reset: bool = False,
) -> None:
    from snowflake.connector.pandas_tools import write_pandas

    conn = _snowflake_connection()
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")

        if reset:
            print(f"RESET: dropping {TARGET_TABLE} and {AUDIT_TABLE}.", flush=True)
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )
            cur.execute(f"DROP TABLE IF EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE}")

        cur.execute(usage_ddl())
        cur.execute(audit_ddl())

        # Month-idempotent load: skip months already present in AUVIK_USAGE
        cur.execute(
            f"SELECT DISTINCT TO_CHAR(BILLING_MONTH, 'YYYY-MM') "
            f"FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
            "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
            (TARGET_VENDOR,),
        )
        existing_months: set[str] = {row[0] for row in cur.fetchall() if row[0]}

        if not usage_df.empty:
            incoming_months = sorted(
                usage_df["BILLING_MONTH"].dropna().apply(
                    lambda d: d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)[:7]
                ).unique()
            )
            new_months = [m for m in incoming_months if m not in existing_months]
            skipped = [m for m in incoming_months if m in existing_months]
            if skipped:
                print(f"Skipping already-loaded months: {skipped}", flush=True)
            if new_months:
                load_df = usage_df[
                    usage_df["BILLING_MONTH"].apply(
                        lambda d: (d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)[:7])
                        in new_months
                    )
                ].reset_index(drop=True)
                load_df = _fill_missing_prices(load_df, TARGET_VENDOR, conn=conn)
                success, chunks, rows, output = write_pandas(
                    conn,
                    load_df,
                    TARGET_TABLE,
                    database=TARGET_DATABASE,
                    schema=TARGET_SCHEMA,
                    quote_identifiers=False,
                )
                if not success:
                    raise RuntimeError(f"write_pandas failed for {TARGET_TABLE}: {output}")
                print(f"Loaded {rows:,} rows into {TARGET_TABLE} ({chunks} chunk(s)).", flush=True)
            else:
                print("Nothing new to load into AUVIK_USAGE.", flush=True)

        # Audit: always replace (one row per source file per run)
        if not audit_df.empty:
            success, chunks, rows, output = write_pandas(
                conn,
                audit_df,
                AUDIT_TABLE,
                database=TARGET_DATABASE,
                schema=TARGET_SCHEMA,
                quote_identifiers=False,
            )
            if not success:
                raise RuntimeError(f"write_pandas failed for {AUDIT_TABLE}: {output}")
            print(f"Loaded {rows:,} rows into {AUDIT_TABLE}.", flush=True)

        conn.commit()
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Local audit summary
# ---------------------------------------------------------------------------

def summarize_local(usage_df: pd.DataFrame, audit_df: pd.DataFrame) -> None:
    print("\nLOCAL USAGE CONTROLS", flush=True)
    if usage_df.empty:
        print("  (no usage rows)", flush=True)
    else:
        summary = (
            usage_df.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
            .agg(
                rows=("AMOUNT", "size"),
                quantity=("QUANTITY", "sum"),
                total_charge=("AMOUNT", "sum"),
            )
            .reset_index()
            .sort_values(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"])
        )
        print(summary.to_string(index=False), flush=True)

        # Entity totals cross-check
        print("\nENTITY TOTALS", flush=True)
        entity = (
            usage_df.groupby(["BILLING_MONTH", "MODIFIER"], dropna=False)["AMOUNT"]
            .sum()
            .reset_index()
            .sort_values(["BILLING_MONTH", "MODIFIER"])
        )
        print(entity.to_string(index=False), flush=True)

    if not audit_df.empty:
        print("\nAUDIT CHECKS", flush=True)
        checks = audit_df[
            [
                "SOURCE_FILE",
                "LOAD_STATUS",
                "BILLING_MONTH",
                "TOTAL_DATA_ROWS",
                "CW_ROWS",
                "CMS_ROWS",
                "CHARGEABLE_ROWS",
                "ZERO_AMOUNT_ROWS",
                "INGESTED_CW_TOTAL",
                "SUMMARY_CW_TOTAL",
                "CW_TOTAL_DELTA",
                "INGESTED_CMS_TOTAL",
                "SUMMARY_CMS_TOTAL",
                "CMS_TOTAL_DELTA",
                "ENTITY_TOTAL_CHECK",
                "PERIOD_UNIQUE_CHECK",
                "ERROR_MESSAGE",
            ]
        ]
        print(checks.to_string(index=False), flush=True)


def write_local_audit(usage_df: pd.DataFrame, audit_df: pd.DataFrame, label: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    usage_path = OUTPUT_DIR / f"auvik_ingest_usage_{label}.csv"
    audit_path = OUTPUT_DIR / f"auvik_ingest_audit_{label}.csv"
    usage_df.to_csv(usage_path, index=False)
    audit_df.to_csv(audit_path, index=False)
    print(f"Wrote {usage_path}", flush=True)
    print(f"Wrote {audit_path}", flush=True)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Dynamic invoice rate fill (universal safety net)
# ---------------------------------------------------------------------------
def _fill_missing_prices(df, vendor_name, conn=None):
    from invoice_rate_backfill import fill_missing_prices_dynamic

    return fill_missing_prices_dynamic(df=df, vendor_name=vendor_name, conn=conn)

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Ingest Auvik Connectwise Usage Report workbooks into AUVIK_USAGE. "
            "Only the Auvik CW folder is scanned; the CW/CMS entity split is "
            "derived from Partner Name in the data."
        )
    )
    parser.add_argument(
        "--source-root",
        default=str(DEFAULT_SOURCE_ROOT_CW),
        help="Path to the Auvik CW month folder root (default: OneDrive path).",
    )
    # Kept for backwards compatibility with run_full_pipeline.py
    parser.add_argument("--channel", help="Ignored â€” Auvik reads CW only.")
    parser.add_argument("--month", help="Load a single month in YYYY-MM format.")
    parser.add_argument(
        "--all-months", action="store_true", help="Parse every discovered month folder."
    )
    parser.add_argument("--start-month", help="Lower bound (inclusive) for --all-months.")
    parser.add_argument("--end-month", help="Upper bound (inclusive) for --all-months.")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and audit locally without loading to Snowflake.",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Drop and recreate AUVIK_USAGE and AUVIK_USAGE_FILE_AUDIT before loading.",
    )
    args = parser.parse_args()

    source_root = Path(args.source_root)
    if not source_root.exists():
        raise FileNotFoundError(f"Source root does not exist: {source_root}")
    if not args.month and not args.all_months:
        raise SystemExit("Provide --month YYYY-MM or --all-months.")

    month_filter: str | None = args.month

    usage_df, audit_df = build_usage(source_root, month_filter=month_filter)

    # Apply start/end-month bounds when using --all-months
    if args.all_months and (args.start_month or args.end_month):
        def _in_range(d: object) -> bool:
            if d is None:
                return True
            m = d.strftime("%Y-%m") if hasattr(d, "strftime") else str(d)[:7]
            if args.start_month and m < args.start_month:
                return False
            if args.end_month and m > args.end_month:
                return False
            return True

        usage_df = usage_df[usage_df["BILLING_MONTH"].apply(_in_range)].copy()
        audit_df = audit_df[audit_df["BILLING_MONTH"].apply(_in_range)].copy()

    if usage_df.empty:
        raise RuntimeError("No usage rows parsed. Check source root and month filters.")

    summarize_local(usage_df, audit_df)

    label: str
    if args.month:
        label = args.month.replace("-", "_")
    elif args.start_month or args.end_month:
        label = f"{(args.start_month or 'min').replace('-','_')}_to_{(args.end_month or 'max').replace('-','_')}"
    else:
        label = "all_months"

    write_local_audit(usage_df, audit_df, label)

    if args.dry_run:
        print("Dry run complete â€” Snowflake load skipped.", flush=True)
        return

    load_snowflake(usage_df, audit_df, reset=args.reset)


if __name__ == "__main__":
    main()



