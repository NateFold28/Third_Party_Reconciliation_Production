from __future__ import annotations

import csv
import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
PROJECTS = Path(r"C:\Users\Nate.Fold\projects")
MANUAL_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026"
)
OUT_DIR = REPO / "outputs" / f"deep_audit_{datetime.now():%Y%m%d_%H%M%S}"

sys.path.insert(0, str(PROJECTS))
from TEMPLATES.Python.connection import get_snowflake_connection  # noqa: E402


USE = """
USE ROLE DEVELOPER;
USE WAREHOUSE REPORTING_WH;
USE DATABASE ANALYTICS_DEV;
USE SCHEMA DBT_NFOLD_TRANSFORMATION;
"""

VENDOR_FOLDERS = {
    "Acronis": "Acronis",
    "Auvik CMS": "Auvik",
    "Auvik CW": "Auvik",
    "Bitdefender": "Bitdefender",
    "ESET": "ESET",
    "Exium": "Exium",
    "KeepIT": "KeepIT",
    "Proofpoint": "Proofpoint",
    "SentinelOne": "SentinelOne",
    "Webroot CMS": "Webroot",
    "Webroot CW": "Webroot",
}

TARGET_SHEET_RE = re.compile(r"^(data|consolidated data|full data|most_recent_month_full data|control)$", re.I)
NOISY_FILE_PREFIXES = ("~$",)


def query_df(conn, sql: str) -> pd.DataFrame:
    cur = conn.cursor()
    cur.execute(sql)
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]
    cur.close()
    return pd.DataFrame(rows, columns=cols)


def write_df(df: pd.DataFrame, name: str) -> Path:
    path = OUT_DIR / name
    df.to_csv(path, index=False, quoting=csv.QUOTE_MINIMAL)
    return path


def scan_manual_workbooks() -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    if not MANUAL_ROOT.exists():
        return pd.DataFrame([{"manual_root": str(MANUAL_ROOT), "error": "manual root not found"}])

    for folder, vendor in VENDOR_FOLDERS.items():
        folder_path = MANUAL_ROOT / folder
        if not folder_path.exists():
            rows.append({"vendor": vendor, "folder": folder, "error": "folder not found"})
            continue
        for path in sorted(folder_path.rglob("*.xlsx")):
            if path.name.startswith(NOISY_FILE_PREFIXES):
                continue
            try:
                wb = load_workbook(path, read_only=True, data_only=False)
            except Exception as exc:
                rows.append({
                    "vendor": vendor,
                    "folder": folder,
                    "file": str(path),
                    "error": str(exc)[:250],
                })
                continue
            try:
                candidate_sheets = []
                sheet_rows = {}
                sheet_cols = {}
                for ws in wb.worksheets:
                    title = ws.title.strip()
                    if TARGET_SHEET_RE.match(title) or "data" in title.lower():
                        candidate_sheets.append(title)
                    sheet_rows[title] = ws.max_row
                    sheet_cols[title] = ws.max_column
                rows.append({
                    "vendor": vendor,
                    "folder": folder,
                    "file": str(path),
                    "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                    "sheet_count": len(wb.sheetnames),
                    "candidate_target_sheets": " | ".join(candidate_sheets),
                    "all_sheets": " | ".join(wb.sheetnames),
                    "sheet_rows": "; ".join(f"{k}:{v}" for k, v in sheet_rows.items()),
                    "sheet_cols": "; ".join(f"{k}:{v}" for k, v in sheet_cols.items()),
                    "error": "",
                })
            finally:
                wb.close()
    return pd.DataFrame(rows)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database="ANALYTICS_DEV",
        schema="DBT_NFOLD_TRANSFORMATION",
    )
    try:
        conn.execute_string(USE)
        app_freshness = query_df(conn, """
            SELECT TABLE_NAME, ROW_COUNT, LAST_ALTERED
            FROM ANALYTICS_DEV.INFORMATION_SCHEMA.TABLES
            WHERE TABLE_SCHEMA = 'DBT_NFOLD_TRANSFORMATION'
              AND TABLE_NAME IN (
                  'THIRD_PARTY_RECON_DETAIL_PROD',
                  'THIRD_PARTY_RECON_OUTPUT_PROD',
                  'THIRD_PARTY_RECON_SUMMARY_PROD'
              )
            ORDER BY TABLE_NAME
        """)
        write_df(app_freshness, "app_table_freshness.csv")

        vendor_metrics = query_df(conn, """
            WITH loaded_months AS (
                SELECT vendor, billing_month
                FROM THIRD_PARTY_RECON_SUMMARY_PROD
                WHERE data_load_status = 'LOADED'
            )
            SELECT
                o.vendor,
                COUNT(*) AS total_rows_all,
                COUNT_IF(o.exception_type = 'Clear') AS clear_rows_all,
                ROUND(COUNT_IF(o.exception_type = 'Clear') * 100.0 / NULLIF(COUNT(*), 0), 1) AS clear_pct_all,
                COUNT_IF(l.billing_month IS NOT NULL) AS total_rows_loaded,
                COUNT_IF(l.billing_month IS NOT NULL AND o.exception_type = 'Clear') AS clear_rows_loaded,
                ROUND(COUNT_IF(l.billing_month IS NOT NULL AND o.exception_type = 'Clear') * 100.0
                    / NULLIF(COUNT_IF(l.billing_month IS NOT NULL), 0), 1) AS clear_pct_loaded,
                ROUND(SUM(ABS(COALESCE(o.qty_delta, 0))), 0) AS abs_qty_delta_all,
                ROUND(SUM(IFF(l.billing_month IS NOT NULL, ABS(COALESCE(o.qty_delta, 0)), 0)), 0) AS abs_qty_delta_loaded,
                ROUND(SUM(COALESCE(o.vendor_amount, 0)), 2) AS vendor_cost_all,
                ROUND(SUM(COALESCE(o.total_billing_amount, 0)), 2) AS cw_revenue_all,
                ROUND(SUM(COALESCE(o.total_billing_amount, 0)) - SUM(COALESCE(o.vendor_amount, 0)), 2) AS cw_margin_all,
                ROUND((SUM(COALESCE(o.total_billing_amount, 0)) - SUM(COALESCE(o.vendor_amount, 0)))
                    / NULLIF(SUM(COALESCE(o.total_billing_amount, 0)), 0) * 100, 1) AS cw_margin_pct_all,
                ROUND(SUM(IFF(o.exception_type IN (
                    'Vendor Billing, No CW Billing',
                    'Vendor Billing, Insufficient CW Billing',
                    'API Usage Recorded, No CW Billing',
                    'Vendor SKU, No CW SKU'
                ), ABS(COALESCE(o.amount_delta, 0)), 0)), 2) AS finance_leakage_amount
            FROM THIRD_PARTY_RECON_OUTPUT_PROD o
            LEFT JOIN loaded_months l
              ON l.vendor = o.vendor
             AND l.billing_month = o.billing_month
            GROUP BY o.vendor
            ORDER BY clear_pct_loaded DESC NULLS LAST
        """)
        write_df(vendor_metrics, "vendor_metrics.csv")

        month_metrics = query_df(conn, """
            SELECT
                vendor,
                billing_month,
                data_load_status,
                usage_row_count,
                total_rows,
                perfect_match_rows AS clear_rows,
                clear_pct,
                total_vendor_amount AS vendor_cost,
                total_billing_amount AS cw_revenue,
                ROUND(total_billing_amount - total_vendor_amount, 2) AS cw_margin,
                total_leakage_amount
            FROM THIRD_PARTY_RECON_SUMMARY_PROD
            ORDER BY vendor, billing_month
        """)
        write_df(month_metrics, "vendor_month_metrics.csv")

        exception_distribution = query_df(conn, """
            SELECT
                vendor,
                exception_type,
                COUNT(*) AS row_count,
                COUNT(DISTINCT sf_id) AS account_count,
                ROUND(SUM(ABS(COALESCE(qty_delta, 0))), 0) AS abs_qty_delta,
                ROUND(SUM(ABS(COALESCE(amount_delta, 0))), 2) AS abs_amount_delta,
                ROUND(SUM(COALESCE(vendor_amount, 0)), 2) AS vendor_cost,
                ROUND(SUM(COALESCE(total_billing_amount, 0)), 2) AS cw_revenue
            FROM THIRD_PARTY_RECON_OUTPUT_PROD
            GROUP BY vendor, exception_type
            ORDER BY vendor, row_count DESC
        """)
        write_df(exception_distribution, "exception_distribution.csv")

        top_products = query_df(conn, """
            SELECT *
            FROM (
                SELECT
                    vendor,
                    exception_type,
                    vendor_product,
                    sku_match_group,
                    COUNT(*) AS row_count,
                    COUNT(DISTINCT sf_id) AS account_count,
                    ROUND(SUM(ABS(COALESCE(qty_delta, 0))), 0) AS abs_qty_delta,
                    ROUND(SUM(ABS(COALESCE(amount_delta, 0))), 2) AS abs_amount_delta,
                    ROUND(SUM(COALESCE(vendor_amount, 0)), 2) AS vendor_cost,
                    ROUND(SUM(COALESCE(total_billing_amount, 0)), 2) AS cw_revenue,
                    ROW_NUMBER() OVER (
                        PARTITION BY vendor
                        ORDER BY SUM(ABS(COALESCE(amount_delta, 0))) DESC, COUNT(*) DESC
                    ) AS rn
                FROM THIRD_PARTY_RECON_OUTPUT_PROD
                WHERE exception_type <> 'Clear'
                GROUP BY vendor, exception_type, vendor_product, sku_match_group
            )
            WHERE rn <= 25
            ORDER BY vendor, rn
        """)
        write_df(top_products, "top_exception_products.csv")

        mapping_gaps = query_df(conn, """
            SELECT
                vendor,
                exception_type,
                vendor_partner_name,
                sf_id,
                vendor_product,
                COUNT(*) AS row_count,
                ROUND(SUM(ABS(COALESCE(qty_delta, 0))), 0) AS abs_qty_delta,
                ROUND(SUM(ABS(COALESCE(amount_delta, 0))), 2) AS abs_amount_delta,
                ROUND(SUM(COALESCE(vendor_amount, 0)), 2) AS vendor_cost,
                ROUND(SUM(COALESCE(total_billing_amount, 0)), 2) AS cw_revenue
            FROM THIRD_PARTY_RECON_OUTPUT_PROD
            WHERE exception_type IN ('Unmapped Partner', 'Vendor SKU, No CW SKU', 'CW SKU, No Vendor SKU')
            GROUP BY vendor, exception_type, vendor_partner_name, sf_id, vendor_product
            ORDER BY abs_amount_delta DESC, row_count DESC
            LIMIT 250
        """)
        write_df(mapping_gaps, "mapping_gap_candidates.csv")

        negative_margin = query_df(conn, """
            SELECT
                vendor,
                billing_month,
                sf_id,
                vendor_partner_name,
                vendor_product,
                vendor_amount,
                total_billing_amount,
                amount_delta,
                cw_margin_pct,
                exception_type,
                outcome_flag
            FROM THIRD_PARTY_RECON_OUTPUT_PROD
            WHERE COALESCE(vendor_amount, 0) > COALESCE(total_billing_amount, 0)
              AND COALESCE(vendor_amount, 0) > 0
            ORDER BY ABS(COALESCE(amount_delta, 0)) DESC
            LIMIT 500
        """)
        write_df(negative_margin, "negative_margin_candidates.csv")

        eset_quantity_native = query_df(conn, """
            SELECT
                base_outcome_flag,
                COUNT(*) AS row_count,
                ROUND(SUM(ABS(COALESCE(abs_qty_delta, 0))), 0) AS abs_qty_delta,
                ROUND(SUM(COALESCE(contract_cost_basis_amount, 0)), 2) AS contract_cost_basis_amount,
                ROUND(SUM(COALESCE(total_billing_amount, 0)), 2) AS cw_revenue
            FROM ESET_RECON_DETAIL
            GROUP BY base_outcome_flag
            ORDER BY row_count DESC
        """)
        write_df(eset_quantity_native, "eset_quantity_native_metrics.csv")

    finally:
        conn.close()

    manual_inventory = scan_manual_workbooks()
    write_df(manual_inventory, "manual_workbook_inventory.csv")

    readme = OUT_DIR / "README.md"
    with readme.open("w", encoding="utf-8", newline="\n") as f:
        f.write("# Deep Audit Current State\n\n")
        f.write(f"Generated: {datetime.now():%Y-%m-%d %H:%M:%S}\n\n")
        f.write("## Files\n\n")
        for path in sorted(OUT_DIR.glob("*.csv")):
            f.write(f"- `{path.name}`\n")
        f.write("\n## Notes\n\n")
        f.write("- `clear_pct_loaded` excludes vendor-months marked `NOT_LOADED` or `PARTIAL` in `THIRD_PARTY_RECON_SUMMARY_PROD`.\n")
        f.write("- Manual workbook tabs are inventoried as validation targets only; this script does not ingest manual data into production tables.\n")
        f.write("- Optimize in this order: mapping gaps, product/SKU rule gaps, source coverage/stale-month gaps, then pricing/margin calibration.\n")

    print(f"Audit written to: {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
