from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[4]
PIPELINE_ROOT = Path(__file__).resolve().parents[1]
INGEST_PATH = PIPELINE_ROOT / "Ingestion" / "KeepIT_Vendor_Usage_Ingestion_Prod.py"

sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PIPELINE_ROOT / "Ingestion"))

from TEMPLATES.Python.connection import get_snowflake_connection


def load_ingest_module():
    spec = importlib.util.spec_from_file_location("keepit_ingest", INGEST_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {INGEST_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def print_df(title: str, df: pd.DataFrame, max_rows: int = 50) -> None:
    print(f"\n## {title}")
    if df.empty:
        print("(empty)")
        return
    with pd.option_context("display.max_rows", max_rows, "display.max_columns", 40, "display.width", 240):
        print(df.head(max_rows).to_string(index=False))


def run_sql(cur, title: str, sql: str, max_rows: int = 50) -> pd.DataFrame:
    cur.execute(sql)
    df = cur.fetch_pandas_all()
    print_df(title, df, max_rows=max_rows)
    return df


def audit_local_file_selection(ingest) -> None:
    rows: list[dict[str, object]] = []
    source_root = ingest.DEFAULT_SOURCE_ROOT
    for month, folder in ingest.discover_month_folders(source_root).items():
        files = [p for p in folder.iterdir() if p.is_file()]
        selected = ingest.locate_usage_files(folder)
        selected_names = {p.name for _, p in selected}
        promo_candidates = sorted(p for p in files if ingest.is_promo_summary_file(p))
        takeout_xlsx_candidates = sorted(p for p in files if ingest.is_takeout_summary_file(p))
        takeout_pdf_candidates = sorted(
            p for p in files if ingest.is_takeout_invoice_file(p) and ingest.pdf_has_takeout_lines(p)
        )
        row = {
            "month": month,
            "selected": "; ".join(f"{family}:{path.name}" for family, path in selected),
            "promo_candidates": "; ".join(p.name for p in promo_candidates),
            "selected_promo_count": sum(1 for family, _ in selected if family == "PROMO"),
            "promo_candidate_count": len(promo_candidates),
            "takeout_pdf_count": len(takeout_pdf_candidates),
            "takeout_xlsx_count": len(takeout_xlsx_candidates),
            "suppressed_promo_files": "; ".join(p.name for p in promo_candidates if p.name not in selected_names),
        }
        rows.append(row)
    print_df("Local File Selection", pd.DataFrame(rows), max_rows=20)


def audit_candidate_quantities(ingest) -> None:
    rows: list[dict[str, object]] = []
    for month, folder in ingest.discover_month_folders(ingest.DEFAULT_SOURCE_ROOT).items():
        billing_month = ingest.billing_month_from_folder(folder)
        files = [p for p in folder.iterdir() if p.is_file()]
        selected = ingest.locate_usage_files(folder)
        selected_names = {p.name for _, p in selected}
        candidates: list[tuple[str, Path]] = []
        candidates.extend(("PROMO_CANDIDATE", p) for p in sorted(p for p in files if ingest.is_promo_summary_file(p)))
        candidates.extend(("TAKEOUT_XLSX_CANDIDATE", p) for p in sorted(p for p in files if ingest.is_takeout_summary_file(p)))
        candidates.extend(
            ("TAKEOUT_PDF_CANDIDATE", p)
            for p in sorted(p for p in files if ingest.is_takeout_invoice_file(p) and ingest.pdf_has_takeout_lines(p))
        )
        for family, path in candidates:
            try:
                if family == "PROMO_CANDIDATE":
                    parsed, _ = ingest.parse_promo(path, billing_month)
                elif family == "TAKEOUT_XLSX_CANDIDATE":
                    parsed, _ = ingest.parse_takeout_workbook(path, billing_month)
                else:
                    parsed, _ = ingest.parse_takeout_pdf(path, billing_month)
                df = pd.DataFrame(parsed)
                qty = float(df["QUANTITY"].sum()) if not df.empty else 0.0
                amount = float(df["AMOUNT"].sum()) if not df.empty else 0.0
                row_count = len(df)
                error = None
            except Exception as exc:
                qty = amount = 0.0
                row_count = 0
                error = str(exc)[:200]
            rows.append(
                {
                    "month": month,
                    "candidate_type": family,
                    "selected_by_current_ingestion": path.name in selected_names,
                    "file": path.name,
                    "row_count": row_count,
                    "quantity": qty,
                    "amount": amount,
                    "error": error,
                }
            )
    print_df("Candidate Promo/Takeout Quantities", pd.DataFrame(rows), max_rows=80)


def audit_manual_recon_tabs() -> None:
    import openpyxl

    rows: list[dict[str, object]] = []
    root = load_ingest_module().DEFAULT_SOURCE_ROOT
    for path in sorted(root.rglob("*.xlsx")):
        if "recon" not in path.name.lower():
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            rows.append({"file": path.name, "sheets": "", "data_like_sheets": "", "error": str(exc)[:200]})
            continue
        try:
            data_like = [
                s for s in wb.sheetnames
                if s.strip().lower() in {"data", "consolidated data", "control"}
                or "data" in s.strip().lower()
            ]
            rows.append(
                {
                    "file": path.name,
                    "sheets": "; ".join(wb.sheetnames[:15]),
                    "data_like_sheets": "; ".join(data_like),
                    "error": None,
                }
            )
        finally:
            wb.close()
    print_df("Manual Recon Workbook Tabs", pd.DataFrame(rows), max_rows=80)


def main() -> None:
    ingest = load_ingest_module()
    audit_local_file_selection(ingest)
    audit_candidate_quantities(ingest)
    audit_manual_recon_tabs()

    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database="ANALYTICS_DEV",
        schema="DBT_NFOLD_TRANSFORMATION",
    )
    try:
        cur = conn.cursor()
        run_sql(
            cur,
            "Current Loaded KeepIT Usage by Modifier",
            """
            SELECT
                billing_month,
                COALESCE(modifier, '(NULL)') AS modifier,
                COUNT(*) AS row_count,
                SUM(quantity) AS quantity,
                SUM(amount) AS amount
            FROM third_party_recon_vendor_usage_prod
            WHERE vendor = 'KeepIT'
            GROUP BY 1, 2
            ORDER BY 1, 2
            """,
            max_rows=40,
        )
        run_sql(
            cur,
            "Current KeepIT Summary",
            """
            SELECT
                billing_month,
                source_family,
                total_rows,
                perfect_match_pct,
                actionable_clear_pct,
                abs_qty_variance,
                total_vendor_seats,
                total_billing_seats,
                takeout_support_rows,
                unmapped_rows,
                no_billing_rows,
                billing_only_rows
            FROM keepit_recon_summary
            ORDER BY 1, 2
            """,
            max_rows=40,
        )
        run_sql(
            cur,
            "KeepIT Outcome by Source Family",
            """
            SELECT
                billing_month,
                source_family,
                outcome_flag,
                COUNT(*) AS row_count,
                SUM(vendor_quantity) AS vendor_quantity,
                SUM(total_billing_quantity) AS billing_quantity,
                SUM(abs_qty_delta) AS abs_qty_delta,
                SUM(vendor_amount) AS vendor_amount,
                SUM(total_billing_amount) AS billing_amount
            FROM keepit_recon_detail
            GROUP BY 1, 2, 3
            ORDER BY billing_month, source_family, abs_qty_delta DESC
            """,
            max_rows=120,
        )
        run_sql(
            cur,
            "KeepIT SKU Map Review Rows",
            """
            SELECT
                vendor_product,
                vendor_sku,
                cw_sku,
                sku_match_key
            FROM third_party_recon_sku_map_prod
            WHERE vendor ILIKE '%keepit%'
              AND (
                    cw_sku IS NULL
                 OR TRIM(cw_sku) = ''
                 OR vendor_sku ILIKE '%UNCONFIRMED%'
                 OR sku_match_key ILIKE 'KEEPIT_CW_ONLY%'
              )
            ORDER BY vendor_product, cw_sku
            """,
            max_rows=120,
        )
        run_sql(
            cur,
            "Zuora Promo SKU Classification",
            """
            WITH sm AS (
                SELECT DISTINCT
                    UPPER(TRIM(sku_match_key)) AS sku_match_group,
                    UPPER(TRIM(tok.value)) AS cw_sku_token
                FROM third_party_recon_sku_map_prod,
                     LATERAL SPLIT_TO_TABLE(REPLACE(cw_sku, '/', '|'), '|') tok
                WHERE vendor = 'KeepIT'
                  AND sku_match_key IS NOT NULL
                  AND cw_sku IS NOT NULL
                  AND TRIM(tok.value) <> ''
            )
            SELECT
                z.billing_month,
                UPPER(TRIM(z.product_sku)) AS product_sku,
                COALESCE(sm.sku_match_group, UPPER(TRIM(z.product_sku))) AS sku_match_group,
                CASE
                    WHEN COALESCE(sm.sku_match_group, UPPER(TRIM(z.product_sku))) ILIKE 'KEEPIT_PROMO_%' THEN 'PROMO'
                    WHEN COALESCE(sm.sku_match_group, UPPER(TRIM(z.product_sku))) ILIKE 'KEEPIT_TAKEOUT_%' THEN 'TAKEOUT'
                    ELSE 'MAIN'
                END AS current_source_family,
                COUNT(*) AS row_count,
                SUM(z.qty) AS qty,
                SUM(z.charge_amount_usd) AS amount
            FROM third_party_recon_source_zuora_prod z
            LEFT JOIN sm ON sm.cw_sku_token = UPPER(TRIM(z.product_sku))
            WHERE z.vendor = 'KeepIT'
              AND (
                    z.product_sku ILIKE '%PROMO%'
                 OR z.charge_name ILIKE '%PROMO%'
                 OR z.product_name ILIKE '%PROMO%'
              )
            GROUP BY 1, 2, 3, 4
            ORDER BY 1, product_sku
            """,
            max_rows=120,
        )
        run_sql(
            cur,
            "Top Vendor Rows with Zero Billing",
            """
            SELECT
                billing_month,
                source_family,
                vendor_product,
                outcome_flag,
                COUNT(*) AS row_count,
                SUM(vendor_quantity) AS vendor_quantity,
                SUM(total_billing_quantity) AS billing_quantity,
                SUM(abs_qty_delta) AS abs_qty_delta,
                SUM(vendor_amount) AS vendor_amount
            FROM keepit_recon_detail
            WHERE vendor_quantity > 0
              AND COALESCE(total_billing_quantity, 0) = 0
            GROUP BY 1, 2, 3, 4
            ORDER BY abs_qty_delta DESC
            LIMIT 50
            """,
            max_rows=60,
        )
    finally:
        conn.close()


if __name__ == "__main__":
    main()
