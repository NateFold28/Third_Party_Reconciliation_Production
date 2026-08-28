from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[4]
sys.path.insert(0, str(PROJECT_ROOT))

from TEMPLATES.Python.connection import get_snowflake_connection

MANUAL_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026\KeepIT"
)


def print_df(title: str, df: pd.DataFrame, max_rows: int = 80) -> None:
    print(f"\n## {title}")
    if df.empty:
        print("(empty)")
        return
    with pd.option_context("display.max_rows", max_rows, "display.max_columns", 50, "display.width", 260):
        print(df.head(max_rows).to_string(index=False))


def find_header_row(ws, max_rows: int = 30):
    keywords = {"sku", "product", "charge", "qty", "quantity", "vendor", "zuora", "delta", "cw"}
    best = None
    for row_idx in range(1, min(ws.max_row, max_rows) + 1):
        vals = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 80) + 1)]
        normalized = [str(v).strip().lower() for v in vals if v is not None and str(v).strip()]
        score = sum(any(k in v for k in keywords) for v in normalized)
        if best is None or score > best[0]:
            best = (score, row_idx, vals)
    return best


def audit_manual_data_headers() -> None:
    rows = []
    for path in sorted(MANUAL_ROOT.rglob("*.xlsx")):
        if "promo recon" not in path.name.lower() and "keepit recon" not in path.name.lower():
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            rows.append({"file": path.name, "sheet": None, "header_row": None, "headers": None, "error": str(exc)[:200]})
            continue
        try:
            for sheet in wb.sheetnames:
                if sheet.strip().lower() not in {"data", "consolidated data"}:
                    continue
                ws = wb[sheet]
                best = find_header_row(ws)
                headers = []
                if best:
                    _, header_row, vals = best
                    headers = [str(v).strip() for v in vals if v is not None and str(v).strip()]
                else:
                    header_row = None
                rows.append(
                    {
                        "file": path.name,
                        "sheet": sheet,
                        "header_row": header_row,
                        "headers": " | ".join(headers[:30]),
                        "error": None,
                    }
                )
        finally:
            wb.close()
    print_df("Manual Data/Consolidated Headers", pd.DataFrame(rows), max_rows=120)


def run_sql() -> None:
    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database="ANALYTICS_DEV",
        schema="DBT_NFOLD_TRANSFORMATION",
    )
    try:
        cur = conn.cursor()
        queries = {
            "KeepIT Map Rows for Promo Bundle": """
                SELECT vendor_product, vendor_sku, cw_sku, sku_match_key
                FROM third_party_recon_sku_map_prod
                WHERE vendor = 'KeepIT'
                  AND cw_sku ILIKE '%PROMO%BUNDLE%'
                ORDER BY cw_sku, sku_match_key
            """,
            "KeepIT CW SKU Multi-Map Risk": """
                SELECT
                    cw_sku,
                    COUNT(DISTINCT sku_match_key) AS sku_match_key_count,
                    LISTAGG(DISTINCT sku_match_key, ' | ') WITHIN GROUP (ORDER BY sku_match_key) AS sku_match_keys
                FROM third_party_recon_sku_map_prod
                WHERE vendor = 'KeepIT'
                  AND cw_sku IS NOT NULL
                  AND TRIM(cw_sku) <> ''
                GROUP BY 1
                HAVING COUNT(DISTINCT sku_match_key) > 1
                ORDER BY sku_match_key_count DESC, cw_sku
            """,
            "Promo Bundle Zuora Charge Names": """
                SELECT
                    billing_month,
                    product_sku,
                    product_name,
                    charge_name,
                    COUNT(*) AS row_count,
                    SUM(qty) AS qty,
                    SUM(charge_amount_usd) AS amount
                FROM third_party_recon_source_zuora_prod
                WHERE vendor = 'KeepIT'
                  AND product_sku ILIKE '%PROMO%BUNDLE%'
                GROUP BY 1,2,3,4
                ORDER BY 1, charge_name
            """,
            "Current Qualified Mapping for Promo Bundle": """
                WITH keepit_sku_map AS (
                    SELECT DISTINCT
                        UPPER(TRIM(sku_match_key)) AS sku_match_group,
                        UPPER(TRIM(cw_sku)) AS cw_sku
                    FROM third_party_recon_sku_map_prod
                    WHERE vendor = 'KeepIT'
                      AND sku_match_key IS NOT NULL
                      AND cw_sku IS NOT NULL
                ),
                keepit_sku_map_tokens AS (
                    SELECT DISTINCT
                        sm.sku_match_group,
                        sm.cw_sku,
                        UPPER(TRIM(tok.value)) AS cw_sku_token
                    FROM keepit_sku_map sm,
                         LATERAL SPLIT_TO_TABLE(REPLACE(sm.cw_sku, '/', '|'), '|') tok
                    WHERE TRIM(tok.value) <> ''
                ),
                mapped AS (
                    SELECT
                        z.billing_month,
                        z.invoice_number,
                        z.invoice_id,
                        z.product_sku,
                        z.charge_name,
                        z.qty,
                        z.charge_amount_usd,
                        sm.sku_match_group,
                        ROW_NUMBER() OVER (
                            PARTITION BY z.sf_id, z.billing_month::DATE, z.invoice_number, z.invoice_id, z.product_sku, z.charge_name, z.qty, z.charge_amount_usd
                            ORDER BY IFF(sm.cw_sku = UPPER(TRIM(z.product_sku)), 1, 0) DESC,
                                     LENGTH(COALESCE(sm.cw_sku, UPPER(TRIM(z.product_sku)))) ASC,
                                     sm.sku_match_group
                        ) AS rn
                    FROM third_party_recon_source_zuora_prod z
                    LEFT JOIN keepit_sku_map_tokens sm
                      ON sm.cw_sku_token = UPPER(TRIM(z.product_sku))
                    WHERE z.vendor = 'KeepIT'
                      AND z.product_sku ILIKE '%PROMO%BUNDLE%'
                )
                SELECT billing_month, sku_match_group, COUNT(*) AS row_count, SUM(qty) AS qty, SUM(charge_amount_usd) AS amount
                FROM mapped
                WHERE rn = 1
                GROUP BY 1,2
                ORDER BY 1,2
            """,
        }
        for title, query in queries.items():
            cur.execute(query)
            print_df(title, cur.fetch_pandas_all(), max_rows=120)
    finally:
        conn.close()


def main() -> None:
    audit_manual_data_headers()
    run_sql()


if __name__ == "__main__":
    main()
