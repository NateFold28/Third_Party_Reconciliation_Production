"""Inventory manual reconciliation workbooks for calibration evidence.

The manual workbooks remain validation/calibration inputs only. This script
does not load data to Snowflake; it creates a compact CSV inventory of sheets
and likely header rows so each vendor work-block can target the right tab.
"""
from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


MANUAL_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026"
)
REPO = Path(__file__).resolve().parents[1]
OUT_DIR = REPO / "outputs" / f"manual_workbook_inventory_{datetime.now():%Y%m%d_%H%M%S}"
VENDORS = {
    "Acronis": ["Acronis"],
    "Auvik": ["Auvik CMS", "Auvik CW"],
    "Bitdefender": ["Bitdefender"],
    "ESET": ["ESET"],
    "Exium": ["Exium"],
    "KeepIT": ["KeepIT"],
    "Proofpoint": ["Proofpoint"],
    "SentinelOne": ["SentinelOne"],
    "Webroot": ["Webroot CMS", "Webroot CW"],
}


def month_from_path(path: Path) -> str:
    for part in path.parts:
        if re.match(r"^\d{2}_[A-Z]{3}_2026$", part.upper()):
            return part
    return ""


def sample_sheet(path: Path, sheet_name: str) -> tuple[int, str, str]:
    wb = load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet_name]
    best_row = 0
    best_score = -1
    best_values: list[str] = []
    header_terms = {
        "partner",
        "account",
        "sf",
        "sku",
        "qty",
        "quantity",
        "usage",
        "vendor",
        "zuora",
        "comment",
        "status",
        "difference",
    }
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        values = [str(v).strip() for v in row if v is not None and str(v).strip()]
        lowered = [v.lower() for v in values]
        score = sum(any(term in v for term in header_terms) for v in lowered)
        if score > best_score:
            best_row = idx
            best_score = score
            best_values = values[:20]
    wb.close()
    return best_row, " | ".join(best_values), str(best_score)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out_file = OUT_DIR / "manual_workbook_inventory.csv"
    rows: list[list[str]] = []
    for vendor, folders in VENDORS.items():
        for folder in folders:
            root = MANUAL_ROOT / folder
            if not root.exists():
                rows.append([vendor, folder, "", "", "", "", "", "missing folder"])
                continue
            for path in sorted(root.rglob("*.xlsx")):
                if path.name.startswith("~$"):
                    continue
                if "recon" not in path.name.lower() and "reconciliation" not in path.name.lower():
                    continue
                try:
                    wb = load_workbook(path, read_only=True, data_only=False)
                    sheet_names = wb.sheetnames
                    wb.close()
                    priority_sheets = [
                        s for s in sheet_names
                        if s.strip().lower() in {"data", "consolidated data", "control"}
                    ]
                    if not priority_sheets:
                        priority_sheets = [s for s in sheet_names if "data" in s.lower()][:3]
                    if not priority_sheets:
                        priority_sheets = sheet_names[:3]
                    for sheet in priority_sheets:
                        header_row, header_sample, score = sample_sheet(path, sheet)
                        rows.append([
                            vendor,
                            folder,
                            month_from_path(path),
                            str(path),
                            sheet,
                            str(header_row),
                            score,
                            header_sample,
                        ])
                except Exception as exc:  # keep inventory moving across odd workbooks
                    rows.append([vendor, folder, month_from_path(path), str(path), "", "", "", f"ERROR: {exc}"])

    with out_file.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([
            "vendor",
            "manual_folder",
            "month_folder",
            "workbook_path",
            "sheet_name",
            "likely_header_row",
            "header_score",
            "header_sample",
        ])
        writer.writerows(rows)
    print(f"Wrote {out_file.relative_to(REPO)} ({len(rows):,} rows)")


if __name__ == "__main__":
    main()
