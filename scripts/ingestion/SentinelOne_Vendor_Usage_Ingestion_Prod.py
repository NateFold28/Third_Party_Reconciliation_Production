"""Ingest SentinelOne ConnectWise usage XLSX files into Snowflake.

Source layout:
    <SOURCE_ROOT>/MM_MON_YYYY/ConnectWise Usage_*.xlsx

Each workbook holds a per-site paid-detail sheet in wide form (one column per
agent product). This script melts the wide agent columns into long rows and
resolves the canonical vendor product SKU (sku_match_group) in a single step
so the target table has the same grain and schema as all other vendor pipelines.

Canonical vendor usage schema (shared across all vendors):

    BILLING_MONTH, VENDOR, VENDOR_PARTNER_NAME, VENDOR_PRODUCT_SKU, MODIFIER,
    QUANTITY, UNIT_PRICE, AMOUNT, CURRENCY

VENDOR_PRODUCT_SKU resolution rules (baked in at ingestion, no downstream remap):
    * ``Total Active Agents per site`` rows: VENDOR_PRODUCT_SKU = Sku value
      (Complete / Control / Core). The raw "Total Active Agents" label is
      replaced by the tier name so downstream joins use the sku_match_group
      directly.
    * ``Data Retention`` rows: VENDOR_PRODUCT_SKU = "Data Retention - " + Retention Days
      (e.g. "Data Retention - 180 Days").
    * All other agent columns (Ranger, Purple AI, etc.): VENDOR_PRODUCT_SKU =
      cleaned product name (already the sku_match_group in the SKU map).

Semantics:
    * Vendor grand-total footer rows (blank Site Account Name / Sku) are filtered.
    * Partners whose agent columns sum to zero across every row are excluded.
    * UNIT_PRICE is loaded from the governed invoice-rate seed by
      VENDOR_PRODUCT_SKU / sku_match_group, and AMOUNT = QUANTITY * UNIT_PRICE.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import re
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Paths & Snowflake target
# ---------------------------------------------------------------------------

SENTINELONE_ROOT = Path(__file__).resolve().parents[2]
WORKSPACE_ROOT = SENTINELONE_ROOT.parents[2]
OUTPUT_DIR = SENTINELONE_ROOT / "outputs"
SKU_INVOICE_RATES_PATH = (
    SENTINELONE_ROOT
    / "seeds"
    / "sentinelone_sku_invoice_rates.csv"
)

DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\SentinelOne"
)

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
FQN = f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}"
TARGET_VENDOR = "SentinelOne"

# ---------------------------------------------------------------------------
# Vendor workbook contract
# ---------------------------------------------------------------------------

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)

EXPECTED_SHEET_NAME = "Rich CW Paid Site Detail Report"
HEADER_SEARCH_ROWS = 7           # look for the header within the top N rows
SCHEMA_MATCH_THRESHOLD = 3       # >= this many canonical headers must match

SITE_ACCOUNT_COL = "Site Account Name"
SKU_COL = "Sku"
CREATED_DATE_COL = "Created Date"
TOTAL_AGENTS_COL = "Total Active Agents per site"
RETENTION_COL = "Retention Days"

# Retention column labels drift across workbooks; anything starting with
# "retention" (case-insensitive, non-alphanumerics stripped) collapses to the
# canonical column name.
RETENTION_ALIASES: frozenset[str] = frozenset(
    {"retentiondays", "retentiondata", "retentiondescription", "retentiondesc", "retention"}
)

# Cleaned product-label used when the base-agent roll-up is emitted as a row.
TOTAL_AGENTS_PRODUCT_LABEL = "Total Active Agents"

# Agent columns that don't follow the "*Agent(s)*" naming convention.
KNOWN_AGENT_EXTRAS: tuple[str, ...] = ("Ranger AD Full", "Ranger AD Protect Full")

# Canonical vendor usage schema â€” shared across all third-party pipelines.
TEMPLATE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
)


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------

def _norm(value: object) -> str:
    """Lowercase and strip non-alphanumerics so headers match tolerantly."""
    return re.sub(r"[^a-z0-9]", "", str(value if value is not None else "").lower())


def _clean_product_name(col_name: object) -> str:
    """Convert an agent column header into a product/sku label.

    Examples:
        "Ranger Active Agents"          -> "Ranger"
        "Purple AI Agents"              -> "Purple AI"
        "Ranger AD Full"                -> "Ranger AD"
        "Ranger AD Protect Full"        -> "Ranger AD Protect"
        "Total Active Agents per site"  -> "Total Active Agents"
    """
    s = re.sub(r"\s+", " ", str(col_name).replace("\xa0", " ")).strip()
    if _norm(s) == _norm(TOTAL_AGENTS_COL):
        return TOTAL_AGENTS_PRODUCT_LABEL
    s = re.sub(r"\s+Active Agents$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+Agents$", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\s+Full$", "", s, flags=re.IGNORECASE)
    return s.strip()


def _to_first_of_month(value: object) -> dt.date | None:
    """Truncate a Created Date value to the first day of its month."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return dt.date(value.year, value.month, 1)
    parsed = pd.to_datetime(str(value), errors="coerce")
    if pd.isna(parsed):
        return None
    ts: pd.Timestamp = parsed  # type: ignore[assignment]
    return dt.date(ts.year, ts.month, 1)


def _sku_rate_key(vendor_product_sku: object) -> str:
    """Map the ingestion product label to the invoice-rate seed key."""
    product = str(vendor_product_sku if vendor_product_sku is not None else "").strip()
    product_upper = product.upper()
    if product_upper.startswith("DATA RETENTION"):
        match = re.search(r"\b(30|90|180|365)\b", product_upper)
        if match:
            return f"DATA_RETENTION_{match.group(1)}"
    return re.sub(r"_+", "_", re.sub(r"[^A-Z0-9]+", "_", product_upper)).strip("_")


def load_invoice_rate_map(seed_path: Path = SKU_INVOICE_RATES_PATH) -> dict[str, float]:
    """Return sku_match_group -> vendor invoice unit price from the governed seed."""
    if not seed_path.exists():
        raise FileNotFoundError(seed_path)

    rates: dict[str, float] = {}
    with open(seed_path, "r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            key = (row.get("SKU_MATCH_GROUP") or "").strip().upper()
            if not key:
                continue
            raw_rate = (row.get("VENDOR_INVOICE_UNIT_PRICE") or "").strip()
            if raw_rate:
                rates[key] = float(raw_rate)

    return rates


# ---------------------------------------------------------------------------
# Filesystem discovery
# ---------------------------------------------------------------------------

def discover_month_folders(source_root: Path) -> dict[str, Path]:
    """Return ``{YYYY-MM: path}`` for every ``MM_MON_YYYY`` folder under root."""
    folders: dict[str, Path] = {}
    for child in source_root.iterdir():
        if not child.is_dir():
            continue
        m = MONTH_FOLDER_RE.match(child.name)
        if m:
            folders[f"{m.group('yyyy')}-{m.group('mm')}"] = child
    return dict(sorted(folders.items()))


def locate_usage_file(month_folder: Path) -> Path | None:
    """Find the ``ConnectWise Usage_*.xlsx`` file, skipping Excel lock files."""
    for path in sorted(month_folder.glob("ConnectWise Usage_*.xlsx")):
        if not path.name.startswith("~$"):
            return path
    return None


def _read_file_bytes(path: Path) -> bytes:
    """Read the workbook, falling back to a PowerShell copy on OneDrive locks."""
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
            tmp_path = Path(handle.name)
        try:
            subprocess.run(
                [
                    "powershell", "-NoProfile", "-Command",
                    f"Copy-Item -LiteralPath '{path}' -Destination '{tmp_path}' -Force",
                ],
                check=True,
                capture_output=True,
            )
            return tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------

_REQUIRED_HEADERS = {
    _norm(SITE_ACCOUNT_COL),
    _norm(SKU_COL),
    _norm(CREATED_DATE_COL),
    _norm(TOTAL_AGENTS_COL),
}


def _find_usage_sheet(workbook: openpyxl.Workbook, path: Path) -> tuple[str, int, list[str]] | None:
    """Locate the sheet and header row that carry the CW usage schema."""
    sheet_order = [EXPECTED_SHEET_NAME] if EXPECTED_SHEET_NAME in workbook.sheetnames else []
    sheet_order += [name for name in workbook.sheetnames if name not in sheet_order]

    for sheet_name in sheet_order:
        ws = workbook[sheet_name]
        max_col = ws.max_column or 40
        for r in range(1, HEADER_SEARCH_ROWS + 1):
            headers = [
                (str(ws.cell(row=r, column=c).value).strip()
                 if ws.cell(row=r, column=c).value is not None else "")
                for c in range(1, max_col + 1)
            ]
            normalized = {_norm(h) for h in headers if h}
            if len(_REQUIRED_HEADERS & normalized) >= SCHEMA_MATCH_THRESHOLD:
                return sheet_name, r, headers

    print(f"SKIP {path.name}: no sheet matched required ConnectWise usage headers.")
    return None


def _canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename known columns to canonical names (whitespace/case tolerant)."""
    exact = {_norm(c): c for c in (SITE_ACCOUNT_COL, SKU_COL, CREATED_DATE_COL, TOTAL_AGENTS_COL)}
    rename_map: dict[str, str] = {}
    for col in df.columns:
        n = _norm(col)
        canonical = exact.get(n)
        # Retention aliases -- ignore columns whose label mentions "agent".
        if canonical is None and "agent" not in str(col).lower() and (
            n in RETENTION_ALIASES or n.startswith("retention")
        ):
            canonical = RETENTION_COL
        if canonical and col != canonical:
            rename_map[str(col)] = canonical
    return df.rename(columns=rename_map)


def _identify_agent_columns(columns: list[str]) -> list[str]:
    """Return all raw column names that represent per-site agent counts.

    Includes the ``Total Active Agents per site`` roll-up alongside per-product
    columns and the ``Ranger AD Full`` / ``Ranger AD Protect Full`` carve-outs.
    """
    excluded = {_norm(c) for c in (SITE_ACCOUNT_COL, SKU_COL, CREATED_DATE_COL, RETENTION_COL)}
    extras = {_norm(c) for c in KNOWN_AGENT_EXTRAS}

    agents: list[str] = []
    for h in columns:
        if not h:
            continue
        n = _norm(h)
        if n in excluded:
            continue
        if "agent" in str(h).lower() or n in extras:
            agents.append(h)
    return agents


def _read_workbook(path: Path) -> pd.DataFrame:
    """Read the CW usage sheet from ``path`` into a raw DataFrame."""
    wb = openpyxl.load_workbook(io.BytesIO(_read_file_bytes(path)), read_only=True, data_only=True)
    try:
        match = _find_usage_sheet(wb, path)
        if match is None:
            return pd.DataFrame()

        sheet_name, header_row, headers = match
        headers = [h if h else f"col_{i}" for i, h in enumerate(headers)]
        ws = wb[sheet_name]

        records: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(cell is None for cell in row):
                continue
            records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    finally:
        wb.close()

    return pd.DataFrame(records)


def _clean_text(series: pd.Series) -> pd.Series:
    return (
        series.where(series.notna(), "")
        .astype(str)
        .str.strip()
        .replace({"None": "", "nan": "", "NaN": ""})
    )


def parse_usage_workbook(path: Path) -> pd.DataFrame:
    """Parse a CW usage workbook into the ``SENTINELONE_USAGE`` template."""
    raw = _read_workbook(path)
    if raw.empty:
        return pd.DataFrame(columns=list(TEMPLATE_COLUMNS))

    df = _canonicalize_columns(raw)

    missing = [c for c in (SITE_ACCOUNT_COL, SKU_COL, CREATED_DATE_COL) if c not in df.columns]
    if missing:
        print(f"SKIP {path.name}: matched sheet but missing canonical columns: {missing}")
        return pd.DataFrame(columns=list(TEMPLATE_COLUMNS))

    # Retention column may be absent in older workbooks.
    if RETENTION_COL not in df.columns:
        df[RETENTION_COL] = ""

    for text_col in (SITE_ACCOUNT_COL, SKU_COL, RETENTION_COL):
        df[text_col] = _clean_text(df[text_col])

    df["BILLING_MONTH"] = df[CREATED_DATE_COL].apply(_to_first_of_month)

    # Drop vendor grand-total footer rows (blank Site/Sku) and any row missing
    # a billing month.
    df = df[
        df["BILLING_MONTH"].notna()
        & (df[SITE_ACCOUNT_COL] != "")
        & (df[SKU_COL] != "")
    ].copy()
    if df.empty:
        return pd.DataFrame(columns=list(TEMPLATE_COLUMNS))

    # Coerce all agent columns to numeric. Every agent column is retained --
    # the Total Active Agents roll-up is emitted alongside per-product cols.
    agent_cols = _identify_agent_columns(list(df.columns))
    for c in agent_cols:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Drop partners (Site Account Name) whose usage AGGREGATED across every
    # row and every agent column sums to zero. A partner only qualifies as
    # "no usage" if every single agent column on every single one of its rows
    # is zero -- individual zero rows on an otherwise-live partner are kept
    # so add-on-only patterns (e.g. Ranger without Total Active Agents on a
    # given row) are never silently discarded. Per-(site, sku, product) rows
    # with QUANTITY=0 are filtered later by the melt step, so nothing empty
    # is emitted downstream.
    site_totals = df.groupby(SITE_ACCOUNT_COL)[agent_cols].sum()
    live_sites = site_totals[site_totals.sum(axis=1) > 0].index
    pre_sites = df[SITE_ACCOUNT_COL].nunique()
    pre_rows = len(df)
    df = df[df[SITE_ACCOUNT_COL].isin(live_sites)].copy()
    dropped_sites = pre_sites - df[SITE_ACCOUNT_COL].nunique()
    dropped_rows = pre_rows - len(df)
    if dropped_sites:
        print(
            f"  Dropped {dropped_sites:,} partners (rows={dropped_rows:,}) "
            "with zero usage across every agent column."
        )
    if df.empty:
        return pd.DataFrame(columns=list(TEMPLATE_COLUMNS))

    id_cols = ["BILLING_MONTH", SITE_ACCOUNT_COL, SKU_COL, RETENTION_COL]
    melted = df[id_cols + agent_cols].melt(
        id_vars=id_cols,
        value_vars=agent_cols,
        var_name="RAW_PRODUCT",
        value_name="QUANTITY",
    )
    # Keep only strictly positive quantities so we emit:
    #   * one Total Active Agents row per partner+sku (when total > 0)
    #   * per-product add-on rows only where the module is enabled (> 0)
    melted = melted[melted["QUANTITY"] > 0].copy()
    if melted.empty:
        return pd.DataFrame(columns=list(TEMPLATE_COLUMNS))

    melted["_CLEANED_PRODUCT"] = melted["RAW_PRODUCT"].apply(_clean_product_name)

    # -----------------------------------------------------------------------
    # Resolve VENDOR_PRODUCT_SKU (= sku_match_group) at ingestion time.
    # This eliminates the downstream CASE logic that used ENTITY and
    # RETENTION_DESC to re-identify the billing group.
    # -----------------------------------------------------------------------
    is_total_agents = melted["_CLEANED_PRODUCT"] == TOTAL_AGENTS_PRODUCT_LABEL
    is_data_retention = melted["_CLEANED_PRODUCT"].str.contains("Data Retention", case=False, na=False)

    # Default: cleaned product name IS the sku_match_group (Ranger, Purple AI, etc.)
    melted["VENDOR_PRODUCT_SKU"] = melted["_CLEANED_PRODUCT"]

    # Total Active Agents -> use the SKU/entity value (Complete / Control / Core)
    melted.loc[is_total_agents, "VENDOR_PRODUCT_SKU"] = (
        melted.loc[is_total_agents, SKU_COL].astype(str).str.strip()
    )

    # Data Retention -> combine with retention days descriptor
    ret_filled = melted.loc[is_data_retention, RETENTION_COL].astype(str).str.strip().replace("", "Unknown")
    melted.loc[is_data_retention, "VENDOR_PRODUCT_SKU"] = "Data Retention - " + ret_filled

    agg = (
        melted.groupby(
            ["BILLING_MONTH", SITE_ACCOUNT_COL, "VENDOR_PRODUCT_SKU"],
            dropna=False,
        )
        .agg(QUANTITY=("QUANTITY", "sum"))
        .reset_index()
        .rename(columns={SITE_ACCOUNT_COL: "VENDOR_PARTNER_NAME"})
    )
    rates = load_invoice_rate_map()
    agg["VENDOR"] = "SentinelOne"
    agg["MODIFIER"] = None
    agg["UNIT_PRICE"] = agg["VENDOR_PRODUCT_SKU"].map(lambda value: rates.get(_sku_rate_key(value)))
    agg["AMOUNT"] = agg["QUANTITY"] * agg["UNIT_PRICE"]
    agg["CURRENCY"] = "USD"

    return agg[list(TEMPLATE_COLUMNS)].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def parse_month(source_root: Path, month: str) -> pd.DataFrame:
    """Parse a single ``YYYY-MM`` month folder and return the template frame."""
    month_folder = discover_month_folders(source_root).get(month)
    if month_folder is None:
        raise FileNotFoundError(f"No folder found for {month} under {source_root}")

    usage_path = locate_usage_file(month_folder)
    if usage_path is None:
        raise FileNotFoundError(f"No ConnectWise Usage*.xlsx found in {month_folder}")

    try:
        frame = parse_usage_workbook(usage_path)
    except BaseException as exc:
        if isinstance(exc, KeyboardInterrupt):
            # XML parser on a corrupt file can raise KeyboardInterrupt via
            # pyexpat signal handling. Log and skip rather than halt the run.
            print(
                f"[{month}] WARN: XML parse interrupted on {usage_path.name} "
                f"(likely corrupt/malformed XLSX). Skipping â€” replace source file."
            )
            return pd.DataFrame(columns=list(TEMPLATE_COLUMNS))
        raise

    if frame.empty:
        print(f"[{month}] {usage_path.name}: no rows produced.")
        return frame

    resolved = sorted(pd.to_datetime(frame["BILLING_MONTH"]).dt.date.astype(str).unique())
    print(
        f"[{month} folder] {usage_path.name}: "
        f"billing_month(s)={resolved}, rows={len(frame):,}, "
        f"quantity={frame['QUANTITY'].sum():,.0f}, "
        f"partners={frame['VENDOR_PARTNER_NAME'].nunique():,}, "
        f"products={frame['VENDOR_PRODUCT_SKU'].nunique():,}"
    )
    return frame


def write_audit(df: pd.DataFrame, label: str) -> Path:
    """Emit a per-month/product audit CSV alongside the load."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    audit = (
        df.groupby(["BILLING_MONTH", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(
            row_count=("VENDOR", "size"),
            quantity=("QUANTITY", "sum"),
            distinct_partners=("VENDOR_PARTNER_NAME", "nunique"),
        )
        .reset_index()
    )
    path = OUTPUT_DIR / f"sentinelone_usage_ingest_audit_{label}.csv"
    audit.to_csv(path, index=False, quoting=csv.QUOTE_MINIMAL)
    print(f"Wrote audit: {path}")
    return path


# ---------------------------------------------------------------------------
# Snowflake load
# ---------------------------------------------------------------------------

def snowflake_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {FQN} (
    BILLING_MONTH       DATE,
    VENDOR              VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU  VARCHAR,
    MODIFIER            VARCHAR,
    QUANTITY            NUMBER(18, 4),
    UNIT_PRICE          NUMBER(18, 6),
    AMOUNT              NUMBER(18, 4),
    CURRENCY            VARCHAR
);
"""


def load_snowflake(df: pd.DataFrame, *, reset: bool = False) -> None:
    """Idempotent load: skip billing months already present unless ``reset``."""
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from snowflake.connector.pandas_tools import write_pandas
    from TEMPLATES.Python.connection import get_snowflake_connection

    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")

        if reset:
            print(f"RESET: deleting existing {TARGET_VENDOR} rows from {FQN}.")
            cur.execute(
                f"DELETE FROM {FQN} WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )

        cur.execute(snowflake_ddl())

        cur.execute(
            f"SELECT DISTINCT BILLING_MONTH FROM {FQN} "
            "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
            (TARGET_VENDOR,),
        )
        existing = {
            (row[0].isoformat() if hasattr(row[0], "isoformat") else str(row[0]))
            for row in cur.fetchall()
        }
        incoming = sorted(pd.to_datetime(df["BILLING_MONTH"]).dt.date.astype(str).unique())
        new_months = [m for m in incoming if m not in existing]
        skipped = [m for m in incoming if m in existing]

        if skipped:
            print(f"Skipping months already loaded: {skipped}")
        if not new_months:
            print("Nothing to load; every month already exists.")
            return

        load_df = df[df["BILLING_MONTH"].astype(str).isin(new_months)].reset_index(drop=True)
        success, _chunks, rows, output = write_pandas(
            conn, load_df, TARGET_TABLE,
            database=TARGET_DATABASE, schema=TARGET_SCHEMA, quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"Snowflake write_pandas failed: {output}")
        conn.commit()
        print(f"Appended {rows:,} rows for months {new_months} into {FQN}.")
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Ingest SentinelOne monthly vendor usage into Snowflake.",
    )
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT),
                        help="Root folder containing MM_MON_YYYY month folders.")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--month", help="Billing month folder to load, e.g. 2026-05.")
    group.add_argument("--all-months", action="store_true", help="Load every discovered month.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and audit; skip Snowflake load.")
    parser.add_argument("--reset", action="store_true", help=f"Drop {TARGET_TABLE} before loading.")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()

    source_root = Path(args.source_root)
    if not source_root.exists():
        raise FileNotFoundError(f"Source root does not exist: {source_root}")

    months = list(discover_month_folders(source_root).keys()) if args.all_months else [args.month]
    frames = [parse_month(source_root, m) for m in months]
    skipped = [m for m, f in zip(months, frames) if f.empty]
    if skipped:
        print(f"WARNING: {len(skipped)} month(s) skipped (parse failed or no rows): {skipped}")
    non_empty = [f for f in frames if not f.empty]
    all_rows = (
        pd.concat(non_empty, ignore_index=True)
        if non_empty
        else pd.DataFrame(columns=list(TEMPLATE_COLUMNS))
    )

    if all_rows.empty:
        print("No rows produced. Nothing to write.")
        return

    print(
        f"TOTAL rows={len(all_rows):,}, "
        f"quantity={all_rows['QUANTITY'].sum():,.0f}, "
        f"partners={all_rows['VENDOR_PARTNER_NAME'].nunique():,}, "
        f"products={all_rows['VENDOR_PRODUCT_SKU'].nunique():,}"
    )

    label = "all_months" if args.all_months else args.month.replace("-", "_")
    write_audit(all_rows, label)

    if args.dry_run:
        print("Dry run complete. Snowflake load skipped.")
        return

    load_snowflake(all_rows, reset=args.reset)


if __name__ == "__main__":
    main()

