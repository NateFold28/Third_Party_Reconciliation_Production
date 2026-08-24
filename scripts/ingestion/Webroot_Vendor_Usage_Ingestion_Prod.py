"""Ingest Webroot Aggregator Order Details into Snowflake.

Business contract:
  * Webroot CW and Webroot CMS are handled as one vendor usage feed.
  * The only external vendor source ingested here is Aggregator Order Details.
  * Endpoint, DNS/SAT, Zuora export, overage, invoice, and manual recon files
    are evidence or billing-side inputs, not vendor usage inputs.
  * Zero-dollar usage rows are retained and flagged; printed subtotal rows are
    excluded from usage and captured in the file audit controls.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd


WEBROOT_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = WEBROOT_ROOT.parent
WORKSPACE_ROOT = PROJECT_ROOT.parents[1]

DEFAULT_SOURCE_ROOT_CW = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Webroot CW"
)
DEFAULT_SOURCE_ROOT_CMS = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Webroot CMS"
)
MAPPINGS_FILE = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\Mappings\Webroot\Webroot_Mappings.xlsx"
)

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
AUDIT_TABLE = "WEBROOT_USAGE_FILE_AUDIT"
PARTNER_SEED_TABLE = "WEBROOT_PARTNER_MAP_SEED"
TARGET_VENDOR = "Webroot"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)

SOURCE_COLUMNS: tuple[str, ...] = (
    "Billing Date",
    "Customer Code",
    "Company Name",
    "Reseller Customer Code",
    "Reseller Company Name",
    "Webroot Entity",
    "Location Code",
    "State",
    "Keycode",
    "License Type Description",
    "License Seats",
    "Usage Seats",
    "Total Seats",
    "Order Date",
    "License Category Name",
    "Keycode Age",
    "Retail Price",
    "Total Extended Amount",
    "Cap Amount",
    "Total Cap Amount",
)

USAGE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
)

AUDIT_COLUMNS: tuple[str, ...] = (
    "STREAM",
    "CHANNEL",
    "SOURCE_FOLDER",
    "SOURCE_FILE",
    "SOURCE_SHEET",
    "SOURCE_CONTENT_HASH",
    "LOAD_STATUS",
    "BILLING_MONTH",
    "BILLING_DATE",
    "LICENSE_CATEGORY_NAME",
    "DATA_ROW_COUNT",
    "CHARGEABLE_ROW_COUNT",
    "ZERO_AMOUNT_ROW_COUNT",
    "SUBTOTAL_ROW_COUNT",
    "RECOMPUTED_TOTAL_SEATS",
    "PRINTED_TOTAL_SEATS",
    "TOTAL_SEATS_DELTA",
    "RECOMPUTED_TOTAL_EXTENDED_AMOUNT",
    "PRINTED_TOTAL_EXTENDED_AMOUNT",
    "TOTAL_EXTENDED_AMOUNT_DELTA",
    "ERROR_MESSAGE",
    "INGESTED_AT",
)

PARTNER_SEED_COLUMNS: tuple[str, ...] = (
    "VENDOR",
    "PARTNER_NAME",
    "PARENT_COMPANY",
    "SF_ID",
    "CMS_ID",
    "ZUORA_NAME",
)

WEBROOT_PRODUCT_GROUPS = {
    "SAEP": "GSM",
    "SDNS": "DNS",
    "SECA": "SAT",
}


@dataclass(frozen=True)
class SourceFile:
    stream: str
    channel: str
    path: Path


def _read_file_bytes(path: Path) -> bytes:
    """Read bytes, falling back to PowerShell Copy-Item for locked OneDrive files."""
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir) / path.name
            escaped_src = str(path).replace("'", "''")
            escaped_dest = str(tmp).replace("'", "''")
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    f"Copy-Item -LiteralPath '{escaped_src}' -Destination '{escaped_dest}' -Force",
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            return tmp.read_bytes()


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).replace("\xa0", " ").strip()
    return text if text else None


def _to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("$", "").strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_date(value: object) -> dt.date | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return pd.to_datetime(value, errors="coerce").date() if not pd.isna(pd.to_datetime(value, errors="coerce")) else None


def _first_day(value: dt.date | None) -> dt.date | None:
    return dt.date(value.year, value.month, 1) if value else None


def _vendor_product_sku(value: object) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    return WEBROOT_PRODUCT_GROUPS.get(text.upper(), text)


def _month_from_folder(path: Path) -> str | None:
    match = MONTH_FOLDER_RE.match(path.parent.name)
    if not match:
        return None
    return f"{match.group('yyyy')}-{match.group('mm')}"


def _canonical_row(values: Iterable[object]) -> str:
    return "\t".join(_clean_text(v) or "" for v in values)


def _content_hash(rows: list[tuple[object, ...]]) -> str:
    payload = "\n".join(_canonical_row(row[: len(SOURCE_COLUMNS)]) for row in rows)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def discover_source_files(source_root_cw: Path, source_root_cms: Path) -> list[SourceFile]:
    files: list[SourceFile] = []
    for stream, root in (("CW", source_root_cw), ("CMS", source_root_cms)):
        if not root.exists():
            print(f"WARNING: source root does not exist: {root}", flush=True)
            continue
        for path in sorted(root.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            if "Aggregator Order Details" not in path.name:
                continue
            channel = "RESELLER" if "RESELLER" in path.name.upper() else "MSP"
            files.append(SourceFile(stream=stream, channel=channel, path=path))
    return files


def parse_aggregator_file(source: SourceFile, *, ingested_at: dt.datetime) -> tuple[pd.DataFrame, pd.DataFrame]:
    file_bytes = _read_file_bytes(source.path)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        row_iter = ws.iter_rows(values_only=True)
        raw_header = next(row_iter)
        header = tuple(_clean_text(v) for v in raw_header[: len(SOURCE_COLUMNS)])
        if header != SOURCE_COLUMNS:
            audit = pd.DataFrame(
                [
                    {
                        "STREAM": source.stream,
                        "CHANNEL": source.channel,
                        "SOURCE_FOLDER": source.path.parent.name,
                        "SOURCE_FILE": source.path.name,
                        "SOURCE_SHEET": sheet_name,
                        "SOURCE_CONTENT_HASH": hashlib.sha256(file_bytes).hexdigest(),
                        "LOAD_STATUS": "FAILED_SCHEMA",
                        "BILLING_MONTH": None,
                        "BILLING_DATE": None,
                        "LICENSE_CATEGORY_NAME": None,
                        "DATA_ROW_COUNT": 0,
                        "CHARGEABLE_ROW_COUNT": 0,
                        "ZERO_AMOUNT_ROW_COUNT": 0,
                        "SUBTOTAL_ROW_COUNT": 0,
                        "RECOMPUTED_TOTAL_SEATS": None,
                        "PRINTED_TOTAL_SEATS": None,
                        "TOTAL_SEATS_DELTA": None,
                        "RECOMPUTED_TOTAL_EXTENDED_AMOUNT": None,
                        "PRINTED_TOTAL_EXTENDED_AMOUNT": None,
                        "TOTAL_EXTENDED_AMOUNT_DELTA": None,
                        "ERROR_MESSAGE": f"Unexpected header: {header}",
                        "INGESTED_AT": ingested_at,
                    }
                ],
                columns=AUDIT_COLUMNS,
            )
            return pd.DataFrame(columns=USAGE_COLUMNS), audit

        data_records: list[dict[str, object]] = []
        subtotal_records: list[dict[str, object]] = []
        content_rows: list[tuple[object, ...]] = []
        for source_row_number, raw_row in enumerate(row_iter, start=2):
            row = tuple(raw_row[: len(SOURCE_COLUMNS)])
            cleaned = [_clean_text(v) for v in row]
            if not any(cleaned):
                continue
            content_rows.append(row)
            row_dict = dict(zip(SOURCE_COLUMNS, row, strict=True))
            customer_code = _clean_text(row_dict["Customer Code"])
            if customer_code and customer_code.lower() == "subtotal":
                subtotal_records.append(row_dict | {"source_row_number": source_row_number})
                continue
            data_records.append(row_dict | {"source_row_number": source_row_number})
    finally:
        wb.close()

    source_content_hash = _content_hash(content_rows)
    data_df = pd.DataFrame(data_records)
    subtotal_df = pd.DataFrame(subtotal_records)

    if data_df.empty:
        usage_df = pd.DataFrame(columns=USAGE_COLUMNS)
        audit_df = _build_audit(
            source=source,
            sheet_name=sheet_name,
            source_content_hash=source_content_hash,
            data_df=data_df,
            subtotal_df=subtotal_df,
            load_status="LOADED",
            error_message=None,
            ingested_at=ingested_at,
        )
        return usage_df, audit_df

    for col in (
        "License Seats",
        "Usage Seats",
        "Total Seats",
        "Keycode Age",
        "Retail Price",
        "Total Extended Amount",
        "Cap Amount",
        "Total Cap Amount",
    ):
        data_df[col] = data_df[col].map(_to_number)

    billing_dates = data_df["Billing Date"].map(_to_date)
    order_dates = data_df["Order Date"].map(_to_date)
    billing_months = billing_dates.map(_first_day)
    usage_df = pd.DataFrame(
        {
            "BILLING_MONTH": billing_months,
            "VENDOR": "Webroot",
            "VENDOR_PARTNER_NAME": data_df["Company Name"].map(_clean_text),
            "VENDOR_PRODUCT_SKU": data_df["License Category Name"].map(_vendor_product_sku),
            "MODIFIER": source.stream,
            "QUANTITY": data_df["Total Seats"],
            "UNIT_PRICE": data_df["Retail Price"],
            "AMOUNT": data_df["Total Extended Amount"],
            "CURRENCY": "USD",
        },
        columns=USAGE_COLUMNS,
    )
    audit_df = _build_audit(
        source=source,
        sheet_name=sheet_name,
        source_content_hash=source_content_hash,
        data_df=data_df,
        subtotal_df=subtotal_df,
        load_status="LOADED",
        error_message=None,
        ingested_at=ingested_at,
    )
    return usage_df, audit_df


def _build_audit(
    *,
    source: SourceFile,
    sheet_name: str,
    source_content_hash: str,
    data_df: pd.DataFrame,
    subtotal_df: pd.DataFrame,
    load_status: str,
    error_message: str | None,
    ingested_at: dt.datetime,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    all_categories = sorted(
        {
            *[c for c in data_df.get("License Category Name", pd.Series(dtype=object)).map(_clean_text).dropna().unique()],
            *[c for c in subtotal_df.get("License Category Name", pd.Series(dtype=object)).map(_clean_text).dropna().unique()],
        }
    )
    if not all_categories:
        all_categories = [None]

    working_data = data_df.copy()
    working_subtotal = subtotal_df.copy()
    for col in ("Total Seats", "Total Extended Amount"):
        if col in working_data:
            working_data[col] = working_data[col].map(_to_number)
        if col in working_subtotal:
            working_subtotal[col] = working_subtotal[col].map(_to_number)

    for category in all_categories:
        category_mask = (
            working_data["License Category Name"].map(_clean_text).eq(category)
            if not working_data.empty
            else pd.Series(dtype=bool)
        )
        subtotal_mask = (
            working_subtotal["License Category Name"].map(_clean_text).eq(category)
            if not working_subtotal.empty
            else pd.Series(dtype=bool)
        )
        category_data = working_data[category_mask] if not working_data.empty else working_data
        category_subtotals = working_subtotal[subtotal_mask] if not working_subtotal.empty else working_subtotal
        recomputed_seats = category_data["Total Seats"].sum() if "Total Seats" in category_data else None
        recomputed_amount = (
            category_data["Total Extended Amount"].sum() if "Total Extended Amount" in category_data else None
        )
        printed_seats = category_subtotals["Total Seats"].sum() if "Total Seats" in category_subtotals else None
        printed_amount = (
            category_subtotals["Total Extended Amount"].sum()
            if "Total Extended Amount" in category_subtotals
            else None
        )
        billing_date = None
        if not category_data.empty:
            billing_date = _to_date(category_data["Billing Date"].iloc[0])
        billing_month = _first_day(billing_date) or (
            dt.date.fromisoformat(f"{_month_from_folder(source.path)}-01") if _month_from_folder(source.path) else None
        )
        rows.append(
            {
                "STREAM": source.stream,
                "CHANNEL": source.channel,
                "SOURCE_FOLDER": source.path.parent.name,
                "SOURCE_FILE": source.path.name,
                "SOURCE_SHEET": sheet_name,
                "SOURCE_CONTENT_HASH": source_content_hash,
                "LOAD_STATUS": load_status,
                "BILLING_MONTH": billing_month,
                "BILLING_DATE": billing_date,
                "LICENSE_CATEGORY_NAME": category,
                "DATA_ROW_COUNT": int(len(category_data)),
                "CHARGEABLE_ROW_COUNT": int(
                    category_data["Total Extended Amount"].fillna(0).ne(0).sum()
                )
                if "Total Extended Amount" in category_data
                else 0,
                "ZERO_AMOUNT_ROW_COUNT": int(
                    category_data["Total Extended Amount"].fillna(0).eq(0).sum()
                )
                if "Total Extended Amount" in category_data
                else 0,
                "SUBTOTAL_ROW_COUNT": int(len(category_subtotals)),
                "RECOMPUTED_TOTAL_SEATS": recomputed_seats,
                "PRINTED_TOTAL_SEATS": printed_seats,
                "TOTAL_SEATS_DELTA": (recomputed_seats or 0) - (printed_seats or 0)
                if printed_seats is not None
                else None,
                "RECOMPUTED_TOTAL_EXTENDED_AMOUNT": recomputed_amount,
                "PRINTED_TOTAL_EXTENDED_AMOUNT": printed_amount,
                "TOTAL_EXTENDED_AMOUNT_DELTA": (recomputed_amount or 0) - (printed_amount or 0)
                if printed_amount is not None
                else None,
                "ERROR_MESSAGE": error_message,
                "INGESTED_AT": ingested_at,
            }
        )
    return pd.DataFrame(rows, columns=AUDIT_COLUMNS)


def build_usage(source_root_cw: Path, source_root_cms: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    ingested_at = dt.datetime.now(dt.UTC).replace(tzinfo=None)
    files = discover_source_files(source_root_cw, source_root_cms)
    if not files:
        raise FileNotFoundError(f"No Aggregator Order Details workbooks found in {source_root_cw} or {source_root_cms}")

    usage_frames: list[pd.DataFrame] = []
    audit_frames: list[pd.DataFrame] = []
    seen_content_hashes: set[str] = set()
    for source in files:
        usage_df, audit_df = parse_aggregator_file(source, ingested_at=ingested_at)
        content_hash = audit_df["SOURCE_CONTENT_HASH"].iloc[0]
        if content_hash in seen_content_hashes:
            audit_df["LOAD_STATUS"] = "SKIPPED_DUPLICATE_CONTENT"
            usage_df = pd.DataFrame(columns=USAGE_COLUMNS)
        elif not usage_df.empty:
            seen_content_hashes.add(content_hash)
        audit_frames.append(audit_df)
        usage_frames.append(usage_df)

        status = audit_df["LOAD_STATUS"].iloc[0]
        rows = len(usage_df)
        qty = usage_df["QUANTITY"].sum() if not usage_df.empty else 0
        amount = usage_df["AMOUNT"].sum() if not usage_df.empty else 0
        print(
            f"{status}: {source.stream} {source.channel} {source.path.parent.name} {source.path.name} "
            f"rows={rows:,} qty={qty:,.0f} amount={amount:,.2f}",
            flush=True,
        )

    non_empty_usage_frames = [frame for frame in usage_frames if not frame.empty]
    usage_all = (
        pd.concat(non_empty_usage_frames, ignore_index=True)
        if non_empty_usage_frames
        else pd.DataFrame(columns=USAGE_COLUMNS)
    )
    audit_all = pd.concat(audit_frames, ignore_index=True) if audit_frames else pd.DataFrame(columns=AUDIT_COLUMNS)
    if usage_all.empty:
        raise RuntimeError("No Webroot usage rows parsed from eligible source files.")
    usage_all = (
        usage_all.groupby(
            [
                "BILLING_MONTH",
                "VENDOR",
                "VENDOR_PARTNER_NAME",
                "VENDOR_PRODUCT_SKU",
                "MODIFIER",
                "UNIT_PRICE",
                "CURRENCY",
            ],
            dropna=False,
            as_index=False,
        )
        .agg(QUANTITY=("QUANTITY", "sum"), AMOUNT=("AMOUNT", "sum"))
    )
    usage_all = usage_all[
        [
            "BILLING_MONTH",
            "VENDOR",
            "VENDOR_PARTNER_NAME",
            "VENDOR_PRODUCT_SKU",
            "MODIFIER",
            "QUANTITY",
            "UNIT_PRICE",
            "AMOUNT",
            "CURRENCY",
        ]
    ]
    return usage_all, audit_all


def filter_usage_months(usage_df: pd.DataFrame, audit_df: pd.DataFrame, months: list[str] | None) -> tuple[pd.DataFrame, pd.DataFrame]:
    if not months:
        return usage_df, audit_df
    wanted = {dt.date.fromisoformat(f"{month}-01") for month in months}
    filtered_usage = usage_df[usage_df["BILLING_MONTH"].isin(wanted)].copy()
    filtered_audit = audit_df[audit_df["BILLING_MONTH"].isin(wanted)].copy()
    if filtered_usage.empty:
        raise RuntimeError(f"No usage rows found for requested month(s): {', '.join(months)}")
    return filtered_usage, filtered_audit


def usage_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} (
    BILLING_MONTH DATE,
    VENDOR VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU VARCHAR,
    MODIFIER VARCHAR,
    QUANTITY NUMBER(18,4),
    UNIT_PRICE NUMBER(18,6),
    AMOUNT NUMBER(18,4),
    CURRENCY VARCHAR
);
"""


def audit_ddl() -> str:
    return f"""
CREATE OR REPLACE TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} (
    STREAM VARCHAR,
    CHANNEL VARCHAR,
    SOURCE_FOLDER VARCHAR,
    SOURCE_FILE VARCHAR,
    SOURCE_SHEET VARCHAR,
    SOURCE_CONTENT_HASH VARCHAR,
    LOAD_STATUS VARCHAR,
    BILLING_MONTH DATE,
    BILLING_DATE DATE,
    LICENSE_CATEGORY_NAME VARCHAR,
    DATA_ROW_COUNT NUMBER,
    CHARGEABLE_ROW_COUNT NUMBER,
    ZERO_AMOUNT_ROW_COUNT NUMBER,
    SUBTOTAL_ROW_COUNT NUMBER,
    RECOMPUTED_TOTAL_SEATS NUMBER(18,4),
    PRINTED_TOTAL_SEATS NUMBER(18,4),
    TOTAL_SEATS_DELTA NUMBER(18,4),
    RECOMPUTED_TOTAL_EXTENDED_AMOUNT NUMBER(18,4),
    PRINTED_TOTAL_EXTENDED_AMOUNT NUMBER(18,4),
    TOTAL_EXTENDED_AMOUNT_DELTA NUMBER(18,4),
    ERROR_MESSAGE VARCHAR,
    INGESTED_AT TIMESTAMP_NTZ
);
"""


def _snowflake_connection():
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from TEMPLATES.Python.connection import get_snowflake_connection

    return get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )


def load_partner_seed() -> pd.DataFrame:
    if not MAPPINGS_FILE.exists():
        print(f"WARNING: mapping workbook not found: {MAPPINGS_FILE}", flush=True)
        return pd.DataFrame(columns=PARTNER_SEED_COLUMNS)

    file_bytes = _read_file_bytes(MAPPINGS_FILE)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet_name = "PARTNER_MAPPING" if "PARTNER_MAPPING" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        return pd.DataFrame(columns=PARTNER_SEED_COLUMNS)

    header = [_clean_text(value) for value in rows[0]]
    normalized_header = [re.sub(r"[^A-Z0-9]+", "_", value.upper()).strip("_") if value else "" for value in header]
    header_lookup = {name: idx for idx, name in enumerate(normalized_header)}

    def value_at(row: tuple[object, ...], *names: str) -> str | None:
        for name in names:
            idx = header_lookup.get(name)
            if idx is not None and idx < len(row):
                value = _clean_text(row[idx])
                if value is not None:
                    return value
        return None

    records: list[dict[str, object]] = []
    for source_row_number, row in enumerate(rows[1:], start=2):
        partner_name = value_at(row, "VENDOR_PARTNER_NAME", "ACCOUNT_NAME", "ACCOUNT_NAME")
        sf_id = value_at(row, "SF_ID", "SF_ID")
        cms_id = value_at(row, "CMS_ID", "CMS_ID")
        zuora_name = value_at(row, "ZUORA_NAME", "ZUORA_NAME")
        parent_company = value_at(row, "PARENT_CO", "PARENT_COMPANY")
        if not partner_name or not sf_id:
            continue
        records.append(
            {
                "VENDOR": "Webroot",
                "PARTNER_NAME": partner_name,
                "PARENT_COMPANY": parent_company,
                "SF_ID": sf_id,
                "CMS_ID": cms_id,
                "ZUORA_NAME": zuora_name,
            }
        )

    df = pd.DataFrame(records, columns=PARTNER_SEED_COLUMNS).drop_duplicates()
    print(f"Partner seed rows={len(df):,} distinct_sf_ids={df['SF_ID'].nunique():,}", flush=True)
    return df


def load_snowflake(usage_df: pd.DataFrame, audit_df: pd.DataFrame, partner_df: pd.DataFrame) -> None:
    from snowflake.connector.pandas_tools import write_pandas

    conn = _snowflake_connection()
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        cur.execute(usage_ddl())
        cur.execute(audit_ddl())
        cur.execute(
            f"""
CREATE OR REPLACE TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{PARTNER_SEED_TABLE} (
    VENDOR VARCHAR,
    PARTNER_NAME VARCHAR,
    PARENT_COMPANY VARCHAR,
    SF_ID VARCHAR,
    CMS_ID VARCHAR,
    ZUORA_NAME VARCHAR
);
"""
        )
        for df, table in (
            (usage_df, TARGET_TABLE),
            (audit_df, AUDIT_TABLE),
            (partner_df, PARTNER_SEED_TABLE),
        ):
            if df.empty:
                continue
            if table == TARGET_TABLE:
                incoming_months = sorted(
                    pd.to_datetime(df["BILLING_MONTH"]).dt.date.astype(str).unique().tolist()
                )
                month_list = ", ".join(f"'{month}'::DATE" for month in incoming_months)
                cur.execute(
                    f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                    f"WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s) "
                    f"AND BILLING_MONTH IN ({month_list})",
                    (TARGET_VENDOR,),
                )
            df = _fill_missing_prices(df, 'Webroot', conn=conn)
            success, _, rows, output = write_pandas(
                conn,
                df,
                table,
                database=TARGET_DATABASE,
                schema=TARGET_SCHEMA,
                quote_identifiers=False,
            )
            if not success:
                raise RuntimeError(f"write_pandas failed for {table}: {output}")
            print(f"Loaded {rows:,} rows into {table}.", flush=True)
        conn.commit()
    finally:
        conn.close()


def summarize_local(usage_df: pd.DataFrame, audit_df: pd.DataFrame) -> None:
    print("\nLOCAL USAGE CONTROLS", flush=True)
    summary = (
        usage_df.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(
            row_count=("QUANTITY", "size"),
            total_seats=("QUANTITY", "sum"),
            total_amount=("AMOUNT", "sum"),
        )
        .reset_index()
        .sort_values(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"])
    )
    print(summary.to_string(index=False), flush=True)

    bad_audit = audit_df[
        (audit_df["LOAD_STATUS"].eq("LOADED"))
        & (
            audit_df["TOTAL_SEATS_DELTA"].fillna(0).abs().gt(0.0001)
            | audit_df["TOTAL_EXTENDED_AMOUNT_DELTA"].fillna(0).abs().gt(0.0001)
        )
    ]
    if bad_audit.empty:
        print("Subtotal parity: PASS for loaded files.", flush=True)
    else:
        print("Subtotal parity: FAIL", flush=True)
        print(
            bad_audit[
                [
                    "STREAM",
                    "CHANNEL",
                    "SOURCE_FOLDER",
                    "SOURCE_FILE",
                    "LICENSE_CATEGORY_NAME",
                    "TOTAL_SEATS_DELTA",
                    "TOTAL_EXTENDED_AMOUNT_DELTA",
                ]
            ].to_string(index=False),
            flush=True,
        )


def validate_snowflake(local_usage_df: pd.DataFrame, local_audit_df: pd.DataFrame) -> None:
    conn = _snowflake_connection()
    try:
        sql = f"""
SELECT
    BILLING_MONTH,
    MODIFIER,
    VENDOR_PRODUCT_SKU,
    COUNT(*) AS row_count,
    SUM(QUANTITY) AS total_seats,
    SUM(AMOUNT) AS total_amount
FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}
WHERE VENDOR = 'Webroot'
GROUP BY 1,2,3
ORDER BY 1,2,3
"""
        snowflake_usage = pd.read_sql(sql, conn)
        snowflake_usage.columns = [col.lower() for col in snowflake_usage.columns]
        snowflake_usage = snowflake_usage.rename(
            columns={
                "billing_month": "BILLING_MONTH",
                "modifier": "MODIFIER",
                "vendor_product_sku": "VENDOR_PRODUCT_SKU",
            }
        )
        audit_sql = f"""
SELECT
    LOAD_STATUS,
    COUNT(DISTINCT SOURCE_FILE || '|' || SOURCE_FOLDER || '|' || SOURCE_CONTENT_HASH) AS file_count,
    SUM(DATA_ROW_COUNT) AS data_row_count,
    SUM(RECOMPUTED_TOTAL_SEATS) AS recomputed_total_seats,
    SUM(RECOMPUTED_TOTAL_EXTENDED_AMOUNT) AS recomputed_total_extended_amount,
    MAX(ABS(COALESCE(TOTAL_SEATS_DELTA, 0))) AS max_total_seats_delta,
    MAX(ABS(COALESCE(TOTAL_EXTENDED_AMOUNT_DELTA, 0))) AS max_total_extended_amount_delta
FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE}
GROUP BY 1
ORDER BY 1
"""
        snowflake_audit = pd.read_sql(audit_sql, conn)
        snowflake_audit.columns = [col.lower() for col in snowflake_audit.columns]
    finally:
        conn.close()

    local_usage = (
        local_usage_df.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(
            row_count=("QUANTITY", "size"),
            total_seats=("QUANTITY", "sum"),
            total_amount=("AMOUNT", "sum"),
        )
        .reset_index()
    )
    local_usage["BILLING_MONTH"] = pd.to_datetime(local_usage["BILLING_MONTH"])
    snowflake_usage["BILLING_MONTH"] = pd.to_datetime(snowflake_usage["BILLING_MONTH"])
    compare = local_usage.merge(
        snowflake_usage,
        on=["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"],
        how="outer",
        suffixes=("_local", "_snowflake"),
    )
    for col in ("row_count", "total_seats", "total_amount"):
        compare[f"{col}_delta"] = compare[f"{col}_snowflake"].fillna(0) - compare[f"{col}_local"].fillna(0)

    print("\nSNOWFLAKE LOAD PARITY", flush=True)
    print(
        compare[
            [
                "BILLING_MONTH",
                "MODIFIER",
                "VENDOR_PRODUCT_SKU",
                "row_count_delta",
                "total_seats_delta",
                "total_amount_delta",
            ]
        ]
        .sort_values(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"])
        .to_string(index=False),
        flush=True,
    )
    max_abs_delta = compare[
        ["row_count_delta", "total_seats_delta", "total_amount_delta"]
    ].abs().max().max()
    if max_abs_delta > 0.0001:
        raise RuntimeError("Snowflake usage parity failed.")

    print("\nSNOWFLAKE FILE AUDIT", flush=True)
    print(snowflake_audit.to_string(index=False), flush=True)



# ---------------------------------------------------------------------------
# Dynamic invoice rate fill (universal safety net)
# ---------------------------------------------------------------------------
def _fill_missing_prices(df, vendor_name, conn=None):
    """Fill NULL UNIT_PRICE/AMOUNT rows from THIRD_PARTY_RECON_VENDOR_INVOICES.

    Only fires on rows where both UNIT_PRICE and AMOUNT are NULL/0.
    Rows already populated by the source file are never overwritten.
    Exact (billing_month, sku) match first, then carry-forward to most
    recent prior month for same SKU.
    """
    import sys as _sys, pandas as _pd
    _ws = Path(__file__).resolve()
    for _ in range(8):
        _ws = _ws.parent
        if (_ws / "TEMPLATES").exists():
            break
    _sys.path.insert(0, str(_ws))
    try:
        _owned = conn is None
        if _owned:
            from TEMPLATES.Python.connection import get_snowflake_connection as _gc
            _c = _gc(role="DEVELOPER", warehouse="REPORTING_WH",
                     database="ANALYTICS_DEV", schema="DBT_NFOLD_TRANSFORMATION")
        else:
            _c = conn
        _rows = _c.cursor().execute(
            "SELECT BILLING_MONTH, VENDOR_PRODUCT_SKU, AVG(UNIT_PRICE) AS UNIT_PRICE "
            "FROM ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_INVOICES "
            "WHERE VENDOR ILIKE %s AND UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0 "
            "GROUP BY 1, 2 ORDER BY 1 DESC, 2",
            (f"%{vendor_name}%",),
        ).fetchall()
        if _owned:
            _c.close()
    except Exception as _e:
        print(f"[WARN] _fill_missing_prices: {_e}. Prices unchanged.", flush=True)
        return df
    if not _rows:
        return df
    inv = _pd.DataFrame(_rows, columns=["BILLING_MONTH", "VENDOR_PRODUCT_SKU", "UNIT_PRICE"])
    inv["BILLING_MONTH"] = _pd.to_datetime(inv["BILLING_MONTH"]).dt.normalize()
    inv["VENDOR_PRODUCT_SKU"] = inv["VENDOR_PRODUCT_SKU"].astype(str).str.strip()
    inv["UNIT_PRICE"] = _pd.to_numeric(inv["UNIT_PRICE"], errors="coerce")
    inv = inv.dropna(subset=["UNIT_PRICE"])
    exact = {(r.VENDOR_PRODUCT_SKU, r.BILLING_MONTH): float(r.UNIT_PRICE) for r in inv.itertuples(index=False)}
    latest: dict = {}
    for r in inv.itertuples(index=False):
        if r.VENDOR_PRODUCT_SKU not in latest:
            latest[r.VENDOR_PRODUCT_SKU] = float(r.UNIT_PRICE)
    df = df.copy()
    for col in ("UNIT_PRICE", "AMOUNT"):
        if col not in df.columns:
            df[col] = None
    # Normalise for comparison only - do NOT reassign BILLING_MONTH (would break write_pandas DATE cast)
    _billing_norm = _pd.to_datetime(df["BILLING_MONTH"]).dt.normalize()
    df["VENDOR_PRODUCT_SKU"] = df["VENDOR_PRODUCT_SKU"].astype(str).str.strip()
    mask = ((df["UNIT_PRICE"].isna() | (df["UNIT_PRICE"] == 0)) &
            (df["AMOUNT"].isna() | (df["AMOUNT"] == 0)))
    filled = 0
    for _i, idx in enumerate(df[mask].index):
        sku = df.at[idx, "VENDOR_PRODUCT_SKU"]
        mo  = _billing_norm.iloc[df.index.get_loc(idx)]
        price = exact.get((sku, mo))
        if price is None:
            bm, bp = None, None
            for (s, m), p in exact.items():
                if s == sku and m <= mo and (bm is None or m > bm):
                    bm, bp = m, p
            price = bp or latest.get(sku)
        if price:
            df.at[idx, "UNIT_PRICE"] = price
            qty = df.at[idx, "QUANTITY"] if "QUANTITY" in df.columns else None
            if qty is not None and not _pd.isna(qty):
                df.at[idx, "AMOUNT"] = float(qty) * price
            filled += 1
    if filled:
        print(f"[INFO] _fill_missing_prices: filled {filled:,} rows for {vendor_name}.", flush=True)
    return df

def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Webroot vendor usage into Snowflake.")
    parser.add_argument("--source-root-cw", default=str(DEFAULT_SOURCE_ROOT_CW))
    parser.add_argument("--source-root-cms", default=str(DEFAULT_SOURCE_ROOT_CMS))
    parser.add_argument("--month", action="append", help="Billing month YYYY-MM. Can be supplied more than once.")
    parser.add_argument("--all-months", action="store_true", help="Load all discovered months.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and summarize without loading Snowflake.")
    parser.add_argument("--reset", action="store_true", help="Compatibility flag; loads always replace Webroot usage tables.")
    parser.add_argument("--skip-partner-map", action="store_true", help="Do not load WEBROOT_PARTNER_MAP_SEED.")
    parser.add_argument("--skip-snowflake-validation", action="store_true", help="Do not compare local controls to Snowflake.")
    args = parser.parse_args()

    if not args.all_months and not args.month:
        raise SystemExit("Provide --month YYYY-MM or --all-months")

    usage_df, audit_df = build_usage(Path(args.source_root_cw), Path(args.source_root_cms))
    months = None if args.all_months else args.month
    usage_df, audit_df = filter_usage_months(usage_df, audit_df, months)
    partner_df = pd.DataFrame(columns=PARTNER_SEED_COLUMNS) if args.skip_partner_map else load_partner_seed()
    summarize_local(usage_df, audit_df)

    if args.dry_run:
        print("Dry run complete.", flush=True)
        return

    load_snowflake(usage_df, audit_df, partner_df)
    if not args.skip_snowflake_validation:
        validate_snowflake(usage_df, audit_df)


if __name__ == "__main__":
    main()

