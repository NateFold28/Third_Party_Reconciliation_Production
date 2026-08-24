"""Ingest Proofpoint APAC + LLC monthly usage files and load a normalized template into Snowflake.

Source layout (default):
    <SOURCE_ROOT>/MM_MON_YYYY/Connectwise APAC.xlsx
    <SOURCE_ROOT>/MM_MON_YYYY/ConnectWise, LLC.xlsx

Each workbook has its detail header at row 16.

Canonical vendor usage schema (shared across all vendors):

    BILLING_MONTH, VENDOR, VENDOR_PARTNER_NAME, VENDOR_PRODUCT_SKU,
    MODIFIER, QUANTITY, UNIT_PRICE, AMOUNT, CURRENCY

Appends new data to growing table.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd


PROOFPOINT_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = PROOFPOINT_ROOT.parent
WORKSPACE_ROOT = PROJECT_ROOT.parents[1]
# Aligned with the unified third-party recon vendor drop location so every
# vendor pipeline pulls raw files from the same OneDrive tree.
DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Proofpoint"
)
OUTPUT_DIR = PROOFPOINT_ROOT / "outputs"

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
TARGET_VENDOR = "Proofpoint"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)

# 1-based row that contains detail headers in every workbook the vendor sends.
HEADER_ROW = 16

# Minimum count of REQUIRED_HEADER_KEYS that must be present in row 16 for the
# file to be treated as a Proofpoint usage export. Anything below this bar is
# skipped (e.g. summary sheets, unrelated workbooks a partner drops in the folder).
SCHEMA_MATCH_THRESHOLD = 12

# Canonical extract-column name -> normalized vendor header (alphanumeric lowercase).
COLUMN_MAP: dict[str, str] = {
    "parents_parents_name": "parentsparentname",   # PARENT'S PARENT NAME  (grandparent)
    "parent_name": "parentname",                   # PARENT NAME
    "customer": "customer",
    "customer_type": "customertype",
    "active_users": "activeusers",
    "licensed_users": "licensedusers",
    "billed_users": "billedusers",
    "package": "package",
    "start": "start",
    "activation_date": "activationdate",
    "renewal_date": "renewal",
    "date_deactivated": "datedeactivated",
    "currency": "currency",
    "unit_value": "unitvalue",
    "monthly_total": "monthlytotal",
    "nfr_allocation": "nfrallocation",
}

# The subset of COLUMN_MAP values a workbook must expose in row 16 to be
# considered a valid Proofpoint usage export.
REQUIRED_HEADER_KEYS: frozenset[str] = frozenset(COLUMN_MAP.values())

# When the candidate partner name contains any of these tokens (ILIKE),
# fall through to the next level of the recursion (grandparent -> parent -> customer).
VENDOR_PARTNER_EXCLUSION_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"^proofpoint( essentials)?", re.IGNORECASE),
    re.compile(r"connectwise", re.IGNORECASE),
    re.compile(r"cw-", re.IGNORECASE),
)

# Canonical vendor usage schema â€” shared across all third-party pipelines.
TEMPLATE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
)


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------
def normalize(value: object) -> str:
    """Collapse a header cell to its alphanumeric-lowercase form for matching."""
    return re.sub(r"[^a-z0-9]", "", str(value if value is not None else "").strip().lower())


def discover_month_folders(source_root: Path) -> dict[str, Path]:
    """Return {'YYYY-MM': Path} for every MM_MON_YYYY sub-folder under ``source_root``."""
    folders: dict[str, Path] = {}
    for child in source_root.iterdir():
        if not child.is_dir():
            continue
        match = MONTH_FOLDER_RE.match(child.name)
        if not match:
            continue
        folders[f"{match.group('yyyy')}-{match.group('mm')}"] = child
    return dict(sorted(folders.items()))


def locate_raw_files(month_folder: Path) -> list[Path]:
    """Return every candidate ``.xlsx`` in a month folder (schema is validated later)."""
    return [
        path
        for path in sorted(month_folder.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]


# ---------------------------------------------------------------------------
# Workbook I/O
# ---------------------------------------------------------------------------
def _read_file_bytes(path: Path) -> bytes:
    """Read a workbook, falling back to a PowerShell copy if OneDrive holds an exclusive lock."""
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
            tmp_path = Path(handle.name)
        try:
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    f"Copy-Item -LiteralPath '{path}' -Destination '{tmp_path}' -Force",
                ],
                check=True,
                capture_output=True,
            )
            return tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)


def load_worksheet_records(path: Path) -> tuple[str, list[dict[str, object]]] | None:
    """Return ``(sheet_name, records)`` if ``path`` matches the Proofpoint schema, else ``None``.

    A file is considered a match when row ``HEADER_ROW`` exposes at least
    ``SCHEMA_MATCH_THRESHOLD`` of the expected header keys. Non-matching files
    are logged and skipped, so a partner accidentally dropping an unrelated
    workbook in the month folder will never poison the load.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(_read_file_bytes(path)), read_only=True, data_only=True)
    try:
        sheet_name = workbook.sheetnames[0]
        ws = workbook[sheet_name]

        max_col = ws.max_column or len(COLUMN_MAP) * 4
        header_cells = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, max_col + 1)]
        header_index = {normalize(v): idx for idx, v in enumerate(header_cells) if v is not None}

        matched_keys = REQUIRED_HEADER_KEYS & header_index.keys()
        if len(matched_keys) < SCHEMA_MATCH_THRESHOLD:
            print(
                f"SKIP {path.name}: only {len(matched_keys)}/{len(REQUIRED_HEADER_KEYS)} "
                f"expected headers found on row {HEADER_ROW} (threshold={SCHEMA_MATCH_THRESHOLD})."
            )
            return None

        records: list[dict[str, object]] = []
        for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
            if not row or all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
                continue
            record = {
                canonical: (
                    row[header_index[key]]
                    if key in header_index and header_index[key] < len(row)
                    else None
                )
                for canonical, key in COLUMN_MAP.items()
            }
            records.append(record)
        return sheet_name, records
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Template shaping
# ---------------------------------------------------------------------------
def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_excluded_partner(value: object) -> bool:
    text = _clean_text(value)
    if text is None:
        return True
    return any(pattern.search(text) for pattern in VENDOR_PARTNER_EXCLUSION_PATTERNS)


def resolve_vendor_partner_name(record: dict[str, object]) -> str | None:
    """Grandparent -> parent -> customer recursion, skipping Proofpoint/ConnectWise/CW- values."""
    for column in ("parents_parents_name", "parent_name"):
        if not _is_excluded_partner(record.get(column)):
            return _clean_text(record.get(column))
    return _clean_text(record.get("customer"))


def build_template_frame(records: Iterable[dict[str, object]], billing_month: dt.date) -> pd.DataFrame:
    rows = [
        {
            "BILLING_MONTH": billing_month,
            "VENDOR": "Proofpoint",
            "VENDOR_PARTNER_NAME": resolve_vendor_partner_name(record),
            "VENDOR_PRODUCT_SKU": _clean_text(record.get("package")),
            "MODIFIER": None,
            "QUANTITY": record.get("billed_users"),
            "UNIT_PRICE": record.get("unit_value"),
            "AMOUNT": record.get("monthly_total"),
            "CURRENCY": _clean_text(record.get("currency")),
        }
        for record in records
    ]
    df = pd.DataFrame(rows, columns=list(TEMPLATE_COLUMNS))
    for numeric_col in ("QUANTITY", "UNIT_PRICE", "AMOUNT"):
        df[numeric_col] = pd.to_numeric(df[numeric_col], errors="coerce")
    return df


def parse_month(source_root: Path, month: str) -> pd.DataFrame:
    folders = discover_month_folders(source_root)
    month_folder = folders.get(month)
    if month_folder is None:
        raise FileNotFoundError(f"No folder found for {month} under {source_root}")

    raw_files = locate_raw_files(month_folder)
    if not raw_files:
        raise FileNotFoundError(f"No APAC/LLC Proofpoint raw .xlsx files found in {month_folder}")

    billing_month = dt.date.fromisoformat(f"{month}-01")
    frames: list[pd.DataFrame] = []
    for path in raw_files:
        result = load_worksheet_records(path)
        if result is None:
            continue
        sheet_name, records = result
        if not records:
            print(f"[{month}] {path.name}: no data rows below header {HEADER_ROW}; skipping.")
            continue
        frame = build_template_frame(records, billing_month)
        frames.append(frame)
        print(
            f"[{month}] {path.name} sheet='{sheet_name}': "
            f"rows={len(frame):,}, quantity={frame['QUANTITY'].sum(skipna=True):,.2f}, "
            f"amount={frame['AMOUNT'].sum(skipna=True):,.2f}"
        )
    if not frames:
        raise RuntimeError(f"No detail rows parsed for {month}")
    return pd.concat(frames, ignore_index=True)


# ---------------------------------------------------------------------------
# Snowflake
# ---------------------------------------------------------------------------
def snowflake_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} (
    BILLING_MONTH   DATE,
    VENDOR          VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU  VARCHAR,
    MODIFIER        VARCHAR,
    QUANTITY        NUMBER(18, 4),
    UNIT_PRICE      NUMBER(18, 6),
    AMOUNT          NUMBER(18, 4),
    CURRENCY        VARCHAR
);
"""


def load_snowflake(df: pd.DataFrame, *, reset: bool = False) -> None:
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from snowflake.connector.pandas_tools import write_pandas
    from TEMPLATES.Python.connection import get_snowflake_connection

    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")

        if reset:
            print(
                f"RESET: deleting existing {TARGET_VENDOR} rows from "
                f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}."
            )
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )

        cur.execute(snowflake_ddl())

        cur.execute(
            f"SELECT DISTINCT BILLING_MONTH FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
            "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
            (TARGET_VENDOR,),
        )
        existing_months = {
            (row[0].isoformat() if hasattr(row[0], "isoformat") else str(row[0]))
            for row in cur.fetchall()
        }

        incoming_months = sorted(
            pd.to_datetime(df["BILLING_MONTH"]).dt.date.astype(str).unique().tolist()
        )
        new_months = [month for month in incoming_months if month not in existing_months]
        skipped_months = [month for month in incoming_months if month in existing_months]

        if skipped_months:
            print(f"Skipping months already loaded: {skipped_months}")
        if not new_months:
            print("Nothing to load; every month in the batch already exists in the target table.")
            return

        load_df = df[df["BILLING_MONTH"].astype(str).isin(new_months)].reset_index(drop=True)
        load_df = _fill_missing_prices(load_df, 'Proofpoint', conn=conn)
        success, chunks, rows, output = write_pandas(
            conn,
            load_df,
            TARGET_TABLE,
            database=TARGET_DATABASE,
            schema=TARGET_SCHEMA,
            quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"Snowflake write_pandas failed: {output}")
        conn.commit()
        print(
            f"Appended {rows:,} rows for months {new_months} into "
            f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} in {chunks} chunk(s)."
        )
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Local audit
# ---------------------------------------------------------------------------
def write_audit(df: pd.DataFrame, label: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    audit = (
        df.groupby(["BILLING_MONTH", "CURRENCY"], dropna=False)
        .agg(
            row_count=("VENDOR", "size"),
            quantity=("QUANTITY", "sum"),
            amount=("AMOUNT", "sum"),
            distinct_partners=("VENDOR_PARTNER_NAME", "nunique"),
            distinct_packages=("VENDOR_PRODUCT_SKU", "nunique"),
        )
        .reset_index()
    )
    path = OUTPUT_DIR / f"proofpoint_template_ingest_audit_{label}.csv"
    audit.to_csv(path, index=False, quoting=csv.QUOTE_MINIMAL)
    print(f"Wrote audit: {path}")
    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Dynamic invoice rate fill (universal safety net)
# ---------------------------------------------------------------------------
def _fill_missing_prices(df, vendor_name, conn=None):
    """Fill NULL UNIT_PRICE/AMOUNT rows from THIRD_PARTY_RECON_VENDOR_INVOICES.

    Only fires on rows where both UNIT_PRICE and AMOUNT are NULL/0.
    Rows already populated by the source file are never overwritten.
    Exact (billing_month, sku) match first, then carry-forward to most
    recent prior month for same SKU.
    """
    import sys as _sys, pandas as _pd
    _ws = Path(__file__).resolve()
    for _ in range(8):
        _ws = _ws.parent
        if (_ws / "TEMPLATES").exists():
            break
    _sys.path.insert(0, str(_ws))
    try:
        _owned = conn is None
        if _owned:
            from TEMPLATES.Python.connection import get_snowflake_connection as _gc
            _c = _gc(role="DEVELOPER", warehouse="REPORTING_WH",
                     database="ANALYTICS_DEV", schema="DBT_NFOLD_TRANSFORMATION")
        else:
            _c = conn
        _rows = _c.cursor().execute(
            "SELECT BILLING_MONTH, VENDOR_PRODUCT_SKU, AVG(UNIT_PRICE) AS UNIT_PRICE "
            "FROM ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_INVOICES "
            "WHERE VENDOR ILIKE %s AND UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0 "
            "GROUP BY 1, 2 ORDER BY 1 DESC, 2",
            (f"%{vendor_name}%",),
        ).fetchall()
        if _owned:
            _c.close()
    except Exception as _e:
        print(f"[WARN] _fill_missing_prices: {_e}. Prices unchanged.", flush=True)
        return df
    if not _rows:
        return df
    inv = _pd.DataFrame(_rows, columns=["BILLING_MONTH", "VENDOR_PRODUCT_SKU", "UNIT_PRICE"])
    inv["BILLING_MONTH"] = _pd.to_datetime(inv["BILLING_MONTH"]).dt.normalize()
    inv["VENDOR_PRODUCT_SKU"] = inv["VENDOR_PRODUCT_SKU"].astype(str).str.strip()
    inv["UNIT_PRICE"] = _pd.to_numeric(inv["UNIT_PRICE"], errors="coerce")
    inv = inv.dropna(subset=["UNIT_PRICE"])
    exact = {(r.VENDOR_PRODUCT_SKU, r.BILLING_MONTH): float(r.UNIT_PRICE) for r in inv.itertuples(index=False)}
    latest: dict = {}
    for r in inv.itertuples(index=False):
        if r.VENDOR_PRODUCT_SKU not in latest:
            latest[r.VENDOR_PRODUCT_SKU] = float(r.UNIT_PRICE)
    df = df.copy()
    for col in ("UNIT_PRICE", "AMOUNT"):
        if col not in df.columns:
            df[col] = None
    # Normalise for comparison only - do NOT reassign BILLING_MONTH (would break write_pandas DATE cast)
    _billing_norm = _pd.to_datetime(df["BILLING_MONTH"]).dt.normalize()
    df["VENDOR_PRODUCT_SKU"] = df["VENDOR_PRODUCT_SKU"].astype(str).str.strip()
    mask = ((df["UNIT_PRICE"].isna() | (df["UNIT_PRICE"] == 0)) &
            (df["AMOUNT"].isna() | (df["AMOUNT"] == 0)))
    filled = 0
    for _i, idx in enumerate(df[mask].index):
        sku = df.at[idx, "VENDOR_PRODUCT_SKU"]
        mo  = _billing_norm.iloc[df.index.get_loc(idx)]
        price = exact.get((sku, mo))
        if price is None:
            bm, bp = None, None
            for (s, m), p in exact.items():
                if s == sku and m <= mo and (bm is None or m > bm):
                    bm, bp = m, p
            price = bp or latest.get(sku)
        if price:
            df.at[idx, "UNIT_PRICE"] = price
            qty = df.at[idx, "QUANTITY"] if "QUANTITY" in df.columns else None
            if qty is not None and not _pd.isna(qty):
                df.at[idx, "AMOUNT"] = float(qty) * price
            filled += 1
    if filled:
        print(f"[INFO] _fill_missing_prices: filled {filled:,} rows for {vendor_name}.", flush=True)
    return df

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingest Proofpoint APAC/LLC monthly usage files and load a normalized template into Snowflake."
    )
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT), help="Local OneDrive/SharePoint synced root.")
    parser.add_argument("--month", help="Billing month in YYYY-MM format.")
    parser.add_argument("--all-months", action="store_true", help="Parse and load every month folder found.")
    parser.add_argument("--dry-run", action="store_true", help="Parse + audit locally without loading Snowflake.")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="One-time: DROP the target table and recreate it with the template schema before loading.",
    )
    args = parser.parse_args()

    source_root = Path(args.source_root)
    if not source_root.exists():
        raise FileNotFoundError(f"Source root does not exist: {source_root}")
    if not args.month and not args.all_months:
        raise SystemExit("Provide --month YYYY-MM or --all-months")

    months = list(discover_month_folders(source_root).keys()) if args.all_months else [args.month]
    all_rows = pd.concat([parse_month(source_root, month) for month in months], ignore_index=True)

    print(
        f"TOTAL rows={len(all_rows):,}, "
        f"quantity={all_rows['QUANTITY'].sum(skipna=True):,.2f}, "
        f"amount={all_rows['AMOUNT'].sum(skipna=True):,.2f}"
    )
    label = "all_months" if args.all_months else args.month.replace("-", "_")
    write_audit(all_rows, label)

    if args.dry_run:
        print("Dry run complete. Snowflake load skipped.")
        return

    load_snowflake(all_rows, reset=args.reset)


if __name__ == "__main__":
    main()

