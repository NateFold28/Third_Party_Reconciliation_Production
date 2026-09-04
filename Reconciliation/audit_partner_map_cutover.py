"""Audit governed partner-map changes and manual-vs-pipeline clear rates.

This is a read-only audit. It compares the pre-partner-audit Snowflake snapshot
with the Batch 5 staged detail, reads the available June 2026 manual
reconciliation workbooks, and exports cutover evidence. It never publishes or
mutates Snowflake objects.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import sys
from typing import Iterable

import openpyxl
import pandas as pd


def find_workspace_root(start: Path) -> Path:
    for candidate in (start, *start.parents):
        if (candidate / "TEMPLATES").exists():
            return candidate
    raise RuntimeError("Could not find workspace root containing TEMPLATES")


SCRIPT_PATH = Path(__file__).resolve()
PIPELINE_ROOT = SCRIPT_PATH.parents[1]
WORKSPACE_ROOT = find_workspace_root(SCRIPT_PATH)
OUTPUT_DIR = PIPELINE_ROOT / "output" / "partner_map_cutover_audit_20260903"
MANUAL_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files"
)

sys.path.insert(0, str(WORKSPACE_ROOT))
from TEMPLATES.Python.connection import fetch_dataframe, get_snowflake_connection  # noqa: E402

BASELINE_DETAIL = "THIRD_PARTY_RECON_DETAIL_PROD__PARTNER_AUDIT_BASELINE_20260903"
STAGED_DETAIL = "THIRD_PARTY_RECON_DETAIL_STAGED_CUTOVER_20260903"
BASELINE_MAP = "THIRD_PARTY_RECON_PARTNER_MAP_PROD__PARTNER_AUDIT_BASELINE_20260903"
CURRENT_MAP = "THIRD_PARTY_RECON_PARTNER_MAP_PROD"
NAME_NORM = (
    "TRIM(REGEXP_REPLACE(REGEXP_REPLACE(LOWER({column}), "
    "'[^a-z0-9]+', ' '), '\\\\s+', ' '))"
)


@dataclass(frozen=True)
class ManualWorkbook:
    vendor: str
    lane: str
    relative_path: str
    sheet: str
    key_column: str
    status_column: str
    activity_columns: tuple[str, ...]
    product_families: tuple[tuple[str, tuple[str, ...]], ...] = ()


MANUAL_WORKBOOKS = (
    ManualWorkbook(
        "Auvik",
        "CMS",
        r"Auvik CMS\06_JUN_2026\Auvik CMS Recon June'26.xlsx",
        "Data",
        "Account Name",
        "Comments",
        (),
        (
            ("Billable", ("Auvik Billable Qty", "Auvik Billable Amount", "CW Billable Qty", "CW Billable Amt")),
            ("Performance", ("Auvik Performance Qty", "Auvik Performance Amount", "CW Performance Qty", "CW Performance Amt")),
            ("ASM", ("Auvik ASM Qty.", "Auvik ASM Amount", "CW ASM Qty", "CW ASM Amt.")),
        ),
    ),
    ManualWorkbook(
        "Auvik",
        "CW",
        r"Auvik CW\06_JUN_2026\Auvik CW reconciliation June'26.xlsx",
        "DATA",
        "Account Name",
        "Comments",
        (),
        (
            ("Billable", ("Auvik Billable Qty", "Auvik Billable Amount", "CW Billable Qty", "CW Billable Amt")),
            ("Performance", ("Auvik Performance Qty", "Auvik Performance Amount", "CW Performance Qty", "CW Performance Amt")),
            ("ASM", ("Auvik ASM Qty.", "Auvik ASM Amount", "CW ASM Qty", "CW ASM Amt.")),
        ),
    ),
    ManualWorkbook(
        "Bitdefender",
        "Combined",
        r"Bitdefender\06_JUN_2026\Bitdefender reconciliation June'26.xlsx",
        "DATA",
        "Account Name",
        "Comments",
        ("Vendor Qty", "Vendor Amount", "CW Qty", "CW Amount"),
    ),
    ManualWorkbook(
        "Exium",
        "Combined",
        r"Exium\06_JUN_2026\Exium reconciliation June'26.xlsx",
        "DATA",
        "Account name",
        "Comments",
        ("Vendor Qty", "Vendor Amt", "CW Qty", "CW Amt"),
    ),
    ManualWorkbook(
        "KeepIT",
        "Main",
        r"KeepIT\06_JUN_2026\KeepIT Recon June'26.xlsx",
        "Consolidated Data",
        "Account Name",
        "Comments",
        ("Vendor Qty", "Vendor Amt", "CMS Qty", "CMS Amt"),
    ),
    ManualWorkbook(
        "KeepIT",
        "Promo",
        r"KeepIT\06_JUN_2026\KeepIT Promo Recon June'26.xlsx",
        "Data",
        "Account Name",
        "Comments",
        (),
        tuple(
            (
                sku,
                (
                    f"Vendor Qty {sku}",
                    f"Vendor AMT {sku}",
                    f"CW Qty {sku}",
                    f"CW AMT {sku}",
                ),
            )
            for sku in ("KI-AZUR-CSP", "KI-D365-FUL", "KI-GOOG-FUL", "KI-M365-FUL", "KI-SFDC-FUL")
        ),
    ),
)


def normalize_header(value: object) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").strip().lower().split())


def is_active(values: Iterable[object]) -> bool:
    for value in values:
        if value is None or str(value).strip() == "":
            continue
        numeric = pd.to_numeric(value, errors="coerce")
        if pd.notna(numeric) and abs(float(numeric)) > 1e-9:
            return True
    return False


def is_manual_clear(value: object) -> bool:
    return "clear" in str(value or "").strip().lower()


def workbook_rows(spec: ManualWorkbook) -> list[dict[str, object]]:
    path = MANUAL_ROOT / spec.relative_path
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[spec.sheet]
        iterator = worksheet.iter_rows(values_only=True)
        headers = [normalize_header(value) for value in next(iterator)]
        index = {header: position for position, header in enumerate(headers) if header}

        def value(row: tuple[object, ...], column: str) -> object:
            position = index.get(normalize_header(column))
            return row[position] if position is not None and position < len(row) else None

        extracted: list[dict[str, object]] = []
        for row_number, row in enumerate(iterator, start=2):
            account_name = value(row, spec.key_column)
            if account_name is None or not str(account_name).strip():
                continue
            if str(account_name).strip().lower() in {"grand total", "total"}:
                continue
            status = value(row, spec.status_column)
            if spec.product_families:
                for product, columns in spec.product_families:
                    if is_active(value(row, column) for column in columns):
                        extracted.append(
                            {
                                "vendor": spec.vendor,
                                "lane": spec.lane,
                                "source_file": path.name,
                                "sheet": spec.sheet,
                                "workbook_row": row_number,
                                "account_name": str(account_name).strip(),
                                "product": product,
                                "manual_status": str(status or "").strip(),
                                "manual_exact_clear": str(status or "").strip().lower() == "clear",
                                "manual_clear": is_manual_clear(status),
                            }
                        )
            elif is_active(value(row, column) for column in spec.activity_columns):
                extracted.append(
                    {
                        "vendor": spec.vendor,
                        "lane": spec.lane,
                        "source_file": path.name,
                        "sheet": spec.sheet,
                        "workbook_row": row_number,
                        "account_name": str(account_name).strip(),
                        "product": str(value(row, "Product") or value(row, "Product Name") or "Combined").strip(),
                        "manual_status": str(status or "").strip(),
                        "manual_exact_clear": str(status or "").strip().lower() == "clear",
                        "manual_clear": is_manual_clear(status),
                    }
                )
        return extracted
    finally:
        workbook.close()


def manual_clear_rates() -> tuple[pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, object]] = []
    for spec in MANUAL_WORKBOOKS:
        rows.extend(workbook_rows(spec))
    detail = pd.DataFrame(rows)

    # Proofpoint is already loaded from the historical manual recon workbooks.
    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database="ANALYTICS_DEV",
        schema="DBT_NFOLD_TRANSFORMATION",
    )
    try:
        proofpoint = fetch_dataframe(
            """
            SELECT
                'Proofpoint' AS vendor,
                'Combined' AS lane,
                source_file,
                'Snowflake manual consolidation' AS sheet,
                NULL AS workbook_row,
                account_name,
                product,
                COALESCE(comments, '') AS manual_status,
                LOWER(TRIM(COALESCE(comments, ''))) = 'clear' AS manual_exact_clear,
                CONTAINS(LOWER(COALESCE(comments, '')), 'clear') AS manual_clear
            FROM PROOFPOINT_MANUAL_CONSOLIDATED
            WHERE billing_month = '2026-06-01'
            """,
            conn=conn,
        )
    finally:
        conn.close()
    proofpoint.columns = [column.lower() for column in proofpoint.columns]
    detail = pd.concat([detail, proofpoint], ignore_index=True)

    summary = (
        detail.groupby(["vendor", "lane"], as_index=False)
        .agg(
            manual_rows=("manual_clear", "size"),
            manual_exact_clear_rows=("manual_exact_clear", "sum"),
            manual_clear_rows=("manual_clear", "sum"),
        )
    )
    summary["manual_exact_clear_pct"] = (
        100 * summary["manual_exact_clear_rows"] / summary["manual_rows"]
    ).round(2)
    summary["manual_clear_pct"] = (100 * summary["manual_clear_rows"] / summary["manual_rows"]).round(2)
    combined = (
        detail.groupby("vendor", as_index=False)
        .agg(
            manual_rows=("manual_clear", "size"),
            manual_exact_clear_rows=("manual_exact_clear", "sum"),
            manual_clear_rows=("manual_clear", "sum"),
        )
    )
    combined["manual_exact_clear_pct"] = (
        100 * combined["manual_exact_clear_rows"] / combined["manual_rows"]
    ).round(2)
    combined["manual_clear_pct"] = (100 * combined["manual_clear_rows"] / combined["manual_rows"]).round(2)
    combined["lane"] = "ALL_MANUAL_LANES"
    summary = pd.concat([summary, combined[summary.columns]], ignore_index=True)
    return detail, summary.sort_values(["vendor", "lane"])


def export_snowflake_audits() -> dict[str, pd.DataFrame]:
    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database="ANALYTICS_DEV",
        schema="DBT_NFOLD_TRANSFORMATION",
    )
    try:
        pipeline_rates = fetch_dataframe(
            f"""
            WITH versions AS (
                SELECT 'BEFORE' AS version, * FROM {BASELINE_DETAIL}
                UNION ALL
                SELECT 'AFTER' AS version, * FROM {STAGED_DETAIL}
            )
            SELECT
                version,
                vendor,
                COUNT(*) AS pipeline_rows,
                COUNT_IF(UPPER(outcome_flag) = 'CLEAR') AS pipeline_clear_rows,
                ROUND(100 * pipeline_clear_rows / NULLIF(pipeline_rows, 0), 2) AS pipeline_clear_pct,
                COUNT_IF(outcome_flag = 'Unmapped Partner') AS unmapped_rows
            FROM versions
            WHERE billing_month = '2026-06-01'
            GROUP BY 1, 2
            ORDER BY 2, 1
            """,
            conn=conn,
        )

        map_changes = fetch_dataframe(
            f"""
            WITH old_map AS (
                SELECT
                    {NAME_NORM.format(column='partner_name')} AS alias_norm,
                    MIN(partner_name) AS old_partner_name,
                    LISTAGG(DISTINCT COALESCE(sf_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(sf_id, '(NULL)')) AS old_sf_ids,
                    LISTAGG(DISTINCT COALESCE(cms_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(cms_id, '(NULL)')) AS old_cms_ids,
                    COUNT(*) AS old_map_rows
                FROM {BASELINE_MAP}
                GROUP BY 1
            ),
            new_map AS (
                SELECT
                    {NAME_NORM.format(column='partner_name')} AS alias_norm,
                    MIN(partner_name) AS new_partner_name,
                    LISTAGG(DISTINCT COALESCE(sf_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(sf_id, '(NULL)')) AS new_sf_ids,
                    LISTAGG(DISTINCT COALESCE(cms_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(cms_id, '(NULL)')) AS new_cms_ids,
                    COUNT(*) AS new_map_rows
                FROM {CURRENT_MAP}
                GROUP BY 1
            ),
            changes AS (
                SELECT
                    COALESCE(o.alias_norm, n.alias_norm) AS alias_norm,
                    o.old_partner_name,
                    n.new_partner_name,
                    o.old_sf_ids,
                    n.new_sf_ids,
                    o.old_cms_ids,
                    n.new_cms_ids,
                    o.old_map_rows,
                    n.new_map_rows,
                    CASE
                        WHEN o.alias_norm IS NULL THEN 'ADDED'
                        WHEN n.alias_norm IS NULL THEN 'REMOVED'
                        WHEN o.old_sf_ids IS DISTINCT FROM n.new_sf_ids THEN 'SF_ID_CHANGED'
                        WHEN o.old_cms_ids IS DISTINCT FROM n.new_cms_ids THEN 'CMS_ID_CHANGED'
                        ELSE 'UNCHANGED'
                    END AS change_type
                FROM old_map o
                FULL OUTER JOIN new_map n USING (alias_norm)
            ),
            impact AS (
                SELECT
                    {NAME_NORM.format(column='vendor_partner_name')} AS alias_norm,
                    COUNT(*) AS staged_rows,
                    COUNT(DISTINCT vendor) AS staged_vendor_count,
                    LISTAGG(DISTINCT vendor, ', ') WITHIN GROUP (ORDER BY vendor) AS staged_vendors,
                    ROUND(SUM(ABS(COALESCE(vendor_amount, 0))), 2) AS staged_abs_vendor_amount
                FROM {STAGED_DETAIL}
                WHERE NULLIF(TRIM(vendor_partner_name), '') IS NOT NULL
                GROUP BY 1
            )
            SELECT c.*, i.staged_rows, i.staged_vendor_count, i.staged_vendors, i.staged_abs_vendor_amount
            FROM changes c
            LEFT JOIN impact i USING (alias_norm)
            WHERE change_type <> 'UNCHANGED'
            ORDER BY change_type, alias_norm
            """,
            conn=conn,
        )

        same_alias_reassignments = fetch_dataframe(
            f"""
            WITH old_detail AS (
                SELECT
                    vendor,
                    billing_month,
                    {NAME_NORM.format(column='vendor_partner_name')} AS alias_norm,
                    MIN(vendor_partner_name) AS old_partner_name,
                    LISTAGG(DISTINCT COALESCE(sf_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(sf_id, '(NULL)')) AS old_sf_ids,
                    COUNT(*) AS old_rows,
                    COUNT_IF(UPPER(outcome_flag) = 'CLEAR') AS old_clear_rows,
                    ROUND(SUM(COALESCE(vendor_quantity, 0)), 2) AS old_vendor_quantity,
                    ROUND(SUM(COALESCE(vendor_amount, 0)), 2) AS old_vendor_amount
                FROM {BASELINE_DETAIL}
                WHERE NULLIF(TRIM(vendor_partner_name), '') IS NOT NULL
                                    AND COALESCE(vendor_quantity, 0) > 0
                GROUP BY 1, 2, 3
            ),
            new_detail AS (
                SELECT
                    vendor,
                    billing_month,
                    {NAME_NORM.format(column='vendor_partner_name')} AS alias_norm,
                    MIN(vendor_partner_name) AS new_partner_name,
                    LISTAGG(DISTINCT COALESCE(sf_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(sf_id, '(NULL)')) AS new_sf_ids,
                    COUNT(*) AS new_rows,
                    COUNT_IF(UPPER(outcome_flag) = 'CLEAR') AS new_clear_rows,
                    ROUND(SUM(COALESCE(vendor_quantity, 0)), 2) AS new_vendor_quantity,
                    ROUND(SUM(COALESCE(vendor_amount, 0)), 2) AS new_vendor_amount
                FROM {STAGED_DETAIL}
                WHERE NULLIF(TRIM(vendor_partner_name), '') IS NOT NULL
                  AND COALESCE(vendor_quantity, 0) > 0
                GROUP BY 1, 2, 3
            ),
            expected_exact AS (
                SELECT
                    billing_month,
                    partner_name_normalized AS alias_norm,
                    LISTAGG(DISTINCT COALESCE(sf_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(sf_id, '(NULL)')) AS expected_sf_ids,
                    COUNT(DISTINCT sf_id) AS expected_sf_id_count
                FROM RECON_PARTNER_MAP_MONTHLY
                GROUP BY 1, 2
            ),
            expected_composite AS (
                SELECT
                    o.vendor,
                    o.billing_month,
                    o.alias_norm,
                    LISTAGG(DISTINCT COALESCE(pm.sf_id, '(NULL)'), ', ')
                        WITHIN GROUP (ORDER BY COALESCE(pm.sf_id, '(NULL)')) AS expected_sf_ids,
                    COUNT(DISTINCT pm.sf_id) AS expected_sf_id_count
                FROM old_detail o
                JOIN RECON_PARTNER_MAP_MONTHLY pm
                  ON pm.billing_month = o.billing_month
                 AND LENGTH(pm.partner_name_normalized) >= 8
                 AND CONTAINS(o.alias_norm, pm.partner_name_normalized)
                GROUP BY 1, 2, 3
            )
            SELECT
                o.vendor,
                o.billing_month,
                o.alias_norm,
                o.old_partner_name,
                n.new_partner_name,
                o.old_sf_ids,
                n.new_sf_ids,
                COALESCE(e.expected_sf_ids, c.expected_sf_ids) AS expected_sf_ids,
                COALESCE(e.expected_sf_id_count, c.expected_sf_id_count) AS expected_sf_id_count,
                IFF(e.alias_norm IS NOT NULL, 'EXACT_ALIAS', 'COMPOSITE_ALIAS') AS expected_match_method,
                o.old_rows,
                n.new_rows,
                o.old_clear_rows,
                n.new_clear_rows,
                o.old_vendor_quantity,
                n.new_vendor_quantity,
                o.old_vendor_amount,
                n.new_vendor_amount,
                                IFF(
                                        n.new_sf_ids = COALESCE(e.expected_sf_ids, c.expected_sf_ids),
                                        'ALIGNS_TO_GOVERNED_MAP',
                                        'REVIEW'
                                ) AS validation_status
            FROM old_detail o
            JOIN new_detail n USING (vendor, billing_month, alias_norm)
                        LEFT JOIN expected_exact e
              ON e.billing_month = o.billing_month
             AND e.alias_norm = o.alias_norm
                        LEFT JOIN expected_composite c
                            ON c.vendor = o.vendor
                         AND c.billing_month = o.billing_month
                         AND c.alias_norm = o.alias_norm
            WHERE o.old_sf_ids IS DISTINCT FROM n.new_sf_ids
            ORDER BY o.vendor, o.billing_month, o.alias_norm
            """,
            conn=conn,
        )

        blocker_detail = fetch_dataframe(
            f"""
            WITH grouped AS (
                SELECT
                    vendor,
                    outcome_flag,
                    COUNT(*) AS blocker_rows,
                    ROUND(SUM(ABS(COALESCE(vendor_amount, 0))), 2) AS abs_vendor_amount,
                    ROUND(SUM(ABS(COALESCE(amount_delta, 0))), 2) AS abs_amount_delta
                FROM {STAGED_DETAIL}
                WHERE UPPER(outcome_flag) <> 'CLEAR'
                GROUP BY 1, 2
            )
            SELECT
                *,
                ROUND(100 * blocker_rows / SUM(blocker_rows) OVER (PARTITION BY vendor), 2) AS pct_of_vendor_blockers
            FROM grouped
            ORDER BY vendor, blocker_rows DESC
            """,
            conn=conn,
        )

        blocker_summary = fetch_dataframe(
            f"""
            WITH ranked AS (
                SELECT
                    vendor,
                    outcome_flag,
                    COUNT(*) AS blocker_rows,
                    ROW_NUMBER() OVER (PARTITION BY vendor ORDER BY COUNT(*) DESC, outcome_flag) AS priority
                FROM {STAGED_DETAIL}
                WHERE UPPER(outcome_flag) <> 'CLEAR'
                GROUP BY 1, 2
            ), totals AS (
                SELECT
                    vendor,
                    COUNT(*) AS total_rows,
                    COUNT_IF(UPPER(outcome_flag) = 'CLEAR') AS clear_rows,
                    COUNT_IF(UPPER(outcome_flag) <> 'CLEAR') AS blocker_rows
                FROM {STAGED_DETAIL}
                GROUP BY 1
            )
            SELECT
                t.vendor,
                t.total_rows,
                t.clear_rows,
                ROUND(100 * t.clear_rows / NULLIF(t.total_rows, 0), 2) AS clear_pct,
                t.blocker_rows,
                MAX(IFF(r.priority = 1, r.outcome_flag, NULL)) AS top_blocker,
                MAX(IFF(r.priority = 1, r.blocker_rows, NULL)) AS top_blocker_rows,
                MAX(IFF(r.priority = 2, r.outcome_flag, NULL)) AS second_blocker,
                MAX(IFF(r.priority = 2, r.blocker_rows, NULL)) AS second_blocker_rows
            FROM totals t
            LEFT JOIN ranked r USING (vendor)
            GROUP BY 1, 2, 3, 4, 5
            ORDER BY clear_pct
            """,
            conn=conn,
        )
    finally:
        conn.close()

    return {
        "pipeline_clear_rates_june": pipeline_rates,
        "partner_map_changes": map_changes,
        "same_alias_reassignments": same_alias_reassignments,
        "vendor_blocker_detail": blocker_detail,
        "vendor_health_summary": blocker_summary,
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manual_detail, manual_summary = manual_clear_rates()
    audits = export_snowflake_audits()

    manual_detail.to_csv(OUTPUT_DIR / "manual_clear_detail_june.csv", index=False)
    manual_summary.to_csv(OUTPUT_DIR / "manual_clear_rates_june.csv", index=False)
    for name, frame in audits.items():
        frame.to_csv(OUTPUT_DIR / f"{name}.csv", index=False)

    pipeline_after = audits["pipeline_clear_rates_june"]
    pipeline_after = pipeline_after[pipeline_after["VERSION"] == "AFTER"].copy()
    pipeline_after.columns = [column.lower() for column in pipeline_after.columns]
    pipeline_after["pipeline_clear_pct"] = pd.to_numeric(
        pipeline_after["pipeline_clear_pct"], errors="coerce"
    )
    manual_combined = manual_summary[manual_summary["lane"] == "ALL_MANUAL_LANES"].copy()
    comparison = manual_combined.merge(pipeline_after, on="vendor", how="outer")
    comparison["manual_exact_minus_pipeline_pp"] = (
        comparison["manual_exact_clear_pct"] - comparison["pipeline_clear_pct"]
    ).round(2)
    comparison["manual_minus_pipeline_pp"] = (
        comparison["manual_clear_pct"] - comparison["pipeline_clear_pct"]
    ).round(2)
    comparison["comparison_status"] = comparison.apply(
        lambda row: "DIRECTIONAL_JUNE_MANUAL_FILE"
        if pd.notna(row.get("manual_clear_pct"))
        else "NO_CLASSIFIED_JUNE_MANUAL_RECON_FOUND",
        axis=1,
    )
    comparison.to_csv(OUTPUT_DIR / "manual_vs_pipeline_clear_rates_june.csv", index=False)

    map_changes = audits["partner_map_changes"]
    reassignments = audits["same_alias_reassignments"]
    health = audits["vendor_health_summary"]
    report = f"""# Governed partner-map cutover audit — 2026-09-03

## Scope

- Baseline detail: `{BASELINE_DETAIL}`
- Staged detail: `{STAGED_DETAIL}`
- Manual comparison month: June 2026
- Manual `Clear`, `Clear Bundle`, `Clear Internal`, and `MP clear` are treated as the manual clear family.
- Manual rates are available only where a classified June reconciliation workbook exists.

## Evidence generated

- Manual classified rows: **{len(manual_detail):,}**
- Raw-map aliases changed or added: **{len(map_changes):,}**
- Same-alias/month detail populations with a changed SF-ID set: **{len(reassignments):,}**
- Same-alias/month changes aligned exactly to the governed monthly map: **{int((reassignments['VALIDATION_STATUS'] == 'ALIGNS_TO_GOVERNED_MAP').sum()) if not reassignments.empty else 0:,}**

## Decision rule

Do not use clear-rate lift as the acceptance criterion for identity corrections. Correct SF-ID assignment, merge-effective history, no fanout, and quantity/amount preservation are the identity gates. Clear-rate parity is a separate business-policy and source-parity gate.

The shared production output should remain unpublished until the full-refresh quantity and amount movement is explained or a target-only publication is assembled from accepted identity changes.
"""
    (OUTPUT_DIR / "CUTOVER_AUDIT.md").write_text(report, encoding="utf-8")

    print("\nManual June clear rates")
    print(manual_summary.to_string(index=False))
    print("\nManual versus staged pipeline June")
    print(comparison.to_string(index=False))
    print("\nRaw map change types")
    print(map_changes.groupby("CHANGE_TYPE").size().to_string())
    print("\nSame-alias SF-ID reassignments by vendor")
    if reassignments.empty:
        print("None")
    else:
        print(
            reassignments.groupby("VENDOR")
            .agg(
                alias_months=("ALIAS_NORM", "size"),
                aliases=("ALIAS_NORM", "nunique"),
                aligned=("VALIDATION_STATUS", lambda values: (values == "ALIGNS_TO_GOVERNED_MAP").sum()),
                old_clear_rows=("OLD_CLEAR_ROWS", "sum"),
                new_clear_rows=("NEW_CLEAR_ROWS", "sum"),
            )
            .to_string()
        )
    print("\nVendor health summary")
    print(health.to_string(index=False))
    print(f"\nWrote audit outputs to {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
