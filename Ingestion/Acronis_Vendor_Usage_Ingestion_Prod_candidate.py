"""Recreate Acronis manual-team usage files from raw portal CSV exports.

This script scans the raw monthly `Data` and `ASIO Data` folders with the same
guardrails as the manual workbook recreation, then publishes vendor usage at
the standard third-party usage grain:

Snowflake target:
    ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_USAGE_PROD

Published grain:
    BILLING_MONTH x VENDOR x MODIFIER(Status) x VENDOR_PARTNER_NAME x VENDOR_PRODUCT_SKU
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import io
import json
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
from invoice_rate_backfill import fill_missing_prices_dynamic


ACRONIS_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = ACRONIS_ROOT.parent
WORKSPACE_ROOT = next((p for p in (PROJECT_ROOT, *PROJECT_ROOT.parents) if (p / "TEMPLATES").exists()), PROJECT_ROOT)
OUTPUT_DIR = ACRONIS_ROOT / "outputs"
# CSV seed files are no longer used. Unit prices are loaded dynamically from
# THIRD_PARTY_RECON_VENDOR_INVOICES via load_price_seed() below.

DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026\Acronis"
)

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
FQN = f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}"
TARGET_VENDOR = "Acronis"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)
REPORTING_PERIOD_RE = re.compile(r"(?P<yyyy>\d{4})-(?P<mm>\d{2})-\d{2}")
DATE_RANGE_RE = re.compile(
    r"\((?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^)]*\)\s*$",
    re.IGNORECASE,
)

# Layouts observed across months:
#   06_JUN_2026/Data/ + 06_JUN_2026/ASIO Data/
#   02_FEB_2026/Data/ + 02_FEB_2026/Data/ASIO/
# Both are supported: (folder path parts to look under, label used in the scan).
RAW_LAYOUTS: tuple[tuple[tuple[str, ...], str], ...] = (
    (("ASIO Data",), "ASIO Data"),
    (("Data", "ASIO"), "ASIO Data"),
    (("Data",), "Data"),
)

RAW_COLUMNS: tuple[str, ...] = (
    "Tenant name",
    "Identifier",
    "Type",
    "Status",
    "Mode",
    "Service name",
    "Edition",
    "Metric name",
    "Location",
    "Metric unit",
    "Quota",
    "Production usage",
    "Change in production usage",
    "Trial usage",
    "Change in trial usage",
    "Total usage",
    "Change in total usage",
    "Price",
    "Cost",
    "Billing model",
    "Storage type",
    "Pricing tier ID",
    "SKU",
)

OUTPUT_COLUMNS: tuple[str, ...] = ("Entity",) + RAW_COLUMNS
TABLE_COLUMNS: tuple[str, ...] = ("BILLING_MONTH",) + OUTPUT_COLUMNS
LINEAGE_COLUMNS: tuple[str, ...] = (
    "SOURCE_PORTAL",
    "SOURCE_FILE",
    "SOURCE_ROW_NUMBER",
    "SOURCE_HASH",
)
PARSED_COLUMNS: tuple[str, ...] = TABLE_COLUMNS + LINEAGE_COLUMNS
USAGE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
    "ADDITIONAL_INFO",
)
NUMERIC_COMPARE_COLUMNS = {
    "Quota",
    "Production usage",
    "Change in production usage",
    "Trial usage",
    "Change in trial usage",
    "Total usage",
    "Change in total usage",
    "Pricing tier ID",
}

SNOWFLAKE_COLUMNS = USAGE_COLUMNS
BILLABLE_SKU_RE = re.compile(r"^S[A-Z0-9]+$", re.IGNORECASE)

# --- SKU prefix guardrail (Open Question #1 in the spec) -----------------------
# Empirically only S* SKUs are billable in the CW-resold contract. C* SKUs are
# billed directly by Acronis (or sit on partner-owned gateways) and are
# EXCLUDED from every output row across the 2,414-row June workbook. If a new
# prefix appears it must be triaged before we silently keep or drop it.
ALLOWED_SKU_PREFIXES: frozenset[str] = frozenset({"S"})
KNOWN_EXCLUDED_SKU_PREFIXES: frozenset[str] = frozenset({"C"})

# --- Explicit filename -> Entity lookup (spec: maintained constant) -----------
# Key format: (layout_label, normalized_report_title). The normaliser lowercases,
# strips the trailing "(<Month> <Day>, <YYYY>-...)" date-range, collapses
# whitespace, and drops the redundant leading region prefix or "Summary for
# period report -" boilerplate to leave a stable descriptor. Any file that
# does not match a key here raises an alert instead of being silently kept.
FILE_TO_ENTITY: dict[tuple[str, str], str] = {
    # Data folder (device / GB workload reports)
    ("Data", "connectwise au1 per workload"): "AUS_DEV",
    ("Data", "connectwise au1 per gb"): "AUS_GIG",
    ("Data", "capacity gig partners"): "CAN_GIG",
    ("Data", "device workload partners"): "CAN_DEV",
    ("Data", "cw eu2 germany device workload"): "EU2_DEV",
    ("Data", "eu1 connectwise inc"): "EU1_GIG",
    ("Data", "eu8 connectwise inc"): "EU8_Master",
    ("Data", "connectwise sg1 per workload"): "SG1_DEV",
    ("Data", "connectwise sg1 per gigabyte"): "SG1_GIG",
    ("Data", "uk dev connectwise inc 2"): "UK_DEV",
    ("Data", "uk gig connectwise inc"): "UK_GIG",
    ("Data", "us connectwise inc"): "US_Master",
    # Older-month filename convention where US_Master lacks a region prefix
    # (Jan-May 2026 stored the file as 'Summary for period report - ConnectWise, Inc').
    ("Data", "connectwise inc"): "US_Master",
    ("Data", "us5 dev connectwise inc 1"): "US5_DEV",
    ("Data", "us5 gig connectwise inc"): "US5_GIG",
    ("Data", "cw gig us2"): "US2_GIG",
    ("Data", "cw workload us2"): "US2_DEV",
    ("Data", "nz master connectwise inc"): "NZ_Master",
    # ASIO folder (nested ASIO Portal reports)
    ("ASIO Data", "aus dev asio portal"): "AUS_DEV",
    ("ASIO Data", "can dev asio portal"): "CAN_DEV",
    ("ASIO Data", "can gig asio portal"): "CAN_GIG",
    ("ASIO Data", "eu8 asio"): "EU8_Master",
    ("ASIO Data", "uk dev asio portal"): "UK_DEV",
    ("ASIO Data", "uk gig asio portal"): "UK_GIG",
    ("ASIO Data", "us asio"): "US_Master",
    # Older-month ASIO filename that lacked the 'US_' prefix.
    ("ASIO Data", "asio"): "US_Master",
    ("ASIO Data", "us5 dev asio portal"): "US5_DEV",
}

# --- Explicit exclusion list (spec: tenant-level manual exceptions) ----------
# Applied as (entity, tenant_normalized_compact) so future trial or
# double-count tenants are handled by extending this table rather than by
# tweaking regex heuristics. `_norm_compact` collapses whitespace and drops
# all non-alphanumeric characters, so the key for '// TRIALS //' is 'trials'.
EXCLUDED_TENANTS: frozenset[tuple[str, str]] = frozenset({
    # spec section 8 -- seven analyst-hand-strip entries
    ("AUS_GIG", "cwpartnertrials"),
    ("CAN_GIG", "testing"),
    ("US2_DEV", "reporttestpartner"),
    ("US2_GIG", "cwpartnercovid19promo"),
    ("US_Master", "cwpartnercovid19promo"),
    ("US_Master", "cwgigus2"),
    ("US_Master", "cwworkloadus2"),
    # Trial containers that appear as Type=Partner (not Folder) in the raw
    # CSVs and therefore slip past the Folder filter. Confirmed against the
    # JUN 2026 manual workbook.
    ("UK_DEV", "trials"),
    ("UK_GIG", "cwtrials"),
    ("US2_GIG", "trials"),
})

COMMON_DATA_REPORTS = frozenset(
    ("Data", entity)
    for entity in {
        "AUS_DEV",
        "AUS_GIG",
        "CAN_DEV",
        "CAN_GIG",
        "EU1_GIG",
        "EU2_DEV",
        "EU8_Master",
        "UK_DEV",
        "UK_GIG",
        "US2_DEV",
        "US2_GIG",
        "US5_DEV",
        "US5_GIG",
        "US_Master",
    }
)
ASIO_REPORTS = frozenset(
    ("ASIO Data", entity)
    for entity in {
        "AUS_DEV",
        "CAN_DEV",
        "EU8_Master",
        "UK_DEV",
        "UK_GIG",
        "US5_DEV",
        "US_Master",
    }
)
EXPECTED_REPORTS = {
    "2026-05": COMMON_DATA_REPORTS | ASIO_REPORTS | {("Data", "SG1_GIG")},
    "2026-06": COMMON_DATA_REPORTS | ASIO_REPORTS | {("Data", "SG1_DEV")},
    "2026-07": COMMON_DATA_REPORTS
    | ASIO_REPORTS
    | {("Data", "SG1_DEV"), ("Data", "SG1_GIG")},
}


def _norm_compact(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value if value is not None else "").lower())


def _norm_words(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value if value is not None else "").lower()).strip()


def _clean_text(value: object) -> str | None:
    if value is None or bool(pd.isna(value)):
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def normalize_cell(value: object) -> object:
    """Trim whitespace on strings and normalize blanks to None. Non-string
    values pass through unchanged so downstream numeric coercion still works.
    """
    if value is None or bool(pd.isna(value)):
        return None
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text if text else None
    return value


def _normalize_compare_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    try:
        return text.encode("cp1252").decode("utf-8")
    except UnicodeError:
        return text


def _read_file_bytes(path: Path) -> bytes:
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
            tmp_path = Path(handle.name)
        try:
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    "Copy-Item -LiteralPath $args[0] -Destination $args[1] -Force",
                    str(path),
                    str(tmp_path),
                ],
                check=True,
                capture_output=True,
            )
            return tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)


def discover_month_folders(source_root: Path) -> dict[str, Path]:
    folders: dict[str, Path] = {}
    for child in source_root.iterdir():
        if not child.is_dir():
            continue
        match = MONTH_FOLDER_RE.match(child.name)
        if match:
            folders[f"{match.group('yyyy')}-{match.group('mm')}"] = child
    return dict(sorted(folders.items()))


def folder_month(month_folder: Path) -> dt.date:
    match = MONTH_FOLDER_RE.match(month_folder.name)
    if not match:
        raise ValueError(f"Invalid month folder name: {month_folder}")
    return dt.date(int(match.group("yyyy")), int(match.group("mm")), 1)


def locate_raw_files(month_folder: Path) -> list[tuple[str, Path]]:
    """Return (layout_label, path) for every raw CSV in the month folder.

    Supports both raw-file layouts we've seen in 2026:
      06_JUN_2026/Data/ + 06_JUN_2026/ASIO Data/
      02_FEB_2026/Data/ + 02_FEB_2026/Data/ASIO/
    """
    files: list[tuple[str, Path]] = []
    seen: set[Path] = set()
    for parts, layout_label in RAW_LAYOUTS:
        folder = month_folder.joinpath(*parts)
        if not folder.exists() or not folder.is_dir():
            continue
        for path in sorted(folder.iterdir()):
            if path.name.startswith("~$") or not path.is_file() or path in seen:
                continue
            suffixes = [s.lower() for s in path.suffixes]
            if path.suffix.lower() == ".csv" or suffixes[-2:] == [".csv", ".zip"]:
                files.append((layout_label, path))
                seen.add(path)
    return files


def _decode(raw: bytes) -> list[str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return raw.decode(encoding).splitlines()
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8-sig", errors="replace").splitlines()


def iter_csv_payloads(path: Path) -> list[tuple[str, list[str]]]:
    if path.suffix.lower() != ".zip":
        return [(path.name, _decode(_read_file_bytes(path)))]
    payloads: list[tuple[str, list[str]]] = []
    with zipfile.ZipFile(io.BytesIO(_read_file_bytes(path))) as archive:
        for name in archive.namelist():
            if name.lower().endswith(".csv"):
                payloads.append((f"{path.name}::{Path(name).name}", _decode(archive.read(name))))
    return payloads


def find_header_row(rows: list[list[str]]) -> int | None:
    expected = [_norm_compact(c) for c in RAW_COLUMNS]
    for i, row in enumerate(rows[:30]):
        normalized = [_norm_compact(c) for c in row[: len(RAW_COLUMNS)]]
        if normalized == expected:
            return i
    return None


def parse_preamble(rows: list[list[str]], header_idx: int | None) -> dict[str, str]:
    limit = header_idx if header_idx is not None else min(10, len(rows))
    preamble: dict[str, str] = {}
    for row in rows[:limit]:
        if len(row) >= 2 and row[0] and row[1]:
            preamble[str(row[0]).strip()] = str(row[1]).strip()
    return preamble


def reporting_month(preamble: dict[str, str], fallback: dt.date) -> dt.date:
    period = preamble.get("Reporting period")
    if period:
        match = REPORTING_PERIOD_RE.search(period)
        if match:
            return dt.date(int(match.group("yyyy")), int(match.group("mm")), 1)
    return fallback


def normalize_report_key(path: Path) -> str:
    """Reduce a raw CSV filename to a stable descriptor used to look up Entity.

    Rules (empirically derived; produces the same key across Jan-Jul 2026):
      - strip trailing '.csv' / '.csv.zip'
      - strip trailing '(<Month> <Day>, YYYY-<Month> <Day>, YYYY)' date range
      - lowercase
      - drop 'summary for period report' boilerplate
      - drop 'connectwise' when it's clearly a report-title filler
        (kept for 'connectwise, inc' / 'connectwise au1' / 'connectwise sg1'
        as those are the actual portal identifiers)
      - collapse punctuation, whitespace and repeated spaces
    """
    stem = path.name
    # strip .csv or .csv.zip
    if stem.lower().endswith(".csv.zip"):
        stem = stem[: -len(".csv.zip")]
    elif stem.lower().endswith(".csv"):
        stem = stem[: -len(".csv")]
    # strip date-range suffix like " (Jun 1, 2026-Jun 30, 2026)"
    stem = DATE_RANGE_RE.sub("", stem).strip()
    text = stem.lower()
    text = text.replace(",", " ")
    text = text.replace("-", " ")
    text = text.replace("_", " ")
    text = text.replace(".", " ")
    # drop the boilerplate. "summary for period report" and stray "connectwise inc"
    # tokens are always present and never disambiguate one report from another.
    text = re.sub(r"\bsummary for period report\b", " ", text)
    text = re.sub(r"\bportal\b", " portal ", text)
    text = re.sub(r"\s+", " ", text).strip()
    # collapse the specific "connectwise inc" identifier down to just "connectwise inc"
    # (the "1"/"2" region marker after it survives as its own token)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def resolve_entity(path: Path, layout_label: str, portal_tenant: str) -> str | None:
    """Look up the Entity code from an explicit filename -> Entity table.

    The spec (Locating the input files) is explicit: 'code should hold an
    explicit lookup table mapping each source report to its Entity code
    rather than deriving the code from the filename'. If a file cannot be
    matched we return None and the caller emits a warning so the operator
    can extend the table -- silent guesses are unsafe when new regions are
    introduced.
    """
    key = normalize_report_key(path)
    entry = FILE_TO_ENTITY.get((layout_label, key))
    if entry is not None:
        return entry
    # Fallback: try every substring token match against the table for the
    # same layout. This handles minor filename tweaks (extra whitespace,
    # extra "AUS -" prefixes) without silently mapping to a wrong region.
    tokens = set(key.split())
    matches: list[tuple[int, str]] = []
    for (layout, canonical_key), entity in FILE_TO_ENTITY.items():
        if layout != layout_label:
            continue
        canonical_tokens = set(canonical_key.split())
        if canonical_tokens.issubset(tokens):
            matches.append((len(canonical_tokens), entity))
    if not matches:
        return None
    best_score = max(score for score, _ in matches)
    best_entities = {entity for score, entity in matches if score == best_score}
    return next(iter(best_entities)) if len(best_entities) == 1 else None


def normalize_hierarchy_path(value: object) -> str:
    """Spec-compliant path normalization.

    'Strip any leading slashes and prefix exactly one, so that
    "ConnectWise, Inc /2" becomes "/ConnectWise, Inc /2" and "/ ASIO Portal"
    becomes "/ ASIO Portal".'
    """
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    while text.startswith("/"):
        text = text[1:]
    return "/" + text


def is_direct_child_of_root(row_tenant: object, root_tenant: str) -> bool:
    """Return True iff row_tenant is a descendant of root_tenant (not the root
    itself).

    Spec condition 3 combined with condition 2 (Type != 'Folder'): the Folder
    filter already removes intermediate hierarchy rows, so any Partner-typed
    row whose Tenant name starts with the root path is by definition a
    billable leaf. We intentionally DO NOT reject remainders that contain
    additional slashes -- legitimate portal display names like
    'Desert IT Solutions /1', 'Infinity Group / France', and
    'Pure Technology /1' embed slashes verbatim and the manual workbook
    keeps them.
    """
    row_norm = normalize_hierarchy_path(row_tenant)
    root_norm = normalize_hierarchy_path(root_tenant)
    if not root_norm or not row_norm:
        return False
    prefix = root_norm + "/"
    return row_norm.startswith(prefix) and bool(row_norm[len(prefix):])


def clean_tenant_leaf(row_tenant: object, root_tenant: str) -> str | None:
    """Return the tenant name below root (root prefix removed).

    Only invoked after is_direct_child_of_root has returned True. Slashes in
    the remainder are preserved verbatim because the Acronis portal encodes
    display suffixes like '/1', '/2', '/ France' inside the leaf name.
    """
    row_norm = normalize_hierarchy_path(row_tenant)
    root_norm = normalize_hierarchy_path(root_tenant)
    remainder = row_norm[len(root_norm) + 1:]
    return _clean_text(remainder)


def sku_prefix(sku: str) -> str:
    match = re.match(r"^([A-Za-z])", sku)
    return match.group(1).upper() if match else ""


def parse_raw_file(
    *,
    source_folder: str,
    path: Path,
    display_name: str,
    lines: list[str],
    fallback_month: dt.date,
    source_hash: str,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    rows = list(csv.reader(lines))
    header_idx = find_header_row(rows)
    scan = {
        "billing_month": fallback_month.isoformat(),
        "source_folder": source_folder,
        "source_file": display_name,
        "source_hash": source_hash,
        "entity": None,
        "portal_tenant": None,
        "schema_status": "matched" if header_idx is not None else "skipped_schema_mismatch",
        "period_mismatch": False,
        "raw_rows": 0,
        "kept_rows": 0,
        "dropped_non_billable_sku_prefix": 0,
        "dropped_unknown_sku_prefix": 0,
        "dropped_folder_type": 0,
        "dropped_non_direct_child": 0,
        "dropped_root_row": 0,
        "dropped_excluded_tenant": 0,
        "unknown_entity": False,
    }
    if header_idx is None:
        return [], scan

    preamble = parse_preamble(rows, header_idx)
    month = reporting_month(preamble, fallback_month)
    if month != fallback_month:
        scan["period_mismatch"] = True
        raise RuntimeError(
            f"Reporting period {month:%Y-%m} in {display_name!r} does not match "
            f"folder month {fallback_month:%Y-%m}."
        )
    portal_tenant = preamble.get("Tenant name", "")
    entity = resolve_entity(path, source_folder, portal_tenant)
    scan.update({
        "billing_month": month.isoformat(),
        "entity": entity,
        "portal_tenant": portal_tenant,
        "unknown_entity": entity is None,
    })
    if entity is None:
        # Spec: 'raise an alert whenever a SKU prefix appears that is not in the
        # known list.' Same principle applies here for unrecognised reports so
        # the operator can extend FILE_TO_ENTITY instead of the file being
        # silently dropped or mis-routed.
        print(
            f"[WARN] Unknown Entity for report {display_name!r} "
            f"(layout={source_folder}, root={portal_tenant!r}); "
            f"add an entry to FILE_TO_ENTITY."
        )
        return [], scan

    output_rows: list[dict[str, object]] = []
    source_portal = "ASIO" if source_folder == "ASIO Data" else "Legacy"
    for source_row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not row or all(not str(cell).strip() for cell in row):
            continue
        scan["raw_rows"] += 1

        values = {
            RAW_COLUMNS[i]: normalize_cell(row[i]) if i < len(row) else None
            for i in range(len(RAW_COLUMNS))
        }
        raw_tenant = values.get("Tenant name")
        row_type = _clean_text(values.get("Type"))
        sku = _clean_text(values.get("SKU"))

        # Spec condition 1: SKU must begin with 'S'. Blank/non-billable SKUs
        # are dropped by definition. Log any brand-new prefix.
        if not sku:
            scan["dropped_non_billable_sku_prefix"] += 1
            continue
        prefix = sku_prefix(sku)
        if prefix in KNOWN_EXCLUDED_SKU_PREFIXES:
            scan["dropped_non_billable_sku_prefix"] += 1
            continue
        if prefix not in ALLOWED_SKU_PREFIXES:
            scan["dropped_unknown_sku_prefix"] += 1
            print(
                f"[WARN] Unknown SKU prefix {prefix!r} in {display_name!r} "
                f"(SKU={sku!r}); triage before extending ALLOWED_SKU_PREFIXES."
            )
            continue
        if not BILLABLE_SKU_RE.match(sku):
            # Malformed SKU that starts with S but has non-alnum chars.
            scan["dropped_non_billable_sku_prefix"] += 1
            continue

        # Spec condition 2: Type must not be 'Folder' (Folder rows are
        # aggregations of everything beneath them).
        if row_type and row_type.strip().lower() == "folder":
            scan["dropped_folder_type"] += 1
            continue

        # Spec condition 3: tenant must be a DIRECT CHILD of the report root.
        if not is_direct_child_of_root(raw_tenant, portal_tenant):
            row_norm = normalize_hierarchy_path(raw_tenant)
            root_norm = normalize_hierarchy_path(portal_tenant)
            if row_norm == root_norm:
                scan["dropped_root_row"] += 1
            else:
                scan["dropped_non_direct_child"] += 1
            continue

        tenant_name = clean_tenant_leaf(raw_tenant, portal_tenant)
        if not tenant_name:
            scan["dropped_non_direct_child"] += 1
            continue

        # Spec exclusion list: 7 specific (entity, tenant) pairs that the
        # analysts strip by hand every month.
        if (entity, _norm_compact(tenant_name)) in EXCLUDED_TENANTS:
            scan["dropped_excluded_tenant"] += 1
            continue

        values["Tenant name"] = tenant_name
        values["SKU"] = sku.upper()
        record = {
            "BILLING_MONTH": month,
            "Entity": entity,
            "SOURCE_PORTAL": source_portal,
            "SOURCE_FILE": display_name,
            "SOURCE_ROW_NUMBER": source_row_number,
            "SOURCE_HASH": source_hash,
        }
        record.update(values)
        output_rows.append(record)
        scan["kept_rows"] += 1

    return output_rows, scan


def parse_month(source_root: Path, month: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    month_folder = discover_month_folders(source_root).get(month)
    if month_folder is None:
        raise FileNotFoundError(f"No folder found for {month} under {source_root}")

    fallback_month = folder_month(month_folder)
    records: list[dict[str, object]] = []
    scans: list[dict[str, object]] = []
    seen_payloads: dict[str, str] = {}
    for source_folder, path in locate_raw_files(month_folder):
        for display_name, lines in iter_csv_payloads(path):
            source_hash = hashlib.sha256("\n".join(lines).encode("utf-8")).hexdigest()
            duplicate_of = seen_payloads.get(source_hash)
            if duplicate_of is not None:
                raise RuntimeError(
                    f"Duplicate Acronis source content detected: {display_name!r} "
                    f"duplicates {duplicate_of!r}."
                )
            seen_payloads[source_hash] = display_name
            rows, scan = parse_raw_file(
                source_folder=source_folder,
                path=path,
                display_name=display_name,
                lines=lines,
                fallback_month=fallback_month,
                source_hash=source_hash,
            )
            records.extend(rows)
            scans.append(scan)

    df = pd.DataFrame(records, columns=list(PARSED_COLUMNS))
    scan_df = pd.DataFrame(scans)
    print(
        f"[{month}] files={len(scan_df):,}, rows={len(df):,}, "
        f"sku_count={df['SKU'].nunique() if not df.empty else 0:,}, "
        f"entities={df['Entity'].nunique() if not df.empty else 0:,}"
    )
    return df, scan_df


def validate_scan(scan_df: pd.DataFrame) -> None:
    if scan_df.empty:
        raise RuntimeError("No Acronis raw CSV exports were found for the selected month.")

    failures = scan_df[
        scan_df["schema_status"].ne("matched")
        | scan_df["unknown_entity"].fillna(False)
        | scan_df["period_mismatch"].fillna(False)
        | scan_df["dropped_unknown_sku_prefix"].fillna(0).gt(0)
    ]
    if not failures.empty:
        columns = [
            "billing_month",
            "source_folder",
            "source_file",
            "schema_status",
            "unknown_entity",
            "period_mismatch",
            "dropped_unknown_sku_prefix",
        ]
        raise RuntimeError(
            "Acronis ingestion stopped because one or more source files require review:\n"
            + failures[columns].to_string(index=False)
        )

    month = str(scan_df["billing_month"].iloc[0])[:7]
    expected_reports = EXPECTED_REPORTS.get(month)
    if expected_reports is None:
        return
    actual_reports = set(zip(scan_df["source_folder"], scan_df["entity"]))
    if actual_reports != expected_reports:
        raise RuntimeError(
            f"Acronis source manifest mismatch for {month}: "
            f"missing={sorted(expected_reports - actual_reports)}, "
            f"unexpected={sorted(actual_reports - expected_reports)}."
        )
    duplicate_reports = (
        scan_df.groupby(["source_folder", "entity"], dropna=False)
        .size()
        .reset_index(name="file_count")
    )
    duplicate_reports = duplicate_reports[duplicate_reports["file_count"].gt(1)]
    if not duplicate_reports.empty:
        raise RuntimeError(
            "Multiple Acronis reports mapped to the same portal/entity:\n"
            + duplicate_reports.to_string(index=False)
        )


def load_price_seed() -> pd.DataFrame:
    """Load the existing month/SKU price map used by the prior script."""
    import sys as _sys
    _sys.path.insert(0, str(WORKSPACE_ROOT))
    conn = None
    try:
        from TEMPLATES.Python.connection import get_snowflake_connection as _conn
        conn = _conn(role="DEVELOPER", warehouse="REPORTING_WH",
                     database="ANALYTICS_DEV", schema="DBT_NFOLD_TRANSFORMATION")
        monthly_rates = pd.read_sql("""
            SELECT
                BILLING_MONTH,
                VENDOR_PRODUCT_SKU AS VENDOR_SKU,
                CASE
                    WHEN COUNT(IFF(UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0, 1, NULL)) > 0
                        THEN MEDIAN(IFF(UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0, UNIT_PRICE, NULL))
                    WHEN SUM(IFF(QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL, QUANTITY, 0)) > 0
                        THEN SUM(IFF(QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL, AMOUNT, 0))
                             / SUM(IFF(QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL, QUANTITY, 0))
                    ELSE NULL
                END AS UNIT_PRICE
            FROM ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_INVOICES
            WHERE VENDOR ILIKE '%acronis%'
              AND VENDOR_PRODUCT_SKU IS NOT NULL
              AND (
                    (UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0)
                    OR (QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL)
                  )
            GROUP BY 1, 2
        """, conn)
    except Exception as e:
        print(f"[WARN] Could not load Acronis invoice rates from Snowflake ({e}). UNIT_PRICE will be NULL.", flush=True)
        return pd.DataFrame(
            columns=["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE", "CURRENCY"]
        )
    finally:
        if conn is not None:
            conn.close()

    if monthly_rates.empty:
        print("[WARN] No Acronis rows found in THIRD_PARTY_RECON_VENDOR_INVOICES. UNIT_PRICE will be NULL.", flush=True)
        return pd.DataFrame(
            columns=["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE", "CURRENCY"]
        )

    monthly_rates["BILLING_MONTH"] = pd.to_datetime(
        monthly_rates["BILLING_MONTH"]
    ).dt.date
    monthly_rates["VENDOR_SKU"] = monthly_rates["VENDOR_SKU"].map(
        _clean_text
    ).str.upper()
    monthly_rates["UNIT_PRICE"] = pd.to_numeric(
        monthly_rates["UNIT_PRICE"], errors="coerce"
    )
    monthly_rates = monthly_rates[
        monthly_rates["BILLING_MONTH"].notna()
        & monthly_rates["VENDOR_SKU"].notna()
        & monthly_rates["UNIT_PRICE"].notna()
        & monthly_rates["UNIT_PRICE"].gt(0)
    ].copy()
    monthly_rates = monthly_rates.drop_duplicates(
        subset=["BILLING_MONTH", "VENDOR_SKU"],
        keep="first",
    )
    monthly_rates["CURRENCY"] = "USD"
    print(
        f"[INFO] Loaded {len(monthly_rates):,} Acronis SKU rates "
        f"({monthly_rates['BILLING_MONTH'].nunique()} months, "
        f"{monthly_rates['VENDOR_SKU'].nunique()} SKUs).",
        flush=True,
    )
    return monthly_rates


def build_latest_rate_frame(monthly_rates: pd.DataFrame) -> pd.DataFrame:
    if monthly_rates.empty:
        return pd.DataFrame(columns=["VENDOR_PRODUCT_SKU", "FALLBACK_UNIT_PRICE", "FALLBACK_CURRENCY"])
    latest = (
        monthly_rates.sort_values(["VENDOR_PRODUCT_SKU", "BILLING_MONTH"], ascending=[True, False])
        .drop_duplicates(subset=["VENDOR_PRODUCT_SKU"], keep="first")
        [["VENDOR_PRODUCT_SKU", "MONTH_SKU_UNIT_PRICE", "MONTH_SKU_CURRENCY"]]
        .rename(
            columns={
                "MONTH_SKU_UNIT_PRICE": "FALLBACK_UNIT_PRICE",
                "MONTH_SKU_CURRENCY": "FALLBACK_CURRENCY",
            }
        )
    )
    return latest


def _unique_text_values(values: pd.Series) -> list[str]:
    cleaned = {
        text
        for value in values
        if (text := _clean_text(value)) is not None
    }
    return sorted(cleaned)


def build_vendor_usage_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=list(USAGE_COLUMNS))

    monthly_rates = load_price_seed()
    usage = pd.DataFrame(
        {
            "BILLING_MONTH": pd.to_datetime(df["BILLING_MONTH"]).dt.date,
            "VENDOR": "Acronis",
            "VENDOR_PARTNER_NAME": df["Tenant name"].map(_clean_text),
            "VENDOR_PRODUCT_SKU": df["SKU"].map(_clean_text).str.upper(),
            "MODIFIER": df["Status"].map(_clean_text),
            "QUANTITY": pd.to_numeric(df["Total usage"], errors="coerce").fillna(0.0),
            "_ENTITY": df["Entity"].map(_clean_text),
            "_SOURCE_PORTAL": df["SOURCE_PORTAL"].map(_clean_text),
            "_SOURCE_FILE": df["SOURCE_FILE"].map(_clean_text),
            "_IDENTIFIER": df["Identifier"].map(_clean_text),
            "_TYPE": df["Type"].map(_clean_text),
            "_SERVICE_NAME": df["Service name"].map(_clean_text),
            "_EDITION": df["Edition"].map(_clean_text),
            "_METRIC_NAME": df["Metric name"].map(_clean_text),
            "_LOCATION": df["Location"].map(_clean_text),
            "_METRIC_UNIT": df["Metric unit"].map(_clean_text),
            "_BILLING_MODEL": df["Billing model"].map(_clean_text),
            "_STORAGE_TYPE": df["Storage type"].map(_clean_text),
            "_PRICING_TIER_ID": df["Pricing tier ID"].map(_clean_text),
            "_PRODUCTION_USAGE": pd.to_numeric(
                df["Production usage"], errors="coerce"
            ).fillna(0.0),
            "_TRIAL_USAGE": pd.to_numeric(df["Trial usage"], errors="coerce").fillna(0.0),
        }
    )
    usage = usage[
        usage["VENDOR_PARTNER_NAME"].notna()
        & usage["VENDOR_PRODUCT_SKU"].notna()
    ].copy()

    monthly_rates = monthly_rates.rename(
        columns={
            "VENDOR_SKU": "VENDOR_PRODUCT_SKU",
            "UNIT_PRICE": "MONTH_SKU_UNIT_PRICE",
            "CURRENCY": "MONTH_SKU_CURRENCY",
        }
    )
    latest_rates = build_latest_rate_frame(monthly_rates)
    usage = usage.merge(
        monthly_rates,
        on=["BILLING_MONTH", "VENDOR_PRODUCT_SKU"],
        how="left",
    )
    usage = usage.merge(latest_rates, on=["VENDOR_PRODUCT_SKU"], how="left")
    usage["RATE_SOURCE"] = "missing"
    usage.loc[usage["FALLBACK_UNIT_PRICE"].notna(), "RATE_SOURCE"] = "fallback"
    usage.loc[
        usage["MONTH_SKU_UNIT_PRICE"].notna(),
        "RATE_SOURCE",
    ] = "sku_map_month"
    usage["UNIT_PRICE"] = usage["MONTH_SKU_UNIT_PRICE"].fillna(
        usage["FALLBACK_UNIT_PRICE"]
    )
    usage["CURRENCY"] = (
        usage["MONTH_SKU_CURRENCY"]
        .fillna(usage["FALLBACK_CURRENCY"])
        .fillna("USD")
    )

    group_columns = [
        "BILLING_MONTH",
        "VENDOR",
        "VENDOR_PARTNER_NAME",
        "VENDOR_PRODUCT_SKU",
        "MODIFIER",
        "CURRENCY",
    ]
    records: list[dict[str, object]] = []
    for key, group in usage.groupby(group_columns, dropna=False, sort=False):
        quantity = float(group["QUANTITY"].sum())
        unit_prices = group["UNIT_PRICE"].dropna().unique()
        if len(unit_prices) > 1:
            raise RuntimeError(
                f"Multiple SKU-map rates found for {key[0]} / {key[3]}."
            )
        unit_price = float(unit_prices[0]) if len(unit_prices) == 1 else None
        amount = quantity * unit_price if unit_price is not None else None
        additional_info = {
            "entities": _unique_text_values(group["_ENTITY"]),
            "source_portals": _unique_text_values(group["_SOURCE_PORTAL"]),
            "source_files": _unique_text_values(group["_SOURCE_FILE"]),
            "account_types": _unique_text_values(group["_TYPE"]),
            "identifiers": _unique_text_values(group["_IDENTIFIER"]),
            "service_names": _unique_text_values(group["_SERVICE_NAME"]),
            "editions": _unique_text_values(group["_EDITION"]),
            "metric_names": _unique_text_values(group["_METRIC_NAME"]),
            "locations": _unique_text_values(group["_LOCATION"]),
            "metric_units": _unique_text_values(group["_METRIC_UNIT"]),
            "billing_models": _unique_text_values(group["_BILLING_MODEL"]),
            "storage_types": _unique_text_values(group["_STORAGE_TYPE"]),
            "pricing_tier_ids": _unique_text_values(group["_PRICING_TIER_ID"]),
            "production_usage": float(group["_PRODUCTION_USAGE"].sum()),
            "trial_usage": float(group["_TRIAL_USAGE"].sum()),
            "total_usage": quantity,
            "raw_row_count": int(len(group)),
            "rate_sources": _unique_text_values(group["RATE_SOURCE"]),
        }
        records.append(
            {
                "BILLING_MONTH": key[0],
                "VENDOR": key[1],
                "VENDOR_PARTNER_NAME": key[2],
                "VENDOR_PRODUCT_SKU": key[3],
                "MODIFIER": key[4],
                "QUANTITY": quantity,
                "UNIT_PRICE": unit_price,
                "AMOUNT": amount,
                "CURRENCY": key[5],
                "ADDITIONAL_INFO": json.dumps(
                    additional_info,
                    sort_keys=True,
                    separators=(",", ":"),
                ),
            }
        )
    return pd.DataFrame(records, columns=list(USAGE_COLUMNS))


def validate_vendor_usage(vendor_usage: pd.DataFrame, *, require_complete_rates: bool = True) -> None:
    if vendor_usage.empty:
        raise RuntimeError("No Acronis vendor usage rows were produced.")

    grain = [
        "BILLING_MONTH",
        "VENDOR",
        "VENDOR_PARTNER_NAME",
        "VENDOR_PRODUCT_SKU",
        "MODIFIER",
    ]
    duplicates = vendor_usage[vendor_usage.duplicated(grain, keep=False)]
    if not duplicates.empty:
        raise RuntimeError(
            "Acronis output is not unique at the required partner/SKU/status grain:\n"
            + duplicates[grain].drop_duplicates().to_string(index=False)
        )

    if require_complete_rates:
        nonzero_missing_rate = vendor_usage[
            vendor_usage["QUANTITY"].ne(0) & vendor_usage["UNIT_PRICE"].isna()
        ]
        if not nonzero_missing_rate.empty:
            missing = (
                nonzero_missing_rate.groupby(["BILLING_MONTH", "VENDOR_PRODUCT_SKU"])
                .agg(rows=("VENDOR_PARTNER_NAME", "size"), quantity=("QUANTITY", "sum"))
                .reset_index()
            )
            raise RuntimeError(
                "Acronis usage contains non-zero quantities without a supported rate:\n"
                + missing.to_string(index=False)
            )

    priced = vendor_usage["UNIT_PRICE"].notna()
    expected_amount = vendor_usage.loc[priced, "QUANTITY"] * vendor_usage.loc[priced, "UNIT_PRICE"]
    amount_delta = (
        vendor_usage.loc[priced, "AMOUNT"] - expected_amount
    ).abs()
    if amount_delta.gt(0.000001).any():
        raise RuntimeError("One or more Acronis amounts do not equal quantity multiplied by unit price.")


def to_snowflake_frame(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in SNOWFLAKE_COLUMNS:
        if col not in out.columns:
            out[col] = None
    return out[list(SNOWFLAKE_COLUMNS)]


def snowflake_ddl() -> str:
    cols = [
        "BILLING_MONTH DATE",
        "VENDOR VARCHAR",
        "VENDOR_PARTNER_NAME VARCHAR",
        "VENDOR_PRODUCT_SKU VARCHAR",
        "MODIFIER VARCHAR",
        "QUANTITY NUMBER(38, 6)",
        "UNIT_PRICE NUMBER(18, 6)",
        "AMOUNT NUMBER(38, 6)",
        "CURRENCY VARCHAR",
        "ADDITIONAL_INFO VARCHAR",
    ]
    return f"CREATE TABLE IF NOT EXISTS {FQN} ({', '.join(cols)});"


def load_snowflake(df: pd.DataFrame, *, reset: bool) -> None:
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from snowflake.connector.pandas_tools import write_pandas
    from TEMPLATES.Python.connection import get_snowflake_connection

    load_df = to_snowflake_frame(df)
    if load_df.empty:
        print("Nothing to load.")
        return
    if reset:
        print("[INFO] --reset refreshes only the billing months included in this run.")

    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )
    try:
        load_df = fill_missing_prices_dynamic(load_df, TARGET_VENDOR, conn=conn)
        quantity = pd.to_numeric(load_df["QUANTITY"], errors="coerce")
        unit_price = pd.to_numeric(load_df["UNIT_PRICE"], errors="coerce")
        calculated_amount = quantity * unit_price
        load_df["AMOUNT"] = calculated_amount.where(
            unit_price.notna(),
            load_df["AMOUNT"],
        )
        validate_vendor_usage(load_df)
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        cur.execute(snowflake_ddl())
        cur.execute(f"ALTER TABLE {FQN} ADD COLUMN IF NOT EXISTS ADDITIONAL_INFO VARCHAR")
        months = sorted(pd.to_datetime(load_df["BILLING_MONTH"]).dt.date.unique())
        placeholders = ", ".join(["%s"] * len(months))
        cur.execute(
            f"DELETE FROM {FQN} "
            f"WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s) "
            f"AND BILLING_MONTH IN ({placeholders})",
            (TARGET_VENDOR, *months),
        )
        success, chunks, rows, output = write_pandas(
            conn,
            load_df,
            TARGET_TABLE,
            database=TARGET_DATABASE,
            schema=TARGET_SCHEMA,
            quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"write_pandas failed: {output}")
        conn.commit()
        print(f"Loaded {rows:,} rows into {FQN} in {chunks} chunk(s).")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def read_manual_usage_file(path: Path, month: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(_read_file_bytes(path)), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        if _norm_compact(header[0]) not in {"entity", "month"}:
            raise RuntimeError(f"Unexpected first column in {path}: {header[0]}")
        if [_norm_compact(c) for c in header[1 : 1 + len(RAW_COLUMNS)]] != [_norm_compact(c) for c in RAW_COLUMNS]:
            raise RuntimeError(f"Manual usage schema mismatch in {path}: {header}")
        records: list[dict[str, object]] = []
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            rec = {"BILLING_MONTH": dt.date.fromisoformat(f"{month}-01")}
            rec["Entity"] = normalize_cell(row[0]) if len(row) > 0 else None
            for i, col in enumerate(RAW_COLUMNS, start=1):
                rec[col] = normalize_cell(row[i]) if i < len(row) else None
            records.append(rec)
    finally:
        wb.close()
    df = pd.DataFrame(records, columns=list(TABLE_COLUMNS))
    df["SKU"] = df["SKU"].apply(lambda v: str(v).strip().upper() if v is not None else None)
    return df


def locate_manual_usage_files(source_root: Path, month: str) -> list[Path]:
    month_folder = discover_month_folders(source_root).get(month)
    if month_folder is None:
        return []
    files: list[Path] = []
    for pattern in ("Acronis Usage*.xlsx", "Acronis ASIO Usage*.xlsx", "Acronsi ASIO Usage*.xlsx"):
        files.extend(p for p in sorted(month_folder.glob(pattern)) if not p.name.startswith("~$"))
    return files


def comparable_key_frame(df: pd.DataFrame) -> pd.DataFrame:
    cmp = df.copy()
    for col in TABLE_COLUMNS:
        if col not in cmp.columns:
            cmp[col] = None
    cmp = cmp[list(TABLE_COLUMNS)]
    for col in cmp.columns:
        if col == "BILLING_MONTH":
            continue
        if col in NUMERIC_COMPARE_COLUMNS:
            numeric = pd.to_numeric(cmp[col], errors="coerce")
            cmp[col] = numeric.map(
                lambda v: ""
                if pd.isna(v)
                else f"{0.0 if abs(float(v)) == 0 else float(v):.6f}".rstrip("0").rstrip(".")
            )
        else:
            cmp[col] = cmp[col].where(cmp[col].notna(), "").map(_normalize_compare_text)
    cmp["BILLING_MONTH"] = pd.to_datetime(cmp["BILLING_MONTH"]).dt.date.astype(str)
    return (
        cmp.groupby(list(TABLE_COLUMNS), dropna=False)
        .size()
        .reset_index(name="row_count")
    )


def validate_against_manual(source_root: Path, months: list[str], parsed: pd.DataFrame) -> pd.DataFrame:
    summaries: list[dict[str, object]] = []
    detail_records: list[pd.DataFrame] = []
    for month in months:
        manual_files = locate_manual_usage_files(source_root, month)
        manual_frames = [read_manual_usage_file(path, month) for path in manual_files]
        manual = pd.concat(manual_frames, ignore_index=True) if manual_frames else pd.DataFrame(columns=list(TABLE_COLUMNS))
        auto = parsed[pd.to_datetime(parsed["BILLING_MONTH"]).dt.date.astype(str) == f"{month}-01"].copy()

        auto_cmp = comparable_key_frame(auto)
        manual_cmp = comparable_key_frame(manual)
        diff = auto_cmp.merge(
            manual_cmp,
            on=list(TABLE_COLUMNS),
            how="outer",
            suffixes=("_auto", "_manual"),
        )
        diff["row_count_auto"] = diff["row_count_auto"].fillna(0).astype(int)
        diff["row_count_manual"] = diff["row_count_manual"].fillna(0).astype(int)
        diff["row_count_delta"] = diff["row_count_auto"] - diff["row_count_manual"]
        diff = diff[diff["row_count_delta"] != 0].copy()
        diff["month"] = month
        if not diff.empty:
            detail_records.append(diff)

        summaries.append(
            {
                "month": month,
                "manual_files": " | ".join(path.name for path in manual_files),
                "auto_rows": len(auto),
                "manual_rows": len(manual),
                "auto_distinct_rows": len(auto_cmp),
                "manual_distinct_rows": len(manual_cmp),
                "diff_distinct_rows": len(diff),
                "net_row_delta": int(diff["row_count_delta"].sum()) if not diff.empty else 0,
            }
        )

    summary = pd.DataFrame(summaries)
    detail = pd.concat(detail_records, ignore_index=True) if detail_records else pd.DataFrame()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summary.to_csv(OUTPUT_DIR / "acronis_usage_file_recreated_vs_manual_summary.csv", index=False)
    detail.to_csv(OUTPUT_DIR / "acronis_usage_file_recreated_vs_manual_diff.csv", index=False)
    print(summary.to_string(index=False))
    return summary


def write_local_outputs(
    df: pd.DataFrame,
    vendor_usage: pd.DataFrame,
    scan_df: pd.DataFrame,
    label: str,
) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_DIR / f"acronis_usage_file_recreated_{label}.csv", index=False, quoting=csv.QUOTE_MINIMAL)
    vendor_usage.to_csv(
        OUTPUT_DIR / f"acronis_vendor_usage_{label}.csv",
        index=False,
        quoting=csv.QUOTE_MINIMAL,
    )
    scan_df.to_csv(OUTPUT_DIR / f"acronis_usage_file_recreated_scan_{label}.csv", index=False, quoting=csv.QUOTE_MINIMAL)
    audit = (
        df.groupby(["BILLING_MONTH", "Entity"], dropna=False)
        .agg(row_count=("SKU", "size"), sku_rows=("SKU", "count"), distinct_tenants=("Tenant name", "nunique"))
        .reset_index()
    )
    audit.to_csv(OUTPUT_DIR / f"acronis_usage_file_recreated_audit_{label}.csv", index=False)
    rate_audit = (
        vendor_usage.assign(
            rate_missing=vendor_usage["UNIT_PRICE"].isna(),
            amount_missing=vendor_usage["AMOUNT"].isna(),
        )
        .groupby(["BILLING_MONTH", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(
            row_count=("VENDOR_PARTNER_NAME", "size"),
            quantity=("QUANTITY", "sum"),
            missing_rate_rows=("rate_missing", "sum"),
            missing_amount_rows=("amount_missing", "sum"),
        )
        .reset_index()
    )
    rate_audit.to_csv(
        OUTPUT_DIR / f"acronis_vendor_usage_rate_audit_{label}.csv",
        index=False,
    )


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Recreate exact-schema Acronis usage files from raw portal CSVs.")
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT))
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--month", help="YYYY-MM")
    group.add_argument("--all-months", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--reset", action="store_true")
    parser.add_argument("--validate-manual", action="store_true")
    parser.add_argument(
        "--allow-incomplete",
        action="store_true",
        help="Continue despite unknown entities, schema changes, or unknown SKU prefixes.",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    source_root = Path(args.source_root)
    months = list(discover_month_folders(source_root).keys()) if args.all_months else [args.month]

    frames: list[pd.DataFrame] = []
    scans: list[pd.DataFrame] = []
    for month in months:
        df, scan_df = parse_month(source_root, month)
        if not args.allow_incomplete:
            validate_scan(scan_df)
        frames.append(df)
        scans.append(scan_df)

    all_rows = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=list(PARSED_COLUMNS))
    scan_all = pd.concat(scans, ignore_index=True) if scans else pd.DataFrame()
    vendor_usage = build_vendor_usage_frame(all_rows)
    if not args.allow_incomplete:
        # In dry-run, prices have not yet passed through dynamic backfill.
        validate_vendor_usage(vendor_usage, require_complete_rates=not args.dry_run)
    label = "all_months" if args.all_months else args.month.replace("-", "_")
    write_local_outputs(all_rows, vendor_usage, scan_all, label)
    print(f"TOTAL rows={len(all_rows):,}, entities={all_rows['Entity'].nunique():,}, tenants={all_rows['Tenant name'].nunique():,}")

    if args.validate_manual:
        validate_against_manual(source_root, months, all_rows)
    if args.dry_run:
        print("Dry run complete. Snowflake load skipped.")
        return
    load_snowflake(vendor_usage, reset=args.reset)


if __name__ == "__main__":
    main()
