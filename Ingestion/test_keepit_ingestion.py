from __future__ import annotations

import datetime as dt
import sys
import tempfile
from pathlib import Path
from unittest import TestCase, mock

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import KeepIT_Vendor_Usage_Ingestion_Prod as keepit


class KeepITIngestionTests(TestCase):
    def test_main_excludes_aggregate_ms_sheet(self) -> None:
        workbook = openpyxl.Workbook()
        datacenter = workbook.active
        datacenter.title = "us-dc"
        headers = ["companyname2", "fullname2", "description1", "units1", "unit-price1", "price1"]
        datacenter.append(headers)
        datacenter.append(["Partner A", "Partner A", "Google Workspace consumption", 5, 2, 10])

        aggregate = workbook.create_sheet("MS sheet")
        aggregate.append(headers)
        aggregate.append(["Partner A", "Partner A", "Google Workspace consumption", 5, 2, 10])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "main-summary.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_main(path, dt.date(2026, 2, 1))

        self.assertEqual(len(rows), 1)
        self.assertAlmostEqual(sum(float(row["QUANTITY"]) for row in rows), 5.0)
        self.assertAlmostEqual(sum(float(row["AMOUNT"]) for row in rows), 10.0)

    def test_locate_usage_files_uses_excel_filename_routing(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            month = Path(directory)

            main_wb = openpyxl.Workbook()
            main_path = month / "Connectwise-Febr26 Summary.xlsx"
            main_wb.save(main_path)

            takeout_wb = openpyxl.Workbook()
            takeout_path = month / "Takeout Invoice Post year 3 - February 2026 Summary.xlsx"
            takeout_wb.save(takeout_path)

            promo_wb = openpyxl.Workbook()
            promo_path = month / "connectwise-promo-partners-2026-02 Summary.xlsx"
            promo_wb.save(promo_path)

            located = keepit.locate_usage_files(month)

        families = [family for family, _ in located]
        self.assertEqual(families.count("PROMO"), 1)
        self.assertEqual(families.count("TAKEOUT"), 1)
        self.assertEqual(families.count("MAIN"), 1)
        self.assertEqual([path for family, path in located if family == "MAIN"][0].name, main_path.name)
        self.assertEqual([path for family, path in located if family == "PROMO"][0].name, promo_path.name)
        self.assertEqual([path for family, path in located if family == "TAKEOUT"][0].name, takeout_path.name)

    def test_promo_summary_uses_guid_match_and_summary_amount(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.cell(4, 9, "Takeout partners due for invoicing (years 2 and 3 of Promo)")
        sheet.cell(4, 14, "min comm")
        sheet.cell(4, 15, "Consumption")
        sheet.cell(4, 21, "Total per partner")
        sheet.cell(5, 9, "Partner A")
        sheet.cell(5, 10, "abcd12-efgh34-ijkl56")
        sheet.cell(5, 14, 100)
        sheet.cell(5, 15, 8)
        sheet.cell(5, 21, 40)

        dc = workbook.create_sheet("us-dc")
        headers = [
            "account1",
            "fullname2",
            "description1",
            "units1",
            "unit-price1",
            "price1",
        ]
        for column, header in enumerate(headers, start=1):
            dc.cell(1, column, header)
        dc.append(["abcd12-efgh34-ijkl56", "Partner A", "Microsoft 365 Mailbox OneDrive Total consumption (2026-06-01 to 2026-06-30)", 5, 1, 5])
        dc.append(["abcd12-efgh34-ijkl56", "Partner A", "Google Workspace consumption (2026-06-01 to 2026-06-30)", 3, 1, 3])
        dc.append(["xxxx11-yyyy22-zzzz33", "Partner B", "Microsoft 365 Mailbox OneDrive Total consumption (2026-06-01 to 2026-06-30)", 7, 1, 7])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "promo-summary.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_promo(path, dt.date(2026, 7, 1))

        self.assertTrue(rows)
        self.assertEqual({row["VENDOR_PARTNER_NAME"] for row in rows}, {"Partner A"})
        self.assertEqual({row["MODIFIER"] for row in rows}, {"PROMO"})
        self.assertEqual({row["VENDOR_PRODUCT_SKU"] for row in rows}, {"KI-M365-FUL", "KI-GOOG-FUL"})
        self.assertAlmostEqual(sum(float(row["QUANTITY"]) for row in rows), 8.0)
        self.assertAlmostEqual(sum(float(row["AMOUNT"]) for row in rows), 40.0)
        self.assertEqual({row["ADDITIONAL_INFO"] for row in rows}, {"Min Commit 100 units"})

    def test_detail_quantity_and_unit_price_are_preserved_when_summary_consumption_differs(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.cell(4, 9, "Takeout partners due for invoicing (years 2 and 3 of Promo)")
        sheet.cell(4, 14, "Consumption")
        sheet.cell(4, 21, "Total per partner")
        sheet.cell(5, 9, "Partner A")
        sheet.cell(5, 10, "abcd12-efgh34-ijkl56")
        sheet.cell(5, 14, 99)
        sheet.cell(5, 21, 40)

        dc = workbook.create_sheet("us-dc")
        headers = ["account1", "fullname2", "description1", "units1", "unit-price1", "price1"]
        for column, header in enumerate(headers, start=1):
            dc.cell(1, column, header)
        dc.append(["abcd12-efgh34-ijkl56", "Partner A", "Microsoft 365 Mailbox OneDrive Total consumption (2026-06-01 to 2026-06-30)", 5, 2.0, 10])
        dc.append(["abcd12-efgh34-ijkl56", "Partner A", "Google Workspace consumption (2026-06-01 to 2026-06-30)", 3, 1.0, 3])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "promo-summary-mismatch.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_promo(path, dt.date(2026, 7, 1))

        self.assertTrue(rows)
        self.assertAlmostEqual(sum(float(row["QUANTITY"]) for row in rows), 8.0)
        prices = {row["VENDOR_PRODUCT_SKU"]: float(row["UNIT_PRICE"]) for row in rows}
        self.assertAlmostEqual(prices["KI-M365-FUL"], 2.0)
        self.assertAlmostEqual(prices["KI-GOOG-FUL"], 1.0)
        self.assertAlmostEqual(sum(float(row["AMOUNT"]) for row in rows), 40.0)

    def test_takeout_post_year_summary_filters_detail_tabs_by_guid(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.cell(5, 9, "Partners from  yr 4 and onwards")
        sheet.cell(5, 14, "Consumption")
        sheet.cell(5, 15, "Total per partner")
        sheet.cell(7, 9, "Partner X")
        sheet.cell(7, 10, "guid-aaaa-bbbb")
        sheet.cell(7, 14, 11)
        sheet.cell(7, 15, 22)

        dc = workbook.create_sheet("uk-ld")
        headers = [
            "account1",
            "fullname2",
            "description1",
            "units1",
            "unit-price1",
            "price1",
        ]
        for column, header in enumerate(headers, start=1):
            dc.cell(1, column, header)
        dc.append(["guid-aaaa-bbbb", "Partner X", "Google Workspace consumption (2026-06-01 to 2026-06-30)", 11, 1, 11])
        dc.append(["guid-cccc-dddd", "Partner Y", "Google Workspace consumption (2026-06-01 to 2026-06-30)", 13, 1, 13])

        aggregate = workbook.create_sheet("MS sheet")
        for column, header in enumerate(headers, start=1):
            aggregate.cell(1, column, header)
        aggregate.append(["guid-aaaa-bbbb", "Partner X", "Google Workspace consumption (2026-06-01 to 2026-06-30)", 11, 1, 11])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "takeout-post-year.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_takeout_workbook(path, dt.date(2026, 7, 1))

        self.assertTrue(rows)
        self.assertEqual({row["VENDOR_PARTNER_NAME"] for row in rows}, {"Partner X"})
        self.assertEqual({row["VENDOR_PRODUCT_SKU"] for row in rows}, {"KI-GOOG-FUL"})
        self.assertEqual({row["MODIFIER"] for row in rows}, {"TAKEOUT"})
        self.assertAlmostEqual(sum(float(row["QUANTITY"]) for row in rows), 11.0)
        self.assertAlmostEqual(sum(float(row["AMOUNT"]) for row in rows), 11.0)

    def test_matched_guid_keeps_summary_partner_label(self) -> None:
        workbook = openpyxl.Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.cell(4, 9, "Takeout partners due for invoicing (years 2 and 3 of Promo)")
        summary.cell(4, 14, "Consumption")
        summary.cell(4, 21, "Total per partner")
        summary.cell(5, 9, "ATS")
        summary.cell(5, 10, "guid-ats-001")
        summary.cell(5, 14, 10)
        summary.cell(5, 21, 20)

        detail = workbook.create_sheet("us-dc")
        detail.append(["account1", "fullname2", "description1", "units1", "unit-price1", "price1"])
        detail.append(["guid-ats-001", "ATS Communications", "Google Workspace consumption", 10, 2, 20])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "promo-alias-name.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_promo(path, dt.date(2026, 7, 1))

        self.assertTrue(rows)
        self.assertEqual({row["VENDOR_PARTNER_NAME"] for row in rows}, {"ATS"})
        self.assertAlmostEqual(sum(float(row["AMOUNT"]) for row in rows), 20.0)

    def test_unmatched_promo_summary_guid_creates_visible_missing_sku_row(self) -> None:
        workbook = openpyxl.Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.cell(4, 9, "Takeout partners due for invoicing (years 2 and 3 of Promo)")
        summary.cell(4, 14, "Consumption")
        summary.cell(4, 21, "Total per partner")
        summary.cell(5, 9, "Unmatched Partner")
        summary.cell(5, 10, "guid-none-found")
        summary.cell(5, 14, 50)
        summary.cell(5, 21, 100)

        detail = workbook.create_sheet("us-dc")
        detail.append(["account1", "fullname2", "description1", "units1", "unit-price1", "price1"])
        detail.append(["different-guid-here", "Other Partner", "Google Workspace consumption", 5, 1, 5])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "promo-unmatched.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_promo(path, dt.date(2026, 7, 1))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["VENDOR_PARTNER_NAME"], "Unmatched Partner")
        self.assertEqual(rows[0]["VENDOR_PRODUCT_SKU"], "Missing Vendor Usage by SKU")
        self.assertEqual(rows[0]["MODIFIER"], "PROMO")
        self.assertAlmostEqual(float(rows[0]["QUANTITY"]), 50.0)
        self.assertAlmostEqual(float(rows[0]["AMOUNT"]), 100.0)
        self.assertIsNone(rows[0]["UNIT_PRICE"])
        self.assertIn("Summary GUID not matched to Account1 detail", str(rows[0]["ADDITIONAL_INFO"]))

    def test_summary_note_with_consumption_word_does_not_shift_metric_column(self) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Summary"
        sheet.cell(1, 15, "check formulas for consumption in col O")
        sheet.cell(4, 9, "Takeout partners due for invoicing (years 2 and 3 of Promo)")
        sheet.cell(4, 13, "min comm")
        sheet.cell(4, 14, "Consumption")
        sheet.cell(4, 15, "Diff.")
        sheet.cell(4, 26, "Total per partner")
        sheet.cell(5, 9, "Campfire I T Ltd")
        sheet.cell(5, 10, "ovdnzx-3eizfp-96n1bw")
        sheet.cell(5, 13, 100)
        sheet.cell(5, 14, 1)
        sheet.cell(5, 15, -99)
        sheet.cell(5, 26, 75)

        detail = workbook.create_sheet("us-dc")
        headers = ["account1", "fullname2", "description1", "units1", "unit-price1", "price1"]
        for column, header in enumerate(headers, start=1):
            detail.cell(1, column, header)
        detail.append([
            "ovdnzx-3eizfp-96n1bw",
            "Campfire I T Ltd",
            "Google Workspace consumption (2026-07-01 to 2026-07-31)",
            1,
            75,
            75,
        ])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "promo-summary-note.xlsx"
            workbook.save(path)
            rows, _ = keepit.parse_promo(path, dt.date(2026, 7, 1))

        self.assertTrue(rows)
        self.assertAlmostEqual(sum(float(row["QUANTITY"]) for row in rows), 1.0)
        self.assertAlmostEqual(sum(float(row["AMOUNT"]) for row in rows), 75.0)
        self.assertEqual({row["ADDITIONAL_INFO"] for row in rows}, {"Min Commit 100 units"})

    def test_january_yyyy_mm_selector_is_not_truncated(self) -> None:
        empty = pd.DataFrame(columns=keepit.TARGET_COLUMNS)
        with tempfile.TemporaryDirectory() as directory:
            month_folder = Path(directory) / "01_JAN_2026"
            month_folder.mkdir()
            with mock.patch.object(keepit, "load_month", return_value=(empty, pd.DataFrame())) as loader:
                keepit.load_all(Path(directory), months=["2026-01"])

        loader.assert_called_once_with(month_folder)


if __name__ == "__main__":
    import unittest

    unittest.main()
