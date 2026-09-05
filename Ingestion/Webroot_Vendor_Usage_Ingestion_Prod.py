"""Ingest Webroot Aggregator Order Details into Snowflake.

Business contract:
  * Webroot CW and Webroot CMS are handled as one vendor usage feed.
  * The only external vendor source ingested here is Aggregator Order Details.
  * Endpoint, DNS/SAT, Zuora export, overage, invoice, and manual recon files
    are evidence or billing-side inputs, not vendor usage inputs.
  * Zero-dollar usage rows are retained and flagged; printed subtotal rows are
    excluded from usage and captured in the file audit controls.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import json
import re
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
import pandas as pd


WEBROOT_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = WEBROOT_ROOT.parent
WORKSPACE_ROOT = next((p for p in (PROJECT_ROOT, *PROJECT_ROOT.parents) if (p / "TEMPLATES").exists()), PROJECT_ROOT)

DEFAULT_SOURCE_ROOT_CW = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Webroot CW"
)
DEFAULT_SOURCE_ROOT_CMS = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Webroot CMS"
)
MAPPINGS_FILE = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\Mappings\Webroot\Webroot_Mappings.xlsx"
)

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
AUDIT_TABLE = "WEBROOT_USAGE_FILE_AUDIT"
PARTNER_SEED_TABLE = "WEBROOT_PARTNER_MAP_SEED"
TARGET_VENDOR = "Webroot"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)
MONTH_ABBREVIATIONS = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AUG", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC",
}

SOURCE_COLUMNS: tuple[str, ...] = (
    "Billing Date",
    "Customer Code",
    "Company Name",
    "Reseller Customer Code",
    "Reseller Company Name",
    "Webroot Entity",
    "Location Code",
    "State",
    "Keycode",
    "License Type Description",
    "License Seats",
    "Usage Seats",
    "Total Seats",
    "Order Date",
    "License Category Name",
    "Keycode Age",
    "Retail Price",
    "Total Extended Amount",
    "Cap Amount",
    "Total Cap Amount",
)

USAGE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
    "ADDITIONAL_INFO",
)

INTERNAL_USAGE_COLUMNS: tuple[str, ...] = USAGE_COLUMNS + (
    "_CHANNEL",
    "_SOURCE_FILE",
    "_SOURCE_SHEET",
    "_SOURCE_ROW_NUMBER",
    "_SOURCE_CONTENT_HASH",
    "_CUSTOMER_CODE",
    "_RESELLER_CUSTOMER_CODE",
    "_RESELLER_COMPANY_NAME",
    "_KEYCODE",
    "_LICENSE_SEATS",
    "_USAGE_SEATS",
    "_CAP_AMOUNT",
    "_TOTAL_CAP_AMOUNT",
)

AUDIT_COLUMNS: tuple[str, ...] = (
    "STREAM",
    "CHANNEL",
    "SOURCE_FOLDER",
    "SOURCE_FILE",
    "SOURCE_SHEET",
    "SOURCE_CONTENT_HASH",
    "LOAD_STATUS",
    "BILLING_MONTH",
    "BILLING_DATE",
    "LICENSE_CATEGORY_NAME",
    "DATA_ROW_COUNT",
    "CHARGEABLE_ROW_COUNT",
    "ZERO_AMOUNT_ROW_COUNT",
    "SUBTOTAL_ROW_COUNT",
    "IGNORED_NOISE_ROW_COUNT",
    "INVALID_DATA_ROW_COUNT",
    "DUPLICATE_NATURAL_KEY_COUNT",
    "RECOMPUTED_TOTAL_SEATS",
    "PRINTED_TOTAL_SEATS",
    "TOTAL_SEATS_DELTA",
    "RECOMPUTED_TOTAL_EXTENDED_AMOUNT",
    "PRINTED_TOTAL_EXTENDED_AMOUNT",
    "TOTAL_EXTENDED_AMOUNT_DELTA",
    "ERROR_MESSAGE",
    "INGESTED_AT",
)

PARTNER_SEED_COLUMNS: tuple[str, ...] = (
    "VENDOR",
    "PARTNER_NAME",
    "PARENT_COMPANY",
    "SF_ID",
    "CMS_ID",
    "ZUORA_NAME",
)

WEBROOT_PRODUCT_GROUPS = {
    "SAEP": "GSM",
    "SDNS": "DNS",
    "SECA": "SAT",
}
EXPECTED_CATEGORIES = frozenset(WEBROOT_PRODUCT_GROUPS)
EXPECTED_MONTHLY_SOURCES = frozenset({
    ("CW", "MSP"),
    ("CW", "RESELLER"),
    ("CMS", "MSP"),
})


@dataclass(frozen=True)
class SourceFile:
    stream: str
    channel: str
    path: Path


def _read_file_bytes(path: Path) -> bytes:
    """Read bytes, falling back to PowerShell Copy-Item for locked OneDrive files."""
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir) / path.name
            escaped_src = str(path).replace("'", "''")
            escaped_dest = str(tmp).replace("'", "''")
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    f"Copy-Item -LiteralPath '{escaped_src}' -Destination '{escaped_dest}' -Force",
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            return tmp.read_bytes()


def _clean_text(value: object) -> str | None:
    if value is None or bool(pd.isna(value)):
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text if text else None


def _to_number(value: object) -> float | None:
    if value is None or bool(pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("$", "").strip()
    if text == "":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_date(value: object) -> dt.date | None:
    if value is None or bool(pd.isna(value)):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    parsed = pd.to_datetime(str(value), errors="coerce")
    return parsed.date() if not pd.isna(parsed) else None


def _first_day(value: dt.date | None) -> dt.date | None:
    return dt.date(value.year, value.month, 1) if value else None


def _vendor_product_sku(value: object) -> str | None:
    text = _clean_text(value)
    if text is None:
        return None
    return WEBROOT_PRODUCT_GROUPS.get(text.upper(), text)


def _month_from_folder(path: Path) -> str | None:
    for parent in path.parents:
        match = MONTH_FOLDER_RE.match(parent.name)
        if match:
            month_number = int(match.group("mm"))
            expected_abbreviation = MONTH_ABBREVIATIONS.get(month_number)
            actual_abbreviation = parent.name.split("_", 2)[1].upper()
            if expected_abbreviation != actual_abbreviation:
                raise ValueError(f"Invalid month folder name: {parent.name}")
            return f"{match.group('yyyy')}-{month_number:02d}"
    return None


def _billing_month_from_source(path: Path, billing_date: dt.date | None) -> dt.date | None:
    folder_month = _month_from_folder(path)
    if folder_month:
        return dt.date.fromisoformat(f"{folder_month}-01")
    return _first_day(billing_date)


def _source_period_mismatch(path: Path, billing_date: dt.date | None) -> str | None:
    """Describe a folder/source billing-month conflict that must block loading."""
    folder_month = _month_from_folder(path)
    if not folder_month or billing_date is None:
        return None
    billing_date_month = billing_date.strftime("%Y-%m")
    if folder_month == billing_date_month:
        return None
    return (
        f"Source folder month {folder_month} disagrees with embedded billing "
        f"date month {billing_date_month}."
    )


def _canonical_row(values: Iterable[object]) -> str:
    return "\t".join(_clean_text(v) or "" for v in values)


def _content_hash(rows: list[tuple[object, ...]]) -> str:
    payload = "\n".join(_canonical_row(row[: len(SOURCE_COLUMNS)]) for row in rows)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def discover_source_files(
    source_root_cw: Path,
    source_root_cms: Path,
    months: set[str] | None = None,
) -> list[SourceFile]:
    files: list[SourceFile] = []
    for stream, root in (("CW", source_root_cw), ("CMS", source_root_cms)):
        if not root.exists():
            print(f"WARNING: source root does not exist: {root}", flush=True)
            continue
        for path in sorted(root.rglob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            if "Aggregator Order Details" not in path.name:
                continue
            source_month = _month_from_folder(path)
            if source_month is None:
                raise RuntimeError(
                    f"Aggregator Order Details file is outside a monthly folder: {path.name}"
                )
            if months is not None and source_month not in months:
                continue
            channel = "RESELLER" if "RESELLER" in path.name.upper() else "MSP"
            files.append(SourceFile(stream=stream, channel=channel, path=path))
    return files


def _find_source_sheet(
    wb: openpyxl.Workbook,
) -> tuple[str, openpyxl.worksheet.worksheet.Worksheet]:
    matches: list[tuple[str, openpyxl.worksheet.worksheet.Worksheet]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header = tuple(_clean_text(v) for v in raw_header[: len(SOURCE_COLUMNS)])
        if header == SOURCE_COLUMNS:
            matches.append((sheet_name, ws))
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected exactly one Aggregator Order Details sheet, found {len(matches)}."
        )
    return matches[0]


def parse_aggregator_file(source: SourceFile, *, ingested_at: dt.datetime) -> tuple[pd.DataFrame, pd.DataFrame]:
    file_bytes = _read_file_bytes(source.path)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        try:
            sheet_name, ws = _find_source_sheet(wb)
        except RuntimeError as exc:
            audit = pd.DataFrame(
                [{
                    "STREAM": source.stream,
                    "CHANNEL": source.channel,
                    "SOURCE_FOLDER": source.path.parent.name,
                    "SOURCE_FILE": source.path.name,
                    "SOURCE_SHEET": None,
                    "SOURCE_CONTENT_HASH": hashlib.sha256(file_bytes).hexdigest(),
                    "LOAD_STATUS": "FAILED_SCHEMA",
                    "BILLING_MONTH": _billing_month_from_source(source.path, None),
                    "BILLING_DATE": None,
                    "LICENSE_CATEGORY_NAME": None,
                    "DATA_ROW_COUNT": 0,
                    "CHARGEABLE_ROW_COUNT": 0,
                    "ZERO_AMOUNT_ROW_COUNT": 0,
                    "SUBTOTAL_ROW_COUNT": 0,
                    "IGNORED_NOISE_ROW_COUNT": 0,
                    "INVALID_DATA_ROW_COUNT": 0,
                    "DUPLICATE_NATURAL_KEY_COUNT": 0,
                    "RECOMPUTED_TOTAL_SEATS": None,
                    "PRINTED_TOTAL_SEATS": None,
                    "TOTAL_SEATS_DELTA": None,
                    "RECOMPUTED_TOTAL_EXTENDED_AMOUNT": None,
                    "PRINTED_TOTAL_EXTENDED_AMOUNT": None,
                    "TOTAL_EXTENDED_AMOUNT_DELTA": None,
                    "ERROR_MESSAGE": str(exc),
                    "INGESTED_AT": ingested_at,
                }],
                columns=AUDIT_COLUMNS,
            )
            return pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS), audit
        row_iter = ws.iter_rows(values_only=True)
        raw_header = next(row_iter)
        header = tuple(_clean_text(v) for v in raw_header[: len(SOURCE_COLUMNS)])
        if header != SOURCE_COLUMNS:
            audit = pd.DataFrame(
                [
                    {
                        "STREAM": source.stream,
                        "CHANNEL": source.channel,
                        "SOURCE_FOLDER": source.path.parent.name,
                        "SOURCE_FILE": source.path.name,
                        "SOURCE_SHEET": sheet_name,
                        "SOURCE_CONTENT_HASH": hashlib.sha256(file_bytes).hexdigest(),
                        "LOAD_STATUS": "FAILED_SCHEMA",
                        "BILLING_MONTH": None,
                        "BILLING_DATE": None,
                        "LICENSE_CATEGORY_NAME": None,
                        "DATA_ROW_COUNT": 0,
                        "CHARGEABLE_ROW_COUNT": 0,
                        "ZERO_AMOUNT_ROW_COUNT": 0,
                        "SUBTOTAL_ROW_COUNT": 0,
                        "IGNORED_NOISE_ROW_COUNT": 0,
                        "INVALID_DATA_ROW_COUNT": 0,
                        "DUPLICATE_NATURAL_KEY_COUNT": 0,
                        "RECOMPUTED_TOTAL_SEATS": None,
                        "PRINTED_TOTAL_SEATS": None,
                        "TOTAL_SEATS_DELTA": None,
                        "RECOMPUTED_TOTAL_EXTENDED_AMOUNT": None,
                        "PRINTED_TOTAL_EXTENDED_AMOUNT": None,
                        "TOTAL_EXTENDED_AMOUNT_DELTA": None,
                        "ERROR_MESSAGE": f"Unexpected header: {header}",
                        "INGESTED_AT": ingested_at,
                    }
                ],
                columns=AUDIT_COLUMNS,
            )
            return pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS), audit

        data_records: list[dict[str, object]] = []
        subtotal_records: list[dict[str, object]] = []
        content_rows: list[tuple[object, ...]] = []
        for source_row_number, raw_row in enumerate(row_iter, start=2):
            row = tuple(raw_row[: len(SOURCE_COLUMNS)])
            cleaned = [_clean_text(v) for v in row]
            if not any(cleaned):
                continue
            content_rows.append(row)
            row_dict = dict(zip(SOURCE_COLUMNS, row, strict=True))
            customer_code = _clean_text(row_dict["Customer Code"])
            if customer_code and customer_code.lower() == "subtotal":
                subtotal_records.append(row_dict | {"source_row_number": source_row_number})
                continue
            data_records.append(row_dict | {"source_row_number": source_row_number})
    finally:
        wb.close()

    source_content_hash = _content_hash(content_rows)
    data_df = pd.DataFrame(data_records)
    subtotal_df = pd.DataFrame(subtotal_records)

    if data_df.empty:
        usage_df = pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS)
        audit_df = _build_audit(
            source=source,
            sheet_name=sheet_name,
            source_content_hash=source_content_hash,
            data_df=data_df,
            subtotal_df=subtotal_df,
            load_status="FAILED_EMPTY_SOURCE",
            error_message="The source file contains no detail rows.",
            ingested_at=ingested_at,
        )
        return usage_df, audit_df

    for col in (
        "License Seats",
        "Usage Seats",
        "Total Seats",
        "Keycode Age",
        "Retail Price",
        "Total Extended Amount",
        "Cap Amount",
        "Total Cap Amount",
    ):
        data_df[col] = data_df[col].map(_to_number)

    company_names = data_df["Company Name"].map(_clean_text)
    customer_codes = data_df["Customer Code"].map(_clean_text)
    keycodes = data_df["Keycode"].map(_clean_text)
    categories = data_df["License Category Name"].map(_clean_text)
    billing_dates = data_df["Billing Date"].map(_to_date)
    unknown_categories = sorted(set(categories.dropna()) - EXPECTED_CATEGORIES)
    required_valid = (
        billing_dates.notna()
        & customer_codes.notna()
        & company_names.notna()
        & keycodes.notna()
        & categories.notna()
        & data_df["Total Seats"].notna()
        & data_df["Retail Price"].notna()
        & data_df["Total Extended Amount"].notna()
    )
    invalid_data_row_count = int((~required_valid).sum())
    natural_keys = pd.DataFrame(
        {
            "billing_date": billing_dates,
            "customer_code": customer_codes.str.casefold(),
            "keycode": keycodes.str.casefold(),
            "license_category": categories.str.upper(),
        },
        index=data_df.index,
    )
    duplicate_natural_key_count = int(natural_keys.duplicated(keep=False).sum())
    amount_delta = (
        data_df["Total Seats"] * data_df["Retail Price"]
        - data_df["Total Extended Amount"]
    ).abs()
    arithmetic_mismatch_count = int(amount_delta.gt(0.0001).sum())
    unique_billing_dates = sorted(set(billing_dates.dropna()))

    validation_errors: list[str] = []
    if unknown_categories:
        validation_errors.append(f"Unexpected license categories: {unknown_categories}")
    if invalid_data_row_count:
        validation_errors.append(
            f"{invalid_data_row_count} rows are missing required usage fields"
        )
    if duplicate_natural_key_count:
        validation_errors.append(
            f"{duplicate_natural_key_count} rows duplicate the billing/customer/keycode/category key"
        )
    if arithmetic_mismatch_count:
        validation_errors.append(
            f"{arithmetic_mismatch_count} rows do not equal Total Seats x Retail Price"
        )
    if len(unique_billing_dates) != 1:
        validation_errors.append(f"Expected one billing date, found {unique_billing_dates}")

    first_billing_date = unique_billing_dates[0] if len(unique_billing_dates) == 1 else None
    if validation_errors:
        audit_df = _build_audit(
            source=source,
            sheet_name=sheet_name,
            source_content_hash=source_content_hash,
            data_df=data_df,
            subtotal_df=subtotal_df,
            load_status="FAILED_DATA_VALIDATION",
            error_message="; ".join(validation_errors),
            ingested_at=ingested_at,
            invalid_data_row_count=invalid_data_row_count,
            duplicate_natural_key_count=duplicate_natural_key_count,
        )
        return pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS), audit_df
    period_error = _source_period_mismatch(source.path, first_billing_date)
    if period_error:
        audit_df = _build_audit(
            source=source,
            sheet_name=sheet_name,
            source_content_hash=source_content_hash,
            data_df=data_df,
            subtotal_df=subtotal_df,
            load_status="FAILED_PERIOD_MISMATCH",
            error_message=period_error,
            ingested_at=ingested_at,
        )
        return pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS), audit_df

    billing_month = _billing_month_from_source(source.path, first_billing_date)
    billing_months = pd.Series([billing_month] * len(data_df), index=data_df.index)
    usage_df = pd.DataFrame(
        {
            "BILLING_MONTH": billing_months,
            "VENDOR": "Webroot",
            "VENDOR_PARTNER_NAME": data_df["Company Name"].map(_clean_text),
            "VENDOR_PRODUCT_SKU": data_df["License Category Name"].map(_vendor_product_sku),
            "MODIFIER": source.stream,
            "QUANTITY": data_df["Total Seats"],
            "UNIT_PRICE": data_df["Retail Price"],
            "AMOUNT": data_df["Total Extended Amount"],
            "CURRENCY": "USD",
            "ADDITIONAL_INFO": None,
            "_CHANNEL": source.channel,
            "_SOURCE_FILE": source.path.name,
            "_SOURCE_SHEET": sheet_name,
            "_SOURCE_ROW_NUMBER": data_df["source_row_number"],
            "_SOURCE_CONTENT_HASH": source_content_hash,
            "_CUSTOMER_CODE": customer_codes,
            "_RESELLER_CUSTOMER_CODE": data_df["Reseller Customer Code"].map(_clean_text),
            "_RESELLER_COMPANY_NAME": data_df["Reseller Company Name"].map(_clean_text),
            "_KEYCODE": keycodes,
            "_LICENSE_SEATS": data_df["License Seats"],
            "_USAGE_SEATS": data_df["Usage Seats"],
            "_CAP_AMOUNT": data_df["Cap Amount"],
            "_TOTAL_CAP_AMOUNT": data_df["Total Cap Amount"],
        },
        columns=INTERNAL_USAGE_COLUMNS,
    )
    audit_df = _build_audit(
        source=source,
        sheet_name=sheet_name,
        source_content_hash=source_content_hash,
        data_df=data_df,
        subtotal_df=subtotal_df,
        load_status="LOADED",
        error_message=None,
        ingested_at=ingested_at,
        invalid_data_row_count=invalid_data_row_count,
        duplicate_natural_key_count=duplicate_natural_key_count,
    )
    return usage_df, audit_df


def _build_audit(
    *,
    source: SourceFile,
    sheet_name: str,
    source_content_hash: str,
    data_df: pd.DataFrame,
    subtotal_df: pd.DataFrame,
    load_status: str,
    error_message: str | None,
    ingested_at: dt.datetime,
    ignored_noise_row_count: int = 0,
    invalid_data_row_count: int = 0,
    duplicate_natural_key_count: int = 0,
) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    all_categories = sorted(
        {
            *[c for c in data_df.get("License Category Name", pd.Series(dtype=object)).map(_clean_text).dropna().unique()],
            *[c for c in subtotal_df.get("License Category Name", pd.Series(dtype=object)).map(_clean_text).dropna().unique()],
        }
    )
    if not all_categories:
        all_categories = [None]

    working_data = data_df.copy()
    working_subtotal = subtotal_df.copy()
    for col in ("Total Seats", "Total Extended Amount"):
        if col in working_data:
            working_data[col] = working_data[col].map(_to_number)
        if col in working_subtotal:
            working_subtotal[col] = working_subtotal[col].map(_to_number)

    for category in all_categories:
        category_mask = (
            working_data["License Category Name"].map(_clean_text).eq(category)
            if not working_data.empty
            else pd.Series(dtype=bool)
        )
        subtotal_mask = (
            working_subtotal["License Category Name"].map(_clean_text).eq(category)
            if not working_subtotal.empty
            else pd.Series(dtype=bool)
        )
        category_data = working_data[category_mask] if not working_data.empty else working_data
        category_subtotals = working_subtotal[subtotal_mask] if not working_subtotal.empty else working_subtotal
        recomputed_seats = category_data["Total Seats"].sum() if "Total Seats" in category_data else None
        recomputed_amount = (
            category_data["Total Extended Amount"].sum() if "Total Extended Amount" in category_data else None
        )
        printed_seats = category_subtotals["Total Seats"].sum() if "Total Seats" in category_subtotals else None
        printed_amount = (
            category_subtotals["Total Extended Amount"].sum()
            if "Total Extended Amount" in category_subtotals
            else None
        )
        billing_date = None
        if not category_data.empty:
            billing_date = _to_date(category_data["Billing Date"].iloc[0])
        billing_month = _billing_month_from_source(source.path, billing_date)
        rows.append(
            {
                "STREAM": source.stream,
                "CHANNEL": source.channel,
                "SOURCE_FOLDER": source.path.parent.name,
                "SOURCE_FILE": source.path.name,
                "SOURCE_SHEET": sheet_name,
                "SOURCE_CONTENT_HASH": source_content_hash,
                "LOAD_STATUS": load_status,
                "BILLING_MONTH": billing_month,
                "BILLING_DATE": billing_date,
                "LICENSE_CATEGORY_NAME": category,
                "DATA_ROW_COUNT": int(len(category_data)),
                "CHARGEABLE_ROW_COUNT": int(
                    category_data["Total Extended Amount"].fillna(0).ne(0).sum()
                )
                if "Total Extended Amount" in category_data
                else 0,
                "ZERO_AMOUNT_ROW_COUNT": int(
                    category_data["Total Extended Amount"].fillna(0).eq(0).sum()
                )
                if "Total Extended Amount" in category_data
                else 0,
                "SUBTOTAL_ROW_COUNT": int(len(category_subtotals)),
                "IGNORED_NOISE_ROW_COUNT": ignored_noise_row_count,
                "INVALID_DATA_ROW_COUNT": invalid_data_row_count,
                "DUPLICATE_NATURAL_KEY_COUNT": duplicate_natural_key_count,
                "RECOMPUTED_TOTAL_SEATS": recomputed_seats,
                "PRINTED_TOTAL_SEATS": printed_seats,
                "TOTAL_SEATS_DELTA": (recomputed_seats or 0) - (printed_seats or 0)
                if printed_seats is not None
                else None,
                "RECOMPUTED_TOTAL_EXTENDED_AMOUNT": recomputed_amount,
                "PRINTED_TOTAL_EXTENDED_AMOUNT": printed_amount,
                "TOTAL_EXTENDED_AMOUNT_DELTA": (recomputed_amount or 0) - (printed_amount or 0)
                if printed_amount is not None
                else None,
                "ERROR_MESSAGE": error_message,
                "INGESTED_AT": ingested_at,
            }
        )
    return pd.DataFrame(rows, columns=AUDIT_COLUMNS)


def _unique_text(values: pd.Series) -> list[str]:
    return sorted({text for value in values if (text := _clean_text(value)) is not None})


def validate_source_completeness(
    audit_df: pd.DataFrame,
    requested_months: set[str],
) -> None:
    if audit_df.empty:
        raise RuntimeError("No Webroot Aggregator Order Details files were found.")

    file_audit = audit_df[
        ["STREAM", "CHANNEL", "SOURCE_FILE", "LOAD_STATUS", "BILLING_MONTH", "ERROR_MESSAGE"]
    ].drop_duplicates()
    failures = file_audit[
        ~file_audit["LOAD_STATUS"].isin({"LOADED", "SKIPPED_DUPLICATE_CONTENT"})
    ]
    if not failures.empty:
        raise RuntimeError(
            "Webroot source validation failed:\n" + failures.to_string(index=False)
        )

    bad_subtotals = audit_df[
        audit_df["LOAD_STATUS"].eq("LOADED")
        & (
            audit_df["SUBTOTAL_ROW_COUNT"].ne(1)
            | audit_df["TOTAL_SEATS_DELTA"].fillna(0).abs().gt(0.0001)
            | audit_df["TOTAL_EXTENDED_AMOUNT_DELTA"].fillna(0).abs().gt(0.0001)
        )
    ]
    if not bad_subtotals.empty:
        raise RuntimeError(
            "Webroot printed subtotal parity failed:\n"
            + bad_subtotals[
                ["STREAM", "CHANNEL", "SOURCE_FILE", "LICENSE_CATEGORY_NAME",
                 "SUBTOTAL_ROW_COUNT", "TOTAL_SEATS_DELTA", "TOTAL_EXTENDED_AMOUNT_DELTA"]
            ].to_string(index=False)
        )

    file_audit = file_audit.copy()
    file_audit["MONTH"] = pd.to_datetime(
        file_audit["BILLING_MONTH"], errors="coerce"
    ).dt.strftime("%Y-%m")
    for month in sorted(requested_months):
        month_files = file_audit[file_audit["MONTH"].eq(month)]
        source_counts = month_files.groupby(["STREAM", "CHANNEL"]).size().to_dict()
        actual = set(source_counts)
        if actual != EXPECTED_MONTHLY_SOURCES:
            raise RuntimeError(
                f"Webroot source manifest mismatch for {month}: "
                f"missing={sorted(EXPECTED_MONTHLY_SOURCES - actual)}, "
                f"unexpected={sorted(actual - EXPECTED_MONTHLY_SOURCES)}."
            )
        duplicate_slots = {source: count for source, count in source_counts.items() if count != 1}
        if duplicate_slots:
            raise RuntimeError(
                f"Expected one Webroot file per source slot for {month}, found {duplicate_slots}."
            )


def build_usage(
    source_root_cw: Path,
    source_root_cms: Path,
    months: set[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    ingested_at = dt.datetime.now(dt.UTC).replace(tzinfo=None)
    files = discover_source_files(source_root_cw, source_root_cms, months)
    if not files:
        requested = ", ".join(sorted(months)) if months else "any month"
        raise FileNotFoundError(f"No Aggregator Order Details workbooks found for {requested}.")

    usage_frames: list[pd.DataFrame] = []
    audit_frames: list[pd.DataFrame] = []
    seen_content_months: dict[tuple[str, str], object] = {}
    for source in files:
        usage_df, audit_df = parse_aggregator_file(source, ingested_at=ingested_at)
        content_hash = audit_df["SOURCE_CONTENT_HASH"].iloc[0]
        billing_month_key = audit_df["BILLING_MONTH"].iloc[0] if not audit_df.empty else None
        content_key = (content_hash, source.stream)
        prior_month = seen_content_months.get(content_key)
        load_status = audit_df["LOAD_STATUS"].iloc[0]
        if load_status == "LOADED" and prior_month is not None:
            if prior_month == billing_month_key:
                audit_df["LOAD_STATUS"] = "SKIPPED_DUPLICATE_CONTENT"
            else:
                audit_df["LOAD_STATUS"] = "FAILED_DUPLICATE_PERIOD_CONTENT"
                audit_df["ERROR_MESSAGE"] = (
                    f"Content duplicates {source.stream} source data already assigned "
                    f"to {prior_month}; refusing to assign it to {billing_month_key}."
                )
            usage_df = pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS)
        elif not usage_df.empty:
            seen_content_months[content_key] = billing_month_key
        audit_frames.append(audit_df)
        usage_frames.append(usage_df)

        status = audit_df["LOAD_STATUS"].iloc[0]
        rows = len(usage_df)
        qty = usage_df["QUANTITY"].sum() if not usage_df.empty else 0
        amount = usage_df["AMOUNT"].sum() if not usage_df.empty else 0
        print(
            f"{status}: {source.stream} {source.channel} {source.path.parent.name} {source.path.name} "
            f"rows={rows:,} qty={qty:,.0f} amount={amount:,.2f}",
            flush=True,
        )

    non_empty_usage_frames = [frame for frame in usage_frames if not frame.empty]
    usage_all = (
        pd.concat(non_empty_usage_frames, ignore_index=True)
        if non_empty_usage_frames
        else pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS)
    )
    audit_all = pd.concat(audit_frames, ignore_index=True) if audit_frames else pd.DataFrame(columns=AUDIT_COLUMNS)
    if usage_all.empty:
        raise RuntimeError("No Webroot usage rows parsed from eligible source files.")
    if months is not None:
        validate_source_completeness(audit_all, months)

    cross_file_key = [
        "BILLING_MONTH", "MODIFIER", "_CUSTOMER_CODE", "_KEYCODE", "VENDOR_PRODUCT_SKU"
    ]
    cross_file_duplicates = usage_all[usage_all.duplicated(cross_file_key, keep=False)]
    if not cross_file_duplicates.empty:
        raise RuntimeError(
            "Webroot usage contains duplicate natural keys across source files:\n"
            + cross_file_duplicates[
                cross_file_key + ["_CHANNEL", "_SOURCE_FILE", "_SOURCE_ROW_NUMBER"]
            ].sort_values(cross_file_key).to_string(index=False)
        )

    grain = ["BILLING_MONTH", "VENDOR", "VENDOR_PARTNER_NAME", "VENDOR_PRODUCT_SKU", "MODIFIER"]
    output = (
        usage_all.groupby(grain, dropna=False, sort=False)
        .agg(
            QUANTITY=("QUANTITY", "sum"),
            UNIT_PRICE=("UNIT_PRICE", "first"),
            _UNIT_PRICE_COUNT=("UNIT_PRICE", "nunique"),
            AMOUNT=("AMOUNT", "sum"),
            CURRENCY=("CURRENCY", "first"),
            _CURRENCY_COUNT=("CURRENCY", "nunique"),
            _CHANNEL=("_CHANNEL", "first"),
            _CHANNEL_COUNT=("_CHANNEL", "nunique"),
            _SOURCE_FILE=("_SOURCE_FILE", "first"),
            _SOURCE_FILE_COUNT=("_SOURCE_FILE", "nunique"),
            _SOURCE_SHEET=("_SOURCE_SHEET", "first"),
            _SOURCE_CONTENT_HASH=("_SOURCE_CONTENT_HASH", "first"),
            _SOURCE_ROW_MIN=("_SOURCE_ROW_NUMBER", "min"),
            _SOURCE_ROW_MAX=("_SOURCE_ROW_NUMBER", "max"),
            _SOURCE_ROW_COUNT=("_SOURCE_ROW_NUMBER", "size"),
            _KEYCODE_COUNT=("_KEYCODE", "nunique"),
            _LICENSE_SEATS=("_LICENSE_SEATS", "sum"),
            _USAGE_SEATS=("_USAGE_SEATS", "sum"),
            _CAP_AMOUNT=("_CAP_AMOUNT", "sum"),
            _TOTAL_CAP_AMOUNT=("_TOTAL_CAP_AMOUNT", "sum"),
        )
        .reset_index()
    )
    invalid_prices = output[output["_UNIT_PRICE_COUNT"].ne(1)]
    if not invalid_prices.empty:
        raise RuntimeError(
            "Expected one Webroot unit price per partner/SKU group:\n"
            + invalid_prices[grain + ["_UNIT_PRICE_COUNT"]].to_string(index=False)
        )
    invalid_currency = output[
        output["_CURRENCY_COUNT"].ne(1) | output["CURRENCY"].ne("USD")
    ]
    if not invalid_currency.empty:
        raise RuntimeError(
            "Expected USD per Webroot partner/SKU group:\n"
            + invalid_currency[grain + ["CURRENCY", "_CURRENCY_COUNT"]].to_string(index=False)
        )
    amount_delta = (output["AMOUNT"] - output["QUANTITY"] * output["UNIT_PRICE"]).abs()
    if amount_delta.gt(0.0001).any():
        raise RuntimeError(
            "Webroot amount mismatch after partner/SKU aggregation:\n"
            + output.loc[amount_delta.gt(0.0001), grain + ["QUANTITY", "UNIT_PRICE", "AMOUNT"]]
            .to_string(index=False)
        )

    def compact_provenance(row: pd.Series) -> str:
        return json.dumps(
            {
                "channel": row["_CHANNEL"],
                "channel_count": int(row["_CHANNEL_COUNT"]),
                "source_file": row["_SOURCE_FILE"],
                "source_file_count": int(row["_SOURCE_FILE_COUNT"]),
                "source_sheet": row["_SOURCE_SHEET"],
                "source_content_hash": row["_SOURCE_CONTENT_HASH"],
                "source_row_min": int(row["_SOURCE_ROW_MIN"]),
                "source_row_max": int(row["_SOURCE_ROW_MAX"]),
                "source_row_count": int(row["_SOURCE_ROW_COUNT"]),
                "keycode_count": int(row["_KEYCODE_COUNT"]),
                "license_seats": float(row["_LICENSE_SEATS"]),
                "usage_seats": float(row["_USAGE_SEATS"]),
                "total_seats": float(row["QUANTITY"]),
                "cap_amount": float(row["_CAP_AMOUNT"]),
                "total_cap_amount": float(row["_TOTAL_CAP_AMOUNT"]),
            },
            sort_keys=True,
            separators=(",", ":"),
            allow_nan=False,
        )

    output["ADDITIONAL_INFO"] = output.apply(compact_provenance, axis=1)
    output = output[list(USAGE_COLUMNS)]
    duplicates = output[output.duplicated(grain, keep=False)]
    if not duplicates.empty:
        raise RuntimeError(
            "Webroot output is not unique at the required partner/SKU grain:\n"
            + duplicates[grain].drop_duplicates().to_string(index=False)
        )
    return output, audit_all


def usage_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} (
    BILLING_MONTH DATE,
    VENDOR VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU VARCHAR,
    MODIFIER VARCHAR,
    QUANTITY NUMBER(18,4),
    UNIT_PRICE NUMBER(18,6),
    AMOUNT NUMBER(38,6),
    CURRENCY VARCHAR,
    ADDITIONAL_INFO VARCHAR
);
"""


def audit_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} (
    STREAM VARCHAR,
    CHANNEL VARCHAR,
    SOURCE_FOLDER VARCHAR,
    SOURCE_FILE VARCHAR,
    SOURCE_SHEET VARCHAR,
    SOURCE_CONTENT_HASH VARCHAR,
    LOAD_STATUS VARCHAR,
    BILLING_MONTH DATE,
    BILLING_DATE DATE,
    LICENSE_CATEGORY_NAME VARCHAR,
    DATA_ROW_COUNT NUMBER,
    CHARGEABLE_ROW_COUNT NUMBER,
    ZERO_AMOUNT_ROW_COUNT NUMBER,
    SUBTOTAL_ROW_COUNT NUMBER,
    IGNORED_NOISE_ROW_COUNT NUMBER,
    INVALID_DATA_ROW_COUNT NUMBER,
    DUPLICATE_NATURAL_KEY_COUNT NUMBER,
    RECOMPUTED_TOTAL_SEATS NUMBER(18,4),
    PRINTED_TOTAL_SEATS NUMBER(18,4),
    TOTAL_SEATS_DELTA NUMBER(18,4),
    RECOMPUTED_TOTAL_EXTENDED_AMOUNT NUMBER(18,4),
    PRINTED_TOTAL_EXTENDED_AMOUNT NUMBER(18,4),
    TOTAL_EXTENDED_AMOUNT_DELTA NUMBER(18,4),
    ERROR_MESSAGE VARCHAR,
    INGESTED_AT TIMESTAMP_NTZ
);
"""


def _snowflake_connection():
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from TEMPLATES.Python.connection import get_snowflake_connection

    return get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )


def load_partner_seed(mapping_file: Path) -> pd.DataFrame:
    if not mapping_file.exists():
        print(f"WARNING: mapping workbook not found: {mapping_file}", flush=True)
        return pd.DataFrame(columns=PARTNER_SEED_COLUMNS)

    file_bytes = _read_file_bytes(mapping_file)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        if "PARTNER_MAPPING" in wb.sheetnames:
            sheet_names = ["PARTNER_MAPPING"]
        else:
            sheet_names = [name for name in ("Webroot CW", "Webroot CMS") if name in wb.sheetnames]
        if not sheet_names:
            raise RuntimeError(
                "Mapping workbook has neither PARTNER_MAPPING nor Webroot CW/CMS sheets."
            )

        records: list[dict[str, object]] = []
        for sheet_name in sheet_names:
            rows = list(wb[sheet_name].iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, row in enumerate(rows)
                    if "VENDOR_PARTNER_NAME" in {
                        re.sub(r"[^A-Z0-9]+", "_", str(value).upper()).strip("_")
                        for value in row if value is not None
                    }
                ),
                None,
            )
            if header_index is None:
                raise RuntimeError(f"Could not find the partner-map header on {sheet_name}.")
            header = [_clean_text(value) for value in rows[header_index]]
            normalized_header = [
                re.sub(r"[^A-Z0-9]+", "_", value.upper()).strip("_") if value else ""
                for value in header
            ]
            header_lookup = {name: idx for idx, name in enumerate(normalized_header)}

            def value_at(row: tuple[object, ...], *names: str) -> str | None:
                for name in names:
                    idx = header_lookup.get(name)
                    if idx is not None and idx < len(row):
                        value = _clean_text(row[idx])
                        if value is not None:
                            return value
                return None

            for row in rows[header_index + 1:]:
                partner_name = value_at(row, "VENDOR_PARTNER_NAME", "ACCOUNT_NAME")
                sf_id = value_at(row, "SF_ID")
                if not partner_name or not sf_id:
                    continue
                records.append({
                    "VENDOR": "Webroot",
                    "PARTNER_NAME": partner_name,
                    "PARENT_COMPANY": value_at(row, "PARENT_CO", "PARENT_COMPANY"),
                    "SF_ID": sf_id,
                    "CMS_ID": value_at(row, "CMS_ID"),
                    "ZUORA_NAME": value_at(row, "ZUORA_NAME"),
                })
    finally:
        wb.close()

    df = pd.DataFrame(records, columns=PARTNER_SEED_COLUMNS).drop_duplicates()
    print(f"Partner seed rows={len(df):,} distinct_sf_ids={df['SF_ID'].nunique():,}", flush=True)
    return df


def load_snowflake(
    usage_df: pd.DataFrame,
    audit_df: pd.DataFrame,
    partner_df: pd.DataFrame,
    *,
    reset: bool = False,
) -> None:
    from snowflake.connector.pandas_tools import write_pandas

    conn = _snowflake_connection()
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        cur.execute(usage_ddl())
        cur.execute(audit_ddl())
        cur.execute(
            f"ALTER TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
            "ADD COLUMN IF NOT EXISTS ADDITIONAL_INFO VARCHAR"
        )
        for column in (
            "IGNORED_NOISE_ROW_COUNT",
            "INVALID_DATA_ROW_COUNT",
            "DUPLICATE_NATURAL_KEY_COUNT",
        ):
            cur.execute(
                f"ALTER TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} "
                f"ADD COLUMN IF NOT EXISTS {column} NUMBER"
            )
        cur.execute(
            f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{PARTNER_SEED_TABLE} (
    VENDOR VARCHAR,
    PARTNER_NAME VARCHAR,
    PARENT_COMPANY VARCHAR,
    SF_ID VARCHAR,
    CMS_ID VARCHAR,
    ZUORA_NAME VARCHAR
);
"""
        )

        incoming_months = sorted(
            pd.to_datetime(usage_df["BILLING_MONTH"]).dt.date.astype(str).unique().tolist()
        )
        if not incoming_months:
            raise RuntimeError("Refusing to publish Webroot without incoming usage.")
        month_list = ", ".join(f"'{month}'::DATE" for month in incoming_months)

        stage_loads = [
            (usage_df, TARGET_TABLE, "WEBROOT_USAGE_STAGE"),
            (audit_df, AUDIT_TABLE, "WEBROOT_AUDIT_STAGE"),
        ]
        if not partner_df.empty:
            stage_loads.append((partner_df, PARTNER_SEED_TABLE, "WEBROOT_PARTNER_STAGE"))
        for df, target_table, stage_table in stage_loads:
            if df.empty:
                continue
            cur.execute(
                f"CREATE OR REPLACE TEMPORARY TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{stage_table} "
                f"LIKE {TARGET_DATABASE}.{TARGET_SCHEMA}.{target_table}"
            )
            success, _, rows, output = write_pandas(
                conn,
                df,
                stage_table,
                database=TARGET_DATABASE,
                schema=TARGET_SCHEMA,
                quote_identifiers=False,
                use_logical_type=True,
            )
            if not success:
                raise RuntimeError(f"write_pandas failed for {stage_table}: {output}")
            print(f"Staged {rows:,} rows for {target_table}.", flush=True)

        cur.execute("BEGIN")
        usage_columns = ", ".join(USAGE_COLUMNS)
        audit_columns = ", ".join(AUDIT_COLUMNS)
        if reset:
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )
            cur.execute(f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE}")
        else:
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s) "
                f"AND BILLING_MONTH IN ({month_list})",
                (TARGET_VENDOR,),
            )
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} "
                f"WHERE BILLING_MONTH IN ({month_list})"
            )
        cur.execute(
            f"INSERT INTO {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} ({usage_columns}) "
            f"SELECT {usage_columns} FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.WEBROOT_USAGE_STAGE"
        )
        cur.execute(
            f"INSERT INTO {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} ({audit_columns}) "
            f"SELECT {audit_columns} FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.WEBROOT_AUDIT_STAGE"
        )
        if not partner_df.empty:
            partner_columns = ", ".join(PARTNER_SEED_COLUMNS)
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{PARTNER_SEED_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )
            cur.execute(
                f"INSERT INTO {TARGET_DATABASE}.{TARGET_SCHEMA}.{PARTNER_SEED_TABLE} ({partner_columns}) "
                f"SELECT {partner_columns} FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.WEBROOT_PARTNER_STAGE"
            )
        conn.commit()
        print("Published staged Webroot data atomically.", flush=True)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def summarize_local(usage_df: pd.DataFrame, audit_df: pd.DataFrame) -> None:
    print("\nLOCAL USAGE CONTROLS", flush=True)
    summary = (
        usage_df.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(
            row_count=("QUANTITY", "size"),
            total_seats=("QUANTITY", "sum"),
            total_amount=("AMOUNT", "sum"),
        )
        .reset_index()
        .sort_values(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"])
    )
    print(summary.to_string(index=False), flush=True)

    bad_audit = audit_df[
        (audit_df["LOAD_STATUS"].eq("LOADED"))
        & (
            audit_df["TOTAL_SEATS_DELTA"].fillna(0).abs().gt(0.0001)
            | audit_df["TOTAL_EXTENDED_AMOUNT_DELTA"].fillna(0).abs().gt(0.0001)
        )
    ]
    if bad_audit.empty:
        print("Subtotal parity: PASS for loaded files.", flush=True)
    else:
        print("Subtotal parity: FAIL", flush=True)
        print(
            bad_audit[
                [
                    "STREAM",
                    "CHANNEL",
                    "SOURCE_FOLDER",
                    "SOURCE_FILE",
                    "LICENSE_CATEGORY_NAME",
                    "TOTAL_SEATS_DELTA",
                    "TOTAL_EXTENDED_AMOUNT_DELTA",
                ]
            ].to_string(index=False),
            flush=True,
        )


def validate_snowflake(local_usage_df: pd.DataFrame, local_audit_df: pd.DataFrame) -> None:
    incoming_months = sorted(
        pd.to_datetime(local_usage_df["BILLING_MONTH"])
        .dt.date.astype(str).unique().tolist()
    )
    month_list = ", ".join(f"'{month}'::DATE" for month in incoming_months)
    conn = _snowflake_connection()
    try:
        sql = f"""
SELECT
    BILLING_MONTH,
    MODIFIER,
    VENDOR_PRODUCT_SKU,
    COUNT(*) AS row_count,
    SUM(QUANTITY) AS total_seats,
    SUM(AMOUNT) AS total_amount
FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}
WHERE VENDOR = 'Webroot'
    AND BILLING_MONTH IN ({month_list})
GROUP BY 1,2,3
ORDER BY 1,2,3
"""
        snowflake_usage = pd.read_sql(sql, conn)
        snowflake_usage.columns = [col.lower() for col in snowflake_usage.columns]
        snowflake_usage = snowflake_usage.rename(
            columns={
                "billing_month": "BILLING_MONTH",
                "modifier": "MODIFIER",
                "vendor_product_sku": "VENDOR_PRODUCT_SKU",
            }
        )
        audit_sql = f"""
SELECT
    LOAD_STATUS,
    COUNT(DISTINCT SOURCE_FILE || '|' || SOURCE_FOLDER || '|' || SOURCE_CONTENT_HASH) AS file_count,
    SUM(DATA_ROW_COUNT) AS data_row_count,
    SUM(RECOMPUTED_TOTAL_SEATS) AS recomputed_total_seats,
    SUM(RECOMPUTED_TOTAL_EXTENDED_AMOUNT) AS recomputed_total_extended_amount,
    MAX(ABS(COALESCE(TOTAL_SEATS_DELTA, 0))) AS max_total_seats_delta,
    MAX(ABS(COALESCE(TOTAL_EXTENDED_AMOUNT_DELTA, 0))) AS max_total_extended_amount_delta
FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE}
WHERE BILLING_MONTH IN ({month_list})
GROUP BY 1
ORDER BY 1
"""
        snowflake_audit = pd.read_sql(audit_sql, conn)
        snowflake_audit.columns = [col.lower() for col in snowflake_audit.columns]
    finally:
        conn.close()

    local_usage = (
        local_usage_df.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(
            row_count=("QUANTITY", "size"),
            total_seats=("QUANTITY", "sum"),
            total_amount=("AMOUNT", "sum"),
        )
        .reset_index()
    )
    local_usage["BILLING_MONTH"] = pd.to_datetime(local_usage["BILLING_MONTH"])
    snowflake_usage["BILLING_MONTH"] = pd.to_datetime(snowflake_usage["BILLING_MONTH"])
    compare = local_usage.merge(
        snowflake_usage,
        on=["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"],
        how="outer",
        suffixes=("_local", "_snowflake"),
    )
    for col in ("row_count", "total_seats", "total_amount"):
        compare[f"{col}_delta"] = compare[f"{col}_snowflake"].fillna(0) - compare[f"{col}_local"].fillna(0)

    print("\nSNOWFLAKE LOAD PARITY", flush=True)
    print(
        compare[
            [
                "BILLING_MONTH",
                "MODIFIER",
                "VENDOR_PRODUCT_SKU",
                "row_count_delta",
                "total_seats_delta",
                "total_amount_delta",
            ]
        ]
        .sort_values(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"])
        .to_string(index=False),
        flush=True,
    )
    max_abs_delta = compare[
        ["row_count_delta", "total_seats_delta", "total_amount_delta"]
    ].abs().max().max()
    if max_abs_delta > 0.0001:
        raise RuntimeError("Snowflake usage parity failed.")

    print("\nSNOWFLAKE FILE AUDIT", flush=True)
    print(snowflake_audit.to_string(index=False), flush=True)

def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Webroot vendor usage into Snowflake.")
    parser.add_argument("--source-root-cw", default=str(DEFAULT_SOURCE_ROOT_CW))
    parser.add_argument("--source-root-cms", default=str(DEFAULT_SOURCE_ROOT_CMS))
    parser.add_argument("--mappings-file", default=str(MAPPINGS_FILE))
    parser.add_argument("--month", action="append", help="Billing month YYYY-MM. Can be supplied more than once.")
    parser.add_argument("--all-months", action="store_true", help="Load all discovered months.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and summarize without loading Snowflake.")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Replace all Webroot usage history; allowed only with --all-months.",
    )
    parser.add_argument("--skip-partner-map", action="store_true", help="Do not load WEBROOT_PARTNER_MAP_SEED.")
    parser.add_argument("--skip-snowflake-validation", action="store_true", help="Do not compare local controls to Snowflake.")
    args = parser.parse_args()

    if not args.all_months and not args.month:
        raise SystemExit("Provide --month YYYY-MM or --all-months")
    if args.reset and not args.all_months:
        raise SystemExit("--reset requires --all-months")

    requested_months = None if args.all_months else set(args.month)
    usage_df, audit_df = build_usage(
        Path(args.source_root_cw),
        Path(args.source_root_cms),
        requested_months,
    )
    if args.all_months:
        discovered_months = {
            value.strftime("%Y-%m")
            for value in pd.to_datetime(audit_df["BILLING_MONTH"], errors="coerce").dropna()
        }
        validate_source_completeness(audit_df, discovered_months)
    partner_df = (
        pd.DataFrame(columns=PARTNER_SEED_COLUMNS)
        if args.skip_partner_map
        else load_partner_seed(Path(args.mappings_file))
    )
    summarize_local(usage_df, audit_df)

    if args.dry_run:
        print("Dry run complete.", flush=True)
        return

    load_snowflake(usage_df, audit_df, partner_df, reset=args.reset)
    if not args.skip_snowflake_validation:
        validate_snowflake(usage_df, audit_df)


if __name__ == "__main__":
    main()

