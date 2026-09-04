"""Build the normalized KeepIT usage table from invoice-aligned workbooks."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import re
import shutil
import sys
import tempfile
import time
from collections.abc import Iterable
from pathlib import Path

import openpyxl
import pandas as pd
import pdfplumber


KEEPIT_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = KEEPIT_ROOT.parent
WORKSPACE_ROOT = next((p for p in (PROJECT_ROOT, *PROJECT_ROOT.parents) if (p / "TEMPLATES").exists()), PROJECT_ROOT)
OUTPUT_DIR = KEEPIT_ROOT / "outputs"

DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\KeepIT"
)
DEFAULT_MANUAL_ROOT = DEFAULT_SOURCE_ROOT

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
FQN = f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}"
TARGET_VENDOR = "KeepIT"
INVOICE_CONTROL_TABLE = "ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_INVOICES"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)
DATACENTER_TABS = ("dk-co", "au-sy", "us-dc", "ca-tr", "uk-ld")
BLANK_ROW_CUTOFF = 50
SUMMARY_MAX_ROW = 600

TARGET_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
    "ADDITIONAL_INFO",
)

DESCRIPTION_SKU_MAP = {
    "Microsoft 365 Mailbox OneDrive Total consumption": "KI-M365-FUL",
    "Microsoft 365 Exchange consumption": "KI-M365-FUL",
    "Entra ID consumption": "KI-AZUR-CSP",
    "Google Workspace consumption": "KI-GOOG-FUL",
    "GW Gmail consumption": "KI-GOOG-FUL",
    "Salesforce consumption": "KI-SFDC-FUL",
    "Dynamics 365 consumption": "KI-D365-FUL",
    "Dynamics 365 Light consumption": "KI-D365-LGT",
}
TAKEOUT_PLACEHOLDER_SKU = "TAKEOUT-3Y-PROMO-KEEPIT"
PROMO_MODIFIER = "PROMO"
TAKEOUT_MODIFIER = "TAKEOUT"
MISSING_VENDOR_USAGE_SKU = "Missing Vendor Usage by SKU"
GUID_TOKEN_RE = re.compile(r"^[a-z0-9]{4,}(?:-[a-z0-9]{2,}){2,}$", re.IGNORECASE)


def clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def normalize_partner_name(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return text.casefold()


def normalize_lookup_token(value: object) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return re.sub(r"[^a-z0-9]+", "", text.casefold()) or None


def to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    if isinstance(value, str):
        text = value.replace("$", "").replace(" ", "").strip()
        if text == "":
            return None
        # European format: 1.234,56 -> 1234.56
        if re.match(r"^\d{1,3}(\.\d{3})*(,\d+)$", text):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
        value = text
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed)


def to_keepit_invoice_number(value: object) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".")
    parsed = pd.to_numeric(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed)


def read_file_bytes(
    path: Path,
    *,
    allow_copy_fallback: bool = True,
    retries: int = 6,
    base_delay_seconds: float = 0.5,
) -> bytes:
    last_error: PermissionError | None = None
    for attempt in range(retries):
        try:
            return path.read_bytes()
        except PermissionError as exc:
            last_error = exc
            if allow_copy_fallback:
                with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
                    tmp_path = Path(handle.name)
                try:
                    try:
                        shutil.copy2(path, tmp_path)
                        return tmp_path.read_bytes()
                    except PermissionError:
                        pass
                finally:
                    tmp_path.unlink(missing_ok=True)
            if attempt < retries - 1:
                time.sleep(base_delay_seconds * (attempt + 1))
    if last_error:
        raise last_error
    raise PermissionError(f"Unable to read locked file: {path}")


def open_excel_file(path: Path) -> pd.ExcelFile:
    try:
        return pd.ExcelFile(path, engine="calamine")
    except PermissionError:
        # Fall back to in-memory bytes when OneDrive transiently denies direct opens.
        return pd.ExcelFile(io.BytesIO(read_file_bytes(path)), engine="calamine")


def normalize_header(value: object) -> str:
    return str(value if value is not None else "").strip().lower()


def billing_month_from_folder(month_folder: Path) -> dt.date:
    match = MONTH_FOLDER_RE.match(month_folder.name)
    if not match:
        raise ValueError(f"Invalid KeepIT month folder name: {month_folder}")
    return dt.date(int(match.group("yyyy")), int(match.group("mm")), 1)


def discover_month_folders(source_root: Path) -> dict[str, Path]:
    folders: dict[str, Path] = {}
    for child in source_root.iterdir():
        if child.is_dir() and (match := MONTH_FOLDER_RE.match(child.name)):
            folders[f"{match.group('yyyy')}-{match.group('mm')}"] = child
    return dict(sorted(folders.items()))


def strip_description_window(description: object) -> str | None:
    text = clean_text(description)
    if not text:
        return None
    return re.sub(r"\s*\(.*$", "", text).strip()


def derive_vendor_sku(description: object) -> str | None:
    base = strip_description_window(description)
    if not base:
        return None
    return DESCRIPTION_SKU_MAP.get(base)


def is_main_summary_file(path: Path) -> bool:
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and "summary" in name
        and ("connectwise" in name or "main" in name)
        and "promo" not in name
        and "takeout" not in name
        and "zuora" not in name
        and "recon" not in name
    )


def is_promo_summary_file(path: Path) -> bool:
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and ("promo" in name or name.startswith("takeout invoice "))
        and "summary" in name
        and "post year" not in name
        and "post 3rd" not in name
        and "recon" not in name
    )


def is_takeout_summary_file(path: Path) -> bool:
    name = path.name.lower()
    return (
        path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and "takeout" in name
        and ("post year" in name or "post 3rd" in name)
    )


def is_takeout_invoice_file(path: Path) -> bool:
    return path.suffix.lower() == ".pdf" and not path.name.startswith("~$") and "invoice" in path.name.lower()


def _summary_section_flags(path: Path) -> tuple[bool, bool]:
    """Return (has_promo_section, has_post_year_section) from Summary text markers."""
    has_promo = False
    has_post_year = False
    try:
        workbook = load_workbook(path)
    except Exception:
        return has_promo, has_post_year
    try:
        try:
            ws = summary_sheet(workbook)
        except RuntimeError:
            return has_promo, has_post_year
        max_row = min(ws.max_row, 120)
        max_col = min(ws.max_column, 40)
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                text = clean_text(ws.cell(row_idx, col_idx).value)
                if not text:
                    continue
                lower = text.lower()
                if "partners from" in lower and "yr 4" in lower:
                    has_post_year = True
                if (
                    "takeout partners due for invoicing" in lower
                    and ("year 2" in lower or "years 2" in lower or "3 of promo" in lower)
                ):
                    has_promo = True
        return has_promo, has_post_year
    finally:
        workbook.close()


def _is_main_workbook_by_schema(path: Path) -> bool:
    if path.suffix.lower() != ".xlsx" or path.name.startswith("~$"):
        return False
    try:
        workbook = load_workbook(path)
    except Exception:
        return False
    try:
        sheet_lookup = {sheet.lower(): sheet for sheet in workbook.sheetnames}
        present_dc_tabs = [tab for tab in DATACENTER_TABS if tab in sheet_lookup]
        if len(present_dc_tabs) < 3:
            return False
        if "summary" in sheet_lookup or "summay" in sheet_lookup:
            return False
        return True
    finally:
        workbook.close()


def _is_raw_promo_extract(path: Path) -> bool:
    """Detect unpriced raw promo exports (quantity-only) for audit visibility."""
    if path.suffix.lower() != ".xlsx" or path.name.startswith("~$"):
        return False
    try:
        workbook = load_workbook(path)
    except Exception:
        return False
    try:
        if not workbook.sheetnames:
            return False
        ws = workbook[workbook.sheetnames[0]]
        headers = {normalize_header(ws.cell(1, col).value) for col in range(1, min(ws.max_column, 30) + 1)}
        required = {"product_id", "charge_sku", "quantity", "vid"}
        return required.issubset(headers)
    finally:
        workbook.close()


def _inspect_workbook_schema(path: Path) -> dict[str, bool]:
    """Inspect workbook once and return routing flags."""
    flags = {
        "is_main": False,
        "has_promo_summary": False,
        "has_takeout_summary": False,
        "is_raw_promo": False,
    }
    if path.suffix.lower() != ".xlsx" or path.name.startswith("~$"):
        return flags
    try:
        workbook = load_workbook(path)
    except Exception:
        return flags
    try:
        sheet_lookup = {sheet.lower(): sheet for sheet in workbook.sheetnames}
        present_dc_tabs = [tab for tab in DATACENTER_TABS if tab in sheet_lookup]
        if len(present_dc_tabs) >= 3 and "summary" not in sheet_lookup and "summay" not in sheet_lookup:
            flags["is_main"] = True

        if workbook.sheetnames:
            ws0 = workbook[workbook.sheetnames[0]]
            headers = {normalize_header(ws0.cell(1, col).value) for col in range(1, min(ws0.max_column, 30) + 1)}
            flags["is_raw_promo"] = {"product_id", "charge_sku", "quantity", "vid"}.issubset(headers)

        try:
            ws = summary_sheet(workbook)
        except RuntimeError:
            ws = None
        if ws is not None:
            max_row = min(ws.max_row, 120)
            max_col = min(ws.max_column, 40)
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    text = clean_text(ws.cell(row_idx, col_idx).value)
                    if not text:
                        continue
                    lower = text.lower()
                    if "partners from" in lower and "yr 4" in lower:
                        flags["has_takeout_summary"] = True
                    if (
                        "takeout partners due for invoicing" in lower
                        and ("year 2" in lower or "years 2" in lower or "3 of promo" in lower)
                    ):
                        flags["has_promo_summary"] = True
        return flags
    finally:
        workbook.close()


def pdf_has_takeout_lines(path: Path) -> bool:
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages[:1])
    except Exception:
        return False
    return " - Takeout " in text and "SKU Description" in text


def locate_usage_files(month_folder: Path) -> list[tuple[str, Path]]:
    files = sorted(
        path for path in month_folder.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
    )
    selected: dict[str, list[Path]] = {
        "MAIN": [path for path in files if is_main_summary_file(path)],
        "PROMO": [path for path in files if is_promo_summary_file(path)],
        "TAKEOUT": [path for path in files if is_takeout_summary_file(path)],
    }

    # Current-period naming is deterministic and avoids any schema-scan cost.
    if all(len(paths) == 1 for paths in selected.values()):
        return [(family, selected[family][0]) for family in ("MAIN", "PROMO", "TAKEOUT")]

    assigned = {path for paths in selected.values() for path in paths}
    candidates = [
        path for path in files
        if path not in assigned
        and not any(term in path.name.lower() for term in ("recon", "zuora", "saas backup usage"))
    ]
    for path in candidates:
        try:
            with open_excel_file(path) as xls:
                sheet_names = {name.lower(): name for name in xls.sheet_names}
                summary_name = next((sheet_names[name] for name in ("summary", "summay") if name in sheet_names), None)
                if summary_name is None:
                    if len(set(sheet_names) & set(DATACENTER_TABS)) >= 3:
                        selected["MAIN"].append(path)
                    continue
                summary = xls.parse(sheet_name=summary_name, header=None, nrows=120, dtype=object)
                text = " ".join(
                    summary.iloc[:, :40]
                    .astype(str)
                    .to_numpy()
                    .ravel()
                    .tolist()
                ).lower()
                if "partners from" in text and "yr 4" in text:
                    selected["TAKEOUT"].append(path)
                elif "takeout partners due for invoicing" in text:
                    selected["PROMO"].append(path)
        except (OSError, ValueError):
            continue

    ambiguous = {family: paths for family, paths in selected.items() if len(paths) > 1}
    if ambiguous:
        detail = "; ".join(f"{family}={[path.name for path in paths]}" for family, paths in ambiguous.items())
        raise RuntimeError(f"Ambiguous KeepIT Excel workbooks in {month_folder}: {detail}")
    return [(family, selected[family][0]) for family in ("MAIN", "PROMO", "TAKEOUT") if selected[family]]


def make_usage_row(
    *,
    billing_month: dt.date,
    partner: object,
    sku: object,
    quantity: object,
    unit_price: object,
    amount: object,
    modifier: object = None,
    currency: object = "USD",
    additional_info: object = None,
) -> dict[str, object] | None:
    parsed_amount = to_number(amount)
    parsed_quantity = to_number(quantity)
    sku_text = clean_text(sku)
    partner_text = clean_text(partner)
    if parsed_amount is None or parsed_quantity is None or not sku_text:
        return None
    return {
        "BILLING_MONTH": billing_month,
        "VENDOR": "KeepIT",
        "VENDOR_PARTNER_NAME": partner_text,
        "VENDOR_PRODUCT_SKU": sku_text,
        "MODIFIER": clean_text(modifier),
        "QUANTITY": parsed_quantity,
        "UNIT_PRICE": to_number(unit_price),
        "AMOUNT": parsed_amount,
        "CURRENCY": clean_text(currency) or "USD",
        "ADDITIONAL_INFO": clean_text(additional_info),
    }


def load_workbook(path: Path) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        # Only copy a source file if OneDrive is holding a transient lock.
        return openpyxl.load_workbook(io.BytesIO(read_file_bytes(path)), read_only=True, data_only=True)


def parse_main(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rows: list[dict[str, object]] = []
    audit: list[dict[str, object]] = []
    with open_excel_file(path) as xls:
        detail_sheets = [
            name for name in xls.sheet_names
            if name.lower() in DATACENTER_TABS
        ]
        for sheet_name in detail_sheets:
            df = xls.parse(sheet_name=sheet_name, dtype=object)
            normalized_cols = {normalize_header(col): col for col in df.columns}
            required = ["companyname2", "fullname2", "description1", "units1", "unit-price1", "price1"]
            missing = [col for col in required if col not in normalized_cols]
            if missing:
                audit.append({"source_family": "MAIN", "source_file": path.name, "source_sheet": sheet_name, "rows": 0, "reason": "skipped_non_detail_sheet"})
                continue

            work = pd.DataFrame(
                {
                    "company": df[normalized_cols["companyname2"]],
                    "full": df[normalized_cols["fullname2"]],
                    "description": df[normalized_cols["description1"]],
                    "quantity": df[normalized_cols["units1"]],
                    "unit_price": df[normalized_cols["unit-price1"]],
                    "amount": df[normalized_cols["price1"]],
                }
            )
            work["sku"] = work["description"].map(derive_vendor_sku)
            has_data = work[["description", "quantity", "amount"]].notna().any(axis=1)
            unmapped_with_data = work["sku"].isna() & has_data
            if unmapped_with_data.any():
                bad = work.loc[unmapped_with_data, "description"].iloc[0]
                raise RuntimeError(f"Unmapped KeepIT Main description in {path.name}/{sheet_name}: {bad!r}")

            work = work.loc[work["sku"].notna()].copy()
            accepted_before = len(rows)
            if not work.empty:
                partner_series = work["full"].where(work["full"].notna(), work["company"])
                for partner, sku, qty, unit_price, amt in zip(
                    partner_series,
                    work["sku"],
                    work["quantity"],
                    work["unit_price"],
                    work["amount"],
                ):
                    row = make_usage_row(
                        billing_month=billing_month,
                        partner=partner,
                        sku=sku,
                        modifier=None,
                        quantity=qty,
                        unit_price=unit_price,
                        amount=amt,
                    )
                    if row:
                        rows.append(row)
            audit.append({"source_family": "MAIN", "source_file": path.name, "source_sheet": sheet_name, "rows": len(rows) - accepted_before, "reason": "accepted"})
    return rows, audit


def summary_sheet(workbook: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    for name in ("Summary", "Summay"):
        if name in workbook.sheetnames:
            return workbook[name]
    raise RuntimeError(f"Workbook has no Summary/Summay sheet: {workbook.sheetnames}")


def _header_row_for_summary(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    post_year: bool,
) -> tuple[int, int]:
    max_scan_col = min(ws.max_column, 60)
    if post_year:
        strict_terms = ("partners from", "yr 4")
    else:
        strict_terms = ("takeout partners due for invoicing", "year 2")

    # First pass: strict terms to avoid matching generic section labels.
    for row_idx in range(1, min(ws.max_row, 80) + 1):
        for col_idx in range(1, max_scan_col + 1):
            text = clean_text(ws.cell(row_idx, col_idx).value)
            if not text:
                continue
            lower = text.lower()
            if all(term in lower for term in strict_terms):
                return row_idx, col_idx

    # Second pass: backward-compatible loose match.
    partner_header = "partners from" if post_year else "takeout partners due for invoicing"
    for row_idx in range(1, min(ws.max_row, 80) + 1):
        for col_idx in range(1, max_scan_col + 1):
            text = clean_text(ws.cell(row_idx, col_idx).value)
            if not text:
                continue
            lower = text.lower()
            if partner_header in lower:
                return row_idx, col_idx
    raise RuntimeError("Summary sheet missing partner header")


def _format_min_commit(value: object) -> str | None:
    parsed = to_number(value)
    if parsed is None or parsed <= 0:
        return None
    if abs(parsed - round(parsed)) < 0.000001:
        return f"Min Commit {int(round(parsed))} units"
    return f"Min Commit {parsed:g} units"


def _find_summary_metric_column(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    anchor_row: int,
    max_scan_col: int,
    metric: str,
) -> int | None:
    """Find a metric column near the detected header row.

    Prefer exact/clean labels on the anchor row so note text in earlier rows
    (for example a warning sentence mentioning "consumption") is ignored.
    """

    def match_score(text: str) -> int:
        lower = text.lower().strip()
        if metric == "consumption":
            if lower in {"consumption", "consumed", "usage"}:
                return 100
            if (
                "consumption" in lower
                and "unlimited" not in lower
                and "contracts" not in lower
                and len(lower) <= 25
            ):
                return 60
        elif metric == "min_commit":
            if "min comm" in lower:
                return 80
        elif metric == "total":
            if "total per partner" in lower:
                return 100
            if lower.startswith("amount"):
                return 60
        return 0

    search_rows = [anchor_row, anchor_row - 1, anchor_row + 1, anchor_row - 2, anchor_row + 2]
    best_col: int | None = None
    best_score = -1
    for row_idx in search_rows:
        if row_idx < 1 or row_idx > ws.max_row:
            continue
        row_bias = 10 if row_idx == anchor_row else 0
        for col_idx in range(1, max_scan_col + 1):
            text = clean_text(ws.cell(row_idx, col_idx).value)
            if not text:
                continue
            score = match_score(text) + row_bias
            if score > best_score:
                best_score = score
                best_col = col_idx
    return best_col if best_score > 0 else None


def _extract_summary_guid_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    post_year: bool,
) -> list[dict[str, object]]:
    anchor_row, partner_col = _header_row_for_summary(ws, post_year=post_year)
    max_scan_col = min(ws.max_column, 80)
    min_commit_col = _find_summary_metric_column(
        ws,
        anchor_row=anchor_row,
        max_scan_col=max_scan_col,
        metric="min_commit",
    )
    consumption_col = _find_summary_metric_column(
        ws,
        anchor_row=anchor_row,
        max_scan_col=max_scan_col,
        metric="consumption",
    )
    total_col = _find_summary_metric_column(
        ws,
        anchor_row=anchor_row,
        max_scan_col=max_scan_col,
        metric="total",
    )
    if consumption_col is None or total_col is None:
        raise RuntimeError("Summary sheet missing Consumption/Total per partner columns")

    rows: list[dict[str, object]] = []
    blank_rows = 0
    data_start = anchor_row + 1
    for row_idx in range(data_start, min(ws.max_row, SUMMARY_MAX_ROW) + 1):
        partner = clean_text(ws.cell(row_idx, partner_col).value)
        if not partner:
            blank_rows += 1
            if blank_rows >= BLANK_ROW_CUTOFF:
                break
            continue
        blank_rows = 0
        partner_key = normalize_partner_name(partner)
        if not partner_key or partner_key.startswith(("total", "subtotal", "partners from")):
            continue

        guid_value = clean_text(ws.cell(row_idx, partner_col + 1).value)
        guid_key = normalize_lookup_token(guid_value)
        if not guid_value or not guid_key or not GUID_TOKEN_RE.match(guid_value):
            continue

        amount = to_number(ws.cell(row_idx, total_col).value)
        consumption = to_number(ws.cell(row_idx, consumption_col).value)
        min_commit = to_number(ws.cell(row_idx, min_commit_col).value) if min_commit_col else None
        if amount is None:
            continue
        if abs(amount) <= 0.000001:
            continue

        rows.append(
            {
                "PARTNER": partner,
                "GUID": guid_value,
                "GUID_KEY": guid_key,
                "SUMMARY_AMOUNT": float(amount),
                "SUMMARY_CONSUMPTION": float(consumption or 0),
                "MIN_COMMIT": min_commit,
                "ADDITIONAL_INFO": _format_min_commit(min_commit),
            }
        )
    return rows


def _extract_summary_guid_rows_frame(
    raw: pd.DataFrame,
    *,
    post_year: bool,
) -> list[dict[str, object]]:
    """Extract Summary GUID controls from one in-memory dataframe."""
    scan = raw.iloc[:80, :80].map(lambda value: (clean_text(value) or "").lower())
    strict_terms = ("partners from", "yr 4") if post_year else ("takeout partners due for invoicing", "year")
    loose_term = "partners from" if post_year else "takeout partners due for invoicing"

    anchor: tuple[int, int] | None = None
    for strict in (True, False):
        for row_idx, row in scan.iterrows():
            for col_idx, text in enumerate(row.tolist()):
                matched = all(term in text for term in strict_terms) if strict else loose_term in text
                if matched:
                    anchor = (int(row_idx), col_idx)
                    break
            if anchor:
                break
        if anchor:
            break
    if anchor is None:
        raise RuntimeError("Summary sheet missing partner header")
    anchor_row, partner_col = anchor

    def metric_score(text: str, metric: str) -> int:
        if metric == "consumption":
            if text in {"consumption", "consumed", "usage"}:
                return 100
            if "consumption" in text and "unlimited" not in text and "contracts" not in text and len(text) <= 25:
                return 60
        elif metric == "min_commit" and "min comm" in text:
            return 80
        elif metric == "total":
            if "total per partner" in text:
                return 100
            if text.startswith("amount"):
                return 60
        return 0

    def metric_column(metric: str) -> int | None:
        best: tuple[int, int] | None = None
        for row_idx in (anchor_row, anchor_row - 1, anchor_row + 1, anchor_row - 2, anchor_row + 2):
            if row_idx < 0 or row_idx >= len(scan):
                continue
            for col_idx, text in enumerate(scan.iloc[row_idx].tolist()):
                score = metric_score(text, metric) + (10 if row_idx == anchor_row else 0)
                if score > 0 and (best is None or score > best[0]):
                    best = (score, col_idx)
        return best[1] if best else None

    min_commit_col = metric_column("min_commit")
    consumption_col = metric_column("consumption")
    total_col = metric_column("total")
    if consumption_col is None or total_col is None:
        raise RuntimeError("Summary sheet missing Consumption/Total per partner columns")

    data = raw.iloc[anchor_row + 1:min(len(raw), SUMMARY_MAX_ROW), :].copy()
    partner_values = data.iloc[:, partner_col].map(clean_text)
    blank = partner_values.isna().to_numpy()
    consecutive = 0
    cutoff = len(data)
    for offset, is_blank in enumerate(blank):
        consecutive = consecutive + 1 if is_blank else 0
        if consecutive >= BLANK_ROW_CUTOFF:
            cutoff = offset - BLANK_ROW_CUTOFF + 1
            break
    data = data.iloc[:cutoff]
    partner_values = partner_values.iloc[:cutoff]

    guid_values = data.iloc[:, partner_col + 1].map(clean_text)
    guid_keys = guid_values.map(normalize_lookup_token)
    amounts = data.iloc[:, total_col].map(to_number)
    consumption = data.iloc[:, consumption_col].map(to_number).fillna(0.0)
    if min_commit_col is not None:
        min_commit = data.iloc[:, min_commit_col].map(to_number)
    else:
        min_commit = pd.Series([None] * len(data), index=data.index, dtype=object)

    valid = (
        partner_values.notna()
        & guid_values.fillna("").str.match(GUID_TOKEN_RE)
        & guid_keys.notna()
        & amounts.notna()
        & amounts.abs().gt(0.000001)
        & ~partner_values.fillna("").str.casefold().str.startswith(("total", "subtotal", "partners from"))
    )
    result = pd.DataFrame(
        {
            "PARTNER": partner_values,
            "GUID": guid_values,
            "GUID_KEY": guid_keys,
            "SUMMARY_AMOUNT": amounts,
            "SUMMARY_CONSUMPTION": consumption,
            "MIN_COMMIT": min_commit,
        }
    ).loc[valid]
    result["ADDITIONAL_INFO"] = result["MIN_COMMIT"].map(_format_min_commit)
    return result.to_dict("records")


def _aggregate_detail_by_guid(
    path: Path,
    *,
    source_file: str,
    source_family: str,
    xls: pd.ExcelFile | None = None,
) -> tuple[dict[str, list[dict[str, object]]], list[dict[str, object]]]:
    def canonical(header: str) -> str:
        return re.sub(r"\d+$", "", header)

    def detect_header(preview: pd.DataFrame) -> tuple[int, dict[str, int], dict[str, list[int]]] | None:
        max_col = min(preview.shape[1], 80)
        for row_idx in range(min(preview.shape[0], 30)):
            headers = [normalize_header(preview.iat[row_idx, col]) for col in range(max_col)]
            header_index = {header: idx for idx, header in enumerate(headers) if header}
            canonical_index: dict[str, list[int]] = {}
            for header, idx in header_index.items():
                canonical_index.setdefault(canonical(header), []).append(idx)
            if (
                ("description" in canonical_index)
                and ("units" in canonical_index)
                and ("price" in canonical_index)
                and ("account1" in header_index)
            ):
                return row_idx, header_index, canonical_index
        return None

    if xls is None:
        with open_excel_file(path) as opened:
            return _aggregate_detail_by_guid(
                path,
                source_file=source_file,
                source_family=source_family,
                xls=opened,
            )

    by_guid: dict[str, list[dict[str, object]]] = {}
    audits: list[dict[str, object]] = []
    for tab in xls.sheet_names:
            if tab.lower() not in DATACENTER_TABS:
                continue
            raw = xls.parse(sheet_name=tab, header=None, dtype=object)
            detected = detect_header(raw.iloc[:30])
            if not detected:
                audits.append(
                    {
                        "source_family": source_family,
                        "source_file": source_file,
                        "source_sheet": tab,
                        "rows": 0,
                        "reason": "skipped_no_detail_header",
                    }
                )
                continue
            header_row, header_index, canonical_index = detected
            frame = raw.iloc[header_row + 1:].dropna(how="all").reset_index(drop=True)
            description_idx = canonical_index["description"][0]
            units_idx = canonical_index["units"][0]
            price_idx = canonical_index["price"][0]
            unit_price_idx = canonical_index.get("unit-price", [None])[0]
            account_idx = header_index["account1"]

            detail = pd.DataFrame(index=frame.index)
            description_series = frame.iloc[:, description_idx]
            detail["SKU"] = description_series.map(derive_vendor_sku)
            promo_mask = detail["SKU"].isna() & description_series.astype(str).str.contains("promo", case=False, na=False)
            detail.loc[promo_mask, "SKU"] = TAKEOUT_PLACEHOLDER_SKU

            detail["QUANTITY"] = frame.iloc[:, units_idx].map(to_number)
            detail["AMOUNT"] = frame.iloc[:, price_idx].map(to_number).fillna(0.0)
            if unit_price_idx is not None:
                detail["UNIT_PRICE"] = frame.iloc[:, unit_price_idx].map(to_number)
            else:
                detail["UNIT_PRICE"] = None

            detail["GUID_KEY"] = frame.iloc[:, account_idx].map(normalize_lookup_token)

            partner = pd.Series([None] * len(frame), index=frame.index, dtype=object)
            for key in ("fullname2", "companyname2", "fullname3", "companyname3", "fullname", "companyname"):
                idx = header_index.get(key)
                if idx is None and key in canonical_index:
                    idx = canonical_index[key][0]
                if idx is None:
                    continue
                partner_values = frame.iloc[:, idx].map(clean_text)
                partner = partner.where(partner.notna(), partner_values)
            detail["PARTNER"] = partner

            detail = detail.loc[
                detail["GUID_KEY"].notna()
                & detail["SKU"].notna()
                & detail["QUANTITY"].notna()
            ].copy()
            detail["QUANTITY"] = detail["QUANTITY"].astype(float)
            detail["AMOUNT"] = detail["AMOUNT"].astype(float)

            accepted = len(detail)
            if accepted:
                grouped = detail.groupby("GUID_KEY", dropna=False)
                for guid, group in grouped:
                    records = group[["PARTNER", "SKU", "QUANTITY", "AMOUNT", "UNIT_PRICE"]].to_dict("records")
                    by_guid.setdefault(str(guid), []).extend(records)

            audits.append(
                {
                    "source_family": source_family,
                    "source_file": source_file,
                    "source_sheet": tab,
                    "rows": accepted,
                    "reason": "accepted",
                }
            )
    return by_guid, audits


def _rows_from_summary_guid_distribution(
    summary_rows: list[dict[str, object]],
    detail_by_guid: dict[str, list[dict[str, object]]],
    *,
    billing_month: dt.date,
    use_summary_amount: bool,
    modifier: str,
    include_unmatched_summary: bool = False,
) -> list[dict[str, object]]:
    output: list[dict[str, object]] = []
    for summary in summary_rows:
        guid_key = str(summary["GUID_KEY"])
        detail_rows = detail_by_guid.get(guid_key, [])
        if not detail_rows:
            if include_unmatched_summary:
                additional_info_parts = ["Summary GUID not matched to Account1 detail"]
                if summary.get("ADDITIONAL_INFO"):
                    additional_info_parts.append(str(summary["ADDITIONAL_INFO"]))
                row = make_usage_row(
                    billing_month=billing_month,
                    partner=summary["PARTNER"],
                    sku=MISSING_VENDOR_USAGE_SKU,
                    modifier=modifier,
                    quantity=summary["SUMMARY_CONSUMPTION"],
                    unit_price=None,
                    amount=summary["SUMMARY_AMOUNT"],
                    additional_info="; ".join(additional_info_parts),
                )
                if row:
                    output.append(row)
            continue

        sku_rollup: dict[tuple[str, str], dict[str, float | str | int]] = {}
        for row in detail_rows:
            sku = str(row["SKU"])
            # Keep Summary partner identity as the canonical label for all
            # rows derived from this Summary GUID.
            partner = clean_text(summary.get("PARTNER")) or clean_text(row.get("PARTNER")) or "Unknown Partner"
            key = (partner, sku)
            entry = sku_rollup.setdefault(
                key,
                {
                    "partner": partner,
                    "sku": sku,
                    "qty": 0.0,
                    "amt": 0.0,
                    "unit_price_qty": 0.0,
                    "unit_price_amt": 0.0,
                    "unit_price_count": 0,
                },
            )
            entry["qty"] = float(entry["qty"]) + float(row["QUANTITY"])
            entry["amt"] = float(entry["amt"]) + float(row["AMOUNT"])
            parsed_unit_price = to_number(row.get("UNIT_PRICE"))
            parsed_qty = to_number(row.get("QUANTITY")) or 0.0
            if parsed_unit_price is not None and abs(parsed_qty) > 0.000001:
                entry["unit_price_qty"] = float(entry["unit_price_qty"]) + float(parsed_qty)
                entry["unit_price_amt"] = float(entry["unit_price_amt"]) + float(parsed_unit_price) * float(parsed_qty)
                entry["unit_price_count"] = int(entry["unit_price_count"]) + 1

        dist = list(sku_rollup.values())
        if not dist:
            continue
        amount_total = sum(float(item["amt"]) for item in dist)
        qty_total = sum(float(item["qty"]) for item in dist)
        additional_info = summary.get("ADDITIONAL_INFO")

        output_amounts: list[float]
        if not use_summary_amount:
            output_amounts = [float(item["amt"]) for item in dist]
        else:
            summary_amount = float(summary["SUMMARY_AMOUNT"])
            if abs(amount_total) > 0.000001:
                output_amounts = [summary_amount * (float(item["amt"]) / amount_total) for item in dist]
            elif abs(qty_total) > 0.000001:
                output_amounts = [summary_amount * (float(item["qty"]) / qty_total) for item in dist]
            else:
                output_amounts = [summary_amount] + [0.0] * (len(dist) - 1)
            output_amounts[-1] += summary_amount - sum(output_amounts)

        for item, output_amount in zip(dist, output_amounts):
            # Quantity must remain source-detail units1 (not summary consumption).
            quantity = float(item["qty"])
            if abs(float(item["unit_price_qty"])) > 0.000001:
                unit_price = float(item["unit_price_amt"]) / float(item["unit_price_qty"])
            elif abs(quantity) > 0.000001:
                unit_price = float(item["amt"]) / quantity
            else:
                unit_price = None
            row = make_usage_row(
                billing_month=billing_month,
                partner=item["partner"],
                sku=item["sku"],
                modifier=modifier,
                quantity=quantity,
                unit_price=unit_price,
                amount=output_amount,
                additional_info=additional_info,
            )
            if row and abs(float(row["AMOUNT"])) > 0:
                output.append(row)

    return output


def _parse_summary_guid_workbook(
    path: Path,
    billing_month: dt.date,
    *,
    post_year: bool,
    source_family: str,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    with open_excel_file(path) as xls:
        summary_name = next((name for name in xls.sheet_names if name.lower() in {"summary", "summay"}), None)
        if summary_name is None:
            raise RuntimeError(f"Workbook has no Summary/Summay sheet: {xls.sheet_names}")
        summary_raw = xls.parse(sheet_name=summary_name, header=None, dtype=object)
        summary_rows = _extract_summary_guid_rows_frame(summary_raw, post_year=post_year)
        detail_by_guid, audits = _aggregate_detail_by_guid(
            path,
            source_file=path.name,
            source_family=source_family,
            xls=xls,
        )
    for summary in summary_rows:
        if str(summary["GUID_KEY"]) not in detail_by_guid:
            audits.append(
                {
                    "source_family": source_family,
                    "source_file": path.name,
                    "source_sheet": summary_name,
                    "rows": 0,
                    "reason": "summary_guid_not_matched_to_eligible_account1_detail",
                    "partner": summary["PARTNER"],
                    "guid": summary["GUID"],
                    "summary_quantity": summary["SUMMARY_CONSUMPTION"],
                    "summary_amount": summary["SUMMARY_AMOUNT"],
                }
            )
    rows = _rows_from_summary_guid_distribution(
        summary_rows,
        detail_by_guid,
        billing_month=billing_month,
        use_summary_amount=not post_year,
        modifier=TAKEOUT_MODIFIER if post_year else PROMO_MODIFIER,
        include_unmatched_summary=True,
    )
    return rows, audits


def parse_promo(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    return _parse_summary_guid_workbook(
        path,
        billing_month,
        post_year=False,
        source_family="PROMO",
    )


TAKEOUT_LINE_RE = re.compile(
    r"^(?P<sku>KI-[A-Z0-9-]+)\s+For\s+.+?\s+-\s+Takeout\s+"
    r"(?P<quantity>[0-9.]+)\s+"
    r"(?P<unit_price>[0-9]+,[0-9]+)\s+"
    r"(?P<net_amount>[0-9.]+,[0-9]{2})\s+"
    r"(?P<total_amount>[0-9.]+,[0-9]{2})$"
)


def parse_takeout_pdf(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rows: list[dict[str, object]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                match = TAKEOUT_LINE_RE.match(line.strip())
                if not match:
                    continue
                row = make_usage_row(
                    billing_month=billing_month,
                    partner="ConnectWise (Continuum) - Consolidated",
                    sku=match.group("sku"),
                    modifier=TAKEOUT_MODIFIER,
                    quantity=to_keepit_invoice_number(match.group("quantity")),
                    unit_price=to_keepit_invoice_number(match.group("unit_price")),
                    amount=to_keepit_invoice_number(match.group("total_amount")),
                )
                if row:
                    rows.append(row)
    if not rows:
        raise RuntimeError(f"No takeout invoice lines parsed from {path}")
    return rows, [{"source_family": "TAKEOUT", "source_file": path.name, "source_sheet": "PDF", "rows": len(rows), "reason": "accepted"}]


def parse_takeout_workbook(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    return _parse_summary_guid_workbook(
        path,
        billing_month,
        post_year=True,
        source_family="TAKEOUT",
    )


def parse_takeout(path: Path, billing_month: dt.date) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    if path.suffix.lower() == ".pdf":
        raise RuntimeError("KeepIT ingestion is Excel-only. PDF sources are not supported.")
    return parse_takeout_workbook(path, billing_month)


def load_month(month_folder: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    billing_month = billing_month_from_folder(month_folder)
    records: list[dict[str, object]] = []
    audits: list[dict[str, object]] = []
    usage_files = locate_usage_files(month_folder)
    located_families = {source_family for source_family, _ in usage_files}
    for missing_family in sorted({"MAIN", "PROMO", "TAKEOUT"} - located_families):
        audits.append(
            {
                "billing_month": billing_month.isoformat(),
                "source_family": missing_family,
                "source_file": None,
                "source_sheet": None,
                "rows": 0,
                "reason": "missing_workbook",
            }
        )
    raw_promo_files = [
        path for path in month_folder.iterdir()
        if path.is_file() and path.suffix.lower() == ".xlsx" and "promo" in path.name.lower()
    ]
    if not usage_files:
        raise RuntimeError(f"No KeepIT usage workbook found in {month_folder}")
    if raw_promo_files and not any(source_family == "PROMO" for source_family, _ in usage_files):
        for path in sorted(raw_promo_files):
            audits.append(
                {
                    "billing_month": billing_month.isoformat(),
                    "source_family": "PROMO",
                    "source_file": path.name,
                    "source_sheet": "Sheet1",
                    "rows": 0,
                    "reason": "raw_promo_detected_without_priced_summary",
                }
            )
    for source_family, path in usage_files:
        if source_family == "MAIN":
            parsed, audit = parse_main(path, billing_month)
        elif source_family == "PROMO":
            parsed, audit = parse_promo(path, billing_month)
        elif source_family == "TAKEOUT":
            parsed, audit = parse_takeout(path, billing_month)
        else:
            continue
        records.extend(parsed)
        audits.extend({"billing_month": billing_month.isoformat(), **row} for row in audit)
        frame = pd.DataFrame(parsed, columns=TARGET_COLUMNS)
        print(
            f"[{billing_month}] {source_family} {path.name}: rows={len(frame):,}, "
            f"quantity={frame['QUANTITY'].sum(skipna=True) if not frame.empty else 0:,.2f}, "
            f"amount={frame['AMOUNT'].sum(skipna=True) if not frame.empty else 0:,.2f}"
        )
    df = pd.DataFrame(records, columns=TARGET_COLUMNS)
    if df.empty:
        return df, pd.DataFrame(audits)
    grouped = (
        df.groupby(["BILLING_MONTH", "VENDOR", "VENDOR_PARTNER_NAME", "VENDOR_PRODUCT_SKU", "MODIFIER", "UNIT_PRICE", "CURRENCY", "ADDITIONAL_INFO"], dropna=False)
        .agg(QUANTITY=("QUANTITY", "sum"), AMOUNT=("AMOUNT", "sum"))
        .reset_index()
    )
    return grouped[list(TARGET_COLUMNS)], pd.DataFrame(audits)


def load_all(source_root: Path, months: Iterable[str] | None = None) -> tuple[pd.DataFrame, pd.DataFrame]:
    month_folders = discover_month_folders(source_root)
    if months:
        requested = {str(month)[:7] for month in months}
        month_folders = {month: path for month, path in month_folders.items() if month in requested}
    if not month_folders:
        raise RuntimeError(f"No month folders selected under {source_root}")
    frames: list[pd.DataFrame] = []
    audits: list[pd.DataFrame] = []
    for _, folder in month_folders.items():
        frame, audit = load_month(folder)
        frames.append(frame)
        audits.append(audit)
    return pd.concat(frames, ignore_index=True), pd.concat(audits, ignore_index=True)


def snowflake_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {FQN} (
    BILLING_MONTH DATE,
    VENDOR VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU VARCHAR,
    MODIFIER VARCHAR,
    QUANTITY NUMBER(38, 6),
    UNIT_PRICE NUMBER(18, 6),
    AMOUNT NUMBER(38, 6),
    CURRENCY VARCHAR,
    ADDITIONAL_INFO VARCHAR
);
"""


def align_to_cleaned_invoices(df: pd.DataFrame, conn) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Return source-driven usage and an invoice reconciliation audit.

    This mode does not add synthetic balancing rows. It validates that parsed
    usage reconciles to invoice controls by month, SKU, and source family.
    """
    incoming = df.loc[:, list(TARGET_COLUMNS)].copy()
    incoming["BILLING_MONTH"] = pd.to_datetime(incoming["BILLING_MONTH"]).dt.date
    months = sorted(incoming["BILLING_MONTH"].astype(str).unique().tolist())
    month_sql = ", ".join(f"'{month}'::DATE" for month in months)
    invoice_sql = f"""
SELECT
    BILLING_MONTH,
    UPPER(TRIM(VENDOR_PRODUCT_SKU)) AS VENDOR_PRODUCT_SKU,
    CASE WHEN DESCRIPTION ILIKE '%MAIN%' THEN 'MAIN' ELSE 'TAKEOUT' END AS SOURCE_FAMILY,
    SUM(QUANTITY)::FLOAT AS QUANTITY,
    SUM(AMOUNT)::FLOAT AS AMOUNT
FROM {INVOICE_CONTROL_TABLE}
WHERE VENDOR ILIKE '%KEEPIT%'
  AND BILLING_MONTH IN ({month_sql})
GROUP BY 1,2,3
"""
    cur = conn.cursor()
    try:
        cur.execute(invoice_sql)
        invoices = cur.fetch_pandas_all()
    finally:
        cur.close()
    if invoices.empty:
        raise RuntimeError("No cleaned KeepIT invoice controls found for requested months")

    invoices["BILLING_MONTH"] = pd.to_datetime(invoices["BILLING_MONTH"]).dt.date
    invoices["VENDOR_PRODUCT_SKU"] = invoices["VENDOR_PRODUCT_SKU"].astype(str).str.strip().str.upper()

    incoming["VENDOR_PRODUCT_SKU"] = incoming["VENDOR_PRODUCT_SKU"].astype(str).str.strip().str.upper()
    modifier_upper = incoming["MODIFIER"].fillna("").str.upper()
    incoming["SOURCE_FAMILY"] = "MAIN"
    incoming.loc[modifier_upper.isin({PROMO_MODIFIER, TAKEOUT_MODIFIER}), "SOURCE_FAMILY"] = "TAKEOUT"

    actual = (
        incoming.groupby(["BILLING_MONTH", "SOURCE_FAMILY"], dropna=False)
        .agg(ACTUAL_QUANTITY=("QUANTITY", "sum"), ACTUAL_AMOUNT=("AMOUNT", "sum"))
        .reset_index()
    )
    expected = (
        invoices.groupby(["BILLING_MONTH", "SOURCE_FAMILY"], dropna=False)
        .agg(EXPECTED_QUANTITY=("QUANTITY", "sum"), EXPECTED_AMOUNT=("AMOUNT", "sum"))
        .reset_index()
    )
    audit = expected.merge(
        actual,
        on=["BILLING_MONTH", "SOURCE_FAMILY"],
        how="outer",
    ).fillna(0)
    audit["QTY_DELTA"] = audit["ACTUAL_QUANTITY"] - audit["EXPECTED_QUANTITY"]
    audit["AMOUNT_DELTA"] = audit["ACTUAL_AMOUNT"] - audit["EXPECTED_AMOUNT"]
    failures = audit[(audit["QTY_DELTA"].abs() > 0.000001) | (audit["AMOUNT_DELTA"].abs() > 0.000001)]

    expected_sku = (
        invoices.groupby(["BILLING_MONTH", "SOURCE_FAMILY", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(EXPECTED_QUANTITY=("QUANTITY", "sum"), EXPECTED_AMOUNT=("AMOUNT", "sum"))
        .reset_index()
    )
    actual_sku = (
        incoming.groupby(["BILLING_MONTH", "SOURCE_FAMILY", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(ACTUAL_QUANTITY=("QUANTITY", "sum"), ACTUAL_AMOUNT=("AMOUNT", "sum"))
        .reset_index()
    )
    sku_audit = expected_sku.merge(
        actual_sku,
        on=["BILLING_MONTH", "SOURCE_FAMILY", "VENDOR_PRODUCT_SKU"],
        how="outer",
    ).fillna(0)
    sku_audit["QTY_DELTA"] = sku_audit["ACTUAL_QUANTITY"] - sku_audit["EXPECTED_QUANTITY"]
    sku_audit["AMOUNT_DELTA"] = sku_audit["ACTUAL_AMOUNT"] - sku_audit["EXPECTED_AMOUNT"]
    sku_failures = sku_audit[(sku_audit["QTY_DELTA"].abs() > 0.000001) | (sku_audit["AMOUNT_DELTA"].abs() > 0.000001)]

    if not failures.empty:
        print(
            "WARNING: KeepIT source-driven usage does not fully reconcile to invoice controls. "
            "See keepit_invoice_alignment_*.csv for details.\n"
            + failures.sort_values(["BILLING_MONTH", "SOURCE_FAMILY"]).to_string(index=False)
        )
    if not sku_failures.empty:
        top = sku_failures.assign(_abs_amt=lambda d: d["AMOUNT_DELTA"].abs()).sort_values(
            ["_abs_amt", "BILLING_MONTH", "SOURCE_FAMILY", "VENDOR_PRODUCT_SKU"],
            ascending=[False, True, True, True],
        )
        print(
            "\nSKU-level deltas (top 20 by absolute amount delta):\n"
            + top.drop(columns=["_abs_amt"]).head(20).to_string(index=False)
        )
    return incoming.loc[:, list(TARGET_COLUMNS)], audit, sku_audit


def load_snowflake(df: pd.DataFrame, *, reset: bool = False) -> None:
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from snowflake.connector.pandas_tools import write_pandas
    from TEMPLATES.Python.connection import get_snowflake_connection

    load_df = df.loc[:, list(TARGET_COLUMNS)].copy()
    load_df["BILLING_MONTH"] = pd.to_datetime(load_df["BILLING_MONTH"]).dt.date
    conn = get_snowflake_connection(role="DEVELOPER", warehouse="REPORTING_WH", database=TARGET_DATABASE, schema=TARGET_SCHEMA)
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        if reset:
            print(f"RESET: dropping and recreating {FQN}.")
            cur.execute(
                f"DELETE FROM {FQN} WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )
        cur.execute(snowflake_ddl())
        cur.execute(f"ALTER TABLE {FQN} ADD COLUMN IF NOT EXISTS ADDITIONAL_INFO VARCHAR")
        incoming_months = sorted(load_df["BILLING_MONTH"].astype(str).unique().tolist())
        if not reset:
            month_list = ", ".join(f"'{month}'::DATE" for month in incoming_months)
            cur.execute(
                f"DELETE FROM {FQN} "
                f"WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s) "
                f"AND BILLING_MONTH IN ({month_list})",
                (TARGET_VENDOR,),
            )
        success, chunks, rows, output = write_pandas(
            conn,
            load_df,
            TARGET_TABLE,
            database=TARGET_DATABASE,
            schema=TARGET_SCHEMA,
            quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"write_pandas failed: {output}")
        conn.commit()
        print(f"Loaded {rows:,} rows for {incoming_months} into {FQN} in {chunks} chunk(s).")
    finally:
        conn.close()


def write_ingest_audits(df: pd.DataFrame, scan_df: pd.DataFrame, label: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    scan_path = OUTPUT_DIR / f"keepit_source_scan_{label}.csv"
    scan_df.to_csv(scan_path, index=False, quoting=csv.QUOTE_MINIMAL)
    monthly = (
        df.groupby(["BILLING_MONTH", "VENDOR_PRODUCT_SKU", "MODIFIER", "UNIT_PRICE"], dropna=False)
        .agg(row_count=("VENDOR", "size"), quantity=("QUANTITY", "sum"), amount=("AMOUNT", "sum"))
        .reset_index()
    )
    monthly_path = OUTPUT_DIR / f"keepit_usage_ingest_audit_{label}.csv"
    monthly.to_csv(monthly_path, index=False, quoting=csv.QUOTE_MINIMAL)
    print(f"Wrote source scan: {scan_path}")
    print(f"Wrote usage audit: {monthly_path}")


def validate_usage(df: pd.DataFrame) -> None:
    if df["VENDOR_PRODUCT_SKU"].isna().any():
        raise RuntimeError("KEEPIT_USAGE has null VENDOR_PRODUCT_SKU rows.")
    if df["AMOUNT"].isna().any() or df["QUANTITY"].isna().any():
        raise RuntimeError("KEEPIT_USAGE has null quantity or amount rows.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest source-faithful KeepIT Excel usage into Snowflake.")
    parser.add_argument("--source-root", type=Path, default=DEFAULT_SOURCE_ROOT)
    parser.add_argument("--manual-root", type=Path, default=DEFAULT_MANUAL_ROOT)
    parser.add_argument("--month", action="append", help="Month to load as YYYY-MM or YYYY-MM-01. Repeatable.")
    parser.add_argument("--all-months", action="store_true", help="Load every month folder under the source root.")
    parser.add_argument("--reset", action="store_true", help="Drop and recreate KEEPIT_USAGE before loading.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and audit locally without loading Snowflake.")
    parser.add_argument("--validate-manual", action="store_true", help="Retained for CLI compatibility; no extra action.")
    parser.add_argument("--label", default=dt.datetime.now().strftime("%Y%m%d_%H%M%S"))
    return parser.parse_args()



# ---------------------------------------------------------------------------
# Dynamic invoice rate fill (universal safety net)
# ---------------------------------------------------------------------------
def _fill_missing_prices(df, vendor_name, conn=None):
    from invoice_rate_backfill import fill_missing_prices_dynamic

    return fill_missing_prices_dynamic(df=df, vendor_name=vendor_name, conn=conn)

def main() -> None:
    args = parse_args()
    if not args.all_months and not args.month:
        args.all_months = True
    months = None if args.all_months else args.month
    df, scan_df = load_all(args.source_root, months)
    validate_usage(df)
    write_ingest_audits(df, scan_df, args.label)
    if args.dry_run:
        print(f"DRY RUN: parsed {len(df):,} grouped usage rows. Snowflake load skipped.")
        return
    load_snowflake(df, reset=args.reset)


if __name__ == "__main__":
    main()

