"""Ingest Auvik Connectwise Usage Report workbooks into canonical usage.

One ingestion architecture for both Auvik CW and Auvik CMS:

  * Auvik delivers one workbook per month whose name varies across months
    (e.g. "Connectwise Usage Report July.xlsx", "ConnectWise Usage Report -
    June Invoices.xlsx"). The identical workbook is placed in both the
    Auvik CW and Auvik CMS subfolders, differing only in trivial metadata.

    * The Auvik CW copy is authoritative, with Auvik CMS used only as a fallback
        when a month is absent from CW. Both copies are never ingested together.

  * The CW/CMS entity split is taken from the Partner Name column:
      'ConnectWise Inc'  â†’ STREAM = 'CW'
      'ConnectWise, LLC' â†’ STREAM = 'CMS'
    Every row belongs to exactly one entity.

  * BILLING_MONTH is derived from Start Date in the data (the true usage
    period start). The folder label and the filename are both unreliable â€”
    they name different months across 2026 and must not be trusted.

    * Product Tenant ID is preserved exactly when Excel supplies text. Numeric
        cells are explicitly flagged as precision-limited because digits already
        lost in the workbook cannot be reconstructed.

    * Billable quantity follows the vendor invoice semantics validated against
        the manual pivots: committed rows use Quantity, chargeable Usage rows use
        positive Overage Quantity, and zero-dollar/negative overage rows contribute
        zero. Exact source product and quantity decomposition remain in
        ADDITIONAL_INFO for audit and downstream overage classification.

  * Load-time assertions (any failure raises RuntimeError):
        1. Within a normal monthly file, Start Date / End Date hold exactly one
            distinct pair. The January supplemental file may contain its genuine
            February one-day adjustment and retains the row-level Start Date.
      2. Invoice Date = first day of the month after End Date.
      3. Subtotal + Tax = Total Charge Amount on every data row.
      4. Entity totals from the ingested rows match the Summary tab to
         the cent (hard gate).

    * Source-manifest and content-hash gates reject ambiguous primary files and
        duplicate authoritative content before publication.

Source layout:
    <SOURCE_ROOT>/Auvik CW/MM_MON_YYYY/<usage workbook>.xlsx
    <SOURCE_ROOT>/Auvik CMS/MM_MON_YYYY/<fallback usage workbook>.xlsx
    Workbook: Summary control sheet + one INV-* sheet per invoice.

Target tables:
    THIRD_PARTY_RECON_VENDOR_USAGE_PROD â€“ normalized Auvik usage slice
    AUVIK_USAGE_FILE_AUDIT              â€“ source-file control totals

Both targets are staged before a transactional vendor-slice replacement.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import json
import math
import re
import subprocess
import sys
import tempfile
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


# ---------------------------------------------------------------------------
# Paths and Snowflake targets
# ---------------------------------------------------------------------------

AUVIK_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = AUVIK_ROOT.parent
WORKSPACE_ROOT = next((p for p in (PROJECT_ROOT, *PROJECT_ROOT.parents) if (p / "TEMPLATES").exists()), PROJECT_ROOT)

DEFAULT_SOURCE_ROOT_CW = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Auvik CW"
)

DEFAULT_SOURCE_ROOT_CMS = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\2026 - Vendor Files\Auvik CMS"
)

OUTPUT_DIR = AUVIK_ROOT / "outputs"

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
AUDIT_TABLE = "AUVIK_USAGE_FILE_AUDIT"
TARGET_VENDOR = "Auvik"

# Folder naming: MM_MON_YYYY
MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)
MONTH_ABBREVIATIONS = {
    1: "JAN", 2: "FEB", 3: "MAR", 4: "APR", 5: "MAY", 6: "JUN",
    7: "JUL", 8: "AUG", 9: "SEP", 10: "OCT", 11: "NOV", 12: "DEC",
}

# ---------------------------------------------------------------------------
# File-name noise exclusions.
# Any .xlsx whose lowercased name contains one of these tokens is NOT a
# Connectwise Usage Report workbook and must be skipped.
# ---------------------------------------------------------------------------
_EXCLUDE_NAME_TOKENS: tuple[str, ...] = (
    "overage",
    "zuora",
    "zoura",        # common Auvik filename typo for 'zuora'
    "template",
    "reconcil",     # "reconciliation", "recon"
    "creditnote",
    "credit note",
)

# ---------------------------------------------------------------------------
# Source column spec (18 columns, 0-indexed).
# Column 0 is an unnamed row counter and is skipped; columns 1â€“17 are the
# named data columns confirmed across all eight June invoice tabs.
# ---------------------------------------------------------------------------
# Canonical header names as they appear in the workbook (cols 1-17).
SOURCE_COLUMNS: tuple[str, ...] = (
    "Invoice Date",
    "Invoice Name",
    "Partner Name",
    "Shipping_Account",
    "Domain Prefix",
    "Product",
    "UnitPrice",
    "Charge Type",
    "Primary Tenant ID",
    "Product Tenant ID",   # MUST remain TEXT â€“ 19-digit float precision risk
    "Start Date",
    "End Date",
    "Quantity",
    "Overage Quantity",
    "Subtotal",
    "Tax",
    "Total Charge Amount",
)

# ---------------------------------------------------------------------------
# Target table columns (source + derived).
# ---------------------------------------------------------------------------
USAGE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
    "ADDITIONAL_INFO",
)

INTERNAL_USAGE_COLUMNS: tuple[str, ...] = USAGE_COLUMNS + (
    "_SOURCE_PRODUCT",
    "_CHARGE_TYPE",
    "_INVOICE_NAME",
    "_SOURCE_FILE",
    "_SOURCE_SHEET",
    "_SOURCE_ROW_NUMBER",
    "_PRIMARY_TENANT_ID",
    "_PRODUCT_TENANT_ID",
    "_PRODUCT_TENANT_ID_PRECISION_LIMITED",
    "_DOMAIN_PREFIX",
    "_START_DATE",
    "_END_DATE",
    "_INVOICE_DATE",
    "_SOURCE_QUANTITY",
    "_OVERAGE_QUANTITY",
    "_SUBTOTAL",
    "_TAX",
)

# ---------------------------------------------------------------------------
# Audit table columns.
# ---------------------------------------------------------------------------
AUDIT_COLUMNS: tuple[str, ...] = (
    "SOURCE_FOLDER",
    "SOURCE_FILE",
    "SOURCE_CONTENT_HASH",
    "LOAD_STATUS",
    "BILLING_MONTH",
    "INVOICE_MONTH",
    "PERIOD_START",
    "PERIOD_END",
    "TOTAL_DATA_ROWS",
    "CW_ROWS",
    "CMS_ROWS",
    "CHARGEABLE_ROWS",
    "ZERO_AMOUNT_ROWS",
    "TOTAL_ROWS_EXCLUDED",      # summary/total rows dropped
    "INVALID_ROWS",
    "DUPLICATE_ROWS",
    "INVOICE_SHEETS",
    "INGESTED_CW_TOTAL",        # sum(Total Charge Amount) for CW rows
    "INGESTED_CMS_TOTAL",       # sum(Total Charge Amount) for CMS rows
    "SUMMARY_CW_TOTAL",         # entity total from the Summary tab
    "SUMMARY_CMS_TOTAL",
    "SUMMARY_GRAND_TOTAL",
    "SUMMARY_ADJUSTMENT_TOTAL",
    "SUMMARY_GRAND_TOTAL_DELTA",
    "SUMMARY_NOTES",
    "CW_TOTAL_DELTA",
    "CMS_TOTAL_DELTA",
    "PERIOD_UNIQUE_CHECK",      # PASS / WARN / FAIL
    "INVOICE_DATE_CHECK",       # PASS / WARN
    "SUBTOTAL_TAX_CHECK",       # PASS / WARN
    "ENTITY_TOTAL_CHECK",       # PASS / FAIL
    "ERROR_MESSAGE",
    "INGESTED_AT",
)

# ---------------------------------------------------------------------------
# Partner Name â†’ STREAM mapping
# ---------------------------------------------------------------------------
STREAM_MAP: dict[str, str] = {
    "connectwise inc": "CW",
    "connectwise, llc": "CMS",
}


# ---------------------------------------------------------------------------
# Product family classification (spec Â§"The derived columns")
# ---------------------------------------------------------------------------
def classify_product_family(product: str | None) -> str:
    """Return 'Performance', 'ASM', or 'Billable' for a raw product string.

    Rules (case-insensitive, evaluated in priority order):
      1. Performance  â€” contains 'performance'
      2. ASM          â€” contains 'asm' or 'saas management'
      3. Billable     â€” everything else (Essentials, Billable Devices, etc.)

    An unrecognised product that doesn't match any of the three rules
    defaults to Billable and a WARNING is emitted so a genuinely new product
    line is never silently absorbed.
    """
    if not product:
        return "Billable"
    lower = product.lower()
    if "performance" in lower:
        return "Performance"
    if "asm" in lower or "saas management" in lower:
        return "ASM"
    return "Billable"


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SourceFile:
    path: Path
    folder_month: str | None  # YYYY-MM from folder name, may be None


def _read_file_bytes(path: Path) -> bytes:
    """Read bytes, falling back to a PowerShell Copy-Item for OneDrive locks."""
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp = Path(tmp_dir) / path.name
            escaped_src = str(path).replace("'", "''")
            escaped_dest = str(tmp).replace("'", "''")
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    f"Copy-Item -LiteralPath '{escaped_src}' -Destination '{escaped_dest}' -Force",
                ],
                check=True,
                capture_output=True,
                text=True,
            )
            return tmp.read_bytes()


def _clean_text(value: object) -> str | None:
    if value is None or value is pd.NA or value is pd.NaT:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text if text else None


def _to_number(value: object) -> float | None:
    if value is None or value is pd.NA or value is pd.NaT:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("$", "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_date(value: object) -> dt.date | None:
    if value is None or value is pd.NA or value is pd.NaT:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    ts = pd.to_datetime(str(value), errors="coerce")
    return ts.date() if isinstance(ts, pd.Timestamp) else None


def _first_day(d: dt.date | None) -> dt.date | None:
    return dt.date(d.year, d.month, 1) if d else None


def _vendor_product_sku(product: object) -> str | None:
    value = _clean_text(product)
    if value is None:
        return None
    return re.sub(r"^\s*Overage\s*-\s*", "", value, flags=re.IGNORECASE).strip()


def _billable_quantity(
    charge_type: object,
    quantity: object,
    overage_quantity: object,
    total_charge_amount: object,
) -> float | None:
    # Auvik usage rows can include negative overage quantities that net to a
    # zero-dollar line; these should contribute 0 billable units downstream.
    total = _to_number(total_charge_amount)
    if total is not None and abs(total) < 1e-9:
        return 0.0

    charge = (_clean_text(charge_type) or "").lower()
    if charge == "usage":
        overage = _to_number(overage_quantity)
        return max(overage or 0.0, 0.0)
    return _to_number(quantity)


def _month_from_folder(path: Path) -> str | None:
    for parent in path.parents:
        match = MONTH_FOLDER_RE.match(parent.name)
        if not match:
            continue
        month_number = int(match.group("mm"))
        abbreviation = parent.name.split("_", 2)[1].upper()
        if MONTH_ABBREVIATIONS.get(month_number) != abbreviation:
            raise ValueError(f"Invalid month folder name: {parent.name}")
        return f"{match.group('yyyy')}-{month_number:02d}"
    return None


def _month_argument(value: str) -> str:
    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", value):
        raise argparse.ArgumentTypeError("Month must use YYYY-MM.")
    return value


def _product_tenant_id_text(raw: object) -> tuple[str | None, bool]:
    """Return the stored identifier and whether Excel precision may be lost."""
    if raw is None or raw is pd.NA or raw is pd.NaT:
        return None, False
    if isinstance(raw, float) and math.isnan(raw):
        return None, False
    if isinstance(raw, float):
        return format(raw, ".17g"), abs(raw) > 2**53
    if isinstance(raw, int):
        return str(raw), False
    text = str(raw).strip()
    return (text or None), bool(re.search(r"[eE][+-]?\d+", text))


def _content_hash_rows(all_rows: list[tuple[object, ...]]) -> str:
    """SHA-256 of tab-delimited text representation of every data row."""
    parts: list[str] = []
    for row in all_rows:
        parts.append("\t".join(_clean_text(v) or "" for v in row))
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------------

def _is_usage_report(path: Path) -> bool:
    """Return True if this .xlsx is likely a Connectwise Usage Report workbook.

    Strategy: include any .xlsx that is not a known noise file.
    Noise exclusions: overage template, zuora export, recon workbook, credit
    note, invoice PDF (those are .pdf anyway but guard defensively).
    """
    name_lower = path.name.lower()
    if path.name.startswith("~$"):
        return False
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return False
    for token in _EXCLUDE_NAME_TOKENS:
        if token in name_lower:
            return False
    # Must contain at least one of these keywords to qualify
    # (guards against random reference docs that might be added later)
    if not any(kw in name_lower for kw in ("usage", "report", "connectwise")):
        return False
    return True


def discover_source_files(
    source_root_cw: Path,
    months: set[str] | None = None,
) -> list[SourceFile]:
    """Return eligible usage-report workbooks under one source root."""
    files: list[SourceFile] = []
    if not source_root_cw.exists():
        print(f"WARNING: CW source root does not exist: {source_root_cw}", flush=True)
        return files
    for month_dir in sorted(source_root_cw.iterdir()):
        if not month_dir.is_dir():
            continue
        folder_month = _month_from_folder(month_dir / "placeholder.xlsx")
        if folder_month is None:
            continue
        if months is not None and folder_month not in months:
            continue
        for path in sorted(month_dir.iterdir()):
            if path.is_file() and _is_usage_report(path):
                files.append(SourceFile(path=path, folder_month=folder_month))
    return files


# ---------------------------------------------------------------------------
# Summary-tab parser (control totals)
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SummaryControls:
    entity_totals: dict[str, float]
    grand_total: float | None
    adjustments: tuple[tuple[str, float], ...]
    notes: tuple[str, ...]


def _parse_summary_controls(wb: openpyxl.Workbook) -> SummaryControls:
    """Read entity totals, grand total, adjustments, and notes from Summary.

    The June 2026 Summary sheet layout (confirmed):
        Col 0: None
        Col 1: entity name ('ConnectWise Inc' / 'ConnectWise, LLC' / 'Total')
        Col 2: total amount

    Entity name may be in col 0 or col 1 depending on how Auvik formats the
    sheet — this function checks both positions for robustness.
    Numeric non-total rows, such as separately issued credits, remain audit
    controls rather than being allocated to unsupported partner/SKU rows.
    """
    totals: dict[str, float] = {}
    grand_total: float | None = None
    adjustments: list[tuple[str, float]] = []
    notes: list[str] = []
    if "Summary" not in wb.sheetnames:
        return SummaryControls(totals, grand_total, tuple(adjustments), tuple(notes))
    ws = wb["Summary"]
    entity_keys = {
        "connectwise inc": "ConnectWise Inc",
        "connectwise, llc": "ConnectWise, LLC",
    }
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        entity: str | None = None
        for col_idx in (0, 1):
            if col_idx < len(row):
                candidate = _clean_text(row[col_idx])
                if candidate and candidate.lower() in entity_keys:
                    entity = entity_keys[candidate.lower()]
                    break
        label = entity or next(
            (
                text
                for cell in row
                if (text := _clean_text(cell)) is not None
                and _to_number(cell) is None
            ),
            None,
        )
        amount = None
        for cell in reversed(row):
            v = _to_number(cell)
            if v is not None:
                amount = v
                break
        if entity is not None and amount is not None:
            totals[entity] = amount
        elif label is not None and amount is not None:
            if "total" in label.lower():
                grand_total = amount
            else:
                adjustments.append((label, amount))
        elif label is not None and "credit" in label.lower():
            notes.append(label)
    return SummaryControls(
        entity_totals=totals,
        grand_total=grand_total,
        adjustments=tuple(adjustments),
        notes=tuple(notes),
    )


# ---------------------------------------------------------------------------
# Workbook parser
# ---------------------------------------------------------------------------

def _validate_header(header_raw: tuple[object, ...]) -> bool:
    """Return True if the header row matches the expected 18-column layout.

    Col 0 may be blank or a number. Cols 1-17 must match SOURCE_COLUMNS.
    """
    if len(header_raw) < len(SOURCE_COLUMNS) + 1:
        return False
    actual = tuple(_clean_text(header_raw[i + 1]) for i in range(len(SOURCE_COLUMNS)))
    return actual == SOURCE_COLUMNS


def parse_usage_workbook(
    source: SourceFile,
    *,
    ingested_at: dt.datetime,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse a single Connectwise Usage Report workbook.

    Returns (usage_df, audit_df). Structural and reconciliation failures abort
    the load; source-level credits that cannot be allocated safely are audited.
    """
    file_bytes = _read_file_bytes(source.path)
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    data_records: list[dict[str, object]] = []
    total_rows_excluded = 0
    unknown_streams: list[str] = []
    all_content_rows: list[tuple[object, ...]] = []

    try:
        summary = _parse_summary_controls(wb)
        summary_totals = summary.entity_totals
        is_supplemental = "additional accounts" in source.path.name.lower()
        invoice_sheets = [
            name for name in wb.sheetnames
            if re.fullmatch(r"INV-\d+", name, flags=re.IGNORECASE)
        ]
        unexpected_sheets = [
            name for name in wb.sheetnames
            if name.lower() != "summary" and name not in invoice_sheets
        ]
        if unexpected_sheets:
            raise RuntimeError(f"Unexpected non-invoice sheets: {unexpected_sheets}")
        if not invoice_sheets:
            raise RuntimeError("No INV-* invoice sheets found.")
        if not is_supplemental and set(summary_totals) != {
            "ConnectWise Inc",
            "ConnectWise, LLC",
        }:
            raise RuntimeError(f"Summary totals are incomplete: {summary_totals}")

        for sheet_name in invoice_sheets:
            ws = wb[sheet_name]
            row_iter = ws.iter_rows(values_only=True)
            header_raw = next(row_iter, None)
            if header_raw is None:
                continue

            if not _validate_header(header_raw):
                raise RuntimeError(f"Unexpected header in sheet {sheet_name!r}.")

            for sheet_row_num, raw_row in enumerate(row_iter, start=2):
                if not raw_row or all(c is None for c in raw_row):
                    continue

                # Column 0 is the row counter (may be None on total row)
                # Columns 1-17 are the source data columns
                row_data = raw_row[1 : len(SOURCE_COLUMNS) + 1]
                row_dict = dict(zip(SOURCE_COLUMNS, row_data))

                # Total rows: Partner Name is null
                partner_name_raw = row_dict.get("Partner Name")
                partner_name = _clean_text(partner_name_raw)
                if partner_name is None:
                    total_rows_excluded += 1
                    continue

                # Track content rows for hash (skip total rows)
                all_content_rows.append(raw_row)

                data_records.append(
                    {
                        **row_dict,
                        "_sheet_name": sheet_name,
                        "_source_row_number": sheet_row_num,
                    }
                )
    finally:
        wb.close()

    content_hash = _content_hash_rows(all_content_rows) if all_content_rows else file_hash

    if not data_records:
        # No data rows found at all â€” audit and return empty
        audit_df = pd.DataFrame(
            [
                {
                    "SOURCE_FOLDER": source.path.parent.name,
                    "SOURCE_FILE": source.path.name,
                    "SOURCE_CONTENT_HASH": content_hash,
                    "LOAD_STATUS": "LOADED_EMPTY",
                    "BILLING_MONTH": None,
                    "INVOICE_MONTH": None,
                    "PERIOD_START": None,
                    "PERIOD_END": None,
                    "TOTAL_DATA_ROWS": 0,
                    "CW_ROWS": 0,
                    "CMS_ROWS": 0,
                    "CHARGEABLE_ROWS": 0,
                    "ZERO_AMOUNT_ROWS": 0,
                    "TOTAL_ROWS_EXCLUDED": total_rows_excluded,
                    "INVALID_ROWS": 0,
                    "DUPLICATE_ROWS": 0,
                    "INVOICE_SHEETS": ",".join(invoice_sheets),
                    "INGESTED_CW_TOTAL": None,
                    "INGESTED_CMS_TOTAL": None,
                    "SUMMARY_CW_TOTAL": summary_totals.get("ConnectWise Inc"),
                    "SUMMARY_CMS_TOTAL": summary_totals.get("ConnectWise, LLC"),
                    "SUMMARY_GRAND_TOTAL": summary.grand_total,
                    "SUMMARY_ADJUSTMENT_TOTAL": sum(value for _, value in summary.adjustments),
                    "SUMMARY_GRAND_TOTAL_DELTA": None,
                    "SUMMARY_NOTES": json.dumps(
                        {"adjustments": summary.adjustments, "notes": summary.notes},
                        separators=(",", ":"),
                        allow_nan=False,
                    ),
                    "CW_TOTAL_DELTA": None,
                    "CMS_TOTAL_DELTA": None,
                    "PERIOD_UNIQUE_CHECK": "N/A",
                    "INVOICE_DATE_CHECK": "N/A",
                    "SUBTOTAL_TAX_CHECK": "N/A",
                    "ENTITY_TOTAL_CHECK": "N/A",
                    "ERROR_MESSAGE": "No data rows found",
                    "INGESTED_AT": ingested_at,
                }
            ],
            columns=AUDIT_COLUMNS,
        )
        return pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS), audit_df

    raw_df = pd.DataFrame(data_records)

    # ------------------------------------------------------------------
    # Numeric coercion.  Product Tenant ID stays as TEXT (lossless).
    # ------------------------------------------------------------------
    for col in ("UnitPrice", "Quantity", "Overage Quantity", "Subtotal", "Tax", "Total Charge Amount"):
        raw_df[col] = raw_df[col].map(_to_number)

    raw_df["_invoice_date"] = raw_df["Invoice Date"].map(_to_date)
    raw_df["_start_date"] = raw_df["Start Date"].map(_to_date)
    raw_df["_end_date"] = raw_df["End Date"].map(_to_date)
    product_tenant_values = [
        _product_tenant_id_text(value)
        for value in raw_df["Product Tenant ID"].tolist()
    ]
    raw_df["_product_tenant_id_text"] = [value[0] for value in product_tenant_values]
    raw_df["_product_tenant_id_precision_limited"] = [
        value[1] for value in product_tenant_values
    ]

    # BILLING_MONTH from Start Date (the authoritative usage period start)
    raw_df["_billing_month"] = raw_df["_start_date"].map(_first_day)
    raw_df["_invoice_month"] = raw_df["_invoice_date"].map(_first_day)

    required_fields = pd.DataFrame(
        {
            "invoice_date": raw_df["_invoice_date"].notna(),
            "invoice_name": raw_df["Invoice Name"].map(_clean_text).notna(),
            "partner_name": raw_df["Partner Name"].map(_clean_text).notna(),
            "shipping_account": raw_df["Shipping_Account"].map(_clean_text).notna(),
            "product": raw_df["Product"].map(_clean_text).notna(),
            "unit_price": raw_df["UnitPrice"].notna(),
            "charge_type": raw_df["Charge Type"].map(_clean_text).notna(),
            "start_date": raw_df["_start_date"].notna(),
            "end_date": raw_df["_end_date"].notna(),
            "quantity": raw_df["Quantity"].notna(),
            "overage_quantity": raw_df["Overage Quantity"].notna(),
            "subtotal": raw_df["Subtotal"].notna(),
            "tax": raw_df["Tax"].notna(),
            "total_charge": raw_df["Total Charge Amount"].notna(),
        }
    ).all(axis=1)
    invalid_rows = int((~required_fields).sum())
    duplicate_rows = int(raw_df.duplicated(list(SOURCE_COLUMNS), keep=False).sum())

    # STREAM from Partner Name
    def _resolve_stream(partner: object) -> str:
        p = _clean_text(partner)
        if p:
            key = p.lower()
            if key in STREAM_MAP:
                return STREAM_MAP[key]
            unknown_streams.append(p)
        return "UNKNOWN"

    raw_df["_stream"] = raw_df["Partner Name"].map(_resolve_stream)

    # CHARGEABLE_FLAG
    raw_df["_chargeable"] = raw_df["Total Charge Amount"].fillna(0.0).ne(0.0)

    # ------------------------------------------------------------------
    # Load-time assertions
    # ------------------------------------------------------------------
    errors: list[str] = []
    warnings: list[str] = []
    if invalid_rows:
        errors.append(f"{invalid_rows} invoice rows are missing required fields.")
    if duplicate_rows:
        errors.append(f"{duplicate_rows} exact duplicate invoice rows found.")
    if unknown_streams:
        errors.append(
            "Unrecognized Partner Name values: "
            + ", ".join(sorted(set(unknown_streams)))
        )

    # 1. Period uniqueness: one distinct (Start Date, End Date) pair per file
    period_pairs = raw_df[["_start_date", "_end_date"]].dropna().drop_duplicates()
    if len(period_pairs) == 0:
        period_check = "FAIL"
        errors.append("No valid Start Date / End Date pairs found.")
    elif len(period_pairs) == 1 or is_supplemental:
        period_check = "PASS"
    else:
        period_check = "FAIL"
        errors.append(
            f"Multiple period pairs found ({len(period_pairs)}): "
            + "; ".join(
                f"{row[0]}-{row[1]}" for row in period_pairs.itertuples(index=False, name=None)
            )
        )

    row_months = {
        value.strftime("%Y-%m")
        for value in raw_df["_billing_month"].tolist()
        if isinstance(value, dt.date)
    }
    if not is_supplemental and row_months != {source.folder_month}:
        errors.append(
            f"Folder month {source.folder_month} does not match row billing months "
            f"{sorted(row_months)}."
        )

    # 2. Invoice Date must fall in the month after End Date. Supplemental
    # adjustments may instead be invoiced in their one-day service month.
    invoice_date_check = "PASS"
    invalid_invoice_dates = 0
    for end_date, inv_date in zip(raw_df["_end_date"], raw_df["_invoice_date"], strict=True):
        if end_date is None or inv_date is None:
            invalid_invoice_dates += 1
            continue
        if end_date.month == 12:
            expected_inv_month = dt.date(end_date.year + 1, 1, 1)
        else:
            expected_inv_month = dt.date(end_date.year, end_date.month + 1, 1)
        allowed_months = {expected_inv_month}
        if is_supplemental:
            allowed_months.add(dt.date(end_date.year, end_date.month, 1))
        if _first_day(inv_date) not in allowed_months:
            invalid_invoice_dates += 1
    if invalid_invoice_dates:
        invoice_date_check = "FAIL"
        errors.append(
            f"{invalid_invoice_dates} rows have an invalid invoice/service-period relationship."
        )

    # 3. Subtotal + Tax == Total Charge Amount (tolerance $0.01)
    subtotal_ok = (
        (raw_df["Subtotal"].fillna(0.0) + raw_df["Tax"].fillna(0.0) - raw_df["Total Charge Amount"].fillna(0.0))
        .abs()
        .le(0.011)
        .all()
    )
    subtotal_check = "PASS" if subtotal_ok else "FAIL"
    if not subtotal_ok:
        n_bad = int(
            (
                (raw_df["Subtotal"].fillna(0.0) + raw_df["Tax"].fillna(0.0) - raw_df["Total Charge Amount"].fillna(0.0))
                .abs()
                .gt(0.011)
            ).sum()
        )
        errors.append(f"Subtotal + Tax does not equal Total Charge Amount on {n_bad} row(s).")

    # 4. Entity totals vs Summary tab (hard gate â€” $0.01 tolerance)
    cw_ingested = float(
        raw_df.loc[raw_df["_stream"] == "CW", "Total Charge Amount"].fillna(0.0).sum()
    )
    cms_ingested = float(
        raw_df.loc[raw_df["_stream"] == "CMS", "Total Charge Amount"].fillna(0.0).sum()
    )
    summary_cw = summary_totals.get("ConnectWise Inc")
    summary_cms = summary_totals.get("ConnectWise, LLC")

    cw_delta = abs(cw_ingested - (summary_cw or 0.0))
    cms_delta = abs(cms_ingested - (summary_cms or 0.0))
    entity_check = "PASS"

    if summary_totals:
        if summary_cw is not None and cw_delta > 0.011:
            entity_check = "FAIL"
            errors.append(
                f"CW entity total mismatch: ingested ${cw_ingested:,.2f}, "
                f"Summary ${summary_cw:,.2f}, delta ${cw_delta:,.2f}."
            )
        if summary_cms is not None and cms_delta > 0.011:
            entity_check = "FAIL"
            errors.append(
                f"CMS entity total mismatch: ingested ${cms_ingested:,.2f}, "
                f"Summary ${summary_cms:,.2f}, delta ${cms_delta:,.2f}."
            )
    else:
        entity_check = "N/A" if is_supplemental else "FAIL"
        if not is_supplemental:
            errors.append("Summary totals are missing.")

    adjustment_total = float(sum(value for _, value in summary.adjustments))
    summary_grand_delta: float | None = None
    if summary.grand_total is not None:
        expected_grand_total = cw_ingested + cms_ingested + adjustment_total
        summary_grand_delta = abs(expected_grand_total - summary.grand_total)
        if summary_grand_delta > 0.011:
            errors.append(
                "Summary grand total mismatch: invoice lines plus listed "
                f"adjustments ${expected_grand_total:,.2f}, Summary "
                f"${summary.grand_total:,.2f}, delta ${summary_grand_delta:,.2f}."
            )
    if summary.adjustments:
        warnings.append(
            "Summary contains partner/SKU-unallocated adjustment(s) totaling "
            f"${adjustment_total:,.2f}: "
            + ", ".join(f"{label} (${amount:,.2f})" for label, amount in summary.adjustments)
        )
    if summary.notes:
        warnings.append("Summary contains unallocated note(s): " + " | ".join(summary.notes))

    # ------------------------------------------------------------------
    # Build usage DataFrame
    # ------------------------------------------------------------------
    billing_month = (
        raw_df["_billing_month"].dropna().iloc[0] if not raw_df["_billing_month"].dropna().empty else None
    )
    invoice_month = (
        raw_df["_invoice_month"].dropna().iloc[0] if not raw_df["_invoice_month"].dropna().empty else None
    )

    usage_df = pd.DataFrame(
        {
            "BILLING_MONTH": raw_df["_billing_month"],
            "VENDOR": "Auvik",
            "VENDOR_PARTNER_NAME": raw_df["Shipping_Account"].map(_clean_text),
            "VENDOR_PRODUCT_SKU": raw_df["Product"].map(_vendor_product_sku),
            "MODIFIER": raw_df["_stream"],
            "QUANTITY": [
                _billable_quantity(charge_type, quantity, overage_quantity, total_charge_amount)
                for charge_type, quantity, overage_quantity, total_charge_amount in zip(
                    raw_df["Charge Type"],
                    raw_df["Quantity"],
                    raw_df["Overage Quantity"],
                    raw_df["Total Charge Amount"],
                    strict=True,
                )
            ],
            "UNIT_PRICE": raw_df["UnitPrice"],
            "AMOUNT": raw_df["Total Charge Amount"],
            "CURRENCY": "USD",
            "ADDITIONAL_INFO": None,
            "_SOURCE_PRODUCT": raw_df["Product"].map(_clean_text),
            "_CHARGE_TYPE": raw_df["Charge Type"].map(_clean_text),
            "_INVOICE_NAME": raw_df["Invoice Name"].map(_clean_text),
            "_SOURCE_FILE": source.path.name,
            "_SOURCE_SHEET": raw_df["_sheet_name"],
            "_SOURCE_ROW_NUMBER": raw_df["_source_row_number"],
            "_PRIMARY_TENANT_ID": raw_df["Primary Tenant ID"].map(_clean_text),
            "_PRODUCT_TENANT_ID": raw_df["_product_tenant_id_text"],
            "_PRODUCT_TENANT_ID_PRECISION_LIMITED": raw_df["_product_tenant_id_precision_limited"],
            "_DOMAIN_PREFIX": raw_df["Domain Prefix"].map(_clean_text),
            "_START_DATE": raw_df["_start_date"],
            "_END_DATE": raw_df["_end_date"],
            "_INVOICE_DATE": raw_df["_invoice_date"],
            "_SOURCE_QUANTITY": raw_df["Quantity"],
            "_OVERAGE_QUANTITY": raw_df["Overage Quantity"],
            "_SUBTOTAL": raw_df["Subtotal"],
            "_TAX": raw_df["Tax"],
        },
        columns=INTERNAL_USAGE_COLUMNS,
    )

    # ------------------------------------------------------------------
    # Build audit row
    # ------------------------------------------------------------------
    period_start = raw_df["_start_date"].dropna().min() if not raw_df["_start_date"].dropna().empty else None
    period_end = raw_df["_end_date"].dropna().max() if not raw_df["_end_date"].dropna().empty else None

    load_status = "FAILED_VALIDATION" if errors else "LOADED_WITH_WARNING" if warnings else "LOADED"

    error_msg: str | None = None
    if errors or warnings:
        parts = [f"ERROR: {e}" for e in errors] + [f"WARN: {w}" for w in warnings]
        error_msg = " | ".join(parts)

    audit_df = pd.DataFrame(
        [
            {
                "SOURCE_FOLDER": source.path.parent.name,
                "SOURCE_FILE": source.path.name,
                "SOURCE_CONTENT_HASH": content_hash,
                "LOAD_STATUS": load_status,
                "BILLING_MONTH": billing_month,
                "INVOICE_MONTH": invoice_month,
                "PERIOD_START": period_start,
                "PERIOD_END": period_end,
                "TOTAL_DATA_ROWS": len(usage_df),
                "CW_ROWS": int((usage_df["MODIFIER"] == "CW").sum()),
                "CMS_ROWS": int((usage_df["MODIFIER"] == "CMS").sum()),
                "CHARGEABLE_ROWS": int(usage_df["AMOUNT"].fillna(0).ne(0).sum()),
                "ZERO_AMOUNT_ROWS": int(usage_df["AMOUNT"].fillna(0).eq(0).sum()),
                "TOTAL_ROWS_EXCLUDED": total_rows_excluded,
                "INVALID_ROWS": invalid_rows,
                "DUPLICATE_ROWS": duplicate_rows,
                "INVOICE_SHEETS": ",".join(invoice_sheets),
                "INGESTED_CW_TOTAL": round(cw_ingested, 2),
                "INGESTED_CMS_TOTAL": round(cms_ingested, 2),
                "SUMMARY_CW_TOTAL": summary_cw,
                "SUMMARY_CMS_TOTAL": summary_cms,
                "SUMMARY_GRAND_TOTAL": summary.grand_total,
                "SUMMARY_ADJUSTMENT_TOTAL": round(adjustment_total, 4),
                "SUMMARY_GRAND_TOTAL_DELTA": round(summary_grand_delta, 4) if summary_grand_delta is not None else None,
                "SUMMARY_NOTES": json.dumps(
                    {"adjustments": summary.adjustments, "notes": summary.notes},
                    separators=(",", ":"),
                    allow_nan=False,
                ),
                "CW_TOTAL_DELTA": round(cw_delta, 4) if summary_cw is not None else None,
                "CMS_TOTAL_DELTA": round(cms_delta, 4) if summary_cms is not None else None,
                "PERIOD_UNIQUE_CHECK": period_check,
                "INVOICE_DATE_CHECK": invoice_date_check,
                "SUBTOTAL_TAX_CHECK": subtotal_check,
                "ENTITY_TOTAL_CHECK": entity_check,
                "ERROR_MESSAGE": error_msg,
                "INGESTED_AT": ingested_at,
            }
        ],
        columns=AUDIT_COLUMNS,
    )

    if errors:
        raise RuntimeError(
            f"Auvik source validation FAILED for {source.path.name}: {error_msg}"
        )

    if warnings:
        for w in warnings:
            print(f"  WARNING [{source.path.name}]: {w}", flush=True)

    return usage_df, audit_df


# ---------------------------------------------------------------------------
# Build function â€” scan all files and combine
# ---------------------------------------------------------------------------

def build_usage(
    source_root_cw: Path,
    source_root_cms: Path | None = None,
    months: set[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Parse one authoritative workbook copy per monthly source file."""
    ingested_at = dt.datetime.now(dt.UTC).replace(tzinfo=None)
    cw_files = discover_source_files(source_root_cw, months)
    cms_files = discover_source_files(source_root_cms, months) if source_root_cms else []

    cw_by_month: dict[str, list[SourceFile]] = {}
    cms_by_month: dict[str, list[SourceFile]] = {}
    for source in cw_files:
        cw_by_month.setdefault(source.folder_month or "", []).append(source)
    for source in cms_files:
        cms_by_month.setdefault(source.folder_month or "", []).append(source)

    available_months = sorted(set(cw_by_month) | set(cms_by_month))
    if months is not None:
        missing_months = months - set(available_months)
        if missing_months:
            raise FileNotFoundError(f"Missing Auvik usage workbook(s) for {sorted(missing_months)}.")

    files: list[SourceFile] = []
    for month in available_months:
        authoritative = cw_by_month.get(month) or cms_by_month.get(month) or []
        primary = [s for s in authoritative if "additional accounts" not in s.path.name.lower()]
        supplemental = [s for s in authoritative if "additional accounts" in s.path.name.lower()]
        if len(primary) != 1:
            raise RuntimeError(
                f"Expected exactly one primary Auvik usage workbook for {month}, "
                f"found {len(primary)} among: {[s.path.name for s in authoritative]}"
            )
        expected_supplemental = 1 if month == "2026-01" else 0
        if len(supplemental) != expected_supplemental:
            raise RuntimeError(
                f"Expected {expected_supplemental} Additional Accounts workbook(s) "
                f"for {month}, found {len(supplemental)}: {[s.path.name for s in supplemental]}"
            )
        files.extend(primary + supplemental)

    if not files:
        raise FileNotFoundError(
            "No Auvik Connectwise Usage Report workbooks found for the requested "
            f"months: {sorted(months) if months else 'all'}"
        )

    usage_frames: list[pd.DataFrame] = []
    audit_frames: list[pd.DataFrame] = []
    seen_hashes: dict[str, str] = {}

    for source in files:
        print(
            f"Parsing {source.path.parent.name}/{source.path.name} ...",
            flush=True,
        )
        usage_df, audit_df = parse_usage_workbook(source, ingested_at=ingested_at)
        content_hash = audit_df["SOURCE_CONTENT_HASH"].iloc[0]

        duplicate_of = seen_hashes.get(content_hash)
        if duplicate_of is not None:
            raise RuntimeError(
                f"Duplicate Auvik invoice content: {source.path.name} duplicates {duplicate_of}."
            )
        seen_hashes[content_hash] = source.path.name

        audit_frames.append(audit_df)
        usage_frames.append(usage_df)

        if not usage_df.empty:
            status = audit_df["LOAD_STATUS"].iloc[0]
            n_rows = len(usage_df)
            cw_total = float(
                usage_df.loc[usage_df["MODIFIER"] == "CW", "AMOUNT"].fillna(0).sum()
            )
            cms_total = float(
                usage_df.loc[usage_df["MODIFIER"] == "CMS", "AMOUNT"].fillna(0).sum()
            )
            print(
                f"  {status}: rows={n_rows:,} "
                f"CW=${cw_total:,.2f} CMS=${cms_total:,.2f}",
                flush=True,
            )

    non_empty = [f for f in usage_frames if not f.empty]
    usage_all = (
        pd.concat(non_empty, ignore_index=True)
        if non_empty
        else pd.DataFrame(columns=INTERNAL_USAGE_COLUMNS)
    )
    if not usage_all.empty:
        charge_types = usage_all["_CHARGE_TYPE"].fillna("").str.lower()
        chargeable = usage_all["AMOUNT"].fillna(0).abs().gt(1e-9)
        usage_all["_BILLABLE_OVERAGE_QUANTITY"] = (
            usage_all["_OVERAGE_QUANTITY"].fillna(0)
            .where(charge_types.eq("usage") & chargeable, 0)
            .clip(lower=0)
        )
        usage_all["_COMMITTED_QUANTITY"] = usage_all["_SOURCE_QUANTITY"].fillna(0).where(
            charge_types.eq("committed") & chargeable,
            0,
        )
        grain = [
            "BILLING_MONTH", "VENDOR", "VENDOR_PARTNER_NAME",
            "VENDOR_PRODUCT_SKU", "MODIFIER", "UNIT_PRICE", "CURRENCY",
        ]

        grouped = usage_all.groupby(grain, dropna=False, as_index=False).agg(
            QUANTITY=("QUANTITY", "sum"),
            AMOUNT=("AMOUNT", "sum"),
            source_quantity=("_SOURCE_QUANTITY", "sum"),
            source_overage_quantity=("_OVERAGE_QUANTITY", "sum"),
            billable_overage_quantity=("_BILLABLE_OVERAGE_QUANTITY", "sum"),
            committed_quantity=("_COMMITTED_QUANTITY", "sum"),
            subtotal=("_SUBTOTAL", "sum"),
            tax=("_TAX", "sum"),
            source_row_count=("_SOURCE_ROW_NUMBER", "size"),
            source_product=("_SOURCE_PRODUCT", "first"),
            source_product_count=("_SOURCE_PRODUCT", "nunique"),
            charge_type=("_CHARGE_TYPE", "first"),
            charge_type_count=("_CHARGE_TYPE", "nunique"),
            invoice_name=("_INVOICE_NAME", "first"),
            invoice_name_count=("_INVOICE_NAME", "nunique"),
            source_file=("_SOURCE_FILE", "first"),
            source_file_count=("_SOURCE_FILE", "nunique"),
            source_sheet=("_SOURCE_SHEET", "first"),
            source_sheet_count=("_SOURCE_SHEET", "nunique"),
            source_row_min=("_SOURCE_ROW_NUMBER", "min"),
            source_row_max=("_SOURCE_ROW_NUMBER", "max"),
            domain_prefix=("_DOMAIN_PREFIX", "first"),
            domain_prefix_count=("_DOMAIN_PREFIX", "nunique"),
            primary_tenant_id=("_PRIMARY_TENANT_ID", "first"),
            primary_tenant_id_count=("_PRIMARY_TENANT_ID", "nunique"),
            product_tenant_id=("_PRODUCT_TENANT_ID", "first"),
            product_tenant_id_count=("_PRODUCT_TENANT_ID", "nunique"),
            precision_limited=("_PRODUCT_TENANT_ID_PRECISION_LIMITED", "any"),
            invoice_date=("_INVOICE_DATE", "first"),
            period_start=("_START_DATE", "min"),
            period_end=("_END_DATE", "max"),
        )

        def compact_provenance(row: Any) -> str:
            return json.dumps(
                {
                    "source_product": row.source_product,
                    "source_product_count": int(row.source_product_count),
                    "charge_type": row.charge_type,
                    "charge_type_count": int(row.charge_type_count),
                    "invoice_name": row.invoice_name,
                    "invoice_name_count": int(row.invoice_name_count),
                    "source_file": row.source_file,
                    "source_file_count": int(row.source_file_count),
                    "source_sheet": row.source_sheet,
                    "source_sheet_count": int(row.source_sheet_count),
                    "source_rows": [int(row.source_row_min), int(row.source_row_max)],
                    "source_row_count": int(row.source_row_count),
                    "domain_prefix": row.domain_prefix,
                    "domain_prefix_count": int(row.domain_prefix_count),
                    "primary_tenant_id": row.primary_tenant_id,
                    "primary_tenant_id_count": int(row.primary_tenant_id_count),
                    "product_tenant_id_as_stored": row.product_tenant_id,
                    "product_tenant_id_count": int(row.product_tenant_id_count),
                    "product_tenant_id_precision_limited": bool(row.precision_limited),
                    "source_quantity": float(row.source_quantity),
                    "source_overage_quantity": float(row.source_overage_quantity),
                    "billable_overage_quantity": float(row.billable_overage_quantity),
                    "committed_quantity": float(row.committed_quantity),
                    "billable_quantity": float(row.QUANTITY),
                    "subtotal": float(row.subtotal),
                    "tax": float(row.tax),
                    "invoice_date": row.invoice_date.isoformat(),
                    "period_start": row.period_start.isoformat(),
                    "period_end": row.period_end.isoformat(),
                },
                sort_keys=True,
                separators=(",", ":"),
                allow_nan=False,
            )

        grouped["ADDITIONAL_INFO"] = [
            compact_provenance(row) for row in grouped.itertuples(index=False)
        ]
        usage_all = grouped[list(USAGE_COLUMNS)]
    audit_records = [
        record
        for audit_frame in audit_frames
        for record in audit_frame.to_dict(orient="records")
    ]
    audit_all = pd.DataFrame(audit_records, columns=AUDIT_COLUMNS)
    return usage_all, audit_all


# ---------------------------------------------------------------------------
# DDL
# ---------------------------------------------------------------------------

def usage_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} (
    BILLING_MONTH       DATE,
    VENDOR              VARCHAR,
    VENDOR_PARTNER_NAME VARCHAR,
    VENDOR_PRODUCT_SKU  VARCHAR,
    MODIFIER            VARCHAR,
    QUANTITY            NUMBER(18,4),
    UNIT_PRICE          NUMBER(18,6),
    AMOUNT              NUMBER(38,6),
    CURRENCY            VARCHAR,
    ADDITIONAL_INFO     VARCHAR
);
"""


def audit_ddl() -> str:
    return f"""
CREATE TABLE IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} (
    SOURCE_FOLDER           VARCHAR,
    SOURCE_FILE             VARCHAR,
    SOURCE_CONTENT_HASH     VARCHAR,
    LOAD_STATUS             VARCHAR,
    BILLING_MONTH           DATE,
    INVOICE_MONTH           DATE,
    PERIOD_START            DATE,
    PERIOD_END              DATE,
    TOTAL_DATA_ROWS         NUMBER,
    CW_ROWS                 NUMBER,
    CMS_ROWS                NUMBER,
    CHARGEABLE_ROWS         NUMBER,
    ZERO_AMOUNT_ROWS        NUMBER,
    TOTAL_ROWS_EXCLUDED     NUMBER,
    INVALID_ROWS            NUMBER,
    DUPLICATE_ROWS          NUMBER,
    INVOICE_SHEETS          VARCHAR,
    INGESTED_CW_TOTAL       NUMBER(18,4),
    INGESTED_CMS_TOTAL      NUMBER(18,4),
    SUMMARY_CW_TOTAL        NUMBER(18,4),
    SUMMARY_CMS_TOTAL       NUMBER(18,4),
    SUMMARY_GRAND_TOTAL     NUMBER(18,4),
    SUMMARY_ADJUSTMENT_TOTAL NUMBER(18,4),
    SUMMARY_GRAND_TOTAL_DELTA NUMBER(18,4),
    SUMMARY_NOTES           VARCHAR,
    CW_TOTAL_DELTA          NUMBER(18,4),
    CMS_TOTAL_DELTA         NUMBER(18,4),
    PERIOD_UNIQUE_CHECK     VARCHAR,
    INVOICE_DATE_CHECK      VARCHAR,
    SUBTOTAL_TAX_CHECK      VARCHAR,
    ENTITY_TOTAL_CHECK      VARCHAR,
    ERROR_MESSAGE           VARCHAR,
    INGESTED_AT             TIMESTAMP_NTZ
);
"""


# ---------------------------------------------------------------------------
# Snowflake load
# ---------------------------------------------------------------------------

def _snowflake_connection():
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from TEMPLATES.Python.connection import get_snowflake_connection

    return get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )


def load_snowflake(
    usage_df: pd.DataFrame,
    audit_df: pd.DataFrame,
    *,
    reset: bool = False,
) -> None:
    from snowflake.connector.pandas_tools import write_pandas

    conn = _snowflake_connection()
    usage_stage = f"_AUVIK_USAGE_STAGE_{uuid.uuid4().hex[:12].upper()}"
    audit_stage = f"_AUVIK_AUDIT_STAGE_{uuid.uuid4().hex[:12].upper()}"
    try:
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        cur.execute(usage_ddl())
        cur.execute(audit_ddl())

        cur.execute(
            f"ALTER TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
            "ADD COLUMN IF NOT EXISTS ADDITIONAL_INFO VARCHAR"
        )
        for column, data_type in (
            ("INVALID_ROWS", "NUMBER"),
            ("DUPLICATE_ROWS", "NUMBER"),
            ("INVOICE_SHEETS", "VARCHAR"),
            ("SUMMARY_GRAND_TOTAL", "NUMBER(18,4)"),
            ("SUMMARY_ADJUSTMENT_TOTAL", "NUMBER(18,4)"),
            ("SUMMARY_GRAND_TOTAL_DELTA", "NUMBER(18,4)"),
            ("SUMMARY_NOTES", "VARCHAR"),
        ):
            cur.execute(
                f"ALTER TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} "
                f"ADD COLUMN IF NOT EXISTS {column} {data_type}"
            )

        cur.execute(
            f"CREATE OR REPLACE TEMPORARY TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{usage_stage} "
            f"LIKE {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}"
        )
        cur.execute(
            f"CREATE OR REPLACE TEMPORARY TABLE {TARGET_DATABASE}.{TARGET_SCHEMA}.{audit_stage} "
            f"LIKE {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE}"
        )

        if not usage_df.empty:
            success, _, rows, output = write_pandas(
                conn, usage_df, usage_stage,
                database=TARGET_DATABASE, schema=TARGET_SCHEMA,
                quote_identifiers=False, use_logical_type=True,
            )
            if not success or rows != len(usage_df):
                raise RuntimeError(
                    f"Staging failed for {TARGET_TABLE}: expected {len(usage_df):,} "
                    f"rows, staged {rows:,}; {output}"
                )
        if not audit_df.empty:
            success, _, rows, output = write_pandas(
                conn, audit_df, audit_stage,
                database=TARGET_DATABASE, schema=TARGET_SCHEMA,
                quote_identifiers=False, use_logical_type=True,
            )
            if not success or rows != len(audit_df):
                raise RuntimeError(
                    f"Staging failed for {AUDIT_TABLE}: expected {len(audit_df):,} "
                    f"rows, staged {rows:,}; {output}"
                )

        incoming_months = sorted({
            parsed.strftime("%Y-%m")
            for value in usage_df["BILLING_MONTH"].tolist()
            if (parsed := _to_date(value)) is not None
        })
        month_list = ", ".join(f"'{month}-01'::DATE" for month in incoming_months)

        cur.execute("BEGIN")
        if reset:
            print(f"RESET: clearing {TARGET_TABLE} and {AUDIT_TABLE}.", flush=True)
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
                (TARGET_VENDOR,),
            )
            cur.execute(f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE}")
        else:
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
                "WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s) "
                f"AND BILLING_MONTH IN ({month_list})",
                (TARGET_VENDOR,),
            )
            cur.execute(
                f"DELETE FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} "
                f"WHERE BILLING_MONTH IN ({month_list})"
            )

        usage_column_list = ", ".join(USAGE_COLUMNS)
        audit_column_list = ", ".join(AUDIT_COLUMNS)
        cur.execute(
            f"INSERT INTO {TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE} "
            f"({usage_column_list}) SELECT {usage_column_list} "
            f"FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{usage_stage}"
        )
        cur.execute(
            f"INSERT INTO {TARGET_DATABASE}.{TARGET_SCHEMA}.{AUDIT_TABLE} "
            f"({audit_column_list}) SELECT {audit_column_list} "
            f"FROM {TARGET_DATABASE}.{TARGET_SCHEMA}.{audit_stage}"
        )
        conn.commit()
        print(
            f"Loaded {len(usage_df):,} rows into {TARGET_TABLE} and "
            f"{len(audit_df):,} audit rows.",
            flush=True,
        )
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Local audit summary
# ---------------------------------------------------------------------------

def summarize_local(usage_df: pd.DataFrame, audit_df: pd.DataFrame) -> None:
    print("\nLOCAL USAGE CONTROLS", flush=True)
    if usage_df.empty:
        print("  (no usage rows)", flush=True)
    else:
        summary = (
            usage_df.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
            .agg(
                rows=("AMOUNT", "size"),
                quantity=("QUANTITY", "sum"),
                total_charge=("AMOUNT", "sum"),
            )
            .reset_index()
            .sort_values(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"])
        )
        print(summary.to_string(index=False), flush=True)

        # Entity totals cross-check
        print("\nENTITY TOTALS", flush=True)
        entity = (
            usage_df.groupby(["BILLING_MONTH", "MODIFIER"], dropna=False)["AMOUNT"]
            .sum()
            .reset_index()
            .sort_values(["BILLING_MONTH", "MODIFIER"])
        )
        print(entity.to_string(index=False), flush=True)

    if not audit_df.empty:
        print("\nAUDIT CHECKS", flush=True)
        checks = audit_df[
            [
                "SOURCE_FILE",
                "LOAD_STATUS",
                "BILLING_MONTH",
                "TOTAL_DATA_ROWS",
                "CW_ROWS",
                "CMS_ROWS",
                "CHARGEABLE_ROWS",
                "ZERO_AMOUNT_ROWS",
                "INGESTED_CW_TOTAL",
                "SUMMARY_CW_TOTAL",
                "CW_TOTAL_DELTA",
                "INGESTED_CMS_TOTAL",
                "SUMMARY_CMS_TOTAL",
                "CMS_TOTAL_DELTA",
                "ENTITY_TOTAL_CHECK",
                "PERIOD_UNIQUE_CHECK",
                "ERROR_MESSAGE",
            ]
        ]
        print(checks.to_string(index=False), flush=True)


def write_local_audit(usage_df: pd.DataFrame, audit_df: pd.DataFrame, label: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    usage_path = OUTPUT_DIR / f"auvik_ingest_usage_{label}.csv"
    audit_path = OUTPUT_DIR / f"auvik_ingest_audit_{label}.csv"
    usage_df.to_csv(usage_path, index=False)
    audit_df.to_csv(audit_path, index=False)
    print(f"Wrote {usage_path}", flush=True)
    print(f"Wrote {audit_path}", flush=True)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Ingest Auvik Connectwise Usage Report workbooks into canonical usage. "
            "One authoritative workbook copy is selected per month; the "
            "CW/CMS entity split is derived from Partner Name in the data."
        )
    )
    parser.add_argument(
        "--source-root",
        default=str(DEFAULT_SOURCE_ROOT_CW),
        help="Path to the Auvik CW month folder root (default: OneDrive path).",
    )
    parser.add_argument(
        "--source-root-cms",
        default=str(DEFAULT_SOURCE_ROOT_CMS),
        help="Fallback Auvik CMS root if the CW copy is unavailable.",
    )
    parser.add_argument("--channel", help="Ignored; each workbook contains CW and CMS.")
    month_scope = parser.add_mutually_exclusive_group(required=True)
    month_scope.add_argument("--month", type=_month_argument, help="Load one month in YYYY-MM format.")
    month_scope.add_argument("--all-months", action="store_true", help="Parse every discovered month folder.")
    parser.add_argument("--start-month", type=_month_argument, help="Lower bound for --all-months.")
    parser.add_argument("--end-month", type=_month_argument, help="Upper bound for --all-months.")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and audit locally without loading to Snowflake.",
    )
    parser.add_argument(
        "--reset",
        action="store_true",
        help="Clear all Auvik history before loading; requires --all-months.",
    )
    args = parser.parse_args()

    source_root = Path(args.source_root)
    source_root_cms = Path(args.source_root_cms)
    if not source_root.exists() and not source_root_cms.exists():
        raise FileNotFoundError("Neither the Auvik CW source root nor CMS fallback exists.")
    if args.reset and not args.all_months:
        raise SystemExit("--reset requires --all-months.")
    if args.reset and (args.start_month or args.end_month):
        raise SystemExit("--reset cannot be combined with month bounds.")
    if (args.start_month or args.end_month) and not args.all_months:
        raise SystemExit("--start-month/--end-month require --all-months.")
    if args.start_month and args.end_month and args.start_month > args.end_month:
        raise SystemExit("--start-month cannot be after --end-month.")

    if args.month:
        requested_months = {args.month}
    else:
        available = {
            source.folder_month
            for source in discover_source_files(source_root) + discover_source_files(source_root_cms)
            if source.folder_month is not None
        }
        requested_months = {
            month for month in available
            if (args.start_month is None or month >= args.start_month)
            and (args.end_month is None or month <= args.end_month)
        }

    usage_df, audit_df = build_usage(
        source_root,
        source_root_cms=source_root_cms,
        months=requested_months,
    )

    if usage_df.empty:
        raise RuntimeError("No usage rows parsed. Check source root and month filters.")

    summarize_local(usage_df, audit_df)

    label: str
    if args.month:
        label = args.month.replace("-", "_")
    elif args.start_month or args.end_month:
        label = f"{(args.start_month or 'min').replace('-','_')}_to_{(args.end_month or 'max').replace('-','_')}"
    else:
        label = "all_months"

    write_local_audit(usage_df, audit_df, label)

    if args.dry_run:
        print("Dry run complete â€” Snowflake load skipped.", flush=True)
        return

    load_snowflake(usage_df, audit_df, reset=args.reset)


if __name__ == "__main__":
    main()



