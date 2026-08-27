from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from snowflake.connector.pandas_tools import write_pandas


SCRIPT_DIR = Path(__file__).resolve().parent
PACKAGE_ROOT = SCRIPT_DIR.parent
BITDEFENDER_ROOT = PACKAGE_ROOT.parent
WORKSPACE_ROOT = next((p for p in (SCRIPT_DIR, *SCRIPT_DIR.parents) if (p / "TEMPLATES").exists()), None)
sys.path.insert(0, str(WORKSPACE_ROOT))

from TEMPLATES.Python.connection import get_snowflake_connection


DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026\Bitdefender"
)
RAW_TABLE = "BITDEFENDER_ROYALTY_REPORT_RAW_PROD"
TARGET_TABLE = "ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_USAGE_PROD"

# Writes directly into the shared combined usage table (THIRD_PARTY_RECON_VENDOR_USAGE_PROD),
# matching the pattern used by all other vendor ingestion scripts.  No individual
# BITDEFENDER_USAGE_PROD staging table is created.
PUBLISH_SQL = f"""
USE ROLE DEVELOPER;
USE WAREHOUSE REPORTING_WH;
USE DATABASE ANALYTICS_DEV;
USE SCHEMA DBT_NFOLD_TRANSFORMATION;

DELETE FROM {TARGET_TABLE}
WHERE VENDOR = 'Bitdefender';

INSERT INTO {TARGET_TABLE} (
    BILLING_MONTH, VENDOR, VENDOR_PARTNER_NAME, VENDOR_PRODUCT_SKU, MODIFIER,
    QUANTITY, UNIT_PRICE, AMOUNT, CURRENCY
)
SELECT
    BILLING_MONTH,
    'Bitdefender'::VARCHAR AS VENDOR,
    VENDOR_PARTNER_NAME,
    VENDOR_PRODUCT_SKU,
    NULLIF(
        CONCAT_WS(
            ' | ',
            NULLIF(THIRD_PARTY_TYPE, ''),
            NULLIF(CHARGE_OR_CREDIT, ''),
            NULLIF(INVOICE_NUMBER, '')
        ),
        ''
    ) AS MODIFIER,
    SUM(QUANTITY)::NUMBER(18,4) AS QUANTITY,
    CASE
        WHEN COUNT(DISTINCT UNIT_PRICE) = 1 THEN MAX(UNIT_PRICE)::NUMBER(18,6)
        WHEN SUM(QUANTITY) > 0 THEN (SUM(AMOUNT) / SUM(QUANTITY))::NUMBER(18,6)
        ELSE NULL::NUMBER(18,6)
    END AS UNIT_PRICE,
    SUM(AMOUNT)::NUMBER(18,4) AS AMOUNT,
    'USD'::VARCHAR AS CURRENCY
FROM {RAW_TABLE}
GROUP BY
    BILLING_MONTH,
    VENDOR_PARTNER_NAME,
    VENDOR_PRODUCT_SKU,
    NULLIF(
        CONCAT_WS(
            ' | ',
            NULLIF(THIRD_PARTY_TYPE, ''),
            NULLIF(CHARGE_OR_CREDIT, ''),
            NULLIF(INVOICE_NUMBER, '')
        ),
        ''
    );
"""


def format_seconds(seconds: float) -> str:
    minutes, remainder = divmod(seconds, 60)
    if minutes >= 1:
        return f"{int(minutes)}m {remainder:.1f}s"
    return f"{seconds:.1f}s"


def parse_month(path: Path) -> datetime:
    match = re.search(r"_(\d{2})\.(\d{4})$", path.stem)
    if not match:
        raise ValueError(f"Cannot parse billing month from {path.name}")
    return datetime(int(match.group(2)), int(match.group(1)), 1)


def iter_source_files(source_root: Path, month: str | None) -> list[Path]:
    files = []
    for path in source_root.glob("**/Bitdefender_CW_Royalty Report_*.xlsx"):
        if "Bitdefender_unzipped" in str(path):
            continue
        file_month = parse_month(path).strftime("%Y-%m")
        if month and file_month != month:
            continue
        files.append(path)
    return sorted(files)


def number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_royalty_report(path: Path) -> list[dict[str, object]]:
    billing_month = parse_month(path)
    rows: list[dict[str, object]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_month = row[0] if len(row) > 0 else None
            if not isinstance(row_month, datetime):
                continue
            if row_month.year != billing_month.year or row_month.month != billing_month.month:
                continue

            qty = number(row[14] if len(row) > 14 else None)
            amount = number(row[16] if len(row) > 16 else None)
            if qty is None and amount is None:
                continue

            rows.append(
                {
                    "BILLING_MONTH": billing_month.date(),
                    "VENDOR": "Bitdefender",
                    "CHARGE_OR_CREDIT": clean(row[2] if len(row) > 2 else None),
                    "SF_ACCOUNT_NBR": clean(row[3] if len(row) > 3 else None),
                    "THIRD_PARTY_TYPE": clean(row[4] if len(row) > 4 else None),
                    "INVOICE_NUMBER": clean(row[5] if len(row) > 5 else None),
                    "VENDOR_PARTNER_NAME": clean(row[6] if len(row) > 6 else None),
                    "VENDOR_SKU": clean(row[10] if len(row) > 10 else None),
                    "CW_SKU": clean(row[11] if len(row) > 11 else None),
                    "PRODUCT_DESCRIPTION": clean(row[12] if len(row) > 12 else None),
                    "VENDOR_PRODUCT_SKU": clean(row[13] if len(row) > 13 else None),
                    "QUANTITY": qty or 0.0,
                    "UNIT_PRICE": number(row[15] if len(row) > 15 else None),
                    "AMOUNT": amount or 0.0,
                    "SOURCE_FILE": str(path),
                }
            )
    finally:
        wb.close()
    return rows


def load_raw_table(conn, df: pd.DataFrame) -> None:
    cur = conn.cursor()
    try:
        cur.execute(f"""
            CREATE OR REPLACE TABLE {RAW_TABLE} (
                BILLING_MONTH DATE,
                VENDOR VARCHAR,
                CHARGE_OR_CREDIT VARCHAR,
                SF_ACCOUNT_NBR VARCHAR,
                THIRD_PARTY_TYPE VARCHAR,
                INVOICE_NUMBER VARCHAR,
                VENDOR_PARTNER_NAME VARCHAR,
                VENDOR_SKU VARCHAR,
                CW_SKU VARCHAR,
                PRODUCT_DESCRIPTION VARCHAR,
                VENDOR_PRODUCT_SKU VARCHAR,
                QUANTITY NUMBER(18,4),
                UNIT_PRICE NUMBER(18,6),
                AMOUNT NUMBER(18,4),
                SOURCE_FILE VARCHAR
            )
        """)
        success, _, row_count, _ = write_pandas(
            conn,
            df,
            RAW_TABLE,
            database="ANALYTICS_DEV",
            schema="DBT_NFOLD_TRANSFORMATION",
            quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"write_pandas failed for {RAW_TABLE}")
        print(f"Loaded {row_count:,} raw Bitdefender royalty rows into {RAW_TABLE}.", flush=True)
    finally:
        cur.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Load Bitdefender production usage from CW royalty report workbooks. "
            "Portal exportMyUsage files are intentionally ignored because they are not the billable source."
        )
    )
    parser.add_argument("--month", help="Single month to load, YYYY-MM.")
    parser.add_argument("--all-months", action="store_true", help="Load every available royalty report workbook.")
    parser.add_argument("--reset", action="store_true", help="Accepted for runner compatibility; the raw table is always rebuilt.")
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT), help="Root folder containing Bitdefender monthly manual recon folders.")
    args = parser.parse_args()

    start = time.perf_counter()
    source_root = Path(args.source_root)
    source_files = iter_source_files(source_root, args.month)
    if not source_files:
        raise FileNotFoundError(f"No Bitdefender_CW_Royalty Report workbooks found under {source_root}")

    records: list[dict[str, object]] = []
    for source_file in source_files:
        file_rows = read_royalty_report(source_file)
        records.extend(file_rows)
        print(f"Read {len(file_rows):,} rows from {source_file.name}.", flush=True)

    df = pd.DataFrame.from_records(records)
    if df.empty:
        raise RuntimeError("No royalty rows were parsed from the Bitdefender royalty reports.")

    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database="ANALYTICS_DEV",
        schema="DBT_NFOLD_TRANSFORMATION",
    )
    try:
        load_raw_table(conn, df)
        for cursor in conn.execute_string(PUBLISH_SQL):
            _ = cursor.fetchall() if cursor.description else None
        conn.commit()
    finally:
        conn.close()

    print(f"Loaded Bitdefender into {TARGET_TABLE} from royalty workbooks in {format_seconds(time.perf_counter() - start)}.", flush=True)


if __name__ == "__main__":
    main()


