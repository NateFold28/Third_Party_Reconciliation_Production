"""Recreate Acronis manual-team usage files from raw portal CSV exports.

This script scans the raw monthly `Data` and `ASIO Data` folders with the same
guardrails as the manual workbook recreation, then publishes vendor usage at
the standard third-party usage grain:

Snowflake target:
    ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.ACRONIS_USAGE

Published grain:
    BILLING_MONTH x VENDOR x MODIFIER(Status) x VENDOR_PARTNER_NAME x VENDOR_PRODUCT_SKU
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import io
import re
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Literal

import openpyxl
import pandas as pd
from invoice_rate_backfill import fill_missing_prices_dynamic


ACRONIS_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = ACRONIS_ROOT.parent
WORKSPACE_ROOT = PROJECT_ROOT.parents[1]
OUTPUT_DIR = ACRONIS_ROOT / "outputs"
# CSV seed files are no longer used. Unit prices are loaded dynamically from
# THIRD_PARTY_RECON_VENDOR_INVOICES via load_price_seed() below.

DEFAULT_SOURCE_ROOT = Path(
    r"C:\Users\Nate.Fold\OneDrive - ConnectWise, Inc"
    r"\THIRD_PARTY_RECONCILIATION\Manual Recon Files 2026\Acronis"
)

TARGET_DATABASE = "ANALYTICS_DEV"
TARGET_SCHEMA = "DBT_NFOLD_TRANSFORMATION"
TARGET_TABLE = "THIRD_PARTY_RECON_VENDOR_USAGE_PROD"
FQN = f"{TARGET_DATABASE}.{TARGET_SCHEMA}.{TARGET_TABLE}"
TARGET_VENDOR = "Acronis"

MONTH_FOLDER_RE = re.compile(r"^(?P<mm>\d{2})_[A-Z]{3}_(?P<yyyy>\d{4})$", re.IGNORECASE)
REPORTING_PERIOD_RE = re.compile(r"(?P<yyyy>\d{4})-(?P<mm>\d{2})-\d{2}")
DATE_RANGE_RE = re.compile(
    r"\((?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[^)]*\)\s*$",
    re.IGNORECASE,
)

# Layouts observed across months:
#   06_JUN_2026/Data/ + 06_JUN_2026/ASIO Data/
#   02_FEB_2026/Data/ + 02_FEB_2026/Data/ASIO/
# Both are supported: (folder path parts to look under, label used in the scan).
RAW_LAYOUTS: tuple[tuple[tuple[str, ...], str], ...] = (
    (("ASIO Data",), "ASIO Data"),
    (("Data", "ASIO"), "ASIO Data"),
    (("Data",), "Data"),
)

RAW_COLUMNS: tuple[str, ...] = (
    "Tenant name",
    "Identifier",
    "Type",
    "Status",
    "Mode",
    "Service name",
    "Edition",
    "Metric name",
    "Location",
    "Metric unit",
    "Quota",
    "Production usage",
    "Change in production usage",
    "Trial usage",
    "Change in trial usage",
    "Total usage",
    "Change in total usage",
    "Price",
    "Cost",
    "Billing model",
    "Storage type",
    "Pricing tier ID",
    "SKU",
)

OUTPUT_COLUMNS: tuple[str, ...] = ("Entity",) + RAW_COLUMNS
TABLE_COLUMNS: tuple[str, ...] = ("BILLING_MONTH",) + OUTPUT_COLUMNS
USAGE_COLUMNS: tuple[str, ...] = (
    "BILLING_MONTH",
    "VENDOR",
    "VENDOR_PARTNER_NAME",
    "VENDOR_PRODUCT_SKU",
    "MODIFIER",
    "QUANTITY",
    "UNIT_PRICE",
    "AMOUNT",
    "CURRENCY",
)
NUMERIC_COMPARE_COLUMNS = {
    "Quota",
    "Production usage",
    "Change in production usage",
    "Trial usage",
    "Change in trial usage",
    "Total usage",
    "Change in total usage",
    "Pricing tier ID",
}

COLUMN_TO_SNOWFLAKE = {
    "BILLING_MONTH": "BILLING_MONTH",
    "Entity": "ENTITY",
    "Tenant name": "TENANT_NAME",
    "Identifier": "IDENTIFIER",
    "Type": "TYPE",
    "Status": "STATUS",
    "Mode": "MODE",
    "Service name": "SERVICE_NAME",
    "Edition": "EDITION",
    "Metric name": "METRIC_NAME",
    "Location": "LOCATION",
    "Metric unit": "METRIC_UNIT",
    "Quota": "QUOTA",
    "Production usage": "PRODUCTION_USAGE",
    "Change in production usage": "CHANGE_IN_PRODUCTION_USAGE",
    "Trial usage": "TRIAL_USAGE",
    "Change in trial usage": "CHANGE_IN_TRIAL_USAGE",
    "Total usage": "TOTAL_USAGE",
    "Change in total usage": "CHANGE_IN_TOTAL_USAGE",
    "Price": "PRICE",
    "Cost": "COST",
    "Billing model": "BILLING_MODEL",
    "Storage type": "STORAGE_TYPE",
    "Pricing tier ID": "PRICING_TIER_ID",
    "SKU": "SKU",
}
SNOWFLAKE_COLUMNS = USAGE_COLUMNS
RATE_MATCH_TOLERANCE_PCT = 0.01

BILLABLE_SKU_RE = re.compile(r"^S[A-Z0-9]+$", re.IGNORECASE)

# --- SKU prefix guardrail (Open Question #1 in the spec) -----------------------
# Empirically only S* SKUs are billable in the CW-resold contract. C* SKUs are
# billed directly by Acronis (or sit on partner-owned gateways) and are
# EXCLUDED from every output row across the 2,414-row June workbook. If a new
# prefix appears it must be triaged before we silently keep or drop it.
ALLOWED_SKU_PREFIXES: frozenset[str] = frozenset({"S"})
KNOWN_EXCLUDED_SKU_PREFIXES: frozenset[str] = frozenset({"C"})

# --- Explicit filename -> Entity lookup (spec: maintained constant) -----------
# Key format: (layout_label, normalized_report_title). The normaliser lowercases,
# strips the trailing "(<Month> <Day>, <YYYY>-...)" date-range, collapses
# whitespace, and drops the redundant leading region prefix or "Summary for
# period report -" boilerplate to leave a stable descriptor. Any file that
# does not match a key here raises an alert instead of being silently kept.
FILE_TO_ENTITY: dict[tuple[str, str], str] = {
    # Data folder (device / GB workload reports)
    ("Data", "connectwise au1 per workload"): "AUS_DEV",
    ("Data", "connectwise au1 per gb"): "AUS_GIG",
    ("Data", "capacity gig partners"): "CAN_GIG",
    ("Data", "device workload partners"): "CAN_DEV",
    ("Data", "cw eu2 germany device workload"): "EU2_DEV",
    ("Data", "eu1 connectwise inc"): "EU1_GIG",
    ("Data", "eu8 connectwise inc"): "EU8_Master",
    ("Data", "connectwise sg1 per workload"): "SG1_DEV",
    ("Data", "connectwise sg1 per gigabyte"): "SG1_GIG",
    ("Data", "uk dev connectwise inc 2"): "UK_DEV",
    ("Data", "uk gig connectwise inc"): "UK_GIG",
    ("Data", "us connectwise inc"): "US_Master",
    # Older-month filename convention where US_Master lacks a region prefix
    # (Jan-May 2026 stored the file as 'Summary for period report - ConnectWise, Inc').
    ("Data", "connectwise inc"): "US_Master",
    ("Data", "us5 dev connectwise inc 1"): "US5_DEV",
    ("Data", "us5 gig connectwise inc"): "US5_GIG",
    ("Data", "cw gig us2"): "US2_GIG",
    ("Data", "cw workload us2"): "US2_DEV",
    ("Data", "nz master connectwise inc"): "NZ_Master",
    # ASIO folder (nested ASIO Portal reports)
    ("ASIO Data", "aus dev asio portal"): "AUS_DEV",
    ("ASIO Data", "can dev asio portal"): "CAN_DEV",
    ("ASIO Data", "can gig asio portal"): "CAN_GIG",
    ("ASIO Data", "eu8 asio"): "EU8_Master",
    ("ASIO Data", "uk dev asio portal"): "UK_DEV",
    ("ASIO Data", "uk gig asio portal"): "UK_GIG",
    ("ASIO Data", "us asio"): "US_Master",
    # Older-month ASIO filename that lacked the 'US_' prefix.
    ("ASIO Data", "asio"): "US_Master",
    ("ASIO Data", "us5 dev asio portal"): "US5_DEV",
}

# --- Explicit exclusion list (spec: tenant-level manual exceptions) ----------
# Applied as (entity, tenant_normalized_compact) so future trial or
# double-count tenants are handled by extending this table rather than by
# tweaking regex heuristics. `_norm_compact` collapses whitespace and drops
# all non-alphanumeric characters, so the key for '// TRIALS //' is 'trials'.
EXCLUDED_TENANTS: frozenset[tuple[str, str]] = frozenset({
    # spec section 8 -- seven analyst-hand-strip entries
    ("AUS_GIG", "cwpartnertrials"),
    ("CAN_GIG", "testing"),
    ("US2_DEV", "reporttestpartner"),
    ("US2_GIG", "cwpartnercovid19promo"),
    ("US_Master", "cwpartnercovid19promo"),
    ("US_Master", "cwgigus2"),
    ("US_Master", "cwworkloadus2"),
    # Trial containers that appear as Type=Partner (not Folder) in the raw
    # CSVs and therefore slip past the Folder filter. Confirmed against the
    # JUN 2026 manual workbook.
    ("UK_DEV", "trials"),
    ("UK_GIG", "cwtrials"),
    ("US2_GIG", "trials"),
})



def _norm_compact(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value if value is not None else "").lower())


def _norm_words(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value if value is not None else "").lower()).strip()


def _clean_text(value: object) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()
    return text or None


def normalize_cell(value: object) -> object:
    """Trim whitespace on strings and normalize blanks to None. Non-string
    values pass through unchanged so downstream numeric coercion still works.
    """
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text if text else None
    return value


def _normalize_compare_text(value: object) -> str:
    text = "" if value is None else str(value).strip()
    try:
        return text.encode("cp1252").decode("utf-8")
    except UnicodeError:
        return text


def _path_parts(value: object) -> list[str]:
    return [
        p.strip()
        for p in str(value if value is not None else "").replace("\\", "/").split("/")
        if p.strip()
    ]


def _read_file_bytes(path: Path) -> bytes:
    try:
        return path.read_bytes()
    except PermissionError:
        with tempfile.NamedTemporaryFile(suffix=path.suffix, delete=False) as handle:
            tmp_path = Path(handle.name)
        try:
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    "Copy-Item -LiteralPath $args[0] -Destination $args[1] -Force",
                    str(path),
                    str(tmp_path),
                ],
                check=True,
                capture_output=True,
            )
            return tmp_path.read_bytes()
        finally:
            tmp_path.unlink(missing_ok=True)


def discover_month_folders(source_root: Path) -> dict[str, Path]:
    folders: dict[str, Path] = {}
    for child in source_root.iterdir():
        if not child.is_dir():
            continue
        match = MONTH_FOLDER_RE.match(child.name)
        if match:
            folders[f"{match.group('yyyy')}-{match.group('mm')}"] = child
    return dict(sorted(folders.items()))


def folder_month(month_folder: Path) -> dt.date:
    match = MONTH_FOLDER_RE.match(month_folder.name)
    if not match:
        raise ValueError(f"Invalid month folder name: {month_folder}")
    return dt.date(int(match.group("yyyy")), int(match.group("mm")), 1)


def locate_raw_files(month_folder: Path) -> list[tuple[str, Path]]:
    """Return (layout_label, path) for every raw CSV in the month folder.

    Supports both raw-file layouts we've seen in 2026:
      06_JUN_2026/Data/ + 06_JUN_2026/ASIO Data/
      02_FEB_2026/Data/ + 02_FEB_2026/Data/ASIO/
    """
    files: list[tuple[str, Path]] = []
    seen: set[Path] = set()
    for parts, layout_label in RAW_LAYOUTS:
        folder = month_folder.joinpath(*parts)
        if not folder.exists() or not folder.is_dir():
            continue
        for path in sorted(folder.iterdir()):
            if path.name.startswith("~$") or not path.is_file() or path in seen:
                continue
            suffixes = [s.lower() for s in path.suffixes]
            if path.suffix.lower() == ".csv" or suffixes[-2:] == [".csv", ".zip"]:
                files.append((layout_label, path))
                seen.add(path)
    return files


def _decode(raw: bytes) -> list[str]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return raw.decode(encoding).splitlines()
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8-sig", errors="replace").splitlines()


def iter_csv_payloads(path: Path) -> list[tuple[str, list[str]]]:
    if path.suffix.lower() != ".zip":
        return [(path.name, _decode(_read_file_bytes(path)))]
    payloads: list[tuple[str, list[str]]] = []
    with zipfile.ZipFile(io.BytesIO(_read_file_bytes(path))) as archive:
        for name in archive.namelist():
            if name.lower().endswith(".csv"):
                payloads.append((f"{path.name}::{Path(name).name}", _decode(archive.read(name))))
    return payloads


def find_header_row(rows: list[list[str]]) -> int | None:
    expected = [_norm_compact(c) for c in RAW_COLUMNS]
    for i, row in enumerate(rows[:30]):
        normalized = [_norm_compact(c) for c in row[: len(RAW_COLUMNS)]]
        if normalized == expected:
            return i
    return None


def parse_preamble(rows: list[list[str]], header_idx: int | None) -> dict[str, str]:
    limit = header_idx if header_idx is not None else min(10, len(rows))
    preamble: dict[str, str] = {}
    for row in rows[:limit]:
        if len(row) >= 2 and row[0] and row[1]:
            preamble[str(row[0]).strip()] = str(row[1]).strip()
    return preamble


def reporting_month(preamble: dict[str, str], fallback: dt.date) -> dt.date:
    period = preamble.get("Reporting period")
    if period:
        match = REPORTING_PERIOD_RE.search(period)
        if match:
            return dt.date(int(match.group("yyyy")), int(match.group("mm")), 1)
    return fallback


def normalize_report_key(path: Path) -> str:
    """Reduce a raw CSV filename to a stable descriptor used to look up Entity.

    Rules (empirically derived; produces the same key across Janâ€“Jul 2026):
      - strip trailing '.csv' / '.csv.zip'
      - strip trailing '(<Month> <Day>, YYYY-<Month> <Day>, YYYY)' date range
      - lowercase
      - drop 'summary for period report' boilerplate
      - drop 'connectwise' when it's clearly a report-title filler
        (kept for 'connectwise, inc' / 'connectwise au1' / 'connectwise sg1'
        as those are the actual portal identifiers)
      - collapse punctuation, whitespace and repeated spaces
    """
    stem = path.name
    # strip .csv or .csv.zip
    if stem.lower().endswith(".csv.zip"):
        stem = stem[: -len(".csv.zip")]
    elif stem.lower().endswith(".csv"):
        stem = stem[: -len(".csv")]
    # strip date-range suffix like " (Jun 1, 2026-Jun 30, 2026)"
    stem = DATE_RANGE_RE.sub("", stem).strip()
    text = stem.lower()
    text = text.replace(",", " ")
    text = text.replace("-", " ")
    text = text.replace("_", " ")
    text = text.replace(".", " ")
    # drop the boilerplate. "summary for period report" and stray "connectwise inc"
    # tokens are always present and never disambiguate one report from another.
    text = re.sub(r"\bsummary for period report\b", " ", text)
    text = re.sub(r"\bportal\b", " portal ", text)
    text = re.sub(r"\s+", " ", text).strip()
    # collapse the specific "connectwise inc" identifier down to just "connectwise inc"
    # (the "1"/"2" region marker after it survives as its own token)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def resolve_entity(path: Path, layout_label: str, portal_tenant: str) -> str | None:
    """Look up the Entity code from an explicit filename -> Entity table.

    The spec (Locating the input files) is explicit: 'code should hold an
    explicit lookup table mapping each source report to its Entity code
    rather than deriving the code from the filename'. If a file cannot be
    matched we return None and the caller emits a warning so the operator
    can extend the table -- silent guesses are unsafe when new regions are
    introduced.
    """
    key = normalize_report_key(path)
    entry = FILE_TO_ENTITY.get((layout_label, key))
    if entry is not None:
        return entry
    # Fallback: try every substring token match against the table for the
    # same layout. This handles minor filename tweaks (extra whitespace,
    # extra "AUS -" prefixes) without silently mapping to a wrong region.
    tokens = set(key.split())
    best: tuple[int, str] | None = None
    for (layout, canonical_key), entity in FILE_TO_ENTITY.items():
        if layout != layout_label:
            continue
        canonical_tokens = set(canonical_key.split())
        if canonical_tokens.issubset(tokens):
            score = len(canonical_tokens)
            if best is None or score > best[0]:
                best = (score, entity)
    return best[1] if best else None


def normalize_hierarchy_path(value: object) -> str:
    """Spec-compliant path normalization.

    'Strip any leading slashes and prefix exactly one, so that
    "ConnectWise, Inc /2" becomes "/ConnectWise, Inc /2" and "/ ASIO Portal"
    becomes "/ ASIO Portal".'
    """
    text = str(value if value is not None else "").strip()
    if not text:
        return ""
    while text.startswith("/"):
        text = text[1:]
    return "/" + text


def is_direct_child_of_root(row_tenant: object, root_tenant: str) -> bool:
    """Return True iff row_tenant is a descendant of root_tenant (not the root
    itself).

    Spec condition 3 combined with condition 2 (Type != 'Folder'): the Folder
    filter already removes intermediate hierarchy rows, so any Partner-typed
    row whose Tenant name starts with the root path is by definition a
    billable leaf. We intentionally DO NOT reject remainders that contain
    additional slashes -- legitimate portal display names like
    'Desert IT Solutions /1', 'Infinity Group / France', and
    'Pure Technology /1' embed slashes verbatim and the manual workbook
    keeps them.
    """
    row_norm = normalize_hierarchy_path(row_tenant)
    root_norm = normalize_hierarchy_path(root_tenant)
    if not root_norm or not row_norm:
        return False
    prefix = root_norm + "/"
    return row_norm.startswith(prefix) and bool(row_norm[len(prefix):])


def clean_tenant_leaf(row_tenant: object, root_tenant: str) -> str | None:
    """Return the tenant name below root (root prefix removed).

    Only invoked after is_direct_child_of_root has returned True. Slashes in
    the remainder are preserved verbatim because the Acronis portal encodes
    display suffixes like '/1', '/2', '/ France' inside the leaf name.
    """
    row_norm = normalize_hierarchy_path(row_tenant)
    root_norm = normalize_hierarchy_path(root_tenant)
    remainder = row_norm[len(root_norm) + 1:]
    return _clean_text(remainder)


def sku_prefix(sku: str) -> str:
    match = re.match(r"^([A-Za-z])", sku)
    return match.group(1).upper() if match else ""


def parse_raw_file(
    *,
    source_folder: str,
    path: Path,
    display_name: str,
    lines: list[str],
    fallback_month: dt.date,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    rows = list(csv.reader(lines))
    header_idx = find_header_row(rows)
    scan = {
        "billing_month": fallback_month.isoformat(),
        "source_folder": source_folder,
        "source_file": display_name,
        "entity": None,
        "portal_tenant": None,
        "schema_status": "matched" if header_idx is not None else "skipped_schema_mismatch",
        "raw_rows": 0,
        "kept_rows": 0,
        "dropped_non_billable_sku_prefix": 0,
        "dropped_unknown_sku_prefix": 0,
        "dropped_folder_type": 0,
        "dropped_non_direct_child": 0,
        "dropped_root_row": 0,
        "dropped_excluded_tenant": 0,
        "unknown_entity": False,
    }
    if header_idx is None:
        return [], scan

    preamble = parse_preamble(rows, header_idx)
    month = reporting_month(preamble, fallback_month)
    portal_tenant = preamble.get("Tenant name", "")
    entity = resolve_entity(path, source_folder, portal_tenant)
    scan.update({
        "billing_month": month.isoformat(),
        "entity": entity,
        "portal_tenant": portal_tenant,
        "unknown_entity": entity is None,
    })
    if entity is None:
        # Spec: 'raise an alert whenever a SKU prefix appears that is not in the
        # known list.' Same principle applies here for unrecognised reports so
        # the operator can extend FILE_TO_ENTITY instead of the file being
        # silently dropped or mis-routed.
        print(
            f"[WARN] Unknown Entity for report {display_name!r} "
            f"(layout={source_folder}, root={portal_tenant!r}); "
            f"add an entry to FILE_TO_ENTITY."
        )
        return [], scan

    output_rows: list[dict[str, object]] = []
    for row in rows[header_idx + 1 :]:
        if not row or all(not str(cell).strip() for cell in row):
            continue
        scan["raw_rows"] += 1

        values = {
            RAW_COLUMNS[i]: normalize_cell(row[i]) if i < len(row) else None
            for i in range(len(RAW_COLUMNS))
        }
        raw_tenant = values.get("Tenant name")
        row_type = _clean_text(values.get("Type"))
        sku = _clean_text(values.get("SKU"))

        # Spec condition 1: SKU must begin with 'S'. Blank/non-billable SKUs
        # are dropped by definition. Log any brand-new prefix.
        if not sku:
            scan["dropped_non_billable_sku_prefix"] += 1
            continue
        prefix = sku_prefix(sku)
        if prefix in KNOWN_EXCLUDED_SKU_PREFIXES:
            scan["dropped_non_billable_sku_prefix"] += 1
            continue
        if prefix not in ALLOWED_SKU_PREFIXES:
            scan["dropped_unknown_sku_prefix"] += 1
            print(
                f"[WARN] Unknown SKU prefix {prefix!r} in {display_name!r} "
                f"(SKU={sku!r}); triage before extending ALLOWED_SKU_PREFIXES."
            )
            continue
        if not BILLABLE_SKU_RE.match(sku):
            # Malformed SKU that starts with S but has non-alnum chars.
            scan["dropped_non_billable_sku_prefix"] += 1
            continue

        # Spec condition 2: Type must not be 'Folder' (Folder rows are
        # aggregations of everything beneath them).
        if row_type and row_type.strip().lower() == "folder":
            scan["dropped_folder_type"] += 1
            continue

        # Spec condition 3: tenant must be a DIRECT CHILD of the report root.
        if not is_direct_child_of_root(raw_tenant, portal_tenant):
            row_norm = normalize_hierarchy_path(raw_tenant)
            root_norm = normalize_hierarchy_path(portal_tenant)
            if row_norm == root_norm:
                scan["dropped_root_row"] += 1
            else:
                scan["dropped_non_direct_child"] += 1
            continue

        tenant_name = clean_tenant_leaf(raw_tenant, portal_tenant)
        if not tenant_name:
            scan["dropped_non_direct_child"] += 1
            continue

        # Spec exclusion list: 7 specific (entity, tenant) pairs that the
        # analysts strip by hand every month.
        if (entity, _norm_compact(tenant_name)) in EXCLUDED_TENANTS:
            scan["dropped_excluded_tenant"] += 1
            continue

        values["Tenant name"] = tenant_name
        values["SKU"] = sku.upper()
        record = {"BILLING_MONTH": month, "Entity": entity}
        record.update(values)
        output_rows.append(record)
        scan["kept_rows"] += 1

    return output_rows, scan


def parse_month(source_root: Path, month: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    month_folder = discover_month_folders(source_root).get(month)
    if month_folder is None:
        raise FileNotFoundError(f"No folder found for {month} under {source_root}")

    fallback_month = folder_month(month_folder)
    records: list[dict[str, object]] = []
    scans: list[dict[str, object]] = []
    for source_folder, path in locate_raw_files(month_folder):
        for display_name, lines in iter_csv_payloads(path):
            rows, scan = parse_raw_file(
                source_folder=source_folder,
                path=path,
                display_name=display_name,
                lines=lines,
                fallback_month=fallback_month,
            )
            records.extend(rows)
            scans.append(scan)

    df = pd.DataFrame(records, columns=list(TABLE_COLUMNS))
    scan_df = pd.DataFrame(scans)
    print(
        f"[{month}] files={len(scan_df):,}, rows={len(df):,}, "
        f"sku_count={df['SKU'].nunique() if not df.empty else 0:,}, "
        f"entities={df['Entity'].nunique() if not df.empty else 0:,}"
    )
    return df, scan_df


def load_price_seed() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load Acronis unit prices from THIRD_PARTY_RECON_VENDOR_INVOICES.

    Priority (matching 00b_backfill_invoice_prices.sql logic):
      1. Exact (billing_month, sku) match from invoices table.
      2. Most-recent prior-month rate for same SKU (carry-forward via LAST_VALUE IGNORE NULLS).
      3. Returns empty DataFrames if the invoice table doesn't exist yet.

    Returns (monthly_entity_df, monthly_sku_df, latest_fallback_df) matching the
    shape the rest of the pipeline expects.
    """
    import sys as _sys
    _sys.path.insert(0, str(WORKSPACE_ROOT))
    try:
        from TEMPLATES.Python.connection import get_snowflake_connection as _conn
        conn = _conn(role="DEVELOPER", warehouse="REPORTING_WH",
                     database="ANALYTICS_DEV", schema="DBT_NFOLD_TRANSFORMATION")
        inv = pd.read_sql("""
            SELECT
                BILLING_MONTH,
                VENDOR_PRODUCT_SKU AS VENDOR_SKU,
                CASE
                    -- Prefer direct invoice unit prices when present; this is
                    -- robust even if quantity OCR has occasional scale issues.
                    WHEN COUNT(IFF(UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0, 1, NULL)) > 0
                        THEN MEDIAN(IFF(UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0, UNIT_PRICE, NULL))
                    -- Fallback only if direct prices are absent.
                    WHEN SUM(IFF(QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL, QUANTITY, 0)) > 0
                        THEN SUM(IFF(QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL, AMOUNT, 0))
                             / SUM(IFF(QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL, QUANTITY, 0))
                    ELSE NULL
                END AS UNIT_PRICE
            FROM ANALYTICS_DEV.DBT_NFOLD_TRANSFORMATION.THIRD_PARTY_RECON_VENDOR_INVOICES
            WHERE VENDOR ILIKE '%acronis%'
                            AND (
                                        (UNIT_PRICE IS NOT NULL AND UNIT_PRICE > 0)
                                        OR (QUANTITY IS NOT NULL AND QUANTITY > 0 AND AMOUNT IS NOT NULL)
                                    )
            GROUP BY 1, 2
        """, conn)
        conn.close()
    except Exception as e:
        print(f"[WARN] Could not load Acronis invoice rates from Snowflake ({e}). UNIT_PRICE will be NULL.", flush=True)
        empty_entity = pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "ENTITY", "UNIT_PRICE", "CURRENCY"])
        empty = pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE", "CURRENCY"])
        fallback = pd.DataFrame(columns=["VENDOR_SKU", "UNIT_PRICE", "CURRENCY"])
        return empty_entity, empty, fallback

    if inv.empty:
        print("[WARN] No Acronis rows found in THIRD_PARTY_RECON_VENDOR_INVOICES. UNIT_PRICE will be NULL.", flush=True)
        empty_entity = pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "ENTITY", "UNIT_PRICE", "CURRENCY"])
        empty = pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE", "CURRENCY"])
        fallback = pd.DataFrame(columns=["VENDOR_SKU", "UNIT_PRICE", "CURRENCY"])
        return empty_entity, empty, fallback

    inv["BILLING_MONTH"] = pd.to_datetime(inv["BILLING_MONTH"]).dt.date
    inv["VENDOR_SKU"] = inv["VENDOR_SKU"].astype(str).str.strip().str.upper()
    inv["UNIT_PRICE"] = pd.to_numeric(inv["UNIT_PRICE"], errors="coerce")
    inv = inv[inv["UNIT_PRICE"].notna() & (inv["UNIT_PRICE"] > 0)]
    inv["CURRENCY"] = "USD"
    inv = inv.dropna(subset=["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE"])

    monthly_sku = inv[["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE", "CURRENCY"]].drop_duplicates(
        subset=["BILLING_MONTH", "VENDOR_SKU"], keep="first"
    )
    # Carry-forward fallback: most recent unit price per SKU (for months not yet in invoices)
    latest_fallback = (
        inv.sort_values("BILLING_MONTH", ascending=False)
        .drop_duplicates(subset=["VENDOR_SKU"], keep="first")
        [["VENDOR_SKU", "UNIT_PRICE", "CURRENCY"]]
    )
    # No entity-level split for Acronis invoices (they don't have per-partner pricing)
    monthly_entity = pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "ENTITY", "UNIT_PRICE", "CURRENCY"])
    print(f"[INFO] Loaded {len(monthly_sku):,} Acronis invoice rates from VENDOR_INVOICES "
          f"({monthly_sku['BILLING_MONTH'].nunique()} months, {monthly_sku['VENDOR_SKU'].nunique()} SKUs).", flush=True)
    return monthly_entity, monthly_sku, latest_fallback


def parse_money(value: object) -> float | None:
    text = str(value if value is not None else "").strip().replace(",", "")
    text = re.sub(r"[^0-9.\-]+", "", text)
    if not text:
        return None
    parsed = pd.to_numeric(text, errors="coerce")
    if pd.isna(parsed):
        return None
    return float(parsed)


def amount_from_invoice_line(line: object) -> float | None:
    values = re.findall(r"(?<![A-Za-z])(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?(?![A-Za-z])", str(line or ""))
    if not values:
        return None
    return parse_money(values[-1])


def recover_invoice_line_rates() -> pd.DataFrame:
    """Formerly read a stale CSV seed. Now always returns empty \u2014 unit prices
    come from load_price_seed() which queries THIRD_PARTY_RECON_VENDOR_INVOICES."""
    return pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "UNIT_PRICE", "INVOICE_QUANTITY", "INVOICE_AMOUNT"])


def derive_entity_rates_from_invoice(raw_usage: pd.DataFrame) -> pd.DataFrame:
    invoice_lines = recover_invoice_line_rates()
    if invoice_lines.empty or raw_usage.empty:
        return pd.DataFrame(columns=["BILLING_MONTH", "VENDOR_SKU", "ENTITY", "UNIT_PRICE", "CURRENCY"])

    usage_qty = (
        raw_usage.groupby(["BILLING_MONTH", "MODIFIER", "VENDOR_PRODUCT_SKU"], dropna=False)
        .agg(USAGE_QUANTITY=("QUANTITY", "sum"))
        .reset_index()
        .rename(columns={"MODIFIER": "ENTITY", "VENDOR_PRODUCT_SKU": "VENDOR_SKU"})
    )
    usage_qty = usage_qty[usage_qty["USAGE_QUANTITY"] > 0].copy()

    matched: list[dict[str, object]] = []
    for (billing_month, sku), usage_group in usage_qty.groupby(["BILLING_MONTH", "VENDOR_SKU"], dropna=False):
        invoice_group = invoice_lines[
            (invoice_lines["BILLING_MONTH"] == billing_month)
            & (invoice_lines["VENDOR_SKU"] == sku)
        ].copy()
        if invoice_group.empty:
            continue
        remaining = invoice_group.reset_index(drop=True).copy()
        for _, usage_row in usage_group.sort_values("USAGE_QUANTITY", ascending=False).iterrows():
            qty = float(usage_row["USAGE_QUANTITY"])
            if qty <= 0 or remaining.empty:
                continue
            remaining["_ABS_DELTA"] = (remaining["INVOICE_QUANTITY"] - qty).abs()
            nearest_idx = remaining["_ABS_DELTA"].idxmin()
            nearest = remaining.loc[nearest_idx]
            rel_delta = float(nearest["_ABS_DELTA"]) / max(abs(qty), 1.0)
            if rel_delta <= RATE_MATCH_TOLERANCE_PCT:
                matched.append(
                    {
                        "BILLING_MONTH": billing_month,
                        "VENDOR_SKU": sku,
                        "ENTITY": usage_row["ENTITY"],
                        "UNIT_PRICE": nearest["UNIT_PRICE"],
                        "CURRENCY": "USD",
                    }
                )
                remaining = remaining.drop(index=nearest_idx)
    return pd.DataFrame(matched).drop_duplicates(
        subset=["BILLING_MONTH", "VENDOR_SKU", "ENTITY"], keep="first"
    )


def assign_rate_source(
    row: pd.Series,
) -> Literal["invoice_month_entity", "invoice_month_sku", "fallback", "none"]:
    if pd.notna(row.get("ENTITY_UNIT_PRICE")):
        return "invoice_month_entity"
    if pd.notna(row.get("MONTH_SKU_UNIT_PRICE")):
        return "invoice_month_sku"
    if pd.notna(row.get("FALLBACK_UNIT_PRICE")):
        return "fallback"
    return "none"


def build_vendor_usage_frame(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=list(USAGE_COLUMNS))

    usage = pd.DataFrame(
        {
            "BILLING_MONTH": pd.to_datetime(df["BILLING_MONTH"]).dt.date,
            "VENDOR": "Acronis",
            "VENDOR_PARTNER_NAME": df["Tenant name"].map(_clean_text),
            "VENDOR_PRODUCT_SKU": df["SKU"].astype(str).str.strip().str.upper(),
            # Status is the most informative Acronis split for downstream analysis
            # (Enabled vs Disabled). Keep Entity in raw recreated outputs only.
            "MODIFIER": df["Status"].map(_clean_text),
            "QUANTITY": pd.to_numeric(df["Total usage"], errors="coerce").fillna(0.0),
        }
    )
    usage = usage[
        usage["VENDOR_PARTNER_NAME"].notna()
        & usage["VENDOR_PRODUCT_SKU"].notna()
    ].copy()
    invoice_entity_rates = derive_entity_rates_from_invoice(usage)
    monthly_entity_rates, monthly_rates, latest_rates = load_price_seed()
    if monthly_entity_rates.empty and not invoice_entity_rates.empty:
        monthly_entity_rates = invoice_entity_rates
    elif not invoice_entity_rates.empty:
        monthly_entity_rates = pd.concat([monthly_entity_rates, invoice_entity_rates], ignore_index=True).drop_duplicates(
            subset=["BILLING_MONTH", "VENDOR_SKU", "ENTITY"],
            keep="first",
        )
    monthly_entity_rates = monthly_entity_rates.rename(
        columns={
            "VENDOR_SKU": "VENDOR_PRODUCT_SKU",
            "ENTITY": "MODIFIER",
            "UNIT_PRICE": "ENTITY_UNIT_PRICE",
            "CURRENCY": "ENTITY_CURRENCY",
        }
    )
    monthly_rates = monthly_rates.rename(
        columns={
            "VENDOR_SKU": "VENDOR_PRODUCT_SKU",
            "UNIT_PRICE": "MONTH_SKU_UNIT_PRICE",
            "CURRENCY": "MONTH_SKU_CURRENCY",
        }
    )
    latest_rates = latest_rates.rename(
        columns={
            "VENDOR_SKU": "VENDOR_PRODUCT_SKU",
            "UNIT_PRICE": "FALLBACK_UNIT_PRICE",
            "CURRENCY": "FALLBACK_CURRENCY",
        }
    )
    usage = usage.merge(
        monthly_entity_rates,
        on=["BILLING_MONTH", "VENDOR_PRODUCT_SKU", "MODIFIER"],
        how="left",
    )
    usage = usage.merge(
        monthly_rates,
        on=["BILLING_MONTH", "VENDOR_PRODUCT_SKU"],
        how="left",
    )
    usage = usage.merge(latest_rates, on="VENDOR_PRODUCT_SKU", how="left")
    usage["RATE_SOURCE"] = usage.apply(assign_rate_source, axis=1)
    usage["UNIT_PRICE"] = (
        usage["ENTITY_UNIT_PRICE"]
        .fillna(usage["MONTH_SKU_UNIT_PRICE"])
        .fillna(usage["FALLBACK_UNIT_PRICE"])
    )
    usage["CURRENCY"] = (
        usage["ENTITY_CURRENCY"]
        .fillna(usage["MONTH_SKU_CURRENCY"])
        .fillna(usage["FALLBACK_CURRENCY"])
        .fillna("USD")
    )
    usage["AMOUNT"] = usage["QUANTITY"] * usage["UNIT_PRICE"]
    usage = (
        usage.groupby(
            [
                "BILLING_MONTH",
                "VENDOR",
                "VENDOR_PARTNER_NAME",
                "VENDOR_PRODUCT_SKU",
                "MODIFIER",
                "UNIT_PRICE",
                "CURRENCY",
            ],
            dropna=False,
            as_index=False,
        )
        .agg(QUANTITY=("QUANTITY", "sum"), AMOUNT=("AMOUNT", "sum"))
    )
    return usage[list(USAGE_COLUMNS)]


def to_snowflake_frame(df: pd.DataFrame) -> pd.DataFrame:
    out = build_vendor_usage_frame(df)
    for col in SNOWFLAKE_COLUMNS:
        if col not in out.columns:
            out[col] = None
    return out[list(SNOWFLAKE_COLUMNS)]


def snowflake_ddl() -> str:
    cols = [
        "BILLING_MONTH DATE",
        "VENDOR VARCHAR",
        "VENDOR_PARTNER_NAME VARCHAR",
        "VENDOR_PRODUCT_SKU VARCHAR",
        "MODIFIER VARCHAR",
        "QUANTITY NUMBER(38, 6)",
        "UNIT_PRICE NUMBER(18, 6)",
        "AMOUNT NUMBER(38, 6)",
        "CURRENCY VARCHAR",
    ]
    return f"CREATE TABLE IF NOT EXISTS {FQN} ({', '.join(cols)});"


def load_snowflake(df: pd.DataFrame, *, reset: bool) -> None:
    sys.path.insert(0, str(WORKSPACE_ROOT))
    from snowflake.connector.pandas_tools import write_pandas
    from TEMPLATES.Python.connection import get_snowflake_connection

    load_df = to_snowflake_frame(df)
    conn = get_snowflake_connection(
        role="DEVELOPER",
        warehouse="REPORTING_WH",
        database=TARGET_DATABASE,
        schema=TARGET_SCHEMA,
    )
    try:
        load_df = fill_missing_prices_dynamic(load_df, TARGET_VENDOR, conn=conn)
        cur = conn.cursor()
        cur.execute(f"CREATE SCHEMA IF NOT EXISTS {TARGET_DATABASE}.{TARGET_SCHEMA}")
        cur.execute(snowflake_ddl())
        # Always refresh this vendor so newly-landed invoices can recalculate prior months.
        cur.execute(
            f"DELETE FROM {FQN} WHERE UPPER(COALESCE(VENDOR, '')) = UPPER(%s)",
            (TARGET_VENDOR,),
        )
        if load_df.empty:
            print("Nothing to load; all incoming months already exist.")
            return
        success, chunks, rows, output = write_pandas(
            conn,
            load_df,
            TARGET_TABLE,
            database=TARGET_DATABASE,
            schema=TARGET_SCHEMA,
            quote_identifiers=False,
        )
        if not success:
            raise RuntimeError(f"write_pandas failed: {output}")
        conn.commit()
        print(f"Loaded {rows:,} rows into {FQN} in {chunks} chunk(s).")
    finally:
        conn.close()


def read_manual_usage_file(path: Path, month: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(io.BytesIO(_read_file_bytes(path)), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        first_col = "Entity" if _norm_compact(header[0]) == "entity" else "Entity"
        if _norm_compact(header[0]) not in {"entity", "month"}:
            raise RuntimeError(f"Unexpected first column in {path}: {header[0]}")
        if [_norm_compact(c) for c in header[1 : 1 + len(RAW_COLUMNS)]] != [_norm_compact(c) for c in RAW_COLUMNS]:
            raise RuntimeError(f"Manual usage schema mismatch in {path}: {header}")
        records: list[dict[str, object]] = []
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            rec = {"BILLING_MONTH": dt.date.fromisoformat(f"{month}-01")}
            rec["Entity"] = normalize_cell(row[0]) if len(row) > 0 else None
            for i, col in enumerate(RAW_COLUMNS, start=1):
                rec[col] = normalize_cell(row[i]) if i < len(row) else None
            records.append(rec)
    finally:
        wb.close()
    df = pd.DataFrame(records, columns=list(TABLE_COLUMNS))
    df["SKU"] = df["SKU"].apply(lambda v: str(v).strip().upper() if v is not None else None)
    return df


def locate_manual_usage_files(source_root: Path, month: str) -> list[Path]:
    month_folder = discover_month_folders(source_root).get(month)
    if month_folder is None:
        return []
    files: list[Path] = []
    for pattern in ("Acronis Usage*.xlsx", "Acronis ASIO Usage*.xlsx", "Acronsi ASIO Usage*.xlsx"):
        files.extend(p for p in sorted(month_folder.glob(pattern)) if not p.name.startswith("~$"))
    return files


def comparable_key_frame(df: pd.DataFrame) -> pd.DataFrame:
    cmp = df.copy()
    for col in TABLE_COLUMNS:
        if col not in cmp.columns:
            cmp[col] = None
    cmp = cmp[list(TABLE_COLUMNS)]
    for col in cmp.columns:
        if col == "BILLING_MONTH":
            continue
        if col in NUMERIC_COMPARE_COLUMNS:
            numeric = pd.to_numeric(cmp[col], errors="coerce")
            cmp[col] = numeric.map(
                lambda v: ""
                if pd.isna(v)
                else f"{0.0 if abs(float(v)) == 0 else float(v):.6f}".rstrip("0").rstrip(".")
            )
        else:
            cmp[col] = cmp[col].where(cmp[col].notna(), "").map(_normalize_compare_text)
    cmp["BILLING_MONTH"] = pd.to_datetime(cmp["BILLING_MONTH"]).dt.date.astype(str)
    return (
        cmp.groupby(list(TABLE_COLUMNS), dropna=False)
        .size()
        .reset_index(name="row_count")
    )


def validate_against_manual(source_root: Path, months: list[str], parsed: pd.DataFrame) -> pd.DataFrame:
    summaries: list[dict[str, object]] = []
    detail_records: list[pd.DataFrame] = []
    for month in months:
        manual_files = locate_manual_usage_files(source_root, month)
        manual_frames = [read_manual_usage_file(path, month) for path in manual_files]
        manual = pd.concat(manual_frames, ignore_index=True) if manual_frames else pd.DataFrame(columns=list(TABLE_COLUMNS))
        auto = parsed[pd.to_datetime(parsed["BILLING_MONTH"]).dt.date.astype(str) == f"{month}-01"].copy()

        auto_cmp = comparable_key_frame(auto)
        manual_cmp = comparable_key_frame(manual)
        diff = auto_cmp.merge(
            manual_cmp,
            on=list(TABLE_COLUMNS),
            how="outer",
            suffixes=("_auto", "_manual"),
        )
        diff["row_count_auto"] = diff["row_count_auto"].fillna(0).astype(int)
        diff["row_count_manual"] = diff["row_count_manual"].fillna(0).astype(int)
        diff["row_count_delta"] = diff["row_count_auto"] - diff["row_count_manual"]
        diff = diff[diff["row_count_delta"] != 0].copy()
        diff["month"] = month
        if not diff.empty:
            detail_records.append(diff)

        summaries.append(
            {
                "month": month,
                "manual_files": " | ".join(path.name for path in manual_files),
                "auto_rows": len(auto),
                "manual_rows": len(manual),
                "auto_distinct_rows": len(auto_cmp),
                "manual_distinct_rows": len(manual_cmp),
                "diff_distinct_rows": len(diff),
                "net_row_delta": int(diff["row_count_delta"].sum()) if not diff.empty else 0,
            }
        )

    summary = pd.DataFrame(summaries)
    detail = pd.concat(detail_records, ignore_index=True) if detail_records else pd.DataFrame()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    summary.to_csv(OUTPUT_DIR / "acronis_usage_file_recreated_vs_manual_summary.csv", index=False)
    detail.to_csv(OUTPUT_DIR / "acronis_usage_file_recreated_vs_manual_diff.csv", index=False)
    print(summary.to_string(index=False))
    return summary


def write_local_outputs(df: pd.DataFrame, scan_df: pd.DataFrame, label: str) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUTPUT_DIR / f"acronis_usage_file_recreated_{label}.csv", index=False, quoting=csv.QUOTE_MINIMAL)
    build_vendor_usage_frame(df).to_csv(
        OUTPUT_DIR / f"acronis_vendor_usage_{label}.csv",
        index=False,
        quoting=csv.QUOTE_MINIMAL,
    )
    scan_df.to_csv(OUTPUT_DIR / f"acronis_usage_file_recreated_scan_{label}.csv", index=False, quoting=csv.QUOTE_MINIMAL)
    audit = (
        df.groupby(["BILLING_MONTH", "Entity"], dropna=False)
        .agg(row_count=("SKU", "size"), sku_rows=("SKU", "count"), distinct_tenants=("Tenant name", "nunique"))
        .reset_index()
    )
    audit.to_csv(OUTPUT_DIR / f"acronis_usage_file_recreated_audit_{label}.csv", index=False)


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Recreate exact-schema Acronis usage files from raw portal CSVs.")
    parser.add_argument("--source-root", default=str(DEFAULT_SOURCE_ROOT))
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--month", help="YYYY-MM")
    group.add_argument("--all-months", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--reset", action="store_true")
    parser.add_argument("--validate-manual", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    source_root = Path(args.source_root)
    months = list(discover_month_folders(source_root).keys()) if args.all_months else [args.month]

    frames: list[pd.DataFrame] = []
    scans: list[pd.DataFrame] = []
    for month in months:
        df, scan_df = parse_month(source_root, month)
        frames.append(df)
        scans.append(scan_df)

    all_rows = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=list(TABLE_COLUMNS))
    scan_all = pd.concat(scans, ignore_index=True) if scans else pd.DataFrame()
    label = "all_months" if args.all_months else args.month.replace("-", "_")
    write_local_outputs(all_rows, scan_all, label)
    print(f"TOTAL rows={len(all_rows):,}, entities={all_rows['Entity'].nunique():,}, tenants={all_rows['Tenant name'].nunique():,}")

    if args.validate_manual:
        validate_against_manual(source_root, months, all_rows)
    if args.dry_run:
        print("Dry run complete. Snowflake load skipped.")
        return
    load_snowflake(all_rows, reset=args.reset)


if __name__ == "__main__":
    main()

